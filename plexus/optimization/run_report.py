"""Durable, append-only reporting for one periodic optimization run.

The service deliberately owns no optimization orchestration.  Callers execute
the decision stages with ``persist=False`` and publish their returned packets at
well-defined milestones.  A single Task is the lifecycle authority; one Report
is its stable stakeholder location; ReportBlocks contain only compact pointers
to immutable JSON/XLSX revisions.
"""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, replace
from datetime import datetime, timezone
from hashlib import sha256
from io import BytesIO, StringIO
from typing import Any, Callable, Mapping, Optional
from urllib.parse import quote, urlencode, urlparse
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.writer.excel import ExcelWriter

from plexus.optimization.operator_identity import (
    OptimizationOperatorIdentity,
    optimization_operator_identity,
)
from plexus.optimization.decision import normalize_execution_candidate_policy
from plexus.optimization.report_actions import (
    build_action_projection,
    build_decision_summary,
    build_guideline_code_conflict_workstream,
)
from openpyxl.worksheet.table import Table, TableStyleInfo

from plexus.dashboard.api.models.report import Report
from plexus.dashboard.api.models.report_block import ReportBlock
from plexus.dashboard.api.models.task import Task
from plexus.dashboard.api.models.task_stage import TaskStage
from plexus.reports.service import _compact_output_json_for_storage, _get_programmatic_config_id
from plexus.storage.graphql_artifact_store import (
    ArtifactAuthorizationError,
    ArtifactIntegrityError,
    ArtifactTicketError,
    ArtifactTransferError,
    ArtifactTransferRequest,
    GraphQLArtifactStore,
)


LIFECYCLE_VERSION = "optimization-run-report-v1"
_BLOCK_SPECS = (
    (0, "Run Status", "OptimizationRunStatus"),
    (1, "Decision Evidence", "OptimizationDecisionEvidence"),
    (2, "Stakeholder Workbook", "OptimizationStakeholderWorkbook"),
)
SHEET_NAMES = (
    "Overview",
    "Portfolio",
    "Priorities",
    "Feedback Investment",
    "Questions and Issues",
    "Optimization Outcomes",
    "Run Log",
    "Definitions",
)
_STAGES = ("preflight", "ranking", "assessment", "diagnosis", "approval", "optimization", "review", "finalization")
_MILESTONE_STAGE = {
    # Milestones describe the evidence that was just made durable.  Point the
    # dashboard at the work that follows so the Report never appears stuck on
    # a phase whose result is already published.
    "started": "ranking",
    "ranking": "assessment",
    "assessment": "diagnosis",
    "diagnosis": "approval",
    "approval": "approval",
    "optimization": "optimization",
    "optimization_review": "review",
    "finalization": "finalization",
}
_RANKING_PROGRESS_SUPERSEDED_MILESTONES = frozenset(
    milestone for milestone in _MILESTONE_STAGE if milestone != "started"
)
_FINAL_STATES = {
    "complete",
    "completed",
    "complete_with_unresolved_actions",
    "completed_with_unresolved_actions",
    "incomplete",
    "blocked",
    "failed",
}

# These are presentation labels, deliberately separate from attachment names,
# logical IDs, and object keys.  Live progress is stakeholder-facing and must
# never reveal storage or resource identities.
_ARTIFACT_PUBLICATION_KINDS = (
    "decision_evidence",
    "stakeholder_workbook",
    "score_briefs",
    "scorecard_summaries",
    "scorecard_spreadsheets",
    "scorecard_presentations",
    "stakeholder_presentation",
    "revision_manifest",
)

_ROW_COLUMNS: dict[str, tuple[tuple[str, str], ...]] = {
    "portfolio": (
        ("Evidence Rank", "evidence_rank"), ("Eligible Candidate Rank", "candidate_rank"),
        ("Scorecard", "scorecard_name"), ("Score", "score_name"),
        ("Valid Feedback", "valid_feedback_count"),
        ("Reviewed Disagreements", "reviewed_disagreements"),
        ("Disagreement Rate", "disagreement_rate"),
        ("Reviewed Error Opportunity", "reviewed_error_opportunity"),
        ("Policy Disposition", "policy_disposition"), ("Policy Reason", "policy_reason"),
        ("Review Disposition", "review_disposition"), ("Eligible After", "eligibility_timestamp"),
        ("State", "state"), ("Coverage", "coverage_status"),
        ("Recent Trend", "trend"),
        ("Collection State", "collection_state"), ("Guideline State", "guideline_state"),
        ("Feedback/Rubric State", "feedback_rubric_state"), ("Readiness", "readiness"),
        ("Promotion Readiness", "promotion_readiness"), ("Rationale", "rationale"),
        ("Primary Disposition", "primary_disposition"), ("Secondary Issues", "secondary_issue_summary"),
        ("Dispatch Rejection", "dispatch_rejection"),
        ("Automatic Execution", "execution_status"), ("Execution Reason", "execution_reason"),
        ("Execution Authorization", "execution_authorization_source"),
        ("Next Action", "next_action"), ("Dashboard Link", "dashboard_url"),
        ("Semantic Diagnosis Status", "semantic_diagnosis_status"),
        ("Semantic Budget Evidence", "semantic_budget_evidence_reference"),
    ),
    "priorities": (
        ("Evidence Rank", "evidence_rank"), ("Eligible Candidate Rank", "candidate_rank"),
        ("Scorecard", "scorecard_name"), ("Score", "score_name"),
        ("Evidence Count", "evidence_count"), ("Opportunity", "opportunity"),
        ("Disagreement Rate", "disagreement_rate"),
        ("Policy Disposition", "policy_disposition"), ("Policy Reason", "policy_reason"),
        ("Review Disposition", "review_disposition"), ("Eligible After", "eligibility_timestamp"),
        ("State", "state"), ("Coverage", "coverage_status"), ("Recent Trend", "trend"),
        ("Collection State", "collection_state"), ("Readiness", "readiness"),
        ("Promotion Readiness", "promotion_readiness"), ("Rationale", "rationale"),
        ("Primary Disposition", "primary_disposition"), ("Secondary Issues", "secondary_issue_summary"),
        ("Automatic Execution", "execution_status"), ("Execution Reason", "execution_reason"),
        ("Execution Authorization", "execution_authorization_source"),
        ("Next Action", "next_action"), ("Dashboard Link", "dashboard_url"),
        ("Semantic Diagnosis Status", "semantic_diagnosis_status"),
        ("Semantic Budget Evidence", "semantic_budget_evidence_reference"),
    ),
    "feedback_investment": (
        ("Rank", "rank"), ("Scorecard", "scorecard_name"), ("Score", "score_name"),
        ("Evidence Count", "evidence_count"), ("State", "state"),
        ("Coverage", "coverage_status"), ("Recent Trend", "trend"),
        ("Recommendation", "recommendation"), ("Readiness", "readiness"),
        ("Rationale", "rationale"), ("Next Action", "next_action"),
        ("Dashboard Link", "dashboard_url"),
        ("Semantic Diagnosis Status", "semantic_diagnosis_status"),
        ("Semantic Budget Evidence", "semantic_budget_evidence_reference"),
    ),
    "questions_and_issues": (
        ("Rank", "rank"), ("Type", "kind"), ("Scorecard", "scorecard_name"), ("Score", "score_name"),
        ("Evidence Count", "evidence_count"), ("State", "state"),
        ("Coverage", "coverage_status"), ("Guideline State", "guideline_state"),
        ("Feedback/Rubric State", "feedback_rubric_state"),
        ("Issue Flag", "issue_flag"), ("Affected Feedback", "affected_evidence_count"),
        ("Affected Disagreement Rate", "affected_disagreement_rate"),
        ("Evidence References", "evidence_references"),
        ("Question or Issue", "finding"), ("Rationale", "rationale"),
        ("Next Action", "next_action"), ("Dashboard Link", "dashboard_url"),
        ("Semantic Diagnosis Status", "semantic_diagnosis_status"),
        ("Semantic Budget Evidence", "semantic_budget_evidence_reference"),
    ),
    "optimization_outcomes": (
        ("Rank", "rank"), ("Scorecard", "scorecard_name"), ("Score", "score_name"),
        ("Evidence Count", "evidence_count"), ("Outcome", "outcome"),
        ("Evidence Status", "evidence_status"), ("State", "state"),
        ("Coverage", "coverage_status"),
        ("Recent Trend", "trend"), ("Collection State", "collection_state"),
        ("Readiness", "readiness"), ("Promotion Readiness", "promotion_readiness"),
        ("Primary Disposition", "primary_disposition"), ("Secondary Issues", "secondary_issue_summary"),
        ("Rationale", "rationale"), ("Dispatch Rejection", "dispatch_rejection"),
        ("Next Action", "next_action"),
        ("Dashboard Link", "dashboard_url"),
        ("Semantic Diagnosis Status", "semantic_diagnosis_status"),
        ("Semantic Budget Evidence", "semantic_budget_evidence_reference"),
    ),
}
_OVERVIEW_KEYS = {
    "headline", "lifecycle_status", "current_activity", "next_checkpoint",
    "coverage_status", "inventory_coverage_status", "analysis_coverage_status",
    "execution_decision_status",
    "ranking_window", "scorecards_inspected",
    "scorecards_in_scope", "evidence_ranked_score_count",
    "ranked_score_count", "unranked_score_count", "cooldown_excluded_count",
    "assessed_score_count", "assessment_progress", "diagnosis_coverage", "pending_approval_count", "notes",
    "execution_mode", "execution_selected_count", "execution_launched_count", "execution_rejected_count",
    "execution_candidate_policy",
    "execution_named_selected_count", "execution_named_launched_count", "execution_named_rejected_count",
    "execution_detail_coverage", "execution_detail_limitation",
    "ranking_cutoff", "ranking_policy", "priority_display_limit",
    "priority_displayed_count", "priority_cutoff_rank", "priority_cutoff_opportunity",
    "ranked_below_priority_cutoff", "diagnosis_selection_policy",
    "diagnosis_top_priority_count", "diagnosis_monitoring_candidate_count",
    "diagnosis_selected_count", "diagnosis_scheduled_count", "diagnosis_deferred_count",
    "diagnosis_skipped_count", "diagnosis_incomplete_count", "diagnosis_completed_count", "diagnosis_max_count",
    "diagnosis_prerequisite_failure_count", "diagnosis_failure_category",
    "diagnosis_blockers",
    "approved_target_count", "dispatched_optimizer_count", "optimizer_review_count",
    "invalid_run_limit_target_count",
    "primary_disposition_counts", "secondary_issue_counts",
    "semantic_budget_policy_version", "semantic_budget_spec_schema_version",
    "semantic_budget_ledger_schema_version", "semantic_budget_pricing_version",
    "semantic_budget_provider", "semantic_budget_model",
    "semantic_budget_authorized_usd",
    "semantic_budget_settled_actual_usd", "semantic_budget_held_reserved_usd",
    "semantic_budget_available_usd", "semantic_budget_reservation_count",
    "semantic_budget_reserved_count", "semantic_budget_settled_count",
    "semantic_budget_unknown_count",
    "semantic_budget_cancelled_count", "semantic_budget_target_count",
    "semantic_budget_call_site_coverage", "semantic_budget_ledger_revision",
    "semantic_budget_evidence_reference", "semantic_budget_evidence_digest",
    "semantic_budget_deferred_count", "semantic_budget_failure_count",
    "semantic_diagnosis_deferred_after_failure_count",
    "semantic_budget_exhausted_count", "semantic_diagnosis_outcome_unknown_count",
    "semantic_authority_publication_failure_count",
    "semantic_diagnosis_issue_count", "semantic_diagnosis_issue_counts",
    "semantic_diagnosis_issues",
}
_ROW_METADATA_KEYS = {
    "scorecard_ref", "score_ref", "rank", "evidence_rank", "candidate_rank", "policy_disposition",
    "policy_reason", "review_disposition", "eligibility_timestamp",
    "primary_disposition", "secondary_issue_flags", "secondary_issue_summary",
    "issue_flag", "issue_severity", "affected_evidence_count", "affected_disagreement_rate",
    "evidence_references", "evidence_reference_tokens",
    "execution_status", "execution_reason", "execution_authorization_source",
}


class OptimizationRunPublicationError(RuntimeError):
    """A milestone could not be made durable, so the run must stop."""


class OptimizationRunRetryablePublicationError(OptimizationRunPublicationError):
    """A credential or publication interruption left no new committed revision.

    Callers must keep the Task active and replay from ``latest_revision``.
    """


class OptimizationRunIntegrityError(OptimizationRunPublicationError):
    """Committed recovery evidence is corrupt or does not identify this run."""


@dataclass
class OptimizationRunReportState:
    task: Any
    report: Any
    blocks: dict[str, Any]
    stages: dict[str, Any]
    run_spec: Mapping[str, Any]
    attempt_id: str
    operator_identity: OptimizationOperatorIdentity


@dataclass(frozen=True)
class WorkbookArtifact:
    content: bytes
    checksum: str
    row_counts: Mapping[str, int]


@dataclass(frozen=True)
class PublishedRevision:
    number: int
    milestone: str
    published_at: str
    raw_evidence_path: str
    workbook_path: str
    manifest_path: str
    manifest_checksum: str
    manifest_size_bytes: int
    evidence_checksum: str
    workbook_checksum: str
    row_counts: Mapping[str, int]
    overview: Mapping[str, Any]
    artifacts: tuple[Mapping[str, Any], ...] = ()
    detail_status: str = "pending"
    detail_source_revision: Optional[int] = None


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _iso(value: datetime) -> str:
    return value.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def _json(value: Any) -> bytes:
    return json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False, default=str).encode("utf-8")


def _normalize_dashboard_base_url(value: Optional[str]) -> Optional[str]:
    if value is None or not str(value).strip():
        return None
    parsed = urlparse(str(value).strip())
    if (
        parsed.scheme != "https"
        or not parsed.hostname
        or parsed.username
        or parsed.password
        or parsed.path not in {"", "/"}
        or parsed.params
        or parsed.query
        or parsed.fragment
    ):
        raise ValueError("dashboard_base_url must be an HTTPS origin")
    return f"https://{parsed.netloc}"


def dashboard_base_url_from_account_settings(settings: Any) -> Optional[str]:
    parsed_settings = _metadata(settings)
    reporting = parsed_settings.get("reporting")
    if not isinstance(reporting, Mapping):
        return None
    value = reporting.get("dashboardBaseUrl")
    if value is None:
        return None
    if not isinstance(value, str):
        raise ValueError("reporting.dashboardBaseUrl must be a string")
    return _normalize_dashboard_base_url(value)


def _report_artifact_url(
    base_url: str,
    *,
    report_id: str,
    revision_number: int,
    logical_id: str,
) -> str:
    query = urlencode({"revision": revision_number, "artifact": logical_id})
    return f"{base_url}/lab/reports/{quote(report_id, safe='')}?{query}"


def optimization_run_key(account_id: str, run_spec: Mapping[str, Any]) -> str:
    """Stable identity for a frozen account/scope/policy run."""
    if not account_id or not isinstance(run_spec, Mapping):
        raise ValueError("account_id and run_spec are required")
    return "optimization-run-" + sha256(_json({"account_id": account_id, "run_spec": dict(run_spec)})).hexdigest()[:24]


def _same_run_spec(left: Mapping[str, Any], right: Mapping[str, Any]) -> bool:
    # Policy was added after living reports already existed.  Legacy evidence
    # with no field means the conservative default, not an implicit diagnostic
    # opt-in; normalize only that omission before comparing frozen contracts.
    def normalized(value: Mapping[str, Any]) -> dict[str, Any]:
        result = dict(value)
        result["execution_candidate_policy"] = normalize_execution_candidate_policy(
            result.get("execution_candidate_policy")
        )
        return result

    return _json(normalized(left)) == _json(normalized(right))


def _metadata(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return dict(value)
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return dict(parsed) if isinstance(parsed, dict) else {}
    return {}


def _safe_cell(value: Any) -> Any:
    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    if text[:1] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def _validate_view(view: Mapping[str, Any]) -> None:
    allowed_top_level = {"overview", "portfolio", "priorities", "feedback_investment", "questions_and_issues", "optimization_outcomes", "definitions"}
    unknown = set(view) - allowed_top_level
    if unknown:
        raise ValueError(f"stakeholder view keys are not allowed: {sorted(unknown)}")
    overview = view.get("overview", {})
    if not isinstance(overview, Mapping) or set(overview) - _OVERVIEW_KEYS:
        raise ValueError("overview contains keys that are not allowed")
    semantic_issues = overview.get("semantic_diagnosis_issues", [])
    semantic_issue_keys = {
        "scorecard_name", "score_name", "semantic_diagnosis_status",
        "next_action", "rationale",
    }
    if not isinstance(semantic_issues, list) or any(
        not isinstance(issue, Mapping) or set(issue) != semantic_issue_keys
        for issue in semantic_issues
    ):
        raise ValueError("semantic diagnosis issues have an invalid safe projection")
    definitions = view.get("definitions", {})
    if not isinstance(definitions, Mapping):
        raise ValueError("definitions must be a mapping")
    for group, columns in _ROW_COLUMNS.items():
        rows = view.get(group, [])
        if not isinstance(rows, list):
            raise ValueError(f"{group} must be a list")
        allowed = {key for _, key in columns} | _ROW_METADATA_KEYS
        for row in rows:
            if not isinstance(row, Mapping):
                raise ValueError(f"{group} rows must be mappings")
            disallowed = set(row) - allowed
            if disallowed:
                raise ValueError(f"{group} contains keys that are not allowed: {sorted(disallowed)}")


def _write_table(sheet: Any, columns: tuple[tuple[str, str], ...], rows: list[Mapping[str, Any]]) -> None:
    sheet.append([title for title, _ in columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([_safe_cell(row.get(key)) for _, key in columns])
        if "dashboard_url" in row and row.get("dashboard_url"):
            cell = sheet.cell(sheet.max_row, len(columns))
            link = str(row["dashboard_url"])
            if link.startswith("https://"):
                cell.hyperlink = link
                cell.style = "Hyperlink"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    safe_name = "".join(part.title() for part in sheet.title.replace("-", " ").split()) + "Table"
    table = Table(displayName=safe_name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(max(width, 12), 60)


def _canonical_xlsx(workbook: Workbook) -> bytes:
    """Save an xlsx with stable Zip metadata so checksums are reproducible."""
    interim = BytesIO()
    # Workbook.save() replaces the caller-supplied modified timestamp with the
    # wall clock immediately before serialization. Write through ExcelWriter
    # so both core timestamps remain part of the frozen report evidence.
    ExcelWriter(
        workbook,
        ZipFile(interim, "w", compression=ZIP_DEFLATED, allowZip64=True),
    ).save()
    output = BytesIO()
    with ZipFile(interim, "r") as source, ZipFile(output, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
        for name in sorted(source.namelist()):
            info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            target.writestr(info, source.read(name), compress_type=ZIP_DEFLATED, compresslevel=9)
    return output.getvalue()


def _validate_workbook_bytes(content: bytes, expected_row_counts: Mapping[str, int]) -> None:
    """Reopen the final artifact and reject unsafe or divergent workbooks."""
    workbook = load_workbook(BytesIO(content), data_only=False, keep_links=True)
    if workbook.sheetnames != list(SHEET_NAMES):
        raise ValueError("workbook sheet contract changed")
    if getattr(workbook, "_external_links", None):
        raise ValueError("workbook contains external links")
    if any(sheet.sheet_state != "visible" for sheet in workbook.worksheets):
        raise ValueError("workbook contains hidden sheets")
    if any(
        cell.data_type == "f"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ):
        raise ValueError("workbook contains formulas")
    sheet_by_projection = {
        "portfolio": "Portfolio",
        "priorities": "Priorities",
        "feedback_investment": "Feedback Investment",
        "questions_and_issues": "Questions and Issues",
        "optimization_outcomes": "Optimization Outcomes",
        "run_log": "Run Log",
        "definitions": "Definitions",
    }
    for key, expected in expected_row_counts.items():
        actual = max(workbook[sheet_by_projection[key]].max_row - 1, 0)
        if actual != expected:
            raise ValueError(f"workbook row count mismatch for {key}: {actual} != {expected}")


def _stakeholder_execution_projection(
    stakeholder_view: Mapping[str, Any],
    decision_evidence: Mapping[str, Any],
    *,
    expected_execution_mode: str,
) -> dict[str, Any]:
    """Add an automatic-execution projection without copying opaque identities.

    The raw decision evidence is intentionally retained only in the restricted
    evidence artifact.  Stakeholder artifacts receive an explicit mode,
    reconciled aggregate counts, and safe score names/reasons when supplied.
    The frozen run specification is the single authority for execution mode.
    Contradictory decision evidence is an integrity failure, not a display
    preference.  Legacy evidence without a mode is projected under the frozen
    mode but cannot supply automatic target detail.
    """
    if expected_execution_mode not in {"automatic", "approval_required"}:
        raise OptimizationRunIntegrityError("frozen run specification has an invalid execution mode")
    evidence_execution_mode = decision_evidence.get("execution_mode")
    if (
        evidence_execution_mode is not None
        and evidence_execution_mode != expected_execution_mode
    ):
        raise OptimizationRunIntegrityError(
            "decision evidence execution mode conflicts with the frozen run specification"
        )
    decisions = decision_evidence.get("execution_decisions")
    decision_mode = decisions.get("mode") if isinstance(decisions, Mapping) else None
    if decision_mode is not None and decision_mode != expected_execution_mode:
        raise OptimizationRunIntegrityError(
            "execution mode in decisions conflicts with the frozen run specification"
        )
    projected = {
        key: ([dict(row) for row in value] if isinstance(value, list) else dict(value)
              if isinstance(value, Mapping) else value)
        for key, value in stakeholder_view.items()
    }
    execution_mode = expected_execution_mode
    if execution_mode == "approval_required":
        overview = dict(projected.get("overview") or {})
        overview["execution_mode"] = "approval_required"
        projected["overview"] = overview
        return projected
    if not isinstance(decisions, Mapping):
        overview = dict(projected.get("overview") or {})
        overview["execution_mode"] = "automatic"
        projected["overview"] = overview
        return projected

    selected = decisions.get("selected_targets")
    rejected = decisions.get("rejected_targets")
    selected_rows = selected if isinstance(selected, list) else []
    rejected_rows = rejected if isinstance(rejected, list) else []

    def target_identity(row: Any) -> Optional[tuple[str, str]]:
        if not isinstance(row, Mapping):
            return None
        scorecard_id = row.get("scorecard_id")
        score_id = row.get("score_id")
        if (
            not isinstance(scorecard_id, str) or not scorecard_id.strip()
            or not isinstance(score_id, str) or not score_id.strip()
        ):
            return None
        return scorecard_id, score_id

    durable_launched_targets: set[tuple[str, str]] = set()
    dispatch = decision_evidence.get("dispatch")
    children = dispatch.get("children") if isinstance(dispatch, Mapping) else None
    for child in children if isinstance(children, list) else []:
        if not isinstance(child, Mapping):
            continue
        launch_state = child.get("launch_state")
        identity = target_identity(child.get("target"))
        if (
            identity is not None
            and isinstance(child.get("procedure_id"), str) and bool(child.get("procedure_id").strip())
            and isinstance(child.get("task_id"), str) and bool(child.get("task_id").strip())
            and isinstance(launch_state, Mapping)
            and launch_state.get("phase") in {"waiting", "running", "terminal"}
        ):
            durable_launched_targets.add(identity)

    def count(name: str, fallback: int) -> int:
        value = decisions.get(name)
        return value if isinstance(value, int) and not isinstance(value, bool) and value >= 0 else fallback

    selected_count = count("selected_count", len(selected_rows))
    launched_count = count("launched_count", 0)
    rejected_count = count("rejected_count", len(rejected_rows))
    named_selected_count = sum(
        1 for row in selected_rows
        if isinstance(row, Mapping)
        and isinstance(row.get("scorecard_name"), str) and bool(row.get("scorecard_name"))
        and isinstance(row.get("score_name"), str) and bool(row.get("score_name"))
    )
    named_rejected_count = sum(
        1 for row in rejected_rows
        if isinstance(row, Mapping)
        and isinstance(row.get("scorecard_name"), str) and bool(row.get("scorecard_name"))
        and isinstance(row.get("score_name"), str) and bool(row.get("score_name"))
    )
    derived_launched_count = sum(
        1 for row in selected_rows if target_identity(row) in durable_launched_targets
    )
    named_launched_count = sum(
        1 for row in selected_rows
        if target_identity(row) in durable_launched_targets
        and isinstance(row, Mapping)
        and isinstance(row.get("scorecard_name"), str) and bool(row.get("scorecard_name"))
        and isinstance(row.get("score_name"), str) and bool(row.get("score_name"))
    )
    coverage_complete = (
        named_selected_count == selected_count
        and named_rejected_count == rejected_count
        and derived_launched_count == launched_count
        and named_launched_count == launched_count
    )
    limitations: list[str] = []
    if named_selected_count != selected_count:
        limitations.append(f"named detail is available for {named_selected_count} of {selected_count} selected targets")
    if named_rejected_count != rejected_count:
        limitations.append(
            f"named detail is available for {named_rejected_count} of "
            f"{rejected_count} targets not selected by policy"
        )
    if derived_launched_count != launched_count:
        limitations.append(
            f"durable launch evidence reconciles {derived_launched_count} of {launched_count} reported launches"
        )
    elif named_launched_count != launched_count:
        limitations.append(f"named detail is available for {named_launched_count} of {launched_count} launched targets")

    overview = dict(projected.get("overview") or {})
    overview.update({
        "execution_mode": execution_mode,
        "execution_selected_count": selected_count,
        "execution_launched_count": launched_count,
        "execution_rejected_count": rejected_count,
        "execution_named_selected_count": named_selected_count,
        "execution_named_launched_count": named_launched_count,
        "execution_named_rejected_count": named_rejected_count,
        "execution_detail_coverage": "complete" if coverage_complete else "incomplete",
        "execution_detail_limitation": "; ".join(limitations).capitalize() + ("." if limitations else ""),
    })
    projected["overview"] = overview

    execution_by_score: dict[tuple[str, str], dict[str, str]] = {}

    def decision_text(row: Mapping[str, Any], *keys: str) -> str:
        for key in keys:
            value = row.get(key)
            if isinstance(value, str) and value:
                return value
        return ""

    def register(rows: list[Any], status: str) -> None:
        for raw_row in rows:
            if not isinstance(raw_row, Mapping):
                continue
            scorecard_name = raw_row.get("scorecard_name")
            score_name = raw_row.get("score_name")
            if not isinstance(scorecard_name, str) or not scorecard_name or not isinstance(score_name, str) or not score_name:
                continue
            reason = decision_text(
                raw_row, "reason", "decision_reason", "rejection_reason"
            )
            execution_by_score[(scorecard_name, score_name)] = {
                "execution_status": (
                    "automatic_launched"
                    if status == "automatic_selected" and target_identity(raw_row) in durable_launched_targets
                    else "diagnosis_required"
                    if status == "automatic_rejected" and reason == "missing_diagnosis"
                    else "execution_limit_deferred"
                    if status == "automatic_rejected" and reason == "execution_target_limit"
                    else status
                ),
                "execution_reason": reason,
                "execution_authorization_source": decision_text(raw_row, "authorization_source", "authorization"),
            }

    register(selected_rows, "automatic_selected")
    register(rejected_rows, "automatic_rejected")
    for group in ("portfolio", "priorities", "optimization_outcomes"):
        rows = projected.get(group, [])
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            decision = execution_by_score.get((row.get("scorecard_name"), row.get("score_name")))
            if decision:
                row.update(decision)
                if decision["execution_status"] == "diagnosis_required":
                    row.update({
                        "readiness": "incomplete",
                        "state": "incomplete",
                        "coverage_status": "incomplete",
                        "next_action": "await_semantic_diagnosis",
                        "rationale": (
                            "Deterministic assessment found a possible opportunity, "
                            "but semantic diagnosis is not complete. This score is not "
                            "approved or ready for automatic optimization."
                        ),
                    })
                elif decision["execution_status"] == "execution_limit_deferred":
                    row.update({
                        "next_action": "consider_next_portfolio_run",
                        "rationale": (
                            "This target was ready under the automatic policy but fell "
                            "outside this run's frozen top-K execution limit. It was not "
                            "a launch failure and remains eligible for a later run."
                        ),
                    })

    # A policy-rejected target can legitimately have no portfolio row. Keep
    # its safe name and deterministic reason visible in Optimization Outcomes,
    # never its opaque target identity.
    outcomes = projected.setdefault("optimization_outcomes", [])
    if isinstance(outcomes, list):
        existing = {
            (row.get("scorecard_name"), row.get("score_name"))
            for row in outcomes if isinstance(row, Mapping)
        }
        for score_key, decision in execution_by_score.items():
            if decision["execution_status"] == "automatic_rejected" and score_key not in existing:
                outcomes.append({
                    "scorecard_name": score_key[0], "score_name": score_key[1],
                    "outcome": "automatic_rejected", "rationale": decision["execution_reason"],
                    "next_action": "review_execution_decision", **decision,
                })
    return projected


def build_stakeholder_workbook(
    stakeholder_view: Mapping[str, Any], *, revision_number: int, generated_at: datetime
) -> WorkbookArtifact:
    """Create the fixed, macro-free stakeholder workbook from a safe projection only."""
    _validate_view(stakeholder_view)
    generated_at = generated_at.astimezone(timezone.utc)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Plexus"
    workbook.properties.created = generated_at
    workbook.properties.modified = generated_at
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False

    overview = workbook.create_sheet("Overview")
    overview.append(["Optimization Run Overview", "Value"])
    for cell in overview[1]:
        cell.font = Font(bold=True)
    overview.append(["Revision", revision_number])
    overview.append(["Published At", _iso(generated_at)])
    for key in sorted(stakeholder_view.get("overview", {})):
        overview.append([key.replace("_", " ").title(), _safe_cell(stakeholder_view["overview"][key])])
    overview.column_dimensions["A"].width = min(
        max(len(str(cell.value or "")) for cell in overview["A"]) + 2,
        40,
    )
    overview.column_dimensions["B"].width = min(
        max(max(len(str(cell.value or "")) for cell in overview["B"]) + 2, 40),
        100,
    )
    for cell in overview["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    row_counts: dict[str, int] = {}
    table_sheets = (
        ("Portfolio", "portfolio"), ("Priorities", "priorities"),
        ("Feedback Investment", "feedback_investment"), ("Questions and Issues", "questions_and_issues"),
        ("Optimization Outcomes", "optimization_outcomes"),
    )
    for sheet_name, key in table_sheets:
        rows = list(stakeholder_view.get(key, []))
        _write_table(workbook.create_sheet(sheet_name), _ROW_COLUMNS[key], rows)
        row_counts[key] = len(rows)

    run_log = workbook.create_sheet("Run Log")
    _write_table(
        run_log,
        (
            ("Revision", "revision"), ("Published At", "published_at"),
            ("Coverage", "coverage"), ("Status", "status"),
            ("Execution", "execution"),
            ("Semantic Spend", "semantic_spend"),
            ("Semantic Calls", "semantic_calls"),
            ("Semantic Evidence", "semantic_evidence"),
        ),
        [{
            "revision": revision_number, "published_at": _iso(generated_at),
            "coverage": stakeholder_view.get("overview", {}).get("coverage_status", "not provided"),
            "status": "published",
            "execution": _execution_summary_text(stakeholder_view.get("overview", {})),
            "semantic_spend": _semantic_spend_text(stakeholder_view.get("overview", {})),
            "semantic_calls": _semantic_call_text(stakeholder_view.get("overview", {})),
            "semantic_evidence": stakeholder_view.get("overview", {}).get("semantic_budget_evidence_reference"),
        }],
    )
    row_counts["run_log"] = 1

    definitions = workbook.create_sheet("Definitions")
    _write_table(
        definitions,
        (("Term", "term"), ("Definition", "definition")),
        [{"term": term, "definition": definition} for term, definition in sorted(stakeholder_view.get("definitions", {}).items())],
    )
    row_counts["definitions"] = len(stakeholder_view.get("definitions", {}))

    content = _canonical_xlsx(workbook)
    _validate_workbook_bytes(content, row_counts)
    return WorkbookArtifact(content=content, checksum=sha256(content).hexdigest(), row_counts=row_counts)


def _markdown_text(value: Any) -> str:
    return str(value or "").replace("\r", " ").replace("\n", " ").replace("|", "\\|").strip()


def _semantic_spend_text(overview: Any) -> str:
    if not isinstance(overview, Mapping) or not overview.get("semantic_budget_authorized_usd"):
        return "Not configured"
    return (
        f"authorized {overview.get('semantic_budget_authorized_usd')}; "
        f"spent {overview.get('semantic_budget_settled_actual_usd')}; "
        f"held {overview.get('semantic_budget_held_reserved_usd')}; "
        f"remaining {overview.get('semantic_budget_available_usd')}"
    )


def _execution_summary_text(overview: Any) -> str:
    if not isinstance(overview, Mapping) or overview.get("execution_mode") is None:
        return "Not reported"
    if overview.get("execution_mode") != "automatic":
        return "Approval required"
    summary = (
        f"selected {overview.get('execution_selected_count', 0)}; "
        f"launched {overview.get('execution_launched_count', 0)}; "
        f"not selected {overview.get('execution_rejected_count', 0)}; "
        f"detail {overview.get('execution_detail_coverage', 'not reported')}"
    )
    limitation = overview.get("execution_detail_limitation")
    return f"{summary}; {limitation}" if limitation else summary


def _stakeholder_execution_status_text(value: Any) -> str:
    """Translate the execution contract into plain stakeholder language."""
    return {
        "automatic_selected": "selected automatically",
        "automatic_launched": "launched automatically",
        "automatic_rejected": "not selected automatically",
        "diagnosis_required": "requires semantic diagnosis",
        "execution_limit_deferred": "deferred by the execution limit",
    }.get(str(value or ""), str(value or "not applicable"))


def _semantic_call_text(overview: Any) -> str:
    if not isinstance(overview, Mapping) or overview.get("semantic_budget_reservation_count") is None:
        return "Not configured"
    return (
        f"total {overview.get('semantic_budget_reservation_count')}; "
        f"reserved {overview.get('semantic_budget_reserved_count')}; "
        f"settled {overview.get('semantic_budget_settled_count')}; "
        f"unknown {overview.get('semantic_budget_unknown_count')}; "
        f"cancelled {overview.get('semantic_budget_cancelled_count')}; "
        f"deferred {overview.get('semantic_budget_deferred_count')}; "
        f"failed {overview.get('semantic_budget_failure_count')}"
    )


def _artifact_descriptor(
    *,
    logical_id: str,
    kind: str,
    display_name: str,
    scope: str,
    content_type: str,
    content: bytes,
    object_key: str,
    task_id: str,
    source_revision: int,
    scorecard_name: Optional[str] = None,
    score_name: Optional[str] = None,
) -> dict[str, Any]:
    descriptor: dict[str, Any] = {
        "logical_id": logical_id,
        "kind": kind,
        "display_name": display_name,
        "scope": scope,
        "content_type": content_type,
        "size_bytes": len(content),
        "sha256": sha256(content).hexdigest(),
        "task_id": task_id,
        "object_key": object_key,
        "source_revision": source_revision,
    }
    if scorecard_name:
        descriptor["scorecard_name"] = scorecard_name
    if score_name:
        descriptor["score_name"] = score_name
    return descriptor


def _score_issues(
    stakeholder_view: Mapping[str, Any],
    *,
    portfolio_row: Mapping[str, Any],
) -> list[Mapping[str, Any]]:
    return [
        issue for issue in stakeholder_view.get("questions_and_issues", [])
        if isinstance(issue, Mapping)
        and _score_evidence_matches_portfolio_row(portfolio_row, issue)
    ]


def _score_evidence_matches_portfolio_row(
    portfolio_row: Mapping[str, Any], evidence_row: Mapping[str, Any]
) -> bool:
    """Return whether a stakeholder finding identifies this portfolio score.

    Newer projections retain opaque scorecard and score references outside the
    stakeholder-facing tables.  When a finding supplies either reference it is
    authoritative for that component; legacy projections safely retain their
    display-name matching behavior instead.  A score name/reference is always
    required, so scorecard-wide issues do not accidentally create a brief for
    every score in the scorecard.
    """
    evidence_score_ref = str(evidence_row.get("score_ref") or "")
    portfolio_score_ref = str(portfolio_row.get("score_ref") or "")
    evidence_score_name = str(evidence_row.get("score_name") or "")
    portfolio_score_name = str(portfolio_row.get("score_name") or "")
    if evidence_score_ref:
        if not portfolio_score_ref or portfolio_score_ref != evidence_score_ref:
            return False
    elif not evidence_score_name or portfolio_score_name != evidence_score_name:
        return False

    evidence_scorecard_ref = str(evidence_row.get("scorecard_ref") or "")
    portfolio_scorecard_ref = str(portfolio_row.get("scorecard_ref") or "")
    evidence_scorecard_name = str(evidence_row.get("scorecard_name") or "")
    portfolio_scorecard_name = str(portfolio_row.get("scorecard_name") or "")
    if evidence_scorecard_ref:
        return bool(
            portfolio_scorecard_ref
            and portfolio_scorecard_ref == evidence_scorecard_ref
        )
    if evidence_scorecard_name:
        return portfolio_scorecard_name == evidence_scorecard_name
    return True


def _stakeholder_priority_evidence_rows(
    stakeholder_view: Mapping[str, Any],
) -> list[Mapping[str, Any]]:
    """Return only priority rows intentionally displayed to stakeholders."""
    priority_rows = [
        row for row in stakeholder_view.get("priorities", []) if isinstance(row, Mapping)
    ]
    overview = stakeholder_view.get("overview")
    overview = overview if isinstance(overview, Mapping) else {}

    def positive_integer(value: Any) -> int | None:
        if isinstance(value, bool):
            return None
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    cutoff = positive_integer(overview.get("priority_cutoff_rank"))
    if cutoff is None:
        cutoff = positive_integer(overview.get("priority_display_limit"))
    # Older callers supplied only the already-bounded priority display.  Do
    # not reinterpret that legacy shape as an unbounded evidence inventory.
    if cutoff is None:
        return priority_rows

    displayed: list[Mapping[str, Any]] = []
    for position, row in enumerate(priority_rows, start=1):
        rank = positive_integer(row.get("evidence_rank"))
        if rank is None:
            rank = positive_integer(row.get("rank"))
        if rank is None:
            rank = position
        if rank <= cutoff:
            displayed.append(row)
    return displayed


def _score_brief_evidence_rows(
    stakeholder_view: Mapping[str, Any],
) -> list[Mapping[str, Any]]:
    """Return the only stakeholder findings that justify a score brief."""
    questions = [
        row
        for row in stakeholder_view.get("questions_and_issues", [])
        if isinstance(row, Mapping)
    ]
    material_outcomes = [
        row
        for row in stakeholder_view.get("optimization_outcomes", [])
        if isinstance(row, Mapping)
        and str(row.get("outcome") or "").strip().casefold() not in {"", "not_run"}
    ]
    return [
        *_stakeholder_priority_evidence_rows(stakeholder_view),
        *questions,
        *material_outcomes,
    ]


def _score_brief_portfolio_indexes(
    stakeholder_view: Mapping[str, Any],
) -> set[int]:
    """Select portfolio rows that have stakeholder-relevant score evidence.

    This is deliberately the sole score-brief selection contract.  Artifact
    generation and pre-publication progress planning both call it, preventing
    an inaccurate progress total or an unbounded per-score artifact fanout.
    """
    evidence_rows = _score_brief_evidence_rows(stakeholder_view)
    return {
        index
        for index, portfolio_row in enumerate(stakeholder_view.get("portfolio", []))
        if isinstance(portfolio_row, Mapping)
        and any(
            _score_evidence_matches_portfolio_row(portfolio_row, evidence_row)
            for evidence_row in evidence_rows
        )
    }


def _secondary_issue_flags_text(row: Mapping[str, Any]) -> str:
    """Render the canonical flag list while retaining older summary-only rows."""
    flags = row.get("secondary_issue_flags")
    if isinstance(flags, (list, tuple)):
        return ", ".join(str(flag) for flag in flags if str(flag)) or "none"
    return str(row.get("secondary_issue_summary") or "none")


def _score_brief_markdown(
    scorecard_name: str,
    row: Mapping[str, Any],
    stakeholder_view: Mapping[str, Any],
) -> bytes:
    score_name = str(row.get("score_name") or "Unlabeled score")
    lines = [
        f"# {_markdown_text(score_name)}",
        "",
        f"Scorecard: {_markdown_text(scorecard_name)}",
        "",
        "## Current finding",
        "",
        f"- Readiness: {_markdown_text(row.get('readiness') or 'inconclusive')}",
        f"- Feedback collection: {_markdown_text(row.get('collection_state') or 'inconclusive')}",
        f"- Valid feedback: {_markdown_text(row.get('valid_feedback_count') or 0)}",
        f"- Reviewed disagreements: {_markdown_text(row.get('reviewed_disagreements') or 0)}",
        f"- Primary disposition: {_markdown_text(row.get('primary_disposition') or 'not_selected')}",
        f"- Secondary issue flags: {_markdown_text(_secondary_issue_flags_text(row))}",
        f"- Automatic execution: {_markdown_text(_stakeholder_execution_status_text(row.get('execution_status')))}",
        f"- Execution reason: {_markdown_text(row.get('execution_reason') or 'not provided')}",
        f"- Next action: {_markdown_text(row.get('next_action') or 'review')}",
        "",
        _markdown_text(row.get("rationale") or "No stakeholder-safe rationale is available yet."),
    ]
    issues = _score_issues(
        stakeholder_view,
        portfolio_row=row,
    )
    if issues:
        lines.extend(["", "## Questions and issues", ""])
        for issue in issues:
            finding = _markdown_text(
                issue.get("finding") or issue.get("rationale") or "Review required"
            )
            next_action = _markdown_text(issue.get("next_action") or "review")
            lines.append(f"- {finding} Next action: {next_action}.")
    lines.extend([
        "",
        "This brief contains stakeholder-safe findings only. The living Plexus Report is the cover page and lifecycle authority.",
        "",
    ])
    return "\n".join(lines).encode("utf-8")


def _scorecard_summary_markdown(
    scorecard_name: str,
    scorecard_ref: str,
    rows: list[Mapping[str, Any]],
    stakeholder_view: Mapping[str, Any],
) -> bytes:
    readiness_counts: dict[str, int] = {}
    collection_counts: dict[str, int] = {}
    for row in rows:
        readiness = _markdown_text(row.get("readiness") or "inconclusive")
        collection = _markdown_text(row.get("collection_state") or "inconclusive")
        readiness_counts[readiness] = readiness_counts.get(readiness, 0) + 1
        collection_counts[collection] = collection_counts.get(collection, 0) + 1

    def _counts(values: Mapping[str, int]) -> str:
        return ", ".join(f"{key}: {value}" for key, value in sorted(values.items())) or "None"

    def _opportunity(row: Mapping[str, Any]) -> float:
        try:
            return float(row.get("reviewed_error_opportunity") or 0)
        except (TypeError, ValueError):
            return 0.0

    priority_rows = sorted(rows, key=_opportunity, reverse=True)
    lines = [
        f"# {_markdown_text(scorecard_name)}",
        "",
        f"This scorecard has {len(rows)} scored criteria in this optimization report.",
        "",
        f"Readiness: {_counts(readiness_counts)}.",
        f"Feedback investment: {_counts(collection_counts)}.",
        "",
        "## Score details",
        "",
        "| Score | Valid feedback | Reviewed disagreements | Readiness | Primary disposition | Automatic execution | Execution reason | Secondary issue flags | Next action |",
        "|---|---:|---:|---|---|---|---|---|---|",
    ]
    for row in priority_rows:
        lines.append(
            "| " + " | ".join([
                _markdown_text(row.get("score_name") or "Unlabeled score"),
                _markdown_text(row.get("valid_feedback_count")),
                _markdown_text(row.get("reviewed_disagreements")),
                _markdown_text(row.get("readiness") or "inconclusive"),
                _markdown_text(row.get("primary_disposition") or "not_selected"),
                _markdown_text(_stakeholder_execution_status_text(row.get("execution_status"))),
                _markdown_text(row.get("execution_reason") or "not provided"),
                _markdown_text(_secondary_issue_flags_text(row)),
                _markdown_text(row.get("next_action") or "review"),
            ]) + " |"
        )

    issues = [
        row for row in stakeholder_view.get("questions_and_issues", [])
        if (
            row.get("scorecard_ref") == scorecard_ref
            or (
                not row.get("scorecard_ref")
                and row.get("scorecard_name") == scorecard_name
            )
        )
    ]
    if issues:
        lines.extend(["", "## Questions and issues", ""])
        for issue in issues:
            finding = _markdown_text(issue.get("finding") or issue.get("rationale") or "Review required")
            score_name = _markdown_text(issue.get("score_name") or "Score")
            lines.append(f"- {score_name}: {finding}")

    lines.extend([
        "",
        "This summary contains stakeholder-safe findings only. The living Plexus Report is the cover page and lifecycle authority.",
        "",
    ])
    return "\n".join(lines).encode("utf-8")


def _scorecard_csv(rows: list[Mapping[str, Any]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow([title for title, _ in _ROW_COLUMNS["portfolio"]])
    for row in rows:
        writer.writerow([_safe_cell(row.get(key)) for _, key in _ROW_COLUMNS["portfolio"]])
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _scorecard_presentation(
    scorecard_name: str,
    scorecard_ref: str,
    rows: list[Mapping[str, Any]],
    stakeholder_view: Mapping[str, Any],
    score_artifacts: list[list[Mapping[str, Any]]],
) -> bytes:
    issues = [
        dict(row) for row in stakeholder_view.get("questions_and_issues", [])
        if (
            row.get("scorecard_ref") == scorecard_ref
            or (
                not row.get("scorecard_ref")
                and row.get("scorecard_name") == scorecard_name
            )
        )
    ]
    return _json({
        "scorecard_name": scorecard_name,
        "scorecard_ref": scorecard_ref,
        "scores": [
            {**dict(row), "artifacts": [dict(artifact) for artifact in artifacts]}
            for row, artifacts in zip(rows, score_artifacts)
        ],
        "questions_and_issues": issues,
    })


def build_scorecard_artifacts(
    stakeholder_view: Mapping[str, Any],
    *,
    revision_number: int,
    task_id: str,
    uploader: Callable[[str, str, bytes], str],
    publication_id: Optional[str] = None,
    on_artifact_upload: Optional[Callable[[str, bool], None]] = None,
    reuse_artifact: Optional[
        Callable[[str, str, str, bytes, str], Optional[Mapping[str, Any]]]
    ] = None,
    on_artifact_resolved: Optional[
        Callable[[str, Mapping[str, Any]], None]
    ] = None,
) -> list[dict[str, Any]]:
    """Publish one safe Markdown summary and quantitative CSV per scorecard."""
    _validate_view(stakeholder_view)
    relevant_score_indexes = _score_brief_portfolio_indexes(stakeholder_view)
    grouped: dict[tuple[str, str], list[tuple[int, Mapping[str, Any]]]] = {}
    for portfolio_index, row in enumerate(stakeholder_view.get("portfolio", [])):
        scorecard_name = str(row.get("scorecard_name") or "Unlabeled scorecard")
        stable_key = str(row.get("scorecard_ref") or scorecard_name)
        grouped.setdefault((stable_key, scorecard_name), []).append((portfolio_index, row))

    descriptors: list[dict[str, Any]] = []
    publication_suffix = f"-{publication_id}" if publication_id else ""

    def resolve_artifact(
        *,
        progress_kind: str,
        logical_id: str,
        kind: str,
        content_type: str,
        filename: str,
        content: bytes,
    ) -> tuple[str, int]:
        if on_artifact_upload is not None:
            on_artifact_upload(progress_kind, False)
        reusable = (
            reuse_artifact(logical_id, kind, content_type, content, filename)
            if reuse_artifact is not None
            else None
        )
        if isinstance(reusable, Mapping):
            object_key = str(reusable.get("object_key") or "")
            source_revision = int(reusable.get("source_revision") or 0)
            if not object_key or source_revision < 1:
                raise OptimizationRunIntegrityError(
                    "reusable artifact descriptor is incomplete"
                )
        else:
            object_key = uploader(task_id, filename, content)
            source_revision = revision_number
        if on_artifact_upload is not None:
            on_artifact_upload(progress_kind, True)
        return object_key, source_revision

    for (stable_key, scorecard_name), indexed_rows in sorted(grouped.items(), key=lambda item: item[0][1].casefold()):
        rows = [row for _, row in indexed_rows]
        scope_hash = sha256(stable_key.encode("utf-8")).hexdigest()[:16]
        score_artifacts: list[list[Mapping[str, Any]]] = []
        for row_index, (portfolio_index, row) in enumerate(indexed_rows):
            if portfolio_index not in relevant_score_indexes:
                score_artifacts.append([])
                continue
            score_name = str(row.get("score_name") or "Unlabeled score")
            score_hash = sha256(
                f"{stable_key}\0{score_name}\0{row_index}".encode("utf-8")
            ).hexdigest()[:16]
            content = _score_brief_markdown(
                scorecard_name,
                row,
                stakeholder_view,
            )
            filename = (
                f"score-{score_hash}-brief-r{revision_number:04d}"
                f"{publication_suffix}.md"
            )
            logical_id = f"score_brief:{score_hash}"
            object_key, source_revision = resolve_artifact(
                progress_kind="score_briefs",
                logical_id=logical_id,
                kind="score_brief",
                content_type="text/markdown",
                filename=filename,
                content=content,
            )
            descriptor = _artifact_descriptor(
                logical_id=logical_id,
                kind="score_brief",
                display_name="Score brief",
                scope="score",
                content_type="text/markdown",
                content=content,
                object_key=object_key,
                task_id=task_id,
                source_revision=source_revision,
                scorecard_name=scorecard_name,
                score_name=score_name,
            )
            descriptors.append(descriptor)
            if on_artifact_resolved is not None:
                on_artifact_resolved("score_briefs", descriptor)
            score_artifacts.append([descriptor])
        artifacts = (
            (
                "scorecard_summary",
                "Summary",
                "text/markdown",
                f"scorecard-{scope_hash}-summary-r{revision_number:04d}{publication_suffix}.md",
                _scorecard_summary_markdown(
                    scorecard_name,
                    stable_key,
                    rows,
                    stakeholder_view,
                ),
            ),
            (
                "scorecard_portfolio_csv",
                "Quantitative results",
                "text/csv",
                f"scorecard-{scope_hash}-portfolio-r{revision_number:04d}{publication_suffix}.csv",
                _scorecard_csv(rows),
            ),
            (
                "scorecard_presentation",
                "Interactive score details",
                "application/json",
                f"scorecard-{scope_hash}-presentation-r{revision_number:04d}{publication_suffix}.json",
                _scorecard_presentation(
                    scorecard_name,
                    stable_key,
                    rows,
                    stakeholder_view,
                    score_artifacts,
                ),
            ),
        )
        for kind, display_name, content_type, filename, content in artifacts:
            progress_kind = {
                "scorecard_summary": "scorecard_summaries",
                "scorecard_portfolio_csv": "scorecard_spreadsheets",
                "scorecard_presentation": "scorecard_presentations",
            }[kind]
            logical_id = f"{kind}:{scope_hash}"
            object_key, source_revision = resolve_artifact(
                progress_kind=progress_kind,
                logical_id=logical_id,
                kind=kind,
                content_type=content_type,
                filename=filename,
                content=content,
            )
            descriptor = _artifact_descriptor(
                logical_id=logical_id,
                kind=kind,
                display_name=display_name,
                scope="scorecard",
                content_type=content_type,
                content=content,
                object_key=object_key,
                task_id=task_id,
                source_revision=source_revision,
                scorecard_name=scorecard_name,
            )
            descriptors.append(descriptor)
            if on_artifact_resolved is not None:
                on_artifact_resolved(progress_kind, descriptor)
    return descriptors


def _artifact_publication_plan(
    stakeholder_view: Mapping[str, Any],
) -> dict[str, dict[str, int]]:
    """Return safe aggregate artifact counts before a milestone is uploaded.

    The plan intentionally uses only public presentation categories.  It is
    never an artifact index: callers must not add filenames, attachment
    pointers, scorecard references, or any other opaque identifiers here.
    """
    _validate_view(stakeholder_view)
    rows = list(stakeholder_view.get("portfolio") or [])
    scorecard_count = len({
        (
            str(row.get("scorecard_ref") or row.get("scorecard_name") or ""),
            str(row.get("scorecard_name") or ""),
        )
        for row in rows
        if isinstance(row, Mapping)
    })
    totals = {
        "decision_evidence": 1,
        "stakeholder_workbook": 1,
        "score_briefs": len(_score_brief_portfolio_indexes(stakeholder_view)),
        "scorecard_summaries": scorecard_count,
        "scorecard_spreadsheets": scorecard_count,
        "scorecard_presentations": scorecard_count,
        "stakeholder_presentation": 1,
        "revision_manifest": 1,
    }
    return {
        kind: {"completed": 0, "total": int(totals[kind])}
        for kind in _ARTIFACT_PUBLICATION_KINDS
    }


def _primary_decision_category(row: Mapping[str, Any]) -> str:
    """Return the backend's exact disposition; never infer it from prose."""
    value = row.get("primary_disposition")
    return str(value) if isinstance(value, str) and value else "not_selected"


_ATTENTION_SEVERITY = {
    "promotion_ready": 0,
    "stakeholder_decision_required": 1,
    "failed_or_incomplete": 2,
    "awaiting_optimizer_review": 3,
    "awaiting_optimization_approval": 4,
    "stakeholder_clarification_required": 5,
    "guideline_or_code_repair": 6,
    "feedback_curation_review": 7,
    "optimization_in_progress": 8,
    "optimizer_launching": 9,
    "continue_optimization": 10,
    "targeted_feedback_collection": 11,
    "monitoring_or_diminishing_returns": 12,
    "cooldown": 13,
    "insufficient_evidence": 14,
    "not_selected": 15,
}


def build_stakeholder_presentation(
    stakeholder_view: Mapping[str, Any],
    *,
    scorecard_artifacts: list[Mapping[str, Any]],
    detail_status: str = "not_requested",
    detail_source_revision: Optional[int] = None,
) -> dict[str, Any]:
    """Build the safe, deterministic aggregate/card projection for the dashboard."""
    _validate_view(stakeholder_view)
    if detail_status not in {"not_requested", "pending", "complete", "stale"}:
        raise ValueError("unsupported stakeholder detail status")
    if detail_source_revision is not None and detail_source_revision < 1:
        raise ValueError("stakeholder detail source revision must be positive")
    rows = list(stakeholder_view.get("portfolio", []))
    primary_decision_mix: dict[str, int] = {}
    grouped: dict[tuple[str, str], list[Mapping[str, Any]]] = {}
    for row in rows:
        category = _primary_decision_category(row)
        primary_decision_mix[category] = primary_decision_mix.get(category, 0) + 1
        scorecard_name = str(row.get("scorecard_name") or "Unlabeled scorecard")
        scorecard_ref = str(row.get("scorecard_ref") or scorecard_name)
        grouped.setdefault((scorecard_ref, scorecard_name), []).append(row)

    secondary_issue_counts: dict[str, int] = {}
    for row in rows:
        for flag in row.get("secondary_issue_flags") or []:
            flag = str(flag)
            secondary_issue_counts[flag] = secondary_issue_counts.get(flag, 0) + 1

    artifacts_by_ref: dict[str, list[Mapping[str, Any]]] = {}
    for artifact in scorecard_artifacts:
        logical_id = str(artifact.get("logical_id") or "")
        scope_hash = logical_id.rsplit(":", 1)[-1]
        artifacts_by_ref.setdefault(scope_hash, []).append(dict(artifact))

    scorecards: list[dict[str, Any]] = []
    for (scorecard_ref, scorecard_name), score_rows in sorted(
        grouped.items(), key=lambda item: item[0][1].casefold()
    ):
        scope_hash = sha256(scorecard_ref.encode("utf-8")).hexdigest()[:16]
        decision_mix: dict[str, int] = {}
        for row in score_rows:
            category = _primary_decision_category(row)
            decision_mix[category] = decision_mix.get(category, 0) + 1
        scorecards.append({
            "scorecard_ref": scorecard_ref,
            "scorecard_name": scorecard_name,
            "score_count": len(score_rows),
            "primary_disposition_counts": decision_mix,
            # Compatibility alias for the first report presentation revision.
            "primary_decision_mix": decision_mix,
            "reviewed_error_opportunity": sum(
                float(row.get("reviewed_error_opportunity") or 0)
                for row in score_rows
                if isinstance(row.get("reviewed_error_opportunity"), (int, float))
            ),
            # A core revision has deliberately not uploaded drill-down
            # artifacts yet.  The dashboard must not render its empty
            # artifact list as a complete scorecard detail view.
            "detail_status": detail_status,
            "detail_source_revision": detail_source_revision,
            "artifacts": artifacts_by_ref.get(scope_hash, []),
        })

    overview = dict(stakeholder_view.get("overview") or {})
    try:
        priority_display_limit = max(0, int(overview.get("priority_display_limit") or 10))
    except (TypeError, ValueError):
        priority_display_limit = 10
    priorities = sorted(
        stakeholder_view.get("priorities", []),
        key=lambda row: float(row.get("opportunity") or 0)
        if isinstance(row.get("opportunity"), (int, float)) else 0.0,
        reverse=True,
    )[:priority_display_limit]
    opportunity_distribution = sorted(
        ({
            "evidence_rank": row.get("evidence_rank", row.get("rank")),
            "candidate_rank": row.get("candidate_rank"),
            "scorecard_name": row.get("scorecard_name"),
            "score_name": row.get("score_name"),
            "opportunity": row.get("reviewed_error_opportunity"),
            "valid_feedback_count": row.get("valid_feedback_count"),
            "disagreement_rate": row.get("disagreement_rate"),
            "review_disposition": row.get("review_disposition", "eligible_below_selection"),
            "policy_disposition": row.get("policy_disposition", "eligible"),
            "policy_reason": row.get("policy_reason", "meets_rank_policy"),
            "eligibility_timestamp": row.get("eligibility_timestamp"),
            "primary_disposition": _primary_decision_category(row),
            "secondary_issue_flags": list(row.get("secondary_issue_flags") or []),
            "next_action": row.get("next_action"),
            "dashboard_url": row.get("dashboard_url"),
        } for row in rows),
        key=lambda row: int(row.get("evidence_rank") or 10**9),
    )
    attention_queue = sorted(
        ({
            "scorecard_name": row.get("scorecard_name"),
            "score_name": row.get("score_name"),
            "primary_disposition": _primary_decision_category(row),
            "evidence_count": row.get("valid_feedback_count"),
            "severity": _ATTENTION_SEVERITY.get(_primary_decision_category(row), 99),
            "rationale": row.get("rationale"),
            "next_action": row.get("next_action"),
            "dashboard_url": row.get("dashboard_url"),
        } for row in rows),
        key=lambda row: (
            int(row["severity"]),
            -int(row.get("evidence_count") or 0),
            str(row.get("scorecard_name") or "").casefold(),
            str(row.get("score_name") or "").casefold(),
        ),
    )
    action_projection = build_action_projection(rows)
    guideline_code_conflict_workstream = build_guideline_code_conflict_workstream(
        [
            issue for issue in stakeholder_view.get("questions_and_issues", [])
            if isinstance(issue, Mapping)
        ]
    )
    return {
        "overview": overview,
        "score_count": len(rows),
        "scorecard_count": len(scorecards),
        "detail_status": detail_status,
        "detail_source_revision": detail_source_revision,
        "decision_summary": build_decision_summary(overview, primary_decision_mix),
        "action_counts": action_projection["action_counts"],
        "action_workstreams": action_projection["action_workstreams"],
        "guideline_code_conflict_workstream": guideline_code_conflict_workstream,
        "primary_disposition_counts": primary_decision_mix,
        # Compatibility alias for report views published before dispositions.
        "primary_decision_mix": primary_decision_mix,
        "secondary_issue_counts": secondary_issue_counts,
        "attention_queue": attention_queue,
        "opportunity_distribution": opportunity_distribution,
        "top_priorities": [dict(row) for row in priorities],
        "scorecards": scorecards,
        # Keep the operational presentation self-contained: the dashboard
        # reads this artifact directly rather than joining it back to the
        # workbook projection. ``contradictions`` remains a compatibility
        # alias for consumers that adopted the earlier terminology.
        "questions_and_issues": [
            dict(row) for row in stakeholder_view.get("questions_and_issues", [])
        ],
        "contradictions": [
            dict(row) for row in stakeholder_view.get("questions_and_issues", [])
        ],
        "optimization_outcomes": [
            dict(row) for row in stakeholder_view.get("optimization_outcomes", [])
        ],
    }


class OptimizationRunReportService:
    """Publish a periodic run to one stable Report and append-only revisions."""

    def __init__(
        self,
        *,
        client: Any,
        account_id: str,
        run_key: str,
        report_configuration_id: Optional[str] = None,
        dashboard_base_url: Optional[str] = None,
        existing_task: Any = None,
        now: Callable[[], datetime] = _utc_now,
        task_lookup: Optional[Callable[[str], Any]] = None,
        report_lookup: Optional[Callable[[Any], Any]] = None,
        block_lookup: Optional[Callable[[Any], list[Any]]] = None,
        stage_lookup: Optional[Callable[[Any], list[Any]]] = None,
        artifact_uploader: Optional[Callable[[str, str, bytes], str]] = None,
        artifact_store: Optional[Any] = None,
        verify_uploaded_artifacts: bool = True,
        attempt_id_factory: Callable[[], str] = lambda: str(uuid4()),
        publication_id_factory: Callable[[], str] = lambda: uuid4().hex,
    ) -> None:
        if not account_id or not run_key:
            raise ValueError("account_id and run_key are required")
        self.client = client
        self.account_id = account_id
        self.run_key = run_key
        self.report_configuration_id = report_configuration_id
        self.dashboard_base_url = _normalize_dashboard_base_url(dashboard_base_url)
        self._existing_task = existing_task
        self._uses_existing_task = existing_task is not None
        self.now = now
        self._task_lookup = task_lookup or self._find_task
        self._report_lookup = report_lookup or self._find_report
        self._block_lookup = block_lookup or self._find_blocks
        self._stage_lookup = stage_lookup or self._find_stages
        self._artifact_store = artifact_store
        if self._artifact_store is None and artifact_uploader is None:
            self._artifact_store = GraphQLArtifactStore(client)
        self._verify_uploaded_artifacts = bool(verify_uploaded_artifacts)
        self._attempt_id_factory = attempt_id_factory
        self._publication_id_factory = publication_id_factory
        self._artifact_uploader = artifact_uploader or self._upload_task_attachment
        self._state: Optional[OptimizationRunReportState] = None
        self._verified_artifact_reads: set[tuple[str, str, int, str]] = set()
        self._committed_artifact_index_cache: Optional[
            tuple[int, dict[str, dict[str, Any]]]
        ] = None

    def start_or_resume(self, run_spec: Mapping[str, Any]) -> OptimizationRunReportState:
        if not isinstance(run_spec, Mapping):
            raise ValueError("run_spec must be a mapping")
        # The run specification, not later decision evidence, defines the
        # execution contract.  Older direct callers did not supply a mode, so
        # freeze the conservative contract rather than infer it downstream.
        run_spec = dict(run_spec)
        execution_mode = run_spec.setdefault("execution_mode", "approval_required")
        if execution_mode not in {"automatic", "approval_required"}:
            raise ValueError(
                "run_spec.execution_mode must be 'automatic' or 'approval_required'"
            )
        run_spec["execution_candidate_policy"] = normalize_execution_candidate_policy(
            run_spec.get("execution_candidate_policy")
        )
        if self._state is not None:
            if not _same_run_spec(self._state.run_spec, run_spec):
                raise OptimizationRunIntegrityError(
                    "a run key cannot be reused with a different frozen run specification"
                )
            return self._state
        operator_identity = optimization_operator_identity(
            scope=run_spec.get("scope") if isinstance(run_spec, Mapping) else None,
        )
        try:
            task = (
                self._existing_task
                if self._uses_existing_task
                else self._task_lookup(self.run_key)
            )
        except Exception as exc:
            raise OptimizationRunRetryablePublicationError(
                "Could not look up the optimization run Task"
            ) from exc
        created_new_attempt = False
        predecessor: dict[str, Any] = {}
        recovered_linked_report: Any = None
        if task is not None:
            existing_metadata = _metadata(getattr(task, "metadata", {}))
            existing_run_key = str(existing_metadata.get("optimization_run_key") or "")
            if self._uses_existing_task and existing_run_key and existing_run_key != self.run_key:
                raise ValueError("existing Procedure Task is already claimed by another optimization run")
            if self._uses_existing_task and not existing_metadata.get("optimization_run_key"):
                if str(getattr(task, "accountId", self.account_id)) != self.account_id:
                    raise ValueError("existing Procedure Task belongs to a different account")
                try:
                    recovered_linked_report = self._report_lookup(task)
                except Exception as exc:
                    raise OptimizationRunRetryablePublicationError(
                        "Could not look up the Task-linked optimization Report"
                    ) from exc
                linked_run = (
                    _metadata(getattr(recovered_linked_report, "parameters", {}))
                    .get("optimization_run")
                    if recovered_linked_report is not None
                    else None
                )
                if recovered_linked_report is not None:
                    if not isinstance(linked_run, Mapping):
                        raise OptimizationRunIntegrityError(
                            "Task-linked Report is missing optimization run identity"
                        )
                    linked_spec = linked_run.get("run_spec")
                    linked_attempt_id = linked_run.get("attempt_id")
                    if (
                        linked_run.get("run_key") != self.run_key
                        or not isinstance(linked_attempt_id, str)
                        or not linked_attempt_id
                        or not isinstance(linked_spec, Mapping)
                        or not _same_run_spec(linked_spec, run_spec)
                    ):
                        raise OptimizationRunIntegrityError(
                            "Task-linked Report conflicts with the requested run identity"
                        )
                    attempt_id = linked_attempt_id
                    existing_metadata.update({
                        "optimization_run_key": self.run_key,
                        "attempt_id": attempt_id,
                        "lifecycle_version": linked_run.get("lifecycle_version") or LIFECYCLE_VERSION,
                        "run_spec": dict(linked_spec),
                        "operator_identity": linked_run.get("operator_identity") or operator_identity.as_dict(),
                    })
                else:
                    attempt_id = self._attempt_id_factory()
                    if not isinstance(attempt_id, str) or not attempt_id:
                        raise ValueError("attempt_id_factory must return a nonempty string")
                    existing_metadata.update({
                        "optimization_run_key": self.run_key,
                        "attempt_id": attempt_id,
                        "lifecycle_version": LIFECYCLE_VERSION,
                        "run_spec": dict(run_spec),
                        "operator_identity": operator_identity.as_dict(),
                    })
                    created_new_attempt = True
                try:
                    task.update(
                        metadata=json.dumps(existing_metadata),
                        description=(
                            f"{operator_identity.display_title} — "
                            f"{operator_identity.display_scope}"
                        ),
                    )
                except Exception as exc:
                    raise OptimizationRunRetryablePublicationError(
                        "Could not claim the existing Procedure Task"
                    ) from exc
            existing_spec = existing_metadata.get("run_spec")
            if not created_new_attempt and (
                not isinstance(existing_spec, Mapping)
                or not _same_run_spec(existing_spec, run_spec)
            ):
                message = "existing Task does not preserve the exact frozen run specification"
                self._fail_uninitialized_integrity(task, message)
                raise OptimizationRunIntegrityError(message)
            final_status = str(existing_metadata.get("optimization_run_final_status") or "").lower()
            if str(getattr(task, "status", "")).upper() == "FAILED" or final_status in {"failed", "blocked"}:
                if self._uses_existing_task:
                    raise ValueError("a terminal Procedure Task cannot be reused for another optimization attempt")
                try:
                    previous_report = self._report_lookup(task)
                except Exception as exc:
                    raise OptimizationRunRetryablePublicationError(
                        "Could not look up the previous optimization Report"
                    ) from exc
                predecessor = {
                    "previous_attempt_id": existing_metadata.get("attempt_id"),
                    "previous_task_id": task.id,
                    "previous_report_id": getattr(previous_report, "id", None),
                }
                task = None
        if task is None:
            attempt_id = self._attempt_id_factory()
            if not isinstance(attempt_id, str) or not attempt_id:
                raise ValueError("attempt_id_factory must return a nonempty string")
            task_metadata = {
                "optimization_run_key": self.run_key,
                "attempt_id": attempt_id,
                "lifecycle_version": LIFECYCLE_VERSION,
                "run_spec": dict(run_spec),
                "operator_identity": operator_identity.as_dict(),
                **{key: value for key, value in predecessor.items() if value},
            }
            try:
                task = Task.create(
                    client=self.client, accountId=self.account_id, type="OptimizationRunReport",
                    target=f"optimization/run/{self.run_key}", command="optimization portfolio run",
                    description=(
                        f"{operator_identity.display_title} — {operator_identity.display_scope}"
                    ),
                    status="RUNNING", dispatchStatus="LOCAL",
                    startedAt=_iso(self.now()), metadata=json.dumps(task_metadata),
                )
            except Exception as exc:
                raise OptimizationRunRetryablePublicationError(
                    "Could not create the optimization run Task"
                ) from exc
            created_new_attempt = True
        else:
            task_metadata = _metadata(getattr(task, "metadata", {}))
            attempt_id = str(task_metadata.get("attempt_id") or "")
            if not attempt_id:
                raise ValueError("existing optimization run attempt is missing attempt_id")
        report = None
        try:
            # Publish the operator-facing cover before the fixed TaskStages and
            # ReportBlocks incur their sequential remote setup cost.
            report = recovered_linked_report or (
                None if created_new_attempt else self._report_lookup(task)
            )
            if report is None:
                config_id = self.report_configuration_id or _get_programmatic_config_id(self.account_id, self.client)
                parameters = {
                    "_display_title": operator_identity.display_title,
                    "_display_subtitle": self._display_subtitle(operator_identity),
                    "optimization_run": {
                    "run_key": self.run_key,
                    "attempt_id": attempt_id,
                    "lifecycle_version": LIFECYCLE_VERSION,
                    "run_spec": dict(run_spec),
                    "operator_identity": operator_identity.as_dict(),
                    "latest_revision": None,
                    "revisions": [],
                    **{key: value for key, value in predecessor.items() if value},
                }}
                report = Report.create(
                    client=self.client, accountId=self.account_id, taskId=task.id,
                    name=operator_identity.display_title,
                    reportConfigurationId=config_id,
                    parameters=parameters,
                    output=self._render_report_manifest(
                        "running", None, identity=operator_identity,
                        execution_mode=run_spec.get("execution_mode"),
                    ),
                )
            elif not created_new_attempt:
                report_spec = (
                    (_metadata(getattr(report, "parameters", {})).get("optimization_run") or {})
                    .get("run_spec")
                )
                if not isinstance(report_spec, Mapping) or not _same_run_spec(
                    report_spec, run_spec
                ):
                    message = "existing Report does not preserve the exact frozen run specification"
                    self._fail_uninitialized_integrity(task, message)
                    raise OptimizationRunIntegrityError(message)
            stages = (
                list(self._stage_lookup(task))
                if self._uses_existing_task
                else self._ensure_fixed_stages(task)
            )
            blocks = self._ensure_fixed_blocks(report)
            self._state = OptimizationRunReportState(
                task=task, report=report, blocks=blocks,
                stages={
                    str(getattr(stage, "name", "")).strip().lower(): stage
                    for stage in stages
                },
                run_spec=dict(run_spec), attempt_id=attempt_id,
                operator_identity=operator_identity,
            )
            running_stage = next(
                (stage for stage in stages if str(getattr(stage, "status", "")).upper() == "RUNNING"),
                stages[0] if stages else None,
            )
            if (
                not self._uses_existing_task
                and running_stage is not None
                and getattr(task, "currentStageId", None) != running_stage.id
            ):
                task.update(currentStageId=running_stage.id)
            self._ensure_initial_envelopes(self._state)
        except OptimizationRunIntegrityError as exc:
            try:
                if self._state is not None:
                    self.fail(f"Optimization run integrity failure: {exc}")
                else:
                    self._fail_uninitialized_integrity(task, str(exc))
            except Exception:
                pass
            raise
        except Exception as exc:
            # Remote auth and write failures can occur after the Task/Report
            # identity has been allocated.  No committed revision exists yet,
            # so preserve that attempt for request replay.
            raise OptimizationRunRetryablePublicationError(
                "Could not initialize optimization run report"
            ) from exc
        return self._state

    def _latest_committed_artifact_index(
        self,
    ) -> dict[str, dict[str, Any]]:
        """Load the last committed manifest as the only artifact-reuse authority."""
        state = self._require_state()
        latest = self._latest_revision(state.report)
        if latest is None:
            return {}
        if not isinstance(latest, Mapping):
            raise OptimizationRunIntegrityError(
                "Latest optimization revision metadata is malformed"
            )
        revision_number = int(latest.get("number") or 0)
        if (
            self._committed_artifact_index_cache is not None
            and self._committed_artifact_index_cache[0] == revision_number
        ):
            return dict(self._committed_artifact_index_cache[1])
        if self._artifact_store is None:
            return {}

        manifest_descriptor = latest.get("manifest")
        if not isinstance(manifest_descriptor, Mapping):
            raise OptimizationRunIntegrityError(
                "Latest optimization revision is missing its manifest descriptor"
            )
        try:
            manifest_bytes = self._download_task_attachment(
                state.task.id, manifest_descriptor,
            )
            manifest = json.loads(manifest_bytes.decode("utf-8"))
        except OptimizationRunIntegrityError:
            raise
        except Exception as exc:
            raise OptimizationRunRetryablePublicationError(
                "Could not load the latest committed artifact manifest"
            ) from exc
        if not isinstance(manifest, Mapping):
            raise OptimizationRunIntegrityError(
                "Latest optimization artifact manifest is malformed"
            )
        if int(manifest.get("revision") or 0) != revision_number:
            raise OptimizationRunIntegrityError(
                "Latest optimization artifact manifest has the wrong revision"
            )
        if str(manifest.get("milestone") or "") != str(latest.get("milestone") or ""):
            raise OptimizationRunIntegrityError(
                "Latest optimization artifact manifest has the wrong milestone"
            )

        index: dict[str, dict[str, Any]] = {}
        for descriptor in manifest.get("artifacts") or []:
            if not isinstance(descriptor, Mapping):
                raise OptimizationRunIntegrityError(
                    "Latest optimization artifact manifest contains an invalid descriptor"
                )
            logical_id = str(descriptor.get("logical_id") or "")
            if not logical_id or logical_id in index:
                raise OptimizationRunIntegrityError(
                    "Latest optimization artifact manifest has invalid logical identities"
                )
            index[logical_id] = dict(descriptor)

        # A finalization core manifest remains immutable after it is
        # committed.  Its separately immutable detail manifest is the only
        # authority for score-level reuse after a process restart.  Without
        # loading it here, a fresh worker would needlessly republish every
        # score/scorecard artifact despite having verified, completed details.
        if latest.get("detail_status") == "complete":
            detail_descriptor = latest.get("detail_manifest")
            if not isinstance(detail_descriptor, Mapping):
                raise OptimizationRunIntegrityError(
                    "Completed optimization detail enrichment is missing its manifest"
                )
            try:
                detail_bytes = self._download_task_attachment(
                    state.task.id, detail_descriptor,
                )
                detail_manifest = json.loads(detail_bytes.decode("utf-8"))
            except OptimizationRunIntegrityError:
                raise
            except Exception as exc:
                raise OptimizationRunRetryablePublicationError(
                    "Could not load the latest completed optimization detail manifest"
                ) from exc
            if not isinstance(detail_manifest, Mapping):
                raise OptimizationRunIntegrityError(
                    "Latest optimization detail manifest is malformed"
                )
            if int(detail_manifest.get("source_revision") or 0) != revision_number:
                raise OptimizationRunIntegrityError(
                    "Latest optimization detail manifest has the wrong source revision"
                )
            if str(detail_manifest.get("milestone") or "") != str(
                latest.get("milestone") or ""
            ):
                raise OptimizationRunIntegrityError(
                    "Latest optimization detail manifest has the wrong milestone"
                )
            for descriptor in detail_manifest.get("artifacts") or []:
                if not isinstance(descriptor, Mapping):
                    raise OptimizationRunIntegrityError(
                        "Latest optimization detail manifest contains an invalid descriptor"
                    )
                logical_id = str(descriptor.get("logical_id") or "")
                if not logical_id or logical_id in index:
                    raise OptimizationRunIntegrityError(
                        "Latest optimization detail manifest has invalid logical identities"
                    )
                index[logical_id] = dict(descriptor)
        self._committed_artifact_index_cache = (revision_number, index)
        return dict(index)

    def _reuse_committed_artifact(
        self,
        artifact_index: Mapping[str, Mapping[str, Any]],
        *,
        logical_id: str,
        kind: str,
        content_type: str,
        content: bytes,
    ) -> Optional[dict[str, Any]]:
        """Return a checksum-identical committed artifact after safe verification."""
        state = self._require_state()
        candidate = artifact_index.get(logical_id)
        if not isinstance(candidate, Mapping):
            return None
        digest = sha256(content).hexdigest()
        expected = {
            "logical_id": logical_id,
            "kind": kind,
            "content_type": content_type,
            "size_bytes": len(content),
            "sha256": digest,
            "task_id": state.task.id,
        }
        if any(candidate.get(field) != value for field, value in expected.items()):
            return None
        object_key = str(candidate.get("object_key") or "")
        source_revision = int(candidate.get("source_revision") or 0)
        if not object_key or source_revision < 1:
            return None
        verification_key = (object_key, digest, len(content), content_type)
        if verification_key not in self._verified_artifact_reads:
            try:
                downloaded = self._download_task_attachment(state.task.id, candidate)
            except Exception:
                return None
            if downloaded != content:
                return None
        return dict(candidate)

    def _reuse_uncommitted_task_attachment(
        self,
        *,
        logical_id: str,
        kind: str,
        content_type: str,
        content: bytes,
        filename: str,
        publication_id: str,
        source_revision: int,
    ) -> Optional[dict[str, Any]]:
        """Recover a verified artifact left by an interrupted publication."""
        state = self._require_state()
        stem, separator, extension = filename.rpartition(".")
        marker = f"-{publication_id}"
        if not separator or not stem.endswith(marker):
            return None
        stable_prefix = stem[: -len(marker)] + "-"
        stable_suffix = f".{extension}"
        digest = sha256(content).hexdigest()
        for attachment in reversed(list(getattr(state.task, "attachedFiles", None) or [])):
            object_key = str(attachment or "")
            candidate_name = object_key.rsplit("/", 1)[-1]
            if not (
                candidate_name.startswith(stable_prefix)
                and candidate_name.endswith(stable_suffix)
            ):
                continue
            candidate = {
                "logical_id": logical_id,
                "kind": kind,
                "content_type": content_type,
                "size_bytes": len(content),
                "sha256": digest,
                "task_id": state.task.id,
                "object_key": object_key,
                "source_revision": source_revision,
            }
            try:
                downloaded = self._download_task_attachment(state.task.id, candidate)
            except Exception:
                continue
            if downloaded == content:
                return candidate
        return None

    def _publication_generated_at(
        self,
        *,
        revision_number: int,
        milestone: str,
        evidence_checksum: str,
        stakeholder_view_checksum: str,
    ) -> datetime:
        """Freeze workbook time so an interrupted revision is byte-replayable."""
        state = self._require_state()
        task_metadata = _metadata(getattr(state.task, "metadata", {}))
        draft = task_metadata.get("optimization_publication_draft")
        expected = {
            "revision": revision_number,
            "milestone": milestone,
            "evidence_checksum": evidence_checksum,
            "stakeholder_view_checksum": stakeholder_view_checksum,
        }
        if isinstance(draft, Mapping) and all(
            draft.get(field) == value for field, value in expected.items()
        ):
            frozen_value = str(draft.get("generated_at") or "")
            try:
                frozen = datetime.fromisoformat(frozen_value.replace("Z", "+00:00"))
                if frozen.tzinfo is not None:
                    return frozen.astimezone(timezone.utc)
            except ValueError:
                pass

        generated_at = self.now().astimezone(timezone.utc)
        task_metadata["optimization_publication_draft"] = {
            **expected,
            "generated_at": _iso(generated_at),
        }
        state.task.update(metadata=json.dumps(task_metadata))
        return generated_at

    def _load_publication_draft_artifacts(
        self,
        *,
        revision_number: int,
        milestone: str,
        evidence_checksum: str,
        stakeholder_view_checksum: str,
        generated_at: datetime,
    ) -> dict[str, dict[str, Any]]:
        """Recover the newest checksum-valid partial score-artifact checkpoint."""
        state = self._require_state()
        if self._artifact_store is None:
            return {}
        pattern = re.compile(
            rf"^optimization-publication-draft-r{revision_number:04d}-.+"
            r"-c\d{6}-s(\d+)-h([0-9a-f]{64})\.json$"
        )
        for attachment in reversed(list(getattr(state.task, "attachedFiles", None) or [])):
            object_key = str(attachment or "")
            filename = object_key.rsplit("/", 1)[-1]
            match = pattern.fullmatch(filename)
            if match is None:
                continue
            descriptor = {
                "task_id": state.task.id,
                "object_key": object_key,
                "content_type": "application/json",
                "size_bytes": int(match.group(1)),
                "sha256": match.group(2),
            }
            try:
                payload = json.loads(
                    self._download_task_attachment(state.task.id, descriptor).decode("utf-8")
                )
            except Exception:
                continue
            if not isinstance(payload, Mapping):
                continue
            if (
                int(payload.get("revision") or 0) != revision_number
                or str(payload.get("milestone") or "") != milestone
                or str(payload.get("evidence_checksum") or "") != evidence_checksum
                or str(payload.get("stakeholder_view_checksum") or "")
                != stakeholder_view_checksum
                or str(payload.get("generated_at") or "") != _iso(generated_at)
            ):
                continue
            index: dict[str, dict[str, Any]] = {}
            valid = True
            for artifact in payload.get("artifacts") or []:
                if not isinstance(artifact, Mapping):
                    valid = False
                    break
                logical_id = str(artifact.get("logical_id") or "")
                if not logical_id or logical_id in index:
                    valid = False
                    break
                object_key = str(artifact.get("object_key") or "")
                digest = str(artifact.get("sha256") or "")
                content_type = str(artifact.get("content_type") or "")
                size_bytes = artifact.get("size_bytes")
                if (
                    not object_key
                    or artifact.get("task_id") != state.task.id
                    or not isinstance(size_bytes, int)
                    or size_bytes < 0
                    or re.fullmatch(r"[0-9a-f]{64}", digest) is None
                    or not content_type
                ):
                    valid = False
                    break
                index[logical_id] = dict(artifact)
            if valid:
                # Every draft descriptor is added only after upload read-back
                # or checksum-identical reuse, and this draft manifest has just
                # passed its own size/checksum verification. Preserve that
                # durable proof across process replay instead of issuing one
                # GraphQL ticket and HTTPS read per already-verified artifact.
                for artifact in index.values():
                    self._verified_artifact_reads.add((
                        str(artifact["object_key"]),
                        str(artifact["sha256"]),
                        int(artifact["size_bytes"]),
                        str(artifact["content_type"]),
                    ))
                return index
        return {}

    def _persist_publication_draft_artifacts(
        self,
        *,
        revision_number: int,
        milestone: str,
        publication_id: str,
        evidence_checksum: str,
        stakeholder_view_checksum: str,
        generated_at: datetime,
        artifacts: Mapping[str, Mapping[str, Any]],
    ) -> None:
        """Durably index partial score artifacts without attaching every child."""
        state = self._require_state()
        payload = {
            "schema_version": "optimization-publication-draft-v1",
            "revision": revision_number,
            "milestone": milestone,
            "evidence_checksum": evidence_checksum,
            "stakeholder_view_checksum": stakeholder_view_checksum,
            "generated_at": _iso(generated_at),
            "artifacts": [
                dict(artifacts[logical_id]) for logical_id in sorted(artifacts)
            ],
        }
        content = _json(payload)
        digest = sha256(content).hexdigest()
        filename = (
            f"optimization-publication-draft-r{revision_number:04d}-{publication_id}"
            f"-c{len(artifacts):06d}-s{len(content)}-h{digest}.json"
        )
        path = self._artifact_uploader(state.task.id, filename, content)
        self._attach_task_file(state.task, path)

    def publish_milestone(
        self, milestone: str, decision_evidence: Mapping[str, Any], *, stakeholder_view: Mapping[str, Any]
    ) -> PublishedRevision:
        state = self._require_state()
        if not milestone or not isinstance(decision_evidence, Mapping):
            raise ValueError("milestone and decision_evidence mapping are required")
        revision_number = self._latest_revision_number(state.report) + 1
        publication_counts: dict[str, dict[str, int]] = {}
        active_artifact_kind = "revision_manifest"
        active_operation_category = "prepare_core"
        publication_started_at: Optional[datetime] = None
        core_committed = False

        def _publication_label(kind: str) -> str:
            return kind.replace("_", " ")

        def _set_active_artifact_kind(kind: str) -> None:
            nonlocal active_artifact_kind
            if kind not in _ARTIFACT_PUBLICATION_KINDS:
                raise ValueError("unsupported artifact publication kind")
            active_artifact_kind = kind

        def _publish_artifact_progress(
            kind: str,
            *,
            completed: bool = False,
            progress_state: str = "active",
            failure_class: str | None = None,
        ) -> None:
            """Publish one compact, safe live status without an evidence revision."""
            if kind not in publication_counts:
                raise ValueError("unsupported artifact publication kind")
            if completed:
                publication_counts[kind]["completed"] = publication_counts[kind]["total"]
            elapsed = 0
            if publication_started_at is not None:
                elapsed = max(0, int((self.now() - publication_started_at).total_seconds()))
            incomplete = progress_state in {"failed", "incomplete"}
            message = (
                f"Could not publish the {milestone} milestone's {_publication_label(kind)}; "
                "the milestone is incomplete."
                if incomplete
                else f"Publishing {milestone} milestone artifacts: {_publication_label(kind)}."
            )
            current_count = sum(item["completed"] for item in publication_counts.values())
            total_count = sum(item["total"] for item in publication_counts.values())
            failure = None
            if failure_class is not None:
                safe_failure_class = (
                    failure_class
                    if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]{0,127}", failure_class)
                    else "Exception"
                )
                failure = {
                    "exception_class": safe_failure_class,
                    "operation_category": kind,
                    "retry_classification": "retryable",
                    "completed": current_count,
                    "total": total_count,
                }
            self.publish_progress(
                phase="publication",
                current=current_count,
                total=total_count,
                unit="artifacts",
                state=progress_state,
                elapsed_seconds=elapsed,
                next_checkpoint=f"Publishing the {milestone} milestone.",
                message=message,
                artifact_counts={
                    safe_kind: {"completed": count["completed"], "total": count["total"]}
                    for safe_kind, count in publication_counts.items()
                },
                failure=failure,
            )

        def _record_artifact_upload(kind: str, completed: bool) -> None:
            _set_active_artifact_kind(kind)
            if not completed:
                return
            count = publication_counts[kind]
            count["completed"] += 1
            if (
                count["completed"] % 25 == 0
                or count["completed"] == count["total"]
            ):
                _publish_artifact_progress(kind)

        try:
            active_operation_category = "decision_projection"
            publication_id = self._publication_id_factory()
            if (
                not isinstance(publication_id, str)
                or not publication_id
                or any(not (character.isalnum() or character in "-_.") for character in publication_id)
            ):
                raise ValueError(
                    "publication_id_factory must return a filesystem-safe nonempty string"
                )
            stakeholder_view = _stakeholder_execution_projection(
                stakeholder_view,
                decision_evidence,
                expected_execution_mode=str(state.run_spec.get("execution_mode") or ""),
            )
            active_operation_category = "stakeholder_view_validation"
            _validate_view(stakeholder_view)
            active_operation_category = "artifact_reuse_index"
            committed_artifacts = self._latest_committed_artifact_index()
            draft_artifacts: dict[str, dict[str, Any]] = {}

            def _reuse_artifact(
                logical_id: str,
                kind: str,
                content_type: str,
                content: bytes,
                filename: str,
            ) -> Optional[Mapping[str, Any]]:
                committed = self._reuse_committed_artifact(
                    committed_artifacts,
                    logical_id=logical_id,
                    kind=kind,
                    content_type=content_type,
                    content=content,
                )
                if committed is not None:
                    return committed
                draft = self._reuse_committed_artifact(
                    draft_artifacts,
                    logical_id=logical_id,
                    kind=kind,
                    content_type=content_type,
                    content=content,
                )
                if draft is not None:
                    return draft
                return self._reuse_uncommitted_task_attachment(
                    logical_id=logical_id,
                    kind=kind,
                    content_type=content_type,
                    content=content,
                    filename=filename,
                    publication_id=publication_id,
                    source_revision=revision_number,
                )

            publication_counts = _artifact_publication_plan(stakeholder_view)
            if milestone != "finalization":
                for detail_kind in (
                    "score_briefs",
                    "scorecard_summaries",
                    "scorecard_spreadsheets",
                    "scorecard_presentations",
                ):
                    publication_counts[detail_kind] = {"completed": 0, "total": 0}
            publication_started_at = self.now()
            safe_overview = dict(stakeholder_view.get("overview") or {})
            raw_evidence = _json(decision_evidence)
            evidence_checksum = sha256(raw_evidence).hexdigest()
            stakeholder_view_checksum = sha256(_json(stakeholder_view)).hexdigest()
            active_operation_category = "publication_checkpoint"
            generated_at = self._publication_generated_at(
                revision_number=revision_number,
                milestone=milestone,
                evidence_checksum=evidence_checksum,
                stakeholder_view_checksum=stakeholder_view_checksum,
            )
            active_operation_category = "draft_recovery"
            draft_artifacts = self._load_publication_draft_artifacts(
                revision_number=revision_number,
                milestone=milestone,
                evidence_checksum=evidence_checksum,
                stakeholder_view_checksum=stakeholder_view_checksum,
                generated_at=generated_at,
            )
            active_operation_category = "stakeholder_workbook"
            active_artifact_kind = "stakeholder_workbook"
            _publish_artifact_progress("stakeholder_workbook")
            workbook = build_stakeholder_workbook(stakeholder_view, revision_number=revision_number, generated_at=generated_at)
            active_operation_category = "decision_evidence"
            active_artifact_kind = "decision_evidence"
            _publish_artifact_progress("decision_evidence")
            raw_filename = (
                f"optimization-evidence-r{revision_number:04d}-{publication_id}.json"
            )
            reusable_raw = _reuse_artifact(
                "run_evidence", "run_evidence", "application/json",
                raw_evidence, raw_filename,
            )
            if isinstance(reusable_raw, Mapping):
                raw_path = str(reusable_raw.get("object_key") or "")
                raw_source_revision = int(reusable_raw.get("source_revision") or 0)
            else:
                raw_path = self._artifact_uploader(
                    state.task.id, raw_filename, raw_evidence,
                )
                raw_source_revision = revision_number
            if not raw_path or raw_source_revision < 1:
                raise OptimizationRunIntegrityError(
                    "decision evidence artifact descriptor is incomplete"
                )
            self._attach_task_file(state.task, raw_path)
            _publish_artifact_progress("decision_evidence", completed=True)
            active_artifact_kind = "stakeholder_workbook"
            active_operation_category = "stakeholder_workbook"
            workbook_filename = (
                f"optimization-workbook-r{revision_number:04d}-{publication_id}.xlsx"
            )
            reusable_workbook = _reuse_artifact(
                "stakeholder_workbook",
                "stakeholder_workbook",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                workbook.content,
                workbook_filename,
            )
            if isinstance(reusable_workbook, Mapping):
                workbook_path = str(reusable_workbook.get("object_key") or "")
                workbook_source_revision = int(
                    reusable_workbook.get("source_revision") or 0
                )
            else:
                workbook_path = self._artifact_uploader(
                    state.task.id, workbook_filename, workbook.content,
                )
                workbook_source_revision = revision_number
            if not workbook_path or workbook_source_revision < 1:
                raise OptimizationRunIntegrityError(
                    "stakeholder workbook artifact descriptor is incomplete"
                )
            self._attach_task_file(state.task, workbook_path)
            _publish_artifact_progress("stakeholder_workbook", completed=True)
            artifacts = [
                _artifact_descriptor(
                    logical_id="run_evidence",
                    kind="run_evidence",
                    display_name="Decision evidence",
                    scope="run",
                    content_type="application/json",
                    content=raw_evidence,
                    object_key=raw_path,
                    task_id=state.task.id,
                    source_revision=raw_source_revision,
                ),
                _artifact_descriptor(
                    logical_id="stakeholder_workbook",
                    kind="stakeholder_workbook",
                    display_name="Stakeholder workbook",
                    scope="run",
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    content=workbook.content,
                    object_key=workbook_path,
                    task_id=state.task.id,
                    source_revision=workbook_source_revision,
                ),
            ]
            if milestone == "finalization":
                # This safe, immutable input lets a restarted local worker
                # resume detail enrichment without rebuilding or re-running
                # any analysis. It is deliberately not a scorecard artifact.
                detail_input = _json({
                    "schema_version": "optimization-run-detail-input-v1",
                    "source_revision": revision_number,
                    "milestone": milestone,
                    "evidence_checksum": evidence_checksum,
                    "stakeholder_view_checksum": stakeholder_view_checksum,
                    "stakeholder_view": stakeholder_view,
                })
                detail_input_path = self._artifact_uploader(
                    state.task.id,
                    f"optimization-detail-input-r{revision_number:04d}-{publication_id}.json",
                    detail_input,
                )
                self._attach_task_file(state.task, detail_input_path)
                artifacts.append(_artifact_descriptor(
                    logical_id="detail_input",
                    kind="detail_input",
                    display_name="Final detail enrichment input",
                    scope="run",
                    content_type="application/json",
                    content=detail_input,
                    object_key=detail_input_path,
                    task_id=state.task.id,
                    source_revision=revision_number,
                ))
            # Core publication deliberately excludes per-score and per-scorecard
            # material.  Ranking, assessment, diagnosis, and optimizer launch
            # can continue as soon as this compact revision is authoritative.
            active_artifact_kind = "stakeholder_presentation"
            active_operation_category = "stakeholder_presentation"
            _publish_artifact_progress("stakeholder_presentation")
            core_detail_status = "pending"
            presentation = build_stakeholder_presentation(
                stakeholder_view,
                scorecard_artifacts=[],
                detail_status=core_detail_status,
                detail_source_revision=None,
            )
            presentation_bytes = _json(presentation)
            reusable_presentation = _reuse_artifact(
                "stakeholder_presentation",
                "stakeholder_presentation",
                "application/json",
                presentation_bytes,
                f"optimization-presentation-r{revision_number:04d}-{publication_id}.json",
            )
            if isinstance(reusable_presentation, Mapping):
                presentation_path = str(reusable_presentation.get("object_key") or "")
                presentation_source_revision = int(
                    reusable_presentation.get("source_revision") or 0
                )
                if not presentation_path or presentation_source_revision < 1:
                    raise OptimizationRunIntegrityError(
                        "reusable stakeholder presentation descriptor is incomplete"
                    )
            else:
                presentation_path = self._artifact_uploader(
                    state.task.id,
                    f"optimization-presentation-r{revision_number:04d}-{publication_id}.json",
                    presentation_bytes,
                )
                presentation_source_revision = revision_number
            self._attach_task_file(state.task, presentation_path)
            _publish_artifact_progress("stakeholder_presentation", completed=True)
            presentation_artifact = _artifact_descriptor(
                logical_id="stakeholder_presentation",
                kind="stakeholder_presentation",
                display_name="Stakeholder presentation data",
                scope="run",
                content_type="application/json",
                content=presentation_bytes,
                object_key=presentation_path,
                task_id=state.task.id,
                source_revision=presentation_source_revision,
            )
            artifacts.append(presentation_artifact)
            if self.dashboard_base_url:
                for artifact in artifacts:
                    artifact["dashboard_url"] = _report_artifact_url(
                        self.dashboard_base_url,
                        report_id=state.report.id,
                        revision_number=revision_number,
                        logical_id=str(artifact["logical_id"]),
                    )
            manifest = {
                "revision": revision_number, "milestone": milestone, "published_at": _iso(generated_at),
                "coverage_complete": bool((decision_evidence.get("coverage") or {}).get("complete", decision_evidence.get("coverage_complete", False))),
                "evidence_checksum": evidence_checksum, "workbook_checksum": workbook.checksum,
                "workbook_path": workbook_path, "row_counts": dict(workbook.row_counts),
                "scorecard_count": 0,
                "score_count": len(stakeholder_view.get("portfolio", [])),
                "artifacts": artifacts,
                "overview": safe_overview,
                "detail_status": core_detail_status,
                "detail_source_revision": None,
            }
            manifest_bytes = _json(manifest)
            manifest_checksum = sha256(manifest_bytes).hexdigest()
            active_artifact_kind = "revision_manifest"
            active_operation_category = "revision_manifest"
            _publish_artifact_progress("revision_manifest")
            manifest_path = self._artifact_uploader(
                state.task.id,
                f"optimization-revision-r{revision_number:04d}-{publication_id}.json",
                manifest_bytes,
            )
            self._attach_task_file(state.task, manifest_path)
            _publish_artifact_progress("revision_manifest", completed=True)
            self._update_block(state.blocks["evidence"], raw_path, {"revision": revision_number, "milestone": milestone, "checksum": evidence_checksum})
            self._update_block(state.blocks["workbook"], workbook_path, {"revision": revision_number, "milestone": milestone, "checksum": workbook.checksum, "row_counts": dict(workbook.row_counts)})
            self._update_block(state.blocks["status"], manifest_path, {
                "type": "optimization_run_status",
                "status": "published",
                "summary": {
                    "revision": revision_number,
                    "milestone": milestone,
                    "overview": safe_overview,
                    "presentation": presentation_artifact,
                    "detail_status": manifest["detail_status"],
                    "detail_source_revision": None,
                },
            })
            revision = PublishedRevision(
                number=revision_number,
                milestone=milestone,
                published_at=_iso(generated_at),
                raw_evidence_path=raw_path,
                workbook_path=workbook_path,
                manifest_path=manifest_path,
                manifest_checksum=manifest_checksum,
                manifest_size_bytes=len(manifest_bytes),
                evidence_checksum=evidence_checksum,
                workbook_checksum=workbook.checksum,
                row_counts=workbook.row_counts,
                overview=safe_overview,
                artifacts=tuple(artifacts),
                detail_status=manifest["detail_status"],
            )
            self._record_latest_revision(state, revision)
            self._committed_artifact_index_cache = (
                revision_number,
                {
                    str(artifact["logical_id"]): dict(artifact)
                    for artifact in artifacts
                },
            )
            # TaskStage is the dashboard's lifecycle projection. Advance it
            # only after the immutable evidence, workbook, block pointers, and
            # Report cover page have all been made durable.
            self._advance_stage_for_milestone(milestone)
            core_committed = True
            if milestone != "finalization":
                return revision
            return self._publish_final_detail_enrichment(
                core_revision=revision,
                stakeholder_view=stakeholder_view,
                committed_artifacts=committed_artifacts,
                publication_id=publication_id,
                evidence_checksum=evidence_checksum,
                stakeholder_view_checksum=stakeholder_view_checksum,
                generated_at=generated_at,
            )
        except OptimizationRunIntegrityError as exc:
            try:
                self.fail(f"Optimization run integrity failure: {exc}")
            except Exception:
                pass
            raise
        except Exception as exc:
            if not core_committed:
                # The procedure must never continue while its authoritative
                # stakeholder state is behind.  Partial core attachments are
                # uncommitted evidence, so preserve them for audit but fail the
                # attempt instead of scheduling an unbounded publication loop.
                try:
                    if publication_counts:
                        _publish_artifact_progress(
                            active_artifact_kind,
                            progress_state="failed",
                            failure_class=type(exc).__name__,
                        )
                except Exception:
                    pass
                try:
                    self.fail(
                        f"Core optimization Report publication failed during {milestone}"
                    )
                except Exception:
                    pass
                safe_exception_class = type(exc).__name__
                if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]{0,127}", safe_exception_class) is None:
                    safe_exception_class = "Exception"
                raise OptimizationRunPublicationError(
                    f"Could not publish core optimization milestone {milestone} "
                    f"(operation={active_operation_category}, error={safe_exception_class})"
                ) from exc
            # The immutable finalization core is already authoritative.  A
            # later detail failure must preserve optimizer evidence and remain
            # retryable until stakeholder drill-down artifacts reconcile.
            raise OptimizationRunRetryablePublicationError(
                f"Could not publish final optimization detail for {milestone}"
            ) from exc

    def _publish_final_detail_enrichment(
        self,
        *,
        core_revision: PublishedRevision,
        stakeholder_view: Mapping[str, Any],
        committed_artifacts: Mapping[str, Mapping[str, Any]],
        publication_id: str,
        evidence_checksum: str,
        stakeholder_view_checksum: str,
        generated_at: datetime,
    ) -> PublishedRevision:
        """Attach final score details after the finalization core is durable.

        The core revision remains immutable and is the replay checkpoint.  This
        method adds an immutable detail manifest and advances only the mutable
        Report/Task pointers that describe whether that core revision now has a
        reconciled stakeholder-detail projection.
        """
        state = self._require_state()
        revision_number = core_revision.number
        current_artifacts = self._latest_committed_artifact_index()
        reusable_artifacts = {**dict(committed_artifacts), **current_artifacts}
        full_plan = _artifact_publication_plan(stakeholder_view)
        detail_counts = {
            kind: {"completed": 0, "total": full_plan[kind]["total"]}
            for kind in (
                "score_briefs", "scorecard_summaries", "scorecard_spreadsheets",
                "scorecard_presentations",
            )
        }
        active_detail_kind = "score_briefs"

        def publish_detail_progress(
            *, state_value: str = "active", failure_class: str | None = None,
        ) -> None:
            current = sum(item["completed"] for item in detail_counts.values())
            total = sum(item["total"] for item in detail_counts.values())
            failure = None
            if failure_class:
                failure = {
                    "exception_class": failure_class if re.fullmatch(
                        r"[A-Za-z_][A-Za-z0-9_]{0,127}", failure_class,
                    ) else "Exception",
                    "operation_category": active_detail_kind,
                    "retry_classification": "retryable",
                    "completed": current,
                    "total": total,
                }
            # Final-detail work is still artifact publication.  Keep it on
            # the existing public progress contract, but give it its own
            # message and counts so operators do not confuse it with the
            # compact core revision that has already been committed.
            self.publish_progress(
                phase="publication",
                current=current,
                total=total,
                unit="artifacts",
                state=state_value,
                message=(
                    "Final score and scorecard details are incomplete and will retry."
                    if state_value == "failed"
                    else "Publishing final score and scorecard details."
                ),
                next_checkpoint="Completing final stakeholder detail artifacts.",
                artifact_counts=detail_counts,
                failure=failure,
            )

        def record_progress_kind(kind: str, completed: bool) -> None:
            nonlocal active_detail_kind
            active_detail_kind = kind
            if not completed:
                return
            if kind not in detail_counts:
                raise OptimizationRunIntegrityError("unknown final detail artifact category")
            detail_counts[kind]["completed"] += 1
            if (
                detail_counts[kind]["completed"] % 25 == 0
                or detail_counts[kind]["completed"] == detail_counts[kind]["total"]
            ):
                publish_detail_progress()

        publish_detail_progress()
        draft_artifacts = self._load_publication_draft_artifacts(
            revision_number=revision_number,
            milestone=core_revision.milestone,
            evidence_checksum=evidence_checksum,
            stakeholder_view_checksum=stakeholder_view_checksum,
            generated_at=generated_at,
        )
        resolved_draft_artifacts = dict(draft_artifacts)
        draft_dirty_count = 0

        def reuse_artifact(
            logical_id: str,
            kind: str,
            content_type: str,
            content: bytes,
            filename: str,
        ) -> Optional[Mapping[str, Any]]:
            for artifact_index in (reusable_artifacts, draft_artifacts):
                reusable = self._reuse_committed_artifact(
                    artifact_index,
                    logical_id=logical_id,
                    kind=kind,
                    content_type=content_type,
                    content=content,
                )
                if reusable is not None:
                    return reusable
            return self._reuse_uncommitted_task_attachment(
                logical_id=logical_id,
                kind=kind,
                content_type=content_type,
                content=content,
                filename=filename,
                publication_id=publication_id,
                source_revision=revision_number,
            )

        def persist_draft_if_needed(*, force: bool = False) -> None:
            nonlocal draft_dirty_count
            if not force and draft_dirty_count < 25:
                return
            if not resolved_draft_artifacts:
                return
            self._persist_publication_draft_artifacts(
                revision_number=revision_number,
                milestone=core_revision.milestone,
                publication_id=publication_id,
                evidence_checksum=evidence_checksum,
                stakeholder_view_checksum=stakeholder_view_checksum,
                generated_at=generated_at,
                artifacts=resolved_draft_artifacts,
            )
            draft_dirty_count = 0

        def record_artifact(_kind: str, descriptor: Mapping[str, Any]) -> None:
            nonlocal draft_dirty_count
            logical_id = str(descriptor.get("logical_id") or "")
            object_key = str(descriptor.get("object_key") or "")
            if not logical_id or not object_key:
                raise OptimizationRunIntegrityError(
                    "resolved detail artifact descriptor is incomplete"
                )
            normalized = dict(descriptor)
            if resolved_draft_artifacts.get(logical_id) == normalized:
                return
            resolved_draft_artifacts[logical_id] = normalized
            draft_dirty_count += 1
            persist_draft_if_needed()

        try:
            detail_artifacts = build_scorecard_artifacts(
                stakeholder_view,
                revision_number=revision_number,
                task_id=state.task.id,
                uploader=self._artifact_uploader,
                publication_id=publication_id,
                on_artifact_upload=record_progress_kind,
                reuse_artifact=reuse_artifact,
                on_artifact_resolved=record_artifact,
            )
        except Exception as exc:
            try:
                publish_detail_progress(
                    state_value="failed", failure_class=type(exc).__name__,
                )
            except Exception:
                pass
            raise
        persist_draft_if_needed(force=True)
        # Individual detail files may be reused byte-for-byte, but this
        # aggregate projection records completion of this immutable core
        # revision.  Its source must therefore be the current revision.
        detail_source_revision = revision_number
        detail_presentation = build_stakeholder_presentation(
            stakeholder_view,
            scorecard_artifacts=detail_artifacts,
            detail_status="complete",
            detail_source_revision=detail_source_revision,
        )
        detail_presentation_bytes = _json(detail_presentation)
        presentation_filename = (
            f"optimization-detail-presentation-r{revision_number:04d}-{publication_id}.json"
        )
        reusable_presentation = reuse_artifact(
            "stakeholder_detail_presentation",
            "stakeholder_detail_presentation",
            "application/json",
            detail_presentation_bytes,
            presentation_filename,
        )
        if isinstance(reusable_presentation, Mapping):
            presentation_path = str(reusable_presentation.get("object_key") or "")
            presentation_source_revision = int(
                reusable_presentation.get("source_revision") or 0
            )
        else:
            presentation_path = self._artifact_uploader(
                state.task.id, presentation_filename, detail_presentation_bytes,
            )
            presentation_source_revision = revision_number
        if not presentation_path or presentation_source_revision < 1:
            raise OptimizationRunIntegrityError(
                "final stakeholder detail presentation descriptor is incomplete"
            )
        self._attach_task_file(state.task, presentation_path)
        presentation_artifact = _artifact_descriptor(
            logical_id="stakeholder_detail_presentation",
            kind="stakeholder_detail_presentation",
            display_name="Stakeholder detail presentation data",
            scope="run",
            content_type="application/json",
            content=detail_presentation_bytes,
            object_key=presentation_path,
            task_id=state.task.id,
            source_revision=presentation_source_revision,
        )
        detail_artifacts.append(presentation_artifact)
        if self.dashboard_base_url:
            for artifact in detail_artifacts:
                artifact["dashboard_url"] = _report_artifact_url(
                    self.dashboard_base_url,
                    report_id=state.report.id,
                    revision_number=revision_number,
                    logical_id=str(artifact["logical_id"]),
                )
        detail_manifest = {
            "schema_version": "optimization-run-detail-manifest-v1",
            "source_revision": revision_number,
            "source_manifest_checksum": core_revision.manifest_checksum,
            "milestone": core_revision.milestone,
            "published_at": _iso(generated_at),
            "stakeholder_view_checksum": stakeholder_view_checksum,
            "artifacts": detail_artifacts,
        }
        detail_manifest_bytes = _json(detail_manifest)
        detail_manifest_path = self._artifact_uploader(
            state.task.id,
            f"optimization-detail-manifest-r{revision_number:04d}-{publication_id}.json",
            detail_manifest_bytes,
        )
        self._attach_task_file(state.task, detail_manifest_path)
        detail_manifest_descriptor = _artifact_descriptor(
            logical_id="detail_manifest",
            kind="detail_manifest",
            display_name="Final detail artifact manifest",
            scope="run",
            content_type="application/json",
            content=detail_manifest_bytes,
            object_key=detail_manifest_path,
            task_id=state.task.id,
            source_revision=revision_number,
        )
        self._record_detail_enrichment(
            core_revision=core_revision,
            detail_manifest=detail_manifest_descriptor,
            detail_presentation=presentation_artifact,
        )
        self._committed_artifact_index_cache = (
            revision_number,
            {
                str(artifact["logical_id"]): dict(artifact)
                for artifact in (*core_revision.artifacts, *detail_artifacts)
            },
        )
        return replace(
            core_revision,
            artifacts=tuple((*core_revision.artifacts, *detail_artifacts)),
            detail_status="complete",
            detail_source_revision=revision_number,
        )

    def _resume_final_detail_enrichment(
        self,
        *,
        latest: Mapping[str, Any],
        core_manifest: Mapping[str, Any],
    ) -> None:
        """Finish a final detail pass from its immutable core checkpoint."""
        state = self._require_state()
        source_revision = int(latest.get("number") or 0)
        detail_input = next(
            (
                item for item in (core_manifest.get("artifacts") or [])
                if isinstance(item, Mapping) and item.get("kind") == "detail_input"
            ),
            None,
        )
        if not isinstance(detail_input, Mapping):
            raise OptimizationRunIntegrityError(
                "finalization core revision is missing deferred detail input"
            )
        try:
            input_value = json.loads(
                self._download_task_attachment(state.task.id, detail_input).decode("utf-8")
            )
        except OptimizationRunIntegrityError:
            raise
        except Exception as exc:
            raise OptimizationRunRetryablePublicationError(
                "Could not read deferred final detail input"
            ) from exc
        if not isinstance(input_value, Mapping):
            raise OptimizationRunIntegrityError("deferred final detail input is malformed")
        stakeholder_view = input_value.get("stakeholder_view")
        if not isinstance(stakeholder_view, Mapping):
            raise OptimizationRunIntegrityError("deferred final detail input lacks a stakeholder view")
        _validate_view(stakeholder_view)
        stakeholder_view_checksum = str(input_value.get("stakeholder_view_checksum") or "")
        if sha256(_json(stakeholder_view)).hexdigest() != stakeholder_view_checksum:
            raise OptimizationRunIntegrityError("deferred final detail input checksum conflicts with its view")
        evidence_checksum = str(input_value.get("evidence_checksum") or "")
        if evidence_checksum != str(core_manifest.get("evidence_checksum") or ""):
            raise OptimizationRunIntegrityError("deferred final detail input conflicts with core evidence")
        try:
            generated_at = datetime.fromisoformat(
                str(core_manifest.get("published_at") or "").replace("Z", "+00:00")
            ).astimezone(timezone.utc)
        except ValueError as exc:
            raise OptimizationRunIntegrityError(
                "finalization core revision has an invalid publication time"
            ) from exc
        publication_id = self._publication_id_factory()
        if (
            not isinstance(publication_id, str)
            or not publication_id
            or any(not (character.isalnum() or character in "-_.") for character in publication_id)
        ):
            raise OptimizationRunIntegrityError("detail recovery publication ID is invalid")
        core_revision = PublishedRevision(
            number=source_revision,
            milestone="finalization",
            published_at=str(core_manifest.get("published_at") or ""),
            raw_evidence_path=str(latest.get("evidence_path") or ""),
            workbook_path=str(latest.get("workbook_path") or ""),
            manifest_path=str(latest.get("manifest_path") or ""),
            manifest_checksum=str((latest.get("manifest") or {}).get("sha256") or ""),
            manifest_size_bytes=int((latest.get("manifest") or {}).get("size_bytes") or 0),
            evidence_checksum=evidence_checksum,
            workbook_checksum=str(core_manifest.get("workbook_checksum") or ""),
            row_counts=dict(core_manifest.get("row_counts") or {}),
            overview=dict(core_manifest.get("overview") or {}),
            artifacts=tuple(
                dict(item) for item in (core_manifest.get("artifacts") or [])
                if isinstance(item, Mapping)
            ),
            detail_status="pending",
        )
        self._publish_final_detail_enrichment(
            core_revision=core_revision,
            stakeholder_view=stakeholder_view,
            committed_artifacts=self._latest_committed_artifact_index(),
            publication_id=publication_id,
            evidence_checksum=evidence_checksum,
            stakeholder_view_checksum=stakeholder_view_checksum,
            generated_at=generated_at,
        )

    def persist_semantic_budget_ledger(
        self, ledger_value: Mapping[str, Any]
    ) -> dict[str, Any]:
        """Commit one immutable semantic-ledger revision without a workbook.

        Uploading and attaching the JSON may leave an orphan when publication is
        interrupted.  The pointer under ``Report.parameters`` is the sole commit
        point, so recovery deliberately ignores every unreferenced upload.
        """
        from plexus.optimization.semantic_budget import (
            SemanticBudgetLedger,
            SemanticBudgetSpec,
            canonical_json_bytes,
        )

        state = self._require_state()
        parameters = _metadata(getattr(state.report, "parameters", {}))
        run = dict(parameters.get("optimization_run") or {})
        report_spec = run.get("run_spec")
        if not isinstance(report_spec, Mapping) or not _same_run_spec(
            report_spec, state.run_spec
        ):
            raise OptimizationRunIntegrityError(
                "Report does not preserve the frozen semantic budget run specification"
            )
        task_spec = _metadata(getattr(state.task, "metadata", {})).get("run_spec")
        if not isinstance(task_spec, Mapping) or not _same_run_spec(
            task_spec, state.run_spec
        ):
            raise OptimizationRunIntegrityError(
                "Task does not preserve the frozen semantic budget run specification"
            )
        frozen_budget = state.run_spec.get("semantic_budget")
        if not isinstance(frozen_budget, Mapping):
            raise OptimizationRunIntegrityError(
                "frozen run specification has no semantic budget"
            )
        try:
            expected_spec = SemanticBudgetSpec.from_dict(frozen_budget)
            ledger = SemanticBudgetLedger.from_dict(
                ledger_value, expected_spec=expected_spec
            )
        except Exception as exc:
            raise OptimizationRunIntegrityError(
                "semantic budget ledger conflicts with the frozen run specification"
            ) from exc
        if ledger.run_key != self.run_key:
            raise OptimizationRunIntegrityError(
                "semantic budget ledger belongs to another run"
            )

        content = canonical_json_bytes(ledger.to_dict())
        digest = sha256(content).hexdigest()
        latest = run.get("semantic_budget_latest")
        if latest is not None and not isinstance(latest, Mapping):
            raise OptimizationRunIntegrityError(
                "semantic budget commit pointer is malformed"
            )
        if isinstance(latest, Mapping) and latest.get("sha256") == digest:
            if int(latest.get("ledger_revision") or -1) != ledger.revision:
                raise OptimizationRunIntegrityError(
                    "semantic budget commit pointer revision is inconsistent"
                )
            return dict(latest)
        prior_revision = (
            int(latest.get("ledger_revision") or 0)
            if isinstance(latest, Mapping)
            else 0
        )
        if ledger.revision != prior_revision + 1:
            raise OptimizationRunIntegrityError(
                "semantic budget ledger revisions must be committed in order"
            )

        try:
            publication_id = self._publication_id_factory()
            if (
                not isinstance(publication_id, str)
                or not publication_id
                or any(
                    not (character.isalnum() or character in "-_.")
                    for character in publication_id
                )
            ):
                raise ValueError(
                    "publication_id_factory must return a filesystem-safe nonempty string"
                )
            filename = (
                f"optimization-semantic-ledger-r{ledger.revision:06d}-"
                f"{publication_id}.json"
            )
            object_key = self._artifact_uploader(
                state.task.id, filename, content
            )
            self._attach_task_file(state.task, object_key)
            pointer = _artifact_descriptor(
                logical_id="semantic_budget_ledger",
                kind="semantic_budget_ledger",
                display_name="Semantic budget ledger",
                scope="run",
                content_type="application/json",
                content=content,
                object_key=object_key,
                task_id=state.task.id,
                source_revision=ledger.revision,
            )
            pointer.update({
                "ledger_revision": ledger.revision,
                "pricing_version": ledger.spec.pricing_version,
            })
            run["semantic_budget_latest"] = pointer
            parameters["optimization_run"] = run
            state.report.update(parameters=parameters)
            return dict(pointer)
        except OptimizationRunIntegrityError:
            raise
        except Exception as exc:
            raise OptimizationRunRetryablePublicationError(
                "Could not commit semantic budget ledger revision"
            ) from exc

    def load_semantic_budget_ledger(self) -> Optional[dict[str, Any]]:
        """Load and checksum-verify the Report's committed semantic ledger."""
        from plexus.optimization.semantic_budget import (
            SemanticBudgetLedger,
            SemanticBudgetSpec,
            canonical_json_bytes,
        )

        state = self._require_state()
        parameters = _metadata(getattr(state.report, "parameters", {}))
        run = parameters.get("optimization_run")
        if not isinstance(run, Mapping):
            raise OptimizationRunIntegrityError(
                "Report optimization run parameters are malformed"
            )
        report_spec = run.get("run_spec")
        if not isinstance(report_spec, Mapping) or not _same_run_spec(
            report_spec, state.run_spec
        ):
            raise OptimizationRunIntegrityError(
                "Report does not preserve the frozen semantic budget run specification"
            )
        task_spec = _metadata(getattr(state.task, "metadata", {})).get("run_spec")
        if not isinstance(task_spec, Mapping) or not _same_run_spec(
            task_spec, state.run_spec
        ):
            raise OptimizationRunIntegrityError(
                "Task does not preserve the frozen semantic budget run specification"
            )
        pointer = run.get("semantic_budget_latest")
        if pointer is None:
            return None
        if not isinstance(pointer, Mapping):
            raise OptimizationRunIntegrityError(
                "semantic budget commit pointer is malformed"
            )
        frozen_budget = state.run_spec.get("semantic_budget")
        try:
            if (
                pointer.get("kind") != "semantic_budget_ledger"
                or pointer.get("logical_id") != "semantic_budget_ledger"
                or pointer.get("task_id") != state.task.id
                or pointer.get("content_type") != "application/json"
                or pointer.get("source_revision")
                != pointer.get("ledger_revision")
            ):
                raise ValueError("semantic budget descriptor is inconsistent")
            if not isinstance(frozen_budget, Mapping):
                raise ValueError("frozen run specification has no semantic budget")
            content = self._download_task_attachment(state.task.id, pointer)
            parsed = json.loads(content.decode("utf-8"))
            if not isinstance(parsed, Mapping):
                raise ValueError("semantic budget ledger must be an object")
            if canonical_json_bytes(parsed) != content:
                raise ValueError("semantic budget ledger JSON is not canonical")
            expected_spec = SemanticBudgetSpec.from_dict(frozen_budget)
            ledger = SemanticBudgetLedger.from_dict(
                parsed, expected_spec=expected_spec
            )
            if (
                ledger.run_key != self.run_key
                or ledger.revision != int(pointer.get("ledger_revision") or -1)
                or ledger.digest() != pointer.get("sha256")
                or ledger.spec.pricing_version != pointer.get("pricing_version")
            ):
                raise ValueError("semantic budget ledger does not match its pointer")
            return ledger.to_dict()
        except (
            ArtifactTicketError,
            ArtifactAuthorizationError,
            ArtifactTransferError,
        ) as exc:
            raise OptimizationRunRetryablePublicationError(
                "Could not read the semantic budget checkpoint attachment"
            ) from exc
        except OptimizationRunIntegrityError:
            raise
        except Exception as exc:
            raise OptimizationRunIntegrityError(
                "Could not verify the semantic budget checkpoint"
            ) from exc

    def load_latest_checkpoint(self) -> Optional[dict[str, Any]]:
        """Load and verify the last committed evidence revision for replay.

        The Report parameter update performed by ``_record_latest_revision`` is
        the publication commit point.  Unreferenced uploads from an interrupted
        publication are deliberately ignored.  Reads use the same
        GraphQL-authorized Task attachment route as every other artifact.
        """
        state = self._require_state()
        latest = self._latest_revision(state.report)
        if latest is None:
            return None
        if not isinstance(latest, Mapping):
            error = OptimizationRunIntegrityError(
                "Latest optimization revision metadata is malformed"
            )
            try:
                self.fail(f"Optimization run integrity failure: {error}")
            except Exception:
                pass
            raise error
        try:
            manifest_descriptor = latest.get("manifest")
            if not isinstance(manifest_descriptor, Mapping):
                raise ValueError("latest revision is missing its manifest descriptor")
            manifest_bytes = self._download_task_attachment(
                state.task.id, manifest_descriptor,
            )
            manifest = json.loads(manifest_bytes.decode("utf-8"))
            if not isinstance(manifest, Mapping):
                raise ValueError("latest revision manifest must be an object")
            revision_number = int(latest.get("number") or -2)
            if int(manifest.get("revision") or -1) != revision_number:
                raise ValueError("latest revision manifest number does not match")
            milestone = str(manifest.get("milestone") or "").strip()
            if not milestone or milestone != str(latest.get("milestone") or "").strip():
                raise ValueError("latest revision milestone does not match")

            evidence_descriptor = latest.get("evidence")
            if not isinstance(evidence_descriptor, Mapping):
                evidence_descriptor = next(
                    (
                        item for item in (manifest.get("artifacts") or [])
                        if isinstance(item, Mapping) and item.get("kind") == "run_evidence"
                    ),
                    None,
                )
            if not isinstance(evidence_descriptor, Mapping):
                raise ValueError("latest revision is missing its evidence descriptor")
            manifest_evidence = next(
                (
                    item for item in (manifest.get("artifacts") or [])
                    if isinstance(item, Mapping) and item.get("kind") == "run_evidence"
                ),
                None,
            )
            if not isinstance(manifest_evidence, Mapping):
                raise ValueError("latest revision manifest is missing run evidence")
            self._verify_checkpoint_descriptor(
                latest=evidence_descriptor,
                manifest=manifest_evidence,
                task_id=state.task.id,
                revision_number=revision_number,
            )
            evidence_bytes = self._download_task_attachment(
                state.task.id, evidence_descriptor,
            )
            evidence = json.loads(evidence_bytes.decode("utf-8"))
            if not isinstance(evidence, Mapping):
                raise ValueError("latest revision evidence must be an object")
            if str(evidence.get("run_key") or "") != self.run_key:
                raise ValueError("latest revision evidence belongs to a different run key")
            evidence_spec = evidence.get("run_spec")
            if not isinstance(evidence_spec, Mapping):
                raise ValueError("latest revision evidence is missing a frozen run specification")
            if not _same_run_spec(evidence_spec, state.run_spec):
                raise ValueError("latest revision evidence has a different frozen run specification")

            task_metadata = _metadata(getattr(state.task, "metadata", {}))
            persisted_spec = task_metadata.get("run_spec")
            if not isinstance(persisted_spec, Mapping) or not _same_run_spec(persisted_spec, state.run_spec):
                raise ValueError("Task metadata does not preserve the frozen run specification")
            report_spec = ((_metadata(getattr(state.report, "parameters", {})).get("optimization_run") or {}).get("run_spec"))
            if not isinstance(report_spec, Mapping) or not _same_run_spec(report_spec, state.run_spec):
                raise ValueError("Report metadata does not preserve the frozen run specification")
            task_status = str(getattr(state.task, "status", "")).upper()
            final_status = str(task_metadata.get("optimization_run_final_status") or "").upper()
            if (
                milestone == "finalization"
                and latest.get("detail_status") != "complete"
            ):
                self._resume_final_detail_enrichment(
                    latest=latest,
                    core_manifest=manifest,
                )
            return {
                "milestone": milestone,
                "evidence": dict(evidence),
                "task_terminal": task_status in {"COMPLETED", "FAILED"} or bool(final_status),
            }
        except (ArtifactTicketError, ArtifactAuthorizationError, ArtifactTransferError) as exc:
            raise OptimizationRunRetryablePublicationError(
                "Could not read the latest optimization checkpoint attachment"
            ) from exc
        except OptimizationRunRetryablePublicationError:
            raise
        except ArtifactIntegrityError as exc:
            error = OptimizationRunIntegrityError(
                "Latest optimization checkpoint attachment failed integrity verification"
            )
            try:
                self.fail(f"Optimization run integrity failure: {error}")
            except Exception:
                pass
            raise error from exc
        except OptimizationRunIntegrityError as exc:
            try:
                self.fail(f"Optimization run integrity failure: {exc}")
            except Exception:
                pass
            raise
        except Exception as exc:
            error = OptimizationRunIntegrityError(
                "Could not verify the latest optimization checkpoint"
            )
            try:
                self.fail(f"Optimization run integrity failure: {error}")
            except Exception:
                pass
            raise error from exc

    def publish_progress(
        self,
        *,
        phase: str,
        subphase: str | None = None,
        current: int,
        total: int | None,
        message: str,
        unit: str | None = None,
        state: str = "active",
        elapsed_seconds: int | None = None,
        next_checkpoint: str | None = None,
        heartbeat_interval_seconds: int | None = None,
        artifact_counts: Mapping[str, Mapping[str, int]] | None = None,
        failure: Mapping[str, Any] | None = None,
    ) -> None:
        """Publish lightweight live progress without creating an evidence revision.

        Milestones remain the immutable audit trail.  This method updates only
        the existing analysis stage, compact status block, and Report cover so
        operators can see movement during long assessment and diagnosis loops.
        """
        report_state = self._require_state()
        normalized_phase = str(phase or "").strip().lower()
        normalized_subphase = str(subphase or "").strip().lower()
        normalized_message = " ".join(str(message or "").split()).strip()
        normalized_unit = " ".join(str(unit or "").split()).strip().lower()
        normalized_state = str(state or "active").strip().lower()
        normalized_checkpoint = " ".join(str(next_checkpoint or "").split()).strip()
        if normalized_phase not in {"ranking", "assessment", "diagnosis", "publication"}:
            raise ValueError("progress phase must be ranking, assessment, diagnosis, or publication")
        if normalized_subphase and (
            normalized_phase != "ranking"
            or normalized_subphase not in {
                "inventory", "activity_evidence", "feedback_analysis",
            }
        ):
            raise ValueError("ranking progress subphase is unsupported")
        if isinstance(current, bool) or isinstance(total, bool):
            raise ValueError("progress counts must be integers")
        current = int(current)
        if total is not None:
            total = int(total)
        if current < 0 or (total is not None and (total < 0 or current > total)):
            raise ValueError("progress counts must satisfy 0 <= current <= total when total is known")
        if not normalized_message:
            raise ValueError("progress message is required")
        if normalized_state not in {"active", "retrying", "incomplete", "failed"}:
            raise ValueError("progress state must be active, retrying, incomplete, or failed")
        if normalized_unit and normalized_unit not in {
            "scorecards", "scores", "analysis steps", "artifacts",
        }:
            raise ValueError("progress unit is unsupported")
        normalized_artifact_counts: dict[str, dict[str, int]] | None = None
        if artifact_counts is not None:
            if normalized_phase != "publication" or not isinstance(artifact_counts, Mapping):
                raise ValueError("artifact_counts are only supported for publication progress")
            normalized_artifact_counts = {}
            for kind in _ARTIFACT_PUBLICATION_KINDS:
                value = artifact_counts.get(kind)
                if value is None:
                    continue
                if not isinstance(value, Mapping):
                    raise ValueError("artifact publication counts must be mappings")
                completed = value.get("completed")
                artifact_total = value.get("total")
                if (
                    isinstance(completed, bool)
                    or isinstance(artifact_total, bool)
                    or not isinstance(completed, int)
                    or not isinstance(artifact_total, int)
                    or completed < 0
                    or artifact_total < 0
                    or completed > artifact_total
                ):
                    raise ValueError("artifact publication counts must satisfy 0 <= completed <= total")
                normalized_artifact_counts[kind] = {
                    "completed": completed,
                    "total": artifact_total,
                }
            if set(artifact_counts) != set(normalized_artifact_counts):
                raise ValueError("artifact publication kinds are unsupported")
            if sum(value["completed"] for value in normalized_artifact_counts.values()) != current:
                raise ValueError("artifact publication completed counts must match current")
            if sum(value["total"] for value in normalized_artifact_counts.values()) != total:
                raise ValueError("artifact publication totals must match total")
        normalized_failure: dict[str, Any] | None = None
        if failure is not None:
            if (
                normalized_phase != "publication"
                or normalized_state not in {"failed", "retrying"}
                or not isinstance(failure, Mapping)
                or set(failure) != {
                    "exception_class",
                    "operation_category",
                    "retry_classification",
                    "completed",
                    "total",
                }
            ):
                raise ValueError("failure telemetry is only supported for failed publication progress")
            exception_class = failure.get("exception_class")
            operation_category = failure.get("operation_category")
            retry_classification = failure.get("retry_classification")
            if (
                not isinstance(exception_class, str)
                or re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]{0,127}", exception_class) is None
                or operation_category not in _ARTIFACT_PUBLICATION_KINDS
                or retry_classification != "retryable"
                or failure.get("completed") != current
                or failure.get("total") != total
            ):
                raise ValueError("publication failure telemetry is malformed")
            normalized_failure = {
                "exception_class": exception_class,
                "operation_category": operation_category,
                "retry_classification": retry_classification,
                "completed": current,
                "total": total,
            }
        if elapsed_seconds is not None:
            if isinstance(elapsed_seconds, bool) or int(elapsed_seconds) < 0:
                raise ValueError("elapsed_seconds must be a nonnegative integer")
            elapsed_seconds = int(elapsed_seconds)
        if heartbeat_interval_seconds is not None:
            if (
                isinstance(heartbeat_interval_seconds, bool)
                or int(heartbeat_interval_seconds) <= 0
            ):
                raise ValueError("heartbeat_interval_seconds must be a positive integer")
            heartbeat_interval_seconds = int(heartbeat_interval_seconds)

        progress = {
            "phase": normalized_phase,
            "current": current,
            "total": total,
            "message": normalized_message,
            "updated_at": _iso(self.now()),
        }
        if normalized_subphase:
            progress["subphase"] = normalized_subphase
        if normalized_unit:
            progress["unit"] = normalized_unit
        if normalized_state != "active":
            progress["state"] = normalized_state
        if elapsed_seconds is not None:
            progress["elapsed_seconds"] = elapsed_seconds
        if normalized_checkpoint:
            progress["next_checkpoint"] = normalized_checkpoint
        if heartbeat_interval_seconds is not None:
            progress["heartbeat_interval_seconds"] = heartbeat_interval_seconds
        if normalized_artifact_counts is not None:
            progress["artifact_counts"] = normalized_artifact_counts
        if normalized_failure is not None:
            progress["failure"] = normalized_failure
        try:
            # Ranking is synchronous today, but callbacks originate in
            # bounded worker code and must remain safe if they arrive after a
            # later durable milestone. Never let a stale live overlay move the
            # operator view backwards over immutable ranking/assessment/etc.
            latest_revision = self._latest_revision(report_state.report)
            latest_milestone = (
                str(latest_revision.get("milestone") or "").strip().lower()
                if isinstance(latest_revision, Mapping)
                else ""
            )
            if (
                normalized_phase == "ranking"
                and latest_milestone in _RANKING_PROGRESS_SUPERSEDED_MILESTONES
            ):
                return
            stage = (
                report_state.stages.get(normalized_phase)
                or report_state.stages.get("analysis")
            )
            # Some existing direct-run Tasks retain the fixed lifecycle stages
            # but have no generic Analysis stage.  The Report remains the
            # authoritative live publication surface in that case; do not turn
            # an optional TaskStage projection into a publication failure.
            if stage is not None:
                stage_update = {
                    "processedItems": current,
                    "statusMessage": normalized_message[:1000],
                }
                if total is not None:
                    stage_update["totalItems"] = total
                stage.update(**stage_update)

            status_block = report_state.blocks["status"]
            compact = _metadata(getattr(status_block, "output", {}))
            preview = dict(compact.get("preview") or {})
            summary = dict(preview.get("summary") or {})
            summary["live_progress"] = progress
            preview["summary"] = summary
            compact["preview"] = preview
            status_block.update(
                output=json.dumps(compact),
                log="Live progress; see immutable revisions in attachedFiles.",
            )

            report_state.report.update(output=self._render_report_manifest(
                "RUNNING",
                latest_revision,
                identity=report_state.operator_identity,
                live_progress=progress,
                execution_mode=report_state.run_spec.get("execution_mode"),
            ))
        except Exception as exc:
            raise OptimizationRunRetryablePublicationError(
                f"Could not publish live {normalized_phase} progress"
            ) from exc

    def finalize(self, *, status: str = "completed") -> OptimizationRunReportState:
        """Finish a run while preserving its stakeholder-visible terminal state.

        ``Task.status`` has no distinct incomplete/blocked terminal values, so
        it remains ``COMPLETED`` for a safely concluded non-failure run.  The
        precise lifecycle outcome is retained in immutable task metadata and
        the Report cover page rather than being mislabeled as a failure.
        """
        state = self._require_state()
        terminal_status = str(status or "completed").lower()
        if terminal_status not in _FINAL_STATES:
            raise ValueError(f"unsupported optimization run terminal status: {status}")
        if terminal_status == "failed":
            return self.fail("Optimization run finalized as failed")
        try:
            latest_revision = self._latest_revision(state.report)
            if (
                isinstance(latest_revision, Mapping)
                and str(latest_revision.get("milestone") or "") == "finalization"
                and latest_revision.get("detail_status") != "complete"
            ):
                raise OptimizationRunRetryablePublicationError(
                    "Final stakeholder detail enrichment is still pending"
                )
            task_metadata = _metadata(getattr(state.task, "metadata", {}))
            task_metadata["optimization_run_final_status"] = terminal_status
            state.report.update(output=self._render_report_manifest(
                terminal_status,
                self._latest_revision(state.report),
                identity=state.operator_identity,
                execution_mode=state.run_spec.get("execution_mode"),
            ))
            if self._uses_existing_task:
                # The outer Tactus procedure remains the lifecycle authority
                # for its Task and coarse stages.
                state.task.update(metadata=json.dumps(task_metadata))
            else:
                self._complete_all_stages()
                state.task.update(
                    status="COMPLETED",
                    completedAt=_iso(self.now()),
                    metadata=json.dumps(task_metadata),
                )
            return state
        except OptimizationRunIntegrityError as exc:
            try:
                self.fail(f"Optimization run integrity failure: {exc}")
            except Exception:
                pass
            raise
        except Exception as exc:
            # Finalization is replayed from the committed finalization revision.
            # A transport/auth interruption must not replace that active attempt.
            raise OptimizationRunRetryablePublicationError(
                "Could not finalize optimization run report"
            ) from exc

    def fail(self, message: str) -> OptimizationRunReportState:
        state = self._require_state()
        task_metadata = _metadata(getattr(state.task, "metadata", {}))
        task_metadata["optimization_run_final_status"] = "failed"
        self._fail_active_stage(str(message))
        state.task.update(
            status="FAILED", errorMessage=str(message), completedAt=_iso(self.now()),
            metadata=json.dumps(task_metadata),
        )
        try:
            state.report.update(output=self._render_report_manifest(
                "failed",
                self._latest_revision(state.report),
                message,
                identity=state.operator_identity,
                execution_mode=state.run_spec.get("execution_mode"),
            ))
        except Exception:
            # The Task failure is the canonical safety signal if the report view is unavailable.
            pass
        try:
            status_block = state.blocks["status"]
            raw_output = getattr(status_block, "output", None)
            payload = json.loads(raw_output) if isinstance(raw_output, str) else dict(raw_output or {})
            preview = dict(payload.get("preview") or {})
            summary = dict(preview.get("summary") or {})
            summary["run_failure"] = {
                "state": "failure",
                "headline": "The optimization run could not complete",
                "explanation": str(message),
            }
            preview["status"] = "failed"
            preview["summary"] = summary
            payload["status"] = "failed"
            payload["preview"] = preview
            status_block.update(
                output=json.dumps(payload, sort_keys=True, separators=(",", ":")),
                log="Run failed; the latest durable revision remains attached.",
            )
        except Exception:
            # The Task and Report cover page remain the canonical failure
            # signals if the mutable status block cannot be updated.
            pass
        return state

    def _fail_uninitialized_integrity(self, task: Any, message: str) -> None:
        """Best-effort terminal signal for verified identity corruption."""
        metadata = _metadata(getattr(task, "metadata", {}))
        metadata["optimization_run_final_status"] = "failed"
        metadata["failure"] = str(message)
        try:
            task.update(
                status="FAILED",
                errorMessage=str(message),
                completedAt=_iso(self.now()),
                metadata=json.dumps(metadata),
            )
        except Exception:
            # The integrity exception remains authoritative if its terminal
            # projection cannot itself be written.
            pass

    def _advance_stage_for_milestone(self, milestone: str) -> None:
        if self._uses_existing_task:
            return
        target_name = _MILESTONE_STAGE.get(str(milestone))
        if target_name is None:
            return
        state = self._require_state()
        target = state.stages.get(target_name)
        if target is None:
            raise RuntimeError(f"missing TaskStage for milestone {milestone}: {target_name}")
        target_order = int(getattr(target, "order", _STAGES.index(target_name)))
        changed_at = self.now()
        for stage in sorted(state.stages.values(), key=lambda item: int(getattr(item, "order", 0))):
            order = int(getattr(stage, "order", 0))
            status = str(getattr(stage, "status", "")).upper()
            if order < target_order and status not in {"COMPLETED", "FAILED"}:
                stage.update(
                    status="COMPLETED",
                    statusMessage="Complete",
                    startedAt=getattr(stage, "startedAt", None) or changed_at,
                    completedAt=changed_at,
                )
            elif order == target_order and status == "PENDING":
                stage.update(
                    status="RUNNING",
                    statusMessage=f"Publishing {milestone}",
                    startedAt=getattr(stage, "startedAt", None) or changed_at,
                )
        if getattr(state.task, "currentStageId", None) != target.id:
            state.task.update(currentStageId=target.id)

    def _complete_all_stages(self) -> None:
        if self._uses_existing_task:
            return
        state = self._require_state()
        completed_at = self.now()
        for stage in sorted(state.stages.values(), key=lambda item: int(getattr(item, "order", 0))):
            if str(getattr(stage, "status", "")).upper() in {"COMPLETED", "FAILED"}:
                continue
            stage.update(
                status="COMPLETED",
                statusMessage="Complete",
                startedAt=getattr(stage, "startedAt", None) or completed_at,
                completedAt=completed_at,
            )

    def _fail_active_stage(self, message: str) -> None:
        """Best-effort stage failure; the Task remains the canonical signal."""
        if self._uses_existing_task:
            return
        state = self._require_state()
        ordered = sorted(state.stages.values(), key=lambda item: int(getattr(item, "order", 0)))
        active = next(
            (stage for stage in ordered if str(getattr(stage, "status", "")).upper() == "RUNNING"),
            None,
        )
        if active is None:
            active = next(
                (stage for stage in ordered if str(getattr(stage, "status", "")).upper() != "COMPLETED"),
                None,
            )
        if active is None or str(getattr(active, "status", "")).upper() == "FAILED":
            return
        try:
            active.update(
                status="FAILED",
                statusMessage=message[:1000],
                completedAt=self.now(),
            )
        except Exception:
            pass

    def _require_state(self) -> OptimizationRunReportState:
        if self._state is None:
            raise RuntimeError("start_or_resume must be called before publishing")
        return self._state

    def _ensure_fixed_stages(self, task: Any) -> list[Any]:
        existing = {str(getattr(stage, "name", "")): stage for stage in self._stage_lookup(task)}
        stages: list[Any] = []
        for order, name in enumerate(_STAGES):
            stage = existing.get(name)
            if stage is None:
                stage = TaskStage.create(
                    client=self.client,
                    taskId=task.id,
                    name=name,
                    order=order,
                    status="RUNNING" if order == 0 else "PENDING",
                    statusMessage="Current stage" if order == 0 else "Waiting",
                )
            stages.append(stage)
        return stages

    def _ensure_fixed_blocks(self, report: Any) -> dict[str, Any]:
        existing = {str(getattr(block, "type", "")): block for block in self._block_lookup(report)}
        blocks: dict[str, Any] = {}
        for position, name, block_type in _BLOCK_SPECS:
            block = existing.get(block_type)
            if block is None:
                block = ReportBlock.create(
                    client=self.client, reportId=report.id, position=position, type=block_type,
                    name=name, output=json.dumps({"status": "pending", "output_compacted": True}),
                    log="Waiting for the next durable revision.",
                )
            blocks[{"OptimizationRunStatus": "status", "OptimizationDecisionEvidence": "evidence", "OptimizationStakeholderWorkbook": "workbook"}[block_type]] = block
        return blocks

    def _update_block(self, block: Any, attachment_path: str, summary: Mapping[str, Any]) -> None:
        attached = list(getattr(block, "attachedFiles", None) or [])
        if attachment_path not in attached:
            attached.append(attachment_path)
        compact = _compact_output_json_for_storage(summary, attachment_path)
        block.update(output=compact, log="See immutable revision in attachedFiles.", attachedFiles=attached)

    @staticmethod
    def _attach_task_file(task: Any, attachment_path: str) -> None:
        attached = list(getattr(task, "attachedFiles", None) or [])
        if attachment_path not in attached:
            attached.append(attachment_path)
            task.update(attachedFiles=attached)

    def _ensure_initial_envelopes(self, state: OptimizationRunReportState) -> None:
        """Give every fixed block a real compact pointer before a run advances."""
        execution_mode = state.run_spec.get("execution_mode")
        automatic_execution = execution_mode == "automatic"
        initial_execution_overview = {
            "execution_mode": "automatic",
            "execution_selected_count": 0,
            "execution_launched_count": 0,
            "execution_rejected_count": 0,
            "execution_named_selected_count": 0,
            "execution_named_launched_count": 0,
            "execution_named_rejected_count": 0,
            "execution_detail_coverage": "complete",
            "execution_detail_limitation": "",
        } if automatic_execution else (
            {"execution_mode": "approval_required"}
            if execution_mode == "approval_required" else {}
        )
        for key, block in state.blocks.items():
            if list(getattr(block, "attachedFiles", None) or []):
                continue
            payload = {
                "kind": "optimization_run_initialization",
                "run_key": self.run_key,
                "block": key,
                "published_at": _iso(self.now()),
                "status": "running",
                **initial_execution_overview,
            }
            if key == "evidence":
                name = "optimization-run-initial-evidence.json"
                path = self._artifact_uploader(state.task.id, name, _json(payload))
                self._attach_task_file(state.task, path)
            elif key == "workbook":
                name = "optimization-workbook-r0000.xlsx"
                initial_view = {
                    "overview": {
                        "headline": state.operator_identity.display_title,
                        "lifecycle_status": "running",
                        "current_activity": "Preparing exhaustive portfolio discovery.",
                        "next_checkpoint": "Ranking begins after the run and report are durably initialized.",
                        "coverage_status": "pending",
                        "ranking_window": "pending",
                        "scorecards_inspected": 0,
                        "ranked_score_count": 0,
                        "unranked_score_count": 0,
                        "cooldown_excluded_count": 0,
                        "assessment_progress": "Not started",
                        "diagnosis_coverage": "Not started",
                        "pending_approval_count": 0,
                        "notes": "No score, guideline, champion, or feedback setting is changed automatically.",
                        **initial_execution_overview,
                    },
                    "portfolio": [],
                    "priorities": [],
                    "feedback_investment": [],
                    "questions_and_issues": [],
                    "optimization_outcomes": [],
                    "definitions": {},
                }
                workbook = build_stakeholder_workbook(
                    initial_view, revision_number=0, generated_at=self.now(),
                )
                path = self._artifact_uploader(state.task.id, name, workbook.content)
                self._attach_task_file(state.task, path)
                payload = {
                    **payload,
                    "checksum": workbook.checksum,
                    "row_counts": dict(workbook.row_counts),
                }
            else:
                name = "optimization-run-initial-status.json"
                path = self._artifact_uploader(state.task.id, name, _json(payload))
                self._attach_task_file(state.task, path)
            self._update_block(block, path, payload)

    def _record_latest_revision(self, state: OptimizationRunReportState, revision: PublishedRevision) -> None:
        parameters = _metadata(getattr(state.report, "parameters", {}))
        run = dict(parameters.get("optimization_run") or {})
        latest = {
            "number": revision.number, "milestone": revision.milestone,
            "published_at": revision.published_at,
            "manifest_path": revision.manifest_path, "workbook_path": revision.workbook_path,
            "manifest": {
                "task_id": state.task.id,
                "object_key": revision.manifest_path,
                "content_type": "application/json",
                "size_bytes": revision.manifest_size_bytes,
                "sha256": revision.manifest_checksum,
            },
            "evidence_path": revision.raw_evidence_path,
            "evidence_checksum": revision.evidence_checksum, "workbook_checksum": revision.workbook_checksum,
            "row_counts": dict(revision.row_counts),
            "overview": dict(revision.overview),
            "detail_status": revision.detail_status,
            "detail_source_revision": revision.detail_source_revision,
        }
        evidence_descriptor = next(
            (
                dict(item) for item in revision.artifacts
                if isinstance(item, Mapping) and item.get("kind") == "run_evidence"
            ),
            None,
        )
        if evidence_descriptor is not None:
            latest["evidence"] = evidence_descriptor
        revisions = list(run.get("revisions") or [])
        if any(int(item.get("number") or -1) == revision.number for item in revisions if isinstance(item, Mapping)):
            raise ValueError(f"revision {revision.number} already exists")
        revisions.append(latest)
        run["latest_revision"] = latest
        run["revisions"] = revisions
        run["run_key"] = self.run_key
        run["lifecycle_version"] = LIFECYCLE_VERSION
        run["operator_identity"] = state.operator_identity.as_dict()
        parameters["_display_title"] = state.operator_identity.display_title
        parameters["_display_subtitle"] = self._display_subtitle(state.operator_identity)
        parameters["optimization_run"] = run
        state.report.update(
            parameters=parameters,
            output=self._render_report_manifest(
                "RUNNING", latest, identity=state.operator_identity,
                execution_mode=state.run_spec.get("execution_mode"),
            ),
        )
        task_metadata = _metadata(getattr(state.task, "metadata", {}))
        task_metadata["optimization_run_key"] = self.run_key
        task_metadata["latest_revision"] = latest
        task_metadata.pop("optimization_publication_draft", None)
        state.task.update(metadata=json.dumps(task_metadata))

    def _record_detail_enrichment(
        self,
        *,
        core_revision: PublishedRevision,
        detail_manifest: Mapping[str, Any],
        detail_presentation: Mapping[str, Any],
    ) -> None:
        """Mark one already-committed core revision as detail-complete."""
        state = self._require_state()
        parameters = _metadata(getattr(state.report, "parameters", {}))
        run = dict(parameters.get("optimization_run") or {})
        latest = dict(run.get("latest_revision") or {})
        if int(latest.get("number") or -1) != core_revision.number:
            raise OptimizationRunIntegrityError(
                "detail enrichment no longer targets the latest core revision"
            )
        if str(latest.get("milestone") or "") != core_revision.milestone:
            raise OptimizationRunIntegrityError(
                "detail enrichment milestone conflicts with its core revision"
            )
        latest.update({
            "detail_status": "complete",
            "detail_source_revision": core_revision.number,
            "detail_manifest": dict(detail_manifest),
            "detail_presentation": dict(detail_presentation),
        })
        revisions = list(run.get("revisions") or [])
        updated = False
        for index, item in enumerate(revisions):
            if isinstance(item, Mapping) and int(item.get("number") or -1) == core_revision.number:
                revisions[index] = dict(latest)
                updated = True
                break
        if not updated:
            raise OptimizationRunIntegrityError(
                "detail enrichment core revision is absent from Report history"
            )
        run["latest_revision"] = latest
        run["revisions"] = revisions
        parameters["optimization_run"] = run
        self._update_block(state.blocks["status"], str(detail_manifest["object_key"]), {
            "type": "optimization_run_status",
            "status": "details_published",
            "summary": {
                "revision": core_revision.number,
                "milestone": core_revision.milestone,
                "overview": dict(core_revision.overview),
                "presentation": dict(detail_presentation),
                "detail_status": "complete",
                "detail_source_revision": core_revision.number,
            },
        })
        state.report.update(
            parameters=parameters,
            output=self._render_report_manifest(
                "RUNNING",
                latest,
                identity=state.operator_identity,
                execution_mode=state.run_spec.get("execution_mode"),
            ),
        )
        task_metadata = _metadata(getattr(state.task, "metadata", {}))
        task_metadata["optimization_run_key"] = self.run_key
        task_metadata["latest_revision"] = latest
        state.task.update(metadata=json.dumps(task_metadata))

    def _download_task_attachment(
        self, task_id: str, descriptor: Mapping[str, Any]
    ) -> bytes:
        if self._artifact_store is None:
            raise OptimizationRunIntegrityError("GraphQL artifact store is unavailable")
        if str(descriptor.get("task_id") or task_id) != task_id:
            raise ValueError("artifact descriptor belongs to a different Task")
        object_key = str(descriptor.get("object_key") or "")
        filename = object_key.rsplit("/", 1)[-1]
        content_type = str(descriptor.get("content_type") or "")
        digest = str(descriptor.get("sha256") or "")
        size_bytes = descriptor.get("size_bytes")
        if not object_key or not filename or not content_type:
            raise ValueError("artifact descriptor is incomplete")
        request = ArtifactTransferRequest(
            operation="READ",
            resource_type="TASK",
            resource_id=task_id,
            artifact_type="TASK_ATTACHMENT",
            filename=filename,
            content_type=content_type,
            size_bytes=int(size_bytes),
            sha256=digest,
        )
        content = self._artifact_store.download_bytes(request)
        if len(content) != request.size_bytes or sha256(content).hexdigest() != request.sha256:
            raise OptimizationRunIntegrityError("downloaded checkpoint artifact failed integrity verification")
        self._verified_artifact_reads.add(
            (object_key, digest, int(size_bytes), content_type)
        )
        return content

    @staticmethod
    def _verify_checkpoint_descriptor(
        *, latest: Mapping[str, Any], manifest: Mapping[str, Any], task_id: str,
        revision_number: int,
    ) -> None:
        """Require the durable pointers to describe exactly the same artifact."""
        fields = ("task_id", "object_key", "content_type", "size_bytes", "sha256", "kind", "source_revision")
        for field in fields:
            if latest.get(field) != manifest.get(field):
                raise ValueError(f"latest and manifest evidence descriptors differ at {field}")
        if latest.get("task_id") != task_id:
            raise ValueError("checkpoint evidence belongs to another Task")
        if latest.get("kind") != "run_evidence" or latest.get("source_revision") != revision_number:
            raise ValueError("checkpoint evidence has an invalid kind or revision")
        object_key = latest.get("object_key")
        if not isinstance(object_key, str) or not object_key or not object_key.rsplit("/", 1)[-1]:
            raise ValueError("checkpoint evidence has an invalid object key/filename")

    @staticmethod
    def _latest_revision_number(report: Any) -> int:
        return int(((_metadata(getattr(report, "parameters", {})).get("optimization_run") or {}).get("latest_revision") or {}).get("number") or 0)

    @staticmethod
    def _latest_revision(report: Any) -> Any:
        return ((_metadata(getattr(report, "parameters", {})).get("optimization_run") or {}).get("latest_revision"))

    @staticmethod
    def _render_report_manifest(
        status: str,
        revision: Optional[Mapping[str, Any]],
        error: Optional[str] = None,
        *,
        identity: OptimizationOperatorIdentity,
        live_progress: Optional[Mapping[str, Any]] = None,
        execution_mode: Optional[str] = None,
    ) -> str:
        def safe(value: Any) -> str:
            return "" if value is None else " ".join(str(value).split()).strip()

        overview = revision.get("overview") if isinstance(revision, Mapping) else {}
        overview = overview if isinstance(overview, Mapping) else {}
        milestone = safe(revision.get("milestone")) if isinstance(revision, Mapping) else ""
        normalized_status = safe(status) or "running"
        coverage = safe(overview.get("coverage_status")) or "pending"
        inventory_coverage = safe(overview.get("inventory_coverage_status")) or coverage
        analysis_coverage = safe(overview.get("analysis_coverage_status")) or coverage
        inspected = overview.get("scorecards_inspected", 0)
        in_scope = overview.get("scorecards_in_scope", 0)
        ranked = overview.get("ranked_score_count", 0)
        unranked = overview.get("unranked_score_count", 0)
        evidence_ranked = overview.get("evidence_ranked_score_count", ranked + unranked)
        cooldown = overview.get("cooldown_excluded_count", 0)
        # ``overview`` is a stakeholder projection.  It must never become an
        # authority for the execution contract; callers pass the frozen mode.
        mode = execution_mode
        if mode == "automatic":
            introduction = (
                "This living report follows the linked procedure from portfolio analysis "
                "through automatic optimization: safe, policy-selected targets may launch "
                "automatically. Champion promotion remains a separate manual decision."
            )
        elif mode == "approval_required":
            introduction = (
                "This living report follows the linked procedure from portfolio analysis "
                "through the human optimization-approval checkpoint and final outcomes. "
                "Champion promotion remains a separate manual decision."
            )
        else:
            introduction = (
                "This living report follows the linked procedure from portfolio "
                "analysis through human decisions and final outcomes."
            )
        lines = [
            f"# {identity.display_title}",
            "",
            introduction,
        ]
        if revision:
            lines.extend([
                f"Latest durable revision: {revision.get('number')}",
                f"Milestone: {milestone}",
            ])
            detail_status = safe(revision.get("detail_status"))
            if detail_status:
                detail_source_revision = revision.get("detail_source_revision")
                detail_line = f"Score and scorecard details: {detail_status.replace('_', ' ')}"
                if detail_source_revision is not None:
                    detail_line += f" (source revision {detail_source_revision})"
                lines.append(detail_line + ".")
        lines.extend([
            "",
            "```block",
            "class: OptimizationRunStatus",
            "```",
            "",
            f"Status: {normalized_status}",
        ])
        current_activity = safe(overview.get("current_activity"))
        next_checkpoint = safe(overview.get("next_checkpoint"))
        if current_activity:
            lines.extend(["", "## What is happening now", "", current_activity])
        if next_checkpoint:
            lines.extend(["", "Next checkpoint: " + next_checkpoint])
        lines.extend([
            "",
            "## Coverage and progress",
            "",
            f"Portfolio inventory coverage: {inventory_coverage.title()}",
            f"Semantic analysis: {analysis_coverage.title()}",
            (
                f"Portfolio: {in_scope} scorecards in scope; {inspected} account "
                f"scorecards inspected to resolve scope; {evidence_ranked} evidence-ranked "
                f"scores; {ranked} eligible candidates; {unranked} policy-deferred or "
                f"structurally blocked scores, including {cooldown} cooldown deferrals."
            ),
        ])
        if isinstance(live_progress, Mapping):
            phase = safe(live_progress.get("phase")).title()
            subphase = safe(live_progress.get("subphase")).replace("_", " ").title()
            phase_label = f"{phase} / {subphase}" if subphase else phase
            current = live_progress.get("current", 0)
            total = live_progress.get("total")
            state_label = safe(live_progress.get("state"))
            unit = safe(live_progress.get("unit"))
            if not unit:
                unit = "scores assessed" if phase.lower() == "assessment" else (
                    "analysis steps complete" if phase.lower() == "diagnosis" else (
                        "artifacts" if phase.lower() == "publication" else "scorecards"
                    )
                )
            progress_text = (
                f"{phase_label}: {current} {unit} inspected"
                if total is None and phase.lower() == "ranking"
                else f"{phase_label}: {current} {unit}"
                if total is None
                else f"{phase_label}: {current} of {total} {unit}"
            )
            lines.extend([
                progress_text,
                safe(live_progress.get("message")),
            ])
            if state_label and state_label.lower() != "active":
                lines.append(f"Live status: {state_label.title()}")
            elapsed_seconds = live_progress.get("elapsed_seconds")
            if isinstance(elapsed_seconds, int) and not isinstance(elapsed_seconds, bool):
                lines.append(f"Elapsed: {elapsed_seconds} seconds")
            live_next_checkpoint = safe(live_progress.get("next_checkpoint"))
            if live_next_checkpoint:
                lines.append("Next live checkpoint: " + live_next_checkpoint)
        for label, key in (
            ("Assessment", "assessment_progress"),
            ("Diagnosis", "diagnosis_coverage"),
            ("Pending approvals", "pending_approval_count"),
        ):
            value = safe(overview.get(key))
            if value:
                lines.append(f"{label}: {value}")
        semantic_authorized = safe(overview.get("semantic_budget_authorized_usd"))
        if semantic_authorized:
            lines.extend([
                "",
                "## Semantic diagnosis budget",
                "",
                f"Budget policy: {safe(overview.get('semantic_budget_policy_version'))}.",
                f"Budget spec schema: {safe(overview.get('semantic_budget_spec_schema_version'))}.",
                f"Ledger schema: {safe(overview.get('semantic_budget_ledger_schema_version'))}.",
                (
                    f"Model: {safe(overview.get('semantic_budget_provider'))}:"
                    f"{safe(overview.get('semantic_budget_model'))}; pricing: "
                    f"{safe(overview.get('semantic_budget_pricing_version'))}."
                ),
                (
                    f"Authorized: ${semantic_authorized}; spent: ${safe(overview.get('semantic_budget_settled_actual_usd'))}; "
                    f"held: ${safe(overview.get('semantic_budget_held_reserved_usd'))}; "
                    f"remaining: ${safe(overview.get('semantic_budget_available_usd'))}."
                ),
                (
                    f"Reservations: {safe(overview.get('semantic_budget_reservation_count'))} total; "
                    f"{safe(overview.get('semantic_budget_reserved_count'))} reserved; "
                    f"{safe(overview.get('semantic_budget_settled_count'))} settled; "
                    f"{safe(overview.get('semantic_budget_unknown_count'))} outcome unknown; "
                    f"{safe(overview.get('semantic_budget_cancelled_count'))} cancelled."
                ),
                (
                    f"Deferred: {safe(overview.get('semantic_budget_deferred_count'))}; "
                    f"failed: {safe(overview.get('semantic_budget_failure_count'))}."
                ),
                (
                    f"Evidence: {safe(overview.get('semantic_budget_evidence_reference'))} "
                    f"digest {safe(overview.get('semantic_budget_evidence_digest'))}."
                ),
            ])
        semantic_issues = overview.get("semantic_diagnosis_issues")
        semantic_issues = semantic_issues if isinstance(semantic_issues, list) else []
        if semantic_issues:
            issue_counts = overview.get("semantic_diagnosis_issue_counts")
            issue_counts = issue_counts if isinstance(issue_counts, Mapping) else {}
            count_summary = ", ".join(
                f"{safe(category)}: {int(count)}"
                for category, count in sorted(issue_counts.items())
                if isinstance(count, int) and count > 0
            )
            lines.extend([
                "",
                "## Semantic diagnosis issues",
                "",
                (
                    f"Affected scores: {len(semantic_issues)}"
                    + (f" ({count_summary})." if count_summary else ".")
                ),
            ])
            for issue in semantic_issues:
                if not isinstance(issue, Mapping):
                    continue
                lines.extend([
                    "",
                    (
                        f"- {safe(issue.get('scorecard_name'))} — "
                        f"{safe(issue.get('score_name'))}"
                    ),
                    f"  - Status: {safe(issue.get('semantic_diagnosis_status'))}",
                    f"  - Next action: {safe(issue.get('next_action'))}",
                    f"  - Rationale: {safe(issue.get('rationale'))}",
                ])
        notes = safe(overview.get("notes"))
        if (
            notes
            or inventory_coverage.lower() == "incomplete"
            or analysis_coverage.lower() == "incomplete"
        ):
            lines.extend(["", "## Limitations", ""])
            if inventory_coverage.lower() == "incomplete":
                lines.append("Portfolio inventory coverage is incomplete, so rankings are partial rather than exact.")
            if analysis_coverage.lower() == "incomplete":
                lines.append("Semantic analysis is incomplete, so some findings still require review or another run.")
            if notes:
                lines.append(notes)
        lines.extend([
            "",
            "The latest stakeholder workbook and immutable evidence are available in this Report's artifact blocks.",
        ])
        if error:
            lines.extend(["", f"Publication error: {safe(error)}"])
        return "\n".join(lines)

    @staticmethod
    def _display_subtitle(identity: OptimizationOperatorIdentity) -> str:
        if identity.kind == "account_wide_portfolio":
            return "All scorecards"
        if identity.kind == "scorecard_scoped_portfolio":
            return identity.display_scope
        if identity.kind == "single_score":
            return "Focused analysis and optimization of one score"
        return "Living optimization analysis"

    def _find_task(self, run_key: str) -> Any:
        # This is intentionally bounded client-side discovery until a dedicated
        # run-key index/claim exists. The deterministic run key still makes a
        # resumed process reuse the same durable state.
        query = """
        query ListOptimizationRunTasks($accountId: String!, $nextToken: String) {
          listTaskByAccountIdAndUpdatedAt(accountId: $accountId, limit: 100, sortDirection: DESC, nextToken: $nextToken) {
            items { id accountId type status target command metadata createdAt updatedAt startedAt completedAt }
            nextToken
          }
        }
        """
        token = None
        while True:
            result = self.client.execute(query, {"accountId": self.account_id, "nextToken": token})
            page = (result or {}).get("listTaskByAccountIdAndUpdatedAt") or {}
            for item in page.get("items") or []:
                if _metadata(item.get("metadata")).get("optimization_run_key") == run_key:
                    return Task.from_dict(item, self.client)
            token = page.get("nextToken")
            if not token:
                return None

    def _find_report(self, task: Any) -> Any:
        query = """
        query GetTaskLinkedOptimizationReport($id: ID!) {
          getTask(id: $id) {
            id
            report {
              id
              taskId
            }
          }
        }
        """
        response = self.client.execute(query, {"id": task.id})
        task_record = response.get("getTask") if isinstance(response, Mapping) else None
        if not isinstance(task_record, Mapping) or task_record.get("id") != task.id:
            raise OptimizationRunPublicationError(
                "Task-linked Report authority could not be resolved"
            )
        report_ref = task_record.get("report")
        if report_ref is None:
            return None
        if (
            not isinstance(report_ref, Mapping)
            or not isinstance(report_ref.get("id"), str)
            or not report_ref.get("id")
            or report_ref.get("taskId") != task.id
        ):
            raise OptimizationRunIntegrityError(
                "Task-linked Report identity is malformed or conflicting"
            )
        report = Report.get_by_id(str(report_ref["id"]), self.client)
        if report is None or getattr(report, "taskId", task.id) != task.id:
            raise OptimizationRunIntegrityError(
                "Task-linked Report could not be loaded with matching identity"
            )
        return report

    def _find_blocks(self, report: Any) -> list[Any]:
        return ReportBlock.list_by_report_id(report.id, self.client, limit=100, max_items=100)

    @staticmethod
    def _find_stages(task: Any) -> list[Any]:
        return list(task.get_stages())

    def _upload_task_attachment(self, task_id: str, name: str, content: bytes) -> str:
        if self._artifact_store is None:
            raise OptimizationRunPublicationError("GraphQL artifact store is unavailable")
        digest = sha256(content).hexdigest()
        content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if name.endswith(".xlsx")
            else "text/markdown"
            if name.endswith(".md")
            else "text/csv"
            if name.endswith(".csv")
            else "application/json"
        )
        write_request = ArtifactTransferRequest(
            operation="WRITE",
            resource_type="TASK",
            resource_id=task_id,
            artifact_type="TASK_ATTACHMENT",
            filename=name,
            content_type=content_type,
            size_bytes=len(content),
            sha256=digest,
        )
        metadata = self._artifact_store.upload_bytes(write_request, content)
        if not isinstance(metadata, Mapping) or metadata.get("sha256") != digest:
            raise OptimizationRunIntegrityError("artifact upload did not preserve its checksum")
        object_key = metadata.get("_s3_key")
        if not isinstance(object_key, str) or not object_key:
            raise OptimizationRunIntegrityError("artifact upload omitted its object key")
        reader = getattr(self._artifact_store, "download_bytes", None)
        if self._verify_uploaded_artifacts and callable(reader):
            read_request = ArtifactTransferRequest(
                operation="READ",
                resource_type="TASK",
                resource_id=task_id,
                artifact_type="TASK_ATTACHMENT",
                filename=name,
                content_type=content_type,
                size_bytes=len(content),
                sha256=digest,
            )
            downloaded = reader(read_request)
            if downloaded != content or sha256(downloaded).hexdigest() != digest:
                raise OptimizationRunIntegrityError("artifact read-back checksum did not match")
            self._verified_artifact_reads.add(
                (object_key, digest, len(content), content_type)
            )
        return object_key
