"""Outside-in tests for the durable optimization run report lifecycle."""

from __future__ import annotations

from datetime import datetime, timezone
from hashlib import sha256
from io import BytesIO
import csv
from copy import deepcopy
import inspect
import json
from types import SimpleNamespace
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook


class _Task:
    created: list["_Task"] = []

    def __init__(self, identifier: str = "task-1", **values):
        self.id = identifier
        self.accountId = values.get("accountId", "account-1")
        self.type = values.get("type")
        self.description = values.get("description")
        self.status = values.get("status", "RUNNING")
        self.metadata = values.get("metadata", {})
        self.attachedFiles = list(values.get("attachedFiles") or [])
        self.currentStageId = values.get("currentStageId")
        self.updates: list[dict] = []

    @classmethod
    def create(cls, **values):
        task = cls(identifier=f"task-{len(cls.created) + 1}", **values)
        cls.created.append(task)
        return task

    def update(self, **values):
        self.updates.append(values)
        self.description = values.get("description", self.description)
        self.status = values.get("status", self.status)
        self.metadata = values.get("metadata", self.metadata)
        self.attachedFiles = list(values.get("attachedFiles", self.attachedFiles))
        self.currentStageId = values.get("currentStageId", self.currentStageId)
        return self


class _Report:
    created: list["_Report"] = []

    def __init__(self, identifier: str = "report-1", **values):
        self.id = identifier
        self.taskId = values["taskId"]
        self.accountId = values["accountId"]
        self.name = values.get("name")
        self.parameters = values.get("parameters", {})
        self.output = values.get("output")
        self.updates: list[dict] = []

    @classmethod
    def create(cls, **values):
        report = cls(identifier=f"report-{len(cls.created) + 1}", **values)
        cls.created.append(report)
        return report

    def update(self, **values):
        self.updates.append(values)
        self.parameters = values.get("parameters", self.parameters)
        self.output = values.get("output", self.output)
        return self


class _Block:
    created: list["_Block"] = []

    def __init__(self, identifier: str, **values):
        self.id = identifier
        self.reportId = values["reportId"]
        self.position = values["position"]
        self.name = values["name"]
        self.type = values["type"]
        self.output = values.get("output")
        self.attachedFiles = list(values.get("attachedFiles") or [])
        self.updates: list[dict] = []

    @classmethod
    def create(cls, **values):
        block = cls(f"block-{len(cls.created) + 1}", **values)
        cls.created.append(block)
        return block

    def update(self, **values):
        self.updates.append(values)
        self.output = values.get("output", self.output)
        self.attachedFiles = list(values.get("attachedFiles", self.attachedFiles))
        return self


class _TaskStage:
    created: list["_TaskStage"] = []

    def __init__(self, identifier: str, **values):
        self.id = identifier
        self.taskId = values["taskId"]
        self.name = values["name"]
        self.order = values["order"]
        self.status = values["status"]
        self.statusMessage = values.get("statusMessage")
        self.processedItems = values.get("processedItems")
        self.totalItems = values.get("totalItems")
        self.startedAt = values.get("startedAt")
        self.completedAt = values.get("completedAt")
        self.updates: list[dict] = []

    @classmethod
    def create(cls, **values):
        stage = cls(f"stage-{len(cls.created) + 1}", **values)
        cls.created.append(stage)
        return stage

    def update(self, **values):
        self.updates.append(values)
        for key, value in values.items():
            setattr(self, key, value)
        return self


class _ArtifactStore:
    def __init__(self):
        self.uploads = []
        self.downloads = []
        self.content_by_key: dict[str, bytes] = {}

    def upload_bytes(self, request, content, **_kwargs):
        self.uploads.append((request, content))
        object_key = f"{request.resource_type.lower()}s/{request.resource_id}/{request.filename}"
        self.content_by_key[object_key] = bytes(content)
        return {
            "_s3_key": object_key,
            "sha256": request.sha256,
            "size_bytes": request.size_bytes,
            "content_type": request.content_type,
        }

    def download_bytes(self, request):
        self.downloads.append(request)
        object_key = f"{request.resource_type.lower()}s/{request.resource_id}/{request.filename}"
        return self.content_by_key[object_key]


@pytest.fixture(autouse=True)
def _reset_fakes():
    _Task.created = []
    _Report.created = []
    _Block.created = []
    _TaskStage.created = []


def _safe_view():
    return {
        "overview": {"headline": "Daily optimization findings"},
        "portfolio": [
            {
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "valid_feedback_count": 250,
                "reviewed_disagreements": 70,
                "disagreement_rate": 0.28,
                "reviewed_error_opportunity": 70,
                "readiness": "repair_required",
                "next_action": "repair_guidelines",
            }
        ],
        "priorities": [
            {
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "opportunity": 70,
                "rationale": "=must remain literal text",
                "next_action": "repair_guidelines",
                "dashboard_url": "https://dashboard.example/reports/summary",
            }
        ],
        "feedback_investment": [
            {
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "recommendation": "continue_broad_collection",
                "rationale": "More reviewed examples are needed.",
            }
        ],
        "questions_and_issues": [
            {
                "kind": "guideline",
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "finding": "Guidelines need repair.",
                "next_action": "repair_guidelines",
            }
        ],
        "optimization_outcomes": [],
        "definitions": {"Reviewed error opportunity": "Reviewed disagreements in the window."},
    }


def _service(monkeypatch, *, block_class=_Block, dashboard_base_url=None):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", block_class)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    return run_report.OptimizationRunReportService(
        client=SimpleNamespace(),
        account_id="account-1",
        run_key="daily-v1-2026-07-29",
        report_configuration_id="config-1",
        dashboard_base_url=dashboard_base_url,
        now=lambda: datetime(2026, 7, 29, 12, 0, tzinfo=timezone.utc),
        task_lookup=lambda _: None,
        report_lookup=lambda _: None,
        block_lookup=lambda _: [],
        stage_lookup=lambda task: [stage for stage in _TaskStage.created if stage.taskId == task.id],
        artifact_uploader=lambda task_id, name, _: f"tasks/{task_id}/{name}",
        publication_id_factory=lambda: "test",
    )


def _semantic_service(monkeypatch, *, store=None, run_key="semantic-run-1"):
    from plexus.optimization import run_report
    from plexus.optimization.semantic_budget import SemanticBudgetSpec

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    store = store or _ArtifactStore()
    spec = SemanticBudgetSpec(
        max_cost_usd="1.00",
        pricing_version="openai-2025-08-07-v1",
    )
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(),
        account_id="account-1",
        run_key=run_key,
        report_configuration_id="config-1",
        artifact_store=store,
        task_lookup=lambda _: None,
        report_lookup=lambda _: None,
        block_lookup=lambda _: [],
        stage_lookup=lambda task: [
            stage for stage in _TaskStage.created if stage.taskId == task.id
        ],
        publication_id_factory=lambda: "semantic-test",
    )
    state = service.start_or_resume({
        "scope": {},
        "semantic_budget": spec.to_dict(),
    })
    return service, state, store, spec


def test_run_key_is_a_deterministic_fingerprint_of_account_and_frozen_spec():
    from plexus.optimization.run_report import optimization_run_key

    left = optimization_run_key("account-1", {"window": {"start": "2026-04-30T00:00:00Z"}, "scope": ["b", "a"]})
    right = optimization_run_key("account-1", {"scope": ["b", "a"], "window": {"start": "2026-04-30T00:00:00Z"}})
    changed = optimization_run_key("account-1", {"scope": ["a"], "window": {"start": "2026-04-30T00:00:00Z"}})

    assert left == right
    assert left.startswith("optimization-run-")
    assert changed != left


def test_semantic_ledger_is_committed_as_an_immutable_task_attachment_without_a_workbook(
    monkeypatch,
):
    from plexus.optimization.semantic_budget import SemanticBudgetLedger, SemanticCallPlan

    service, state, store, spec = _semantic_service(monkeypatch)
    ledger = SemanticBudgetLedger(run_key=service.run_key, spec=spec)
    ledger.reserve(SemanticCallPlan(
        run_key=service.run_key,
        target_id="scorecard-1:score-1",
        call_site="rubric_consistency",
        attempt=1,
        max_attempts=2,
        provider="openai",
        model="gpt-5-mini-2025-08-07",
        pricing_version=spec.pricing_version,
        max_input_tokens=1_000,
        max_output_tokens=100,
    ))

    pointer = service.persist_semantic_budget_ledger(ledger.to_dict())

    assert pointer["ledger_revision"] == 1
    assert pointer["sha256"] == ledger.digest()
    assert pointer["kind"] == "semantic_budget_ledger"
    filenames = [
        request.filename for request, _content in store.uploads
        if request.filename.startswith("optimization-semantic-ledger-")
    ]
    assert filenames == [
        "optimization-semantic-ledger-r000001-semantic-test.json"
    ]
    assert not any(name.endswith(".xlsx") for name in filenames)
    assert state.report.parameters["optimization_run"]["semantic_budget_latest"] == pointer
    assert service.load_semantic_budget_ledger() == ledger.to_dict()


def test_same_run_key_rejects_a_changed_semantic_budget_or_price_version(monkeypatch):
    from plexus.optimization import run_report

    service, _state, _store, spec = _semantic_service(monkeypatch)
    base = {"scope": {}, "semantic_budget": spec.to_dict()}

    with pytest.raises(run_report.OptimizationRunIntegrityError):
        service.start_or_resume({
            **base,
            "semantic_budget": {**spec.to_dict(), "max_cost_usd": "2"},
        })
    with pytest.raises(run_report.OptimizationRunIntegrityError):
        service.start_or_resume({
            **base,
            "semantic_budget": {
                **spec.to_dict(),
                "pricing_version": "openai-future-v2",
            },
        })


def test_semantic_ledger_commits_before_contact_and_after_settlement_and_replay_is_idempotent(
    monkeypatch,
):
    from plexus.optimization.semantic_budget import (
        SemanticBudgetLedger,
        SemanticCallPlan,
        SemanticUsage,
    )

    service, _state, store, spec = _semantic_service(monkeypatch)
    ledger = SemanticBudgetLedger(run_key=service.run_key, spec=spec)
    reservation = ledger.reserve(SemanticCallPlan(
        run_key=service.run_key,
        target_id="scorecard-1:score-1",
        call_site="rubric_consistency",
        attempt=1,
        max_attempts=2,
        provider="openai",
        model="gpt-5-mini-2025-08-07",
        pricing_version=spec.pricing_version,
        max_input_tokens=1_000,
        max_output_tokens=100,
    ))
    reserved_pointer = service.persist_semantic_budget_ledger(ledger.to_dict())
    assert service.persist_semantic_budget_ledger(ledger.to_dict()) == reserved_pointer

    ledger.settle(
        reservation["reservation_id"],
        SemanticUsage(input_tokens=500, output_tokens=50, provider_request_id="req-1"),
    )
    settled_pointer = service.persist_semantic_budget_ledger(ledger.to_dict())

    assert reserved_pointer["ledger_revision"] == 1
    assert settled_pointer["ledger_revision"] == 2
    assert len([
        request for request, _content in store.uploads
        if request.filename.startswith("optimization-semantic-ledger-")
    ]) == 2
    assert service.load_semantic_budget_ledger() == ledger.to_dict()


def test_semantic_ledger_publication_failure_does_not_advance_the_report_commit_pointer(
    monkeypatch,
):
    from plexus.optimization import run_report
    from plexus.optimization.semantic_budget import SemanticBudgetLedger, SemanticCallPlan

    service, state, _store, spec = _semantic_service(monkeypatch)
    ledger = SemanticBudgetLedger(run_key=service.run_key, spec=spec)
    ledger.reserve(SemanticCallPlan(
        run_key=service.run_key,
        target_id="scorecard-1:score-1",
        call_site="rubric_consistency",
        attempt=1,
        max_attempts=1,
        provider="openai",
        model="gpt-5-mini-2025-08-07",
        pricing_version=spec.pricing_version,
        max_input_tokens=100,
        max_output_tokens=10,
    ))
    original_update = state.report.update

    def interrupted_update(**_values):
        raise RuntimeError("simulated commit interruption")

    state.report.update = interrupted_update
    with pytest.raises(run_report.OptimizationRunRetryablePublicationError):
        service.persist_semantic_budget_ledger(ledger.to_dict())
    assert "semantic_budget_latest" not in state.report.parameters["optimization_run"]

    state.report.update = original_update
    pointer = service.persist_semantic_budget_ledger(ledger.to_dict())
    assert pointer["ledger_revision"] == 1


def test_semantic_ledger_load_rejects_checksum_or_frozen_spec_mismatch(monkeypatch):
    from plexus.optimization import run_report
    from plexus.optimization.semantic_budget import SemanticBudgetLedger, SemanticCallPlan

    service, state, store, spec = _semantic_service(monkeypatch)
    ledger = SemanticBudgetLedger(run_key=service.run_key, spec=spec)
    ledger.reserve(SemanticCallPlan(
        run_key=service.run_key,
        target_id="scorecard-1:score-1",
        call_site="rubric_consistency",
        attempt=1,
        max_attempts=1,
        provider="openai",
        model="gpt-5-mini-2025-08-07",
        pricing_version=spec.pricing_version,
        max_input_tokens=100,
        max_output_tokens=10,
    ))
    pointer = service.persist_semantic_budget_ledger(ledger.to_dict())
    store.content_by_key[pointer["object_key"]] = b"{}"

    with pytest.raises(run_report.OptimizationRunIntegrityError):
        service.load_semantic_budget_ledger()

    state.report.parameters["optimization_run"]["run_spec"]["semantic_budget"] = {
        **spec.to_dict(),
        "max_cost_usd": "2",
    }
    with pytest.raises(run_report.OptimizationRunIntegrityError):
        service.persist_semantic_budget_ledger(ledger.to_dict())


@pytest.mark.parametrize(
    ("transition", "reason_field", "corruption"),
    [
        ("cancel_pre_contact", "cancellation_reason", "missing"),
        ("cancel_pre_contact", "cancellation_reason", "blank"),
        ("mark_outcome_unknown", "outcome_unknown_reason", "missing"),
        ("mark_outcome_unknown", "outcome_unknown_reason", "blank"),
    ],
)
def test_semantic_ledger_checkpoint_load_rejects_evidence_free_states(
    monkeypatch, transition, reason_field, corruption
):
    from plexus.optimization import run_report
    from plexus.optimization.semantic_budget import (
        SemanticBudgetLedger,
        SemanticCallPlan,
        canonical_json_bytes,
    )

    service, state, store, spec = _semantic_service(monkeypatch)
    ledger = SemanticBudgetLedger(run_key=service.run_key, spec=spec)
    reservation = ledger.reserve(SemanticCallPlan(
        run_key=service.run_key,
        target_id="scorecard-1:score-1",
        call_site="rubric_consistency",
        attempt=1,
        max_attempts=1,
        provider="openai",
        model="gpt-5-mini-2025-08-07",
        pricing_version=spec.pricing_version,
        max_input_tokens=100,
        max_output_tokens=10,
    ))
    service.persist_semantic_budget_ledger(ledger.to_dict())
    getattr(ledger, transition)(reservation["reservation_id"], reason="evidence")
    service.persist_semantic_budget_ledger(ledger.to_dict())

    durable_ledger = ledger.to_dict()
    if corruption == "missing":
        durable_ledger["entries"][0].pop(reason_field)
    else:
        durable_ledger["entries"][0][reason_field] = ""
    content = canonical_json_bytes(durable_ledger)
    pointer = state.report.parameters["optimization_run"]["semantic_budget_latest"]
    pointer["size_bytes"] = len(content)
    pointer["sha256"] = sha256(content).hexdigest()
    store.content_by_key[pointer["object_key"]] = content

    with pytest.raises(run_report.OptimizationRunIntegrityError):
        service.load_semantic_budget_ledger()


def test_start_or_resume_uses_one_running_task_report_and_fixed_blocks(monkeypatch):
    service = _service(monkeypatch)

    first = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})
    second = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    assert first.task.id == second.task.id == "task-1"
    assert first.report.id == second.report.id == "report-1"
    assert len(_Task.created) == 1
    assert len(_Report.created) == 1
    assert [(block.position, block.name, block.type) for block in _Block.created] == [
        (0, "Run Status", "OptimizationRunStatus"),
        (1, "Decision Evidence", "OptimizationDecisionEvidence"),
        (2, "Stakeholder Workbook", "OptimizationStakeholderWorkbook"),
    ]
    assert first.task.status == "RUNNING"
    metadata = json.loads(first.task.metadata)
    assert metadata["optimization_run_key"] == "daily-v1-2026-07-29"
    assert metadata["attempt_id"]


def test_start_is_self_identifying_in_the_task_report_list_and_cover(monkeypatch):
    service = _service(monkeypatch)

    state = service.start_or_resume({"scope": {}})

    assert state.task.type == "OptimizationRunReport"
    assert state.task.description == "Account-wide optimization portfolio — All scorecards"
    assert state.report.name == "Account-wide optimization portfolio"
    assert state.report.parameters["_display_title"] == "Account-wide optimization portfolio"
    assert state.report.parameters["_display_subtitle"] == "Periodic analysis across all scorecards"
    assert state.report.parameters["optimization_run"]["operator_identity"] == {
        "kind": "account_wide_portfolio",
        "display_title": "Account-wide optimization portfolio",
        "display_scope": "All scorecards",
    }
    assert state.report.output.startswith("# Account-wide optimization portfolio")
    assert "This living report follows the linked procedure" in state.report.output
    assert "Scope:" not in state.report.output
    assert "Current phase:" not in state.report.output
    assert "Status: running" in state.report.output
    assert "```block\nclass: OptimizationRunStatus\n```" in state.report.output


def test_report_cover_describes_the_execution_mode_without_changing_promotion_authority(monkeypatch):
    automatic = _service(monkeypatch).start_or_resume({"scope": {}, "execution_mode": "automatic"})

    assert "safe, policy-selected targets may launch automatically" in automatic.report.output
    assert "Champion promotion remains a separate manual decision" in automatic.report.output
    assert "human optimization-approval checkpoint" not in automatic.report.output

    _Task.created.clear()
    _Report.created.clear()
    _Block.created.clear()
    approval_required = _service(monkeypatch).start_or_resume(
        {"scope": {}, "execution_mode": "approval_required"}
    )

    assert "human optimization-approval checkpoint" in approval_required.report.output
    assert "safe, policy-selected targets may launch automatically" not in approval_required.report.output
    assert "Champion promotion remains a separate manual decision" in approval_required.report.output


def test_milestone_cover_projects_safe_progress_and_preserves_identity_on_finalize(monkeypatch):
    service = _service(monkeypatch)
    state = service.start_or_resume({"scope": {}})
    view = _safe_view()
    view["overview"] = {
        "headline": "Daily optimization findings",
        "lifecycle_status": "running",
        "current_activity": "Checking deterministic readiness across the ranked portfolio.",
        "next_checkpoint": "Semantic diagnosis begins after assessment is durable.",
        "coverage_status": "incomplete",
        "inventory_coverage_status": "incomplete",
        "analysis_coverage_status": "pending",
        "ranking_window": "2026-05-01 through 2026-07-29 UTC",
        "scorecards_inspected": 56,
        "scorecards_in_scope": 4,
        "evidence_ranked_score_count": 110,
        "ranked_score_count": 18,
        "unranked_score_count": 92,
        "cooldown_excluded_count": 7,
        "assessment_progress": "12 of 18 ranked scores complete",
        "diagnosis_coverage": "0 of 10 selected diagnoses complete; 0 failed",
        "pending_approval_count": 0,
        "notes": "Coverage is incomplete, so priorities are partial rather than exact.",
    }

    service.publish_milestone(
        "assessment",
        {"coverage": {"complete": False}, "restricted": {"score_id": "opaque-score"}},
        stakeholder_view=view,
    )

    cover = state.report.output
    assert cover.startswith("# Account-wide optimization portfolio")
    assert "Current phase:" not in cover
    assert "Checking deterministic readiness" in cover
    assert "Semantic diagnosis begins" in cover
    assert "Portfolio inventory coverage: Incomplete" in cover
    assert "Semantic analysis: Pending" in cover
    assert "4 scorecards in scope" in cover
    assert "56 account scorecards inspected to resolve scope" in cover
    assert "110 evidence-ranked scores" in cover
    assert "18 eligible candidates" in cover
    assert "7 cooldown deferrals" in cover
    assert "12 of 18 ranked scores complete" in cover
    assert "Coverage is incomplete" in cover
    assert "opaque-score" not in cover

    service.finalize(status="incomplete")

    assert state.report.output.startswith("# Account-wide optimization portfolio")
    assert "Status: incomplete" in state.report.output
    assert "Checking deterministic readiness" in state.report.output


def test_report_cover_explains_semantic_budget_and_unknown_outcomes_without_raw_content():
    from plexus.optimization.operator_identity import optimization_operator_identity
    from plexus.optimization.run_report import OptimizationRunReportService

    cover = OptimizationRunReportService._render_report_manifest(
        "incomplete",
        {
            "number": 7,
            "milestone": "finalization",
            "overview": {
                "coverage_status": "complete",
                "inventory_coverage_status": "complete",
                "analysis_coverage_status": "incomplete",
                "semantic_budget_policy_version": "semantic-budget-policy-v1",
                "semantic_budget_spec_schema_version": "semantic-budget-v1",
                "semantic_budget_ledger_schema_version": "semantic-budget-ledger-v1",
                "semantic_budget_pricing_version": "openai-2025-08-07-v1",
                "semantic_budget_provider": "openai",
                "semantic_budget_model": "gpt-5-mini-2025-08-07",
                "semantic_budget_authorized_usd": "1",
                "semantic_budget_settled_actual_usd": "0.000045",
                "semantic_budget_held_reserved_usd": "0.0009",
                "semantic_budget_available_usd": "0.999055",
                "semantic_budget_reservation_count": 4,
                "semantic_budget_reserved_count": 1,
                "semantic_budget_settled_count": 1,
                "semantic_budget_unknown_count": 1,
                "semantic_budget_cancelled_count": 1,
                "semantic_budget_deferred_count": 0,
                "semantic_budget_failure_count": 0,
                "semantic_budget_evidence_reference": "semantic-budget-ledger:r000007",
                "semantic_budget_evidence_digest": "a" * 64,
            },
        },
        identity=optimization_operator_identity(scope={}),
    )

    assert "Semantic diagnosis budget" in cover
    assert "Budget policy: semantic-budget-policy-v1" in cover
    assert "Budget spec schema: semantic-budget-v1" in cover
    assert "Ledger schema: semantic-budget-ledger-v1" in cover
    assert "Model: openai:gpt-5-mini-2025-08-07" in cover
    assert "Authorized: $1; spent: $0.000045; held: $0.0009; remaining: $0.999055." in cover
    assert "Reservations: 4 total; 1 reserved; 1 settled; 1 outcome unknown; 1 cancelled." in cover
    assert "Deferred: 0; failed: 0." in cover
    assert "semantic-budget-ledger:r000007" in cover
    assert "private prompt" not in cover


def test_start_creates_the_fixed_task_stages(monkeypatch):
    service = _service(monkeypatch)

    state = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    assert [(stage.order, stage.name, stage.status) for stage in _TaskStage.created] == [
        (0, "preflight", "RUNNING"),
        (1, "ranking", "PENDING"),
        (2, "assessment", "PENDING"),
        (3, "diagnosis", "PENDING"),
        (4, "approval", "PENDING"),
        (5, "optimization", "PENDING"),
        (6, "review", "PENDING"),
        (7, "finalization", "PENDING"),
    ]
    assert all(stage.taskId == state.task.id for stage in _TaskStage.created)


def test_durable_milestones_advance_visible_task_stages_in_lifecycle_order(monkeypatch):
    service = _service(monkeypatch)
    service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    service.publish_milestone(
        "started", {"coverage": {"complete": False}}, stakeholder_view=_safe_view(),
    )
    assert [stage.status for stage in _TaskStage.created] == [
        "COMPLETED", "RUNNING", "PENDING", "PENDING", "PENDING", "PENDING", "PENDING", "PENDING",
    ]
    assert _Task.created[0].currentStageId == next(
        stage.id for stage in _TaskStage.created if stage.name == "ranking"
    )

    for milestone, running_stage in (
        ("ranking", "assessment"),
        ("assessment", "diagnosis"),
        ("diagnosis", "approval"),
        ("approval", "approval"),
        ("optimization", "optimization"),
        ("optimization_review", "review"),
        ("finalization", "finalization"),
    ):
        service.publish_milestone(
            milestone, {"coverage": {"complete": True}}, stakeholder_view=_safe_view(),
        )
        target = next(stage for stage in _TaskStage.created if stage.name == running_stage)
        assert target.status == "RUNNING"
        assert _Task.created[0].currentStageId == target.id
        assert all(
            stage.status == "COMPLETED"
            for stage in _TaskStage.created
            if stage.order < target.order
        )
        assert all(
            stage.status == "PENDING"
            for stage in _TaskStage.created
            if stage.order > target.order
        )


def test_finalize_completes_every_visible_task_stage(monkeypatch):
    service = _service(monkeypatch)
    service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})
    service.publish_milestone(
        "ranking", {"coverage": {"complete": True}}, stakeholder_view=_safe_view(),
    )

    service.finalize(status="incomplete")

    assert all(stage.status == "COMPLETED" for stage in _TaskStage.created)
    assert all(stage.completedAt is not None for stage in _TaskStage.created)


def test_fixed_blocks_are_immediately_backed_by_compact_artifact_envelopes(monkeypatch):
    service = _service(monkeypatch)

    service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    for block in _Block.created:
        envelope = __import__("json").loads(block.output)
        assert envelope["output_compacted"] is True
        assert envelope["output_attachment"] in block.attachedFiles

    workbook_block = next(block for block in _Block.created if block.name == "Stakeholder Workbook")
    assert workbook_block.attachedFiles == ["tasks/task-1/optimization-workbook-r0000.xlsx"]


def test_resume_rejects_a_different_frozen_scope_for_the_same_run_key(monkeypatch):
    service = _service(monkeypatch)
    service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    from plexus.optimization.run_report import OptimizationRunIntegrityError

    with pytest.raises(OptimizationRunIntegrityError, match="frozen run specification"):
        service.start_or_resume({"window": {"start": "2026-05-01T00:00:00Z"}})


def test_recovery_reuses_existing_task_report_and_fixed_blocks(monkeypatch):
    first_service = _service(monkeypatch)
    first = first_service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})
    recovered = _service(monkeypatch)
    recovered._task_lookup = lambda _: first.task
    recovered._report_lookup = lambda _: first.report
    recovered._block_lookup = lambda _: list(_Block.created)

    resumed = recovered.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    assert resumed.task is first.task
    assert resumed.report is first.report
    assert len(_Task.created) == 1
    assert len(_Report.created) == 1
    assert len(_Block.created) == 3


def test_recovery_restores_lost_task_identity_from_its_existing_report(monkeypatch):
    run_spec = {"window": {"start": "2026-04-30T00:00:00Z"}}
    first_service = _service(monkeypatch)
    first = first_service.start_or_resume(run_spec)
    original_attempt_id = first.report.parameters["optimization_run"]["attempt_id"]
    rewritten_metadata = json.loads(first.task.metadata)
    for key in (
        "optimization_run_key",
        "attempt_id",
        "lifecycle_version",
        "run_spec",
        "operator_identity",
    ):
        rewritten_metadata.pop(key, None)
    rewritten_metadata["procedure_id"] = "procedure-1"
    first.task.metadata = json.dumps(rewritten_metadata)

    recovered = _service(monkeypatch)
    recovered._existing_task = first.task
    recovered._uses_existing_task = True
    recovered._report_lookup = lambda task: first.report if task.id == first.task.id else None
    recovered._block_lookup = lambda _: list(_Block.created)

    resumed = recovered.start_or_resume(run_spec)

    restored_metadata = json.loads(first.task.metadata)
    assert resumed.task is first.task
    assert resumed.report is first.report
    assert restored_metadata["optimization_run_key"] == recovered.run_key
    assert restored_metadata["attempt_id"] == original_attempt_id
    assert restored_metadata["run_spec"] == {
        **run_spec,
        "execution_mode": "approval_required",
    }
    assert restored_metadata["procedure_id"] == "procedure-1"
    assert len(_Report.created) == 1
    assert len(_Block.created) == 3


def test_default_report_lookup_uses_the_exact_task_relationship(monkeypatch):
    from plexus.optimization import run_report

    linked_report = SimpleNamespace(id="report-linked", taskId="task-1")

    class _Client:
        def execute(self, query, variables):
            assert "getTask" in query
            assert variables == {"id": "task-1"}
            return {
                "getTask": {
                    "id": "task-1",
                    "report": {"id": "report-linked", "taskId": "task-1"},
                }
            }

    monkeypatch.setattr(
        run_report.Report,
        "get_by_id",
        lambda report_id, client: linked_report
        if report_id == "report-linked"
        else None,
    )
    monkeypatch.setattr(
        run_report.Report,
        "list_by_account_id",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("account-wide Report scans are not an identity lookup")
        ),
    )
    service = run_report.OptimizationRunReportService(
        client=_Client(),
        account_id="account-1",
        run_key="run-1",
        artifact_uploader=lambda *_args: "unused",
    )

    found = service._find_report(SimpleNamespace(id="task-1"))

    assert found is linked_report


@pytest.mark.parametrize(
    ("record", "mutation"),
    [
        ("task", "missing"),
        ("task", "malformed"),
        ("report", "missing"),
        ("report", "malformed"),
        ("report", "mismatch"),
    ],
)
def test_existing_attempt_requires_exact_frozen_identity_before_recovery(
    monkeypatch, record, mutation,
):
    from plexus.optimization import run_report

    requested_spec = {"scope": {}, "window": {"start": "2026-04-30T00:00:00Z"}}
    first_service = _service(monkeypatch)
    first = first_service.start_or_resume(requested_spec)
    if record == "task":
        metadata = json.loads(first.task.metadata)
        if mutation == "missing":
            metadata.pop("run_spec")
        elif mutation == "malformed":
            metadata["run_spec"] = "not-an-object"
        first.task.metadata = json.dumps(metadata)
    else:
        run = first.report.parameters["optimization_run"]
        if mutation == "missing":
            run.pop("run_spec")
        elif mutation == "malformed":
            run["run_spec"] = ["not", "an", "object"]
        else:
            run["run_spec"] = {"scope": {"scorecard_ids": ["other"]}}

    recovered = _service(monkeypatch)
    recovered._task_lookup = lambda _: first.task
    recovered._report_lookup = lambda _: first.report
    recovered._block_lookup = lambda _: list(_Block.created)

    with pytest.raises(run_report.OptimizationRunIntegrityError):
        recovered.start_or_resume(requested_spec)
    assert recovered._state is None
    assert first.task.status == "FAILED"


def test_recovery_loads_the_latest_durable_evidence_through_the_authorized_task_attachment(monkeypatch):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    store = _ArtifactStore()
    run_spec = {"window": {"start": "2026-04-30T00:00:00Z"}, "scope": {}}
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="daily-v1-2026-07-29",
        report_configuration_id="config-1", artifact_store=store,
        now=lambda: datetime(2026, 7, 29, 12, 0, tzinfo=timezone.utc),
        task_lookup=lambda _: None, report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda task: [stage for stage in _TaskStage.created if stage.taskId == task.id],
    )
    first = service.start_or_resume(run_spec)
    evidence = {
        "run_key": "daily-v1-2026-07-29",
        "run_spec": first.run_spec,
        "coverage": {"complete": True},
        "rank": {"coverage": {"complete": True}, "ranked": []},
        "assessments": [],
    }
    service.publish_milestone("assessment", evidence, stakeholder_view=_safe_view())

    recovered = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="daily-v1-2026-07-29",
        report_configuration_id="config-1", artifact_store=store,
        now=lambda: datetime(2026, 7, 29, 12, 5, tzinfo=timezone.utc),
        task_lookup=lambda _: first.task, report_lookup=lambda _: first.report,
        block_lookup=lambda _: list(_Block.created),
        stage_lookup=lambda task: [stage for stage in _TaskStage.created if stage.taskId == task.id],
    )
    resumed = recovered.start_or_resume(run_spec)

    checkpoint = recovered.load_latest_checkpoint()

    assert resumed.task.id == first.task.id
    assert resumed.report.id == first.report.id
    assert checkpoint["milestone"] == "assessment"
    assert checkpoint["evidence"] == evidence
    assert checkpoint["task_terminal"] is False
    assert store.downloads[-1].operation == "READ"
    assert store.downloads[-1].artifact_type == "TASK_ATTACHMENT"


def test_interrupted_publication_retry_reuses_verified_artifacts_for_the_same_logical_revision(monkeypatch):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)

    class FailFirstManifestStore(_ArtifactStore):
        def __init__(self):
            super().__init__()
            self.failed = False

        def upload_bytes(self, request, content, **kwargs):
            if "optimization-revision-r0001" in request.filename and not self.failed:
                self.failed = True
                raise RuntimeError("simulated interrupted publication")
            return super().upload_bytes(request, content, **kwargs)

    store = FailFirstManifestStore()
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="daily-v1-2026-07-29",
        report_configuration_id="config-1", artifact_store=store,
        publication_id_factory=lambda: "publication-a",
        now=lambda: datetime(2026, 7, 29, 12, 0, tzinfo=timezone.utc),
        task_lookup=lambda _: None, report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda task: [stage for stage in _TaskStage.created if stage.taskId == task.id],
    )
    run_spec = {"scope": {}, "execution_mode": "automatic"}
    state = service.start_or_resume(run_spec)

    with pytest.raises(run_report.OptimizationRunPublicationError):
        service.publish_milestone(
            "ranking", {"run_key": "daily-v1-2026-07-29", "coverage": {"complete": True}},
            stakeholder_view=_safe_view(),
        )
    # A manifest upload interruption has not crossed latest_revision, so this
    # remains the same active attempt rather than a failed/replaced Task.
    assert service._state.task.status == "RUNNING"
    assert json.loads(service._state.task.metadata).get("optimization_run_final_status") is None
    draft = json.loads(service._state.task.metadata)["optimization_publication_draft"]
    assert draft["generated_at"] == "2026-07-29T12:00:00Z"
    assert any(
        path.rsplit("/", 1)[-1].startswith("optimization-publication-draft-r0001-")
        for path in service._state.task.attachedFiles
    )
    recovered = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="daily-v1-2026-07-29",
        report_configuration_id="config-1", artifact_store=store,
        publication_id_factory=lambda: "publication-b",
        now=lambda: datetime(2026, 7, 29, 12, 5, tzinfo=timezone.utc),
        task_lookup=lambda _: state.task,
        report_lookup=lambda _: state.report,
        block_lookup=lambda _: list(_Block.created),
        stage_lookup=lambda task: [stage for stage in _TaskStage.created if stage.taskId == task.id],
    )
    recovered.start_or_resume(run_spec)
    downloads_before_retry = len(store.downloads)
    recovered.publish_milestone(
        "ranking", {"run_key": "daily-v1-2026-07-29", "coverage": {"complete": True}},
        stakeholder_view=_safe_view(),
    )
    assert "optimization_publication_draft" not in json.loads(state.task.metadata)

    evidence_names = [
        request.filename for request, _content in store.uploads
        if request.filename.startswith("optimization-evidence-r0001")
    ]
    assert evidence_names == ["optimization-evidence-r0001-publication-a.json"]
    workbook_names = [
        request.filename for request, _content in store.uploads
        if request.filename.startswith("optimization-workbook-r0001")
    ]
    assert workbook_names == ["optimization-workbook-r0001-publication-a.xlsx"]
    score_artifact_names = [
        request.filename for request, _content in store.uploads
        if request.filename.startswith(("score-", "scorecard-"))
    ]
    assert len(score_artifact_names) == 4
    retry_download_names = [
        request.filename for request in store.downloads[downloads_before_retry:]
    ]
    assert any(
        name.startswith("optimization-publication-draft-r0001-")
        for name in retry_download_names
    )
    assert not any(
        name.startswith(("score-", "scorecard-"))
        for name in retry_download_names
    )


@pytest.mark.parametrize(
    "mutation",
    [
        "missing_run_spec", "frozen_spec_mismatch", "malformed_json",
        "milestone_mismatch", "checksum_mismatch", "descriptor_mismatch",
    ],
)
def test_recovery_rejects_corrupt_committed_evidence(monkeypatch, mutation):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    store = _ArtifactStore()
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="daily-v1-2026-07-29",
        report_configuration_id="config-1", artifact_store=store,
        now=lambda: datetime(2026, 7, 29, 12, 0, tzinfo=timezone.utc),
        task_lookup=lambda _: None, report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda task: [stage for stage in _TaskStage.created if stage.taskId == task.id],
        publication_id_factory=lambda: "test",
    )
    state = service.start_or_resume({"scope": {}})
    service.publish_milestone(
        "ranking",
        {"run_key": service.run_key, "run_spec": {"scope": {}}, "coverage": {"complete": True}},
        stakeholder_view=_safe_view(),
    )
    latest = state.report.parameters["optimization_run"]["latest_revision"]
    evidence_key = latest["evidence"]["object_key"]
    if mutation in {"missing_run_spec", "frozen_spec_mismatch", "malformed_json"}:
        evidence = json.loads(store.content_by_key[evidence_key])
        if mutation == "missing_run_spec":
            evidence.pop("run_spec")
        elif mutation == "frozen_spec_mismatch":
            evidence["run_spec"] = {"scope": {"scorecard_ids": ["other"]}}
        content = (
            b"{malformed-json"
            if mutation == "malformed_json"
            else json.dumps(evidence, sort_keys=True, separators=(",", ":")).encode()
        )
        store.content_by_key[evidence_key] = content
        latest["evidence"]["size_bytes"] = len(content)
        latest["evidence"]["sha256"] = __import__("hashlib").sha256(content).hexdigest()
        manifest_key = latest["manifest"]["object_key"]
        manifest = json.loads(store.content_by_key[manifest_key])
        descriptor = next(item for item in manifest["artifacts"] if item["kind"] == "run_evidence")
        descriptor.update(latest["evidence"])
        manifest_content = json.dumps(manifest, sort_keys=True, separators=(",", ":")).encode()
        store.content_by_key[manifest_key] = manifest_content
        latest["manifest"]["size_bytes"] = len(manifest_content)
        latest["manifest"]["sha256"] = __import__("hashlib").sha256(manifest_content).hexdigest()
    elif mutation == "milestone_mismatch":
        latest["milestone"] = "assessment"
    elif mutation == "checksum_mismatch":
        manifest_key = latest["manifest"]["object_key"]
        manifest = json.loads(store.content_by_key[manifest_key])
        descriptor = next(item for item in manifest["artifacts"] if item["kind"] == "run_evidence")
        latest["evidence"]["sha256"] = "0" * 64
        descriptor["sha256"] = "0" * 64
        manifest_content = json.dumps(manifest, sort_keys=True, separators=(",", ":")).encode()
        store.content_by_key[manifest_key] = manifest_content
        latest["manifest"]["size_bytes"] = len(manifest_content)
        latest["manifest"]["sha256"] = __import__("hashlib").sha256(manifest_content).hexdigest()
    else:
        latest["evidence"]["object_key"] = "tasks/task-1/other-evidence.json"

    with pytest.raises(run_report.OptimizationRunIntegrityError):
        service.load_latest_checkpoint()
    assert state.task.status == "FAILED"
    assert json.loads(state.task.metadata)["optimization_run_final_status"] == "failed"


@pytest.mark.parametrize(
    "error_type_name",
    ["ArtifactTicketError", "ArtifactAuthorizationError", "ArtifactTransferError"],
)
def test_recovery_attachment_transport_errors_are_retryable_and_replayable(
    monkeypatch, error_type_name,
):
    from plexus.optimization import run_report
    from plexus.storage import graphql_artifact_store

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)

    class _InterruptedReadStore(_ArtifactStore):
        interruption = None

        def download_bytes(self, request):
            if self.interruption is not None:
                interruption, self.interruption = self.interruption, None
                raise interruption
            return super().download_bytes(request)

    store = _InterruptedReadStore()
    run_spec = {"scope": {}}
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="replayable-read",
        report_configuration_id="config-1", artifact_store=store,
        task_lookup=lambda _: None, report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda task: [stage for stage in _TaskStage.created if stage.taskId == task.id],
        publication_id_factory=lambda: "test",
    )
    state = service.start_or_resume(run_spec)
    evidence = {
        "run_key": "replayable-read", "run_spec": state.run_spec,
        "coverage": {"complete": True},
    }
    service.publish_milestone("ranking", evidence, stakeholder_view=_safe_view())
    error_type = getattr(graphql_artifact_store, error_type_name)
    store.interruption = error_type("temporary authorized attachment interruption")
    fail_calls: list[str] = []
    service.fail = lambda message: fail_calls.append(str(message))

    with pytest.raises(run_report.OptimizationRunRetryablePublicationError):
        service.load_latest_checkpoint()
    assert fail_calls == []
    assert state.task.status == "RUNNING"
    assert json.loads(state.task.metadata).get("optimization_run_final_status") is None

    checkpoint = service.load_latest_checkpoint()
    assert checkpoint["milestone"] == "ranking"
    assert checkpoint["evidence"] == evidence


def test_artifact_store_integrity_error_terminalizes_recovery(monkeypatch):
    from plexus.optimization import run_report
    from plexus.storage.graphql_artifact_store import ArtifactIntegrityError

    class _IntegrityFailureStore(_ArtifactStore):
        interruption = False

        def download_bytes(self, request):
            if self.interruption:
                raise ArtifactIntegrityError("authorized artifact integrity mismatch")
            return super().download_bytes(request)

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    store = _IntegrityFailureStore()
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="integrity-read",
        report_configuration_id="config-1", artifact_store=store,
        task_lookup=lambda _: None, report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda task: [stage for stage in _TaskStage.created if stage.taskId == task.id],
        publication_id_factory=lambda: "test",
    )
    state = service.start_or_resume({"scope": {}})
    service.publish_milestone(
        "ranking",
        {"run_key": "integrity-read", "run_spec": {"scope": {}}, "coverage": {"complete": True}},
        stakeholder_view=_safe_view(),
    )
    store.interruption = True

    with pytest.raises(run_report.OptimizationRunIntegrityError):
        service.load_latest_checkpoint()
    assert state.task.status == "FAILED"
    assert json.loads(state.task.metadata)["optimization_run_final_status"] == "failed"


def test_retry_after_failed_attempt_creates_a_linked_new_task_and_report(monkeypatch):
    first_service = _service(monkeypatch)
    first = first_service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})
    first_service.fail("publication failed")

    retry = _service(monkeypatch)
    retry._task_lookup = lambda _: first.task
    retry._report_lookup = lambda task: first.report if task.id == first.task.id else None
    retried = retry.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    assert retried.task.id == "task-2"
    assert retried.report.id == "report-2"
    assert first.task.status == "FAILED"
    metadata = json.loads(retried.task.metadata)
    assert metadata["previous_attempt_id"] == json.loads(first.task.metadata)["attempt_id"]
    assert metadata["previous_task_id"] == first.task.id
    assert metadata["previous_report_id"] == first.report.id


def test_retry_after_blocked_attempt_preserves_the_terminal_attempt_and_links_a_successor(monkeypatch):
    first_service = _service(monkeypatch)
    first = first_service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})
    first_service.finalize(status="blocked")
    frozen_updates = list(first.task.updates)

    retry = _service(monkeypatch)
    retry._task_lookup = lambda _: first.task
    retry._report_lookup = lambda task: first.report if task.id == first.task.id else None
    retried = retry.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    assert retried.task.id == "task-2"
    assert first.task.updates == frozen_updates
    assert json.loads(retried.task.metadata)["previous_task_id"] == first.task.id


def test_publish_milestone_keeps_raw_evidence_restricted_and_points_to_latest_immutable_revision(monkeypatch):
    uploaded: list[tuple[str, str]] = []
    service = _service(monkeypatch)
    service._artifact_uploader = lambda task_id, name, _: uploaded.append(("task", name)) or f"tasks/{task_id}/{name}"
    state = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    revision = service.publish_milestone(
        "assessment",
        {"raw_transcript": "do not expose", "score_id": "opaque-id", "coverage": {"complete": True}},
        stakeholder_view=_safe_view(),
    )

    assert revision.number == 1
    assert revision.raw_evidence_path.startswith("tasks/task-1/")
    assert revision.workbook_path.startswith("tasks/task-1/")
    assert revision.manifest_path.startswith("tasks/task-1/")
    assert ("task", "optimization-evidence-r0001-test.json") in uploaded
    assert ("task", "optimization-workbook-r0001-test.xlsx") in uploaded
    assert ("task", "optimization-revision-r0001-test.json") in uploaded
    assert revision.raw_evidence_path in state.task.attachedFiles
    assert state.report.parameters["optimization_run"]["latest_revision"]["number"] == 1
    latest = state.report.parameters["optimization_run"]["latest_revision"]
    assert latest["manifest"] == {
        "task_id": state.task.id,
        "object_key": revision.manifest_path,
        "content_type": "application/json",
        "size_bytes": revision.manifest_size_bytes,
        "sha256": revision.manifest_checksum,
    }
    assert latest["manifest"]["size_bytes"] > 0
    assert len(latest["manifest"]["sha256"]) == 64
    assert [item["number"] for item in state.report.parameters["optimization_run"]["revisions"]] == [1]
    assert "raw_transcript" not in str(_Block.created[0].updates[-1])
    assert "opaque-id" not in str(_Block.created[2].updates[-1])


def test_initial_automatic_report_identifies_the_policy_before_target_selection(monkeypatch):
    uploaded: dict[str, bytes] = {}
    service = _service(monkeypatch)
    service._artifact_uploader = lambda task_id, name, content: (
        uploaded.__setitem__(name, content) or f"tasks/{task_id}/{name}"
    )

    service.start_or_resume({"scope": {}, "execution_mode": "automatic"})

    status = json.loads(uploaded["optimization-run-initial-status.json"])
    assert status["execution_mode"] == "automatic"
    workbook = load_workbook(BytesIO(uploaded["optimization-workbook-r0000.xlsx"]), data_only=False)
    overview = dict(workbook["Overview"].values)
    assert overview["Execution Mode"] == "automatic"
    assert overview["Execution Selected Count"] == 0


def test_automatic_execution_projection_reconciles_counts_and_keeps_opaque_ids_out_of_stakeholder_artifacts(monkeypatch):
    uploaded: dict[str, bytes] = {}
    service = _service(monkeypatch)
    service._artifact_uploader = lambda task_id, name, content: (
        uploaded.__setitem__(name, content) or f"tasks/{task_id}/{name}"
    )
    service.start_or_resume({"scope": {}, "execution_mode": "automatic"})
    view = _safe_view()
    evidence = {
        "execution_mode": "automatic",
        "execution_decisions": {
            "mode": "automatic",
            "selected_count": 1,
            "launched_count": 1,
            "rejected_count": 1,
            "selected_targets": [{
                "target_id": "opaque-selected-id",
                "scorecard_id": "opaque-card-selected",
                "score_id": "opaque-score-selected",
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "reason": "Meets the automatic policy.",
                "authorization_source": "published_policy",
            }],
            "rejected_targets": [{
                "target_id": "opaque-rejected-id",
                "scorecard_name": "Example Portfolio",
                "score_name": "Rejected Score",
                "reason": "Outside the safety cap.",
                "authorization_source": "published_policy",
            }],
        },
        "dispatch": {"children": [{
            "target": {"scorecard_id": "opaque-card-selected", "score_id": "opaque-score-selected"},
            "procedure_id": "procedure-1", "task_id": "task-1",
            "launch_state": {"phase": "waiting"},
        }]},
    }

    service.publish_milestone("approval", evidence, stakeholder_view=view)

    manifest = json.loads(uploaded["optimization-revision-r0001-test.json"])
    assert manifest["overview"]["execution_mode"] == "automatic"
    assert manifest["overview"]["execution_selected_count"] == 1
    assert manifest["overview"]["execution_launched_count"] == 1
    assert manifest["overview"]["execution_rejected_count"] == 1
    presentation = json.loads(uploaded[next(
        name for name in uploaded if name.startswith("optimization-presentation-r0001-")
    )])
    assert presentation["overview"]["execution_mode"] == "automatic"
    selected = presentation["top_priorities"][0]
    assert selected["execution_status"] == "automatic_launched"
    assert selected["execution_reason"] == "Meets the automatic policy."
    rejected = presentation["optimization_outcomes"][0]
    assert rejected["score_name"] == "Rejected Score"
    assert rejected["execution_status"] == "automatic_rejected"
    assert rejected["execution_reason"] == "Outside the safety cap."
    assert "opaque-selected-id" not in json.dumps(presentation)
    assert "opaque-rejected-id" not in json.dumps(presentation)

    workbook = load_workbook(BytesIO(uploaded[next(
        name for name in uploaded if name.startswith("optimization-workbook-r0001-")
    )]), data_only=False)
    overview = dict(workbook["Overview"].values)
    assert overview["Execution Mode"] == "automatic"
    assert overview["Execution Selected Count"] == 1
    execution_summary = next(
        workbook["Run Log"].iter_rows(min_row=2, values_only=True)
    )[4]
    assert "not selected 1" in execution_summary
    assert "rejected" not in execution_summary
    headers = [cell.value for cell in workbook["Portfolio"][1]]
    row = dict(zip(headers, next(workbook["Portfolio"].iter_rows(min_row=2, values_only=True))))
    assert row["Automatic Execution"] == "automatic_launched"
    assert row["Execution Reason"] == "Meets the automatic policy."
    assert "opaque-selected-id" not in "".join(
        str(cell.value or "") for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row
    )
    csv_name = next(name for name in uploaded if name.endswith(".csv"))
    csv_row = next(csv.DictReader(uploaded[csv_name].decode("utf-8-sig").splitlines()))
    assert csv_row["Automatic Execution"] == "automatic_launched"
    assert csv_row["Execution Reason"] == "Meets the automatic policy."
    summary_name = next(name for name in uploaded if name.endswith(".md") and "summary" in name)
    summary = uploaded[summary_name].decode("utf-8")
    assert "launched automatically" in summary
    assert "Meets the automatic policy." in summary
    assert "opaque-selected-id" not in summary


def test_automatic_missing_diagnosis_is_presented_as_analysis_work_not_approved_execution():
    from plexus.optimization.run_report import _stakeholder_execution_projection

    view = _safe_view()
    view["portfolio"][0].update({
        "readiness": "ready_to_optimize",
        "next_action": "run_approved_optimization",
    })
    view["priorities"][0]["next_action"] = "run_approved_optimization"
    view["optimization_outcomes"] = [{
        "scorecard_name": "Example Portfolio",
        "score_name": "Priority Score",
        "readiness": "ready_to_optimize",
        "next_action": "run_approved_optimization",
    }]
    evidence = {
        "execution_mode": "automatic",
        "execution_decisions": {
            "mode": "automatic",
            "selected_count": 0,
            "launched_count": 0,
            "rejected_count": 1,
            "selected_targets": [],
            "rejected_targets": [{
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "reason": "missing_diagnosis",
                "authorization_source": "published_policy",
            }],
        },
    }

    projected = _stakeholder_execution_projection(
        view,
        evidence,
        expected_execution_mode="automatic",
    )

    for section in ("portfolio", "priorities", "optimization_outcomes"):
        row = projected[section][0]
        assert row["execution_status"] == "diagnosis_required"
        assert row["next_action"] == "await_semantic_diagnosis"
        assert "diagnosis" in row["rationale"].lower()


def test_automatic_execution_marks_detail_incomplete_for_unnamed_targets_and_launch_mismatch(monkeypatch):
    uploaded: dict[str, bytes] = {}
    service = _service(monkeypatch)
    service._artifact_uploader = lambda task_id, name, content: (
        uploaded.__setitem__(name, content) or f"tasks/{task_id}/{name}"
    )
    service.start_or_resume({"scope": {}, "execution_mode": "automatic"})
    evidence = {
        "execution_mode": "automatic",
        "execution_decisions": {
            "mode": "automatic", "selected_count": 2, "launched_count": 2, "rejected_count": 2,
            "selected_targets": [{
                "scorecard_id": "named-card", "score_id": "named-score",
                "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
                "reason": "Selected by policy.",
            }, {
                "scorecard_id": "unnamed-card", "score_id": "unnamed-score",
                "reason": "Selected by policy.",
            }],
            "rejected_targets": [{
                "scorecard_id": "rejected-card", "score_id": "rejected-score",
                "scorecard_name": "Other Portfolio", "score_name": "Rejected Score",
                "reason": "Outside the safety cap.",
            }, {
                "scorecard_id": "unnamed-rejected-card", "score_id": "unnamed-rejected-score",
                "reason": "Insufficient evidence.",
            }],
        },
        "dispatch": {"children": [{
            "target": {"scorecard_id": "named-card", "score_id": "named-score"},
            "procedure_id": "procedure-1", "task_id": "task-1",
            "launch_state": {"phase": "running"},
        }, {
            "target": {"scorecard_id": "unnamed-card", "score_id": "unnamed-score"},
            "procedure_id": "", "task_id": "task-2",
            "launch_state": {"phase": "running"},
        }]},
    }

    service.publish_milestone("optimization", evidence, stakeholder_view=_safe_view())

    manifest = json.loads(uploaded["optimization-revision-r0001-test.json"])
    overview = manifest["overview"]
    assert overview["execution_named_selected_count"] == 1
    assert overview["execution_named_launched_count"] == 1
    assert overview["execution_named_rejected_count"] == 1
    assert overview["execution_detail_coverage"] == "incomplete"
    assert "1 of 2 selected" in overview["execution_detail_limitation"]
    assert "durable launch evidence reconciles 1 of 2" in overview["execution_detail_limitation"]
    stakeholder_bytes = b"".join(
        content for name, content in uploaded.items()
        if not name.startswith("optimization-evidence-")
    )
    assert b"unnamed-card" not in stakeholder_bytes
    assert b"unnamed-score" not in stakeholder_bytes


def test_durable_child_launch_matches_exact_target_and_requires_owned_authority(monkeypatch):
    view = _safe_view()
    evidence = {
        "execution_mode": "automatic",
        "execution_decisions": {
            "mode": "automatic", "selected_count": 1, "launched_count": 1, "rejected_count": 0,
            "selected_targets": [{
                "scorecard_id": "card-a", "score_id": "score-a",
                "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
            }], "rejected_targets": [],
        },
        "dispatch": {"children": [{
            "target": {"scorecard_id": "card-a", "score_id": "different-score"},
            "procedure_id": "procedure-1", "task_id": "task-1",
            "launch_state": {"phase": "terminal"},
        }, {
            "target": {"scorecard_id": "card-a", "score_id": "score-a"},
            "procedure_id": " ", "task_id": "task-2",
            "launch_state": {"phase": "running"},
        }]},
    }
    from plexus.optimization.run_report import _stakeholder_execution_projection

    projected = _stakeholder_execution_projection(
        view, evidence, expected_execution_mode="automatic"
    )

    assert projected["portfolio"][0]["execution_status"] == "automatic_selected"
    assert projected["overview"]["execution_named_launched_count"] == 0
    assert projected["overview"]["execution_detail_coverage"] == "incomplete"


def test_execution_projection_keeps_duplicate_score_names_distinct_by_scorecard():
    from plexus.optimization.run_report import _stakeholder_execution_projection

    view = _safe_view()
    second = dict(view["portfolio"][0])
    second["scorecard_name"] = "Second Portfolio"
    view["portfolio"].append(second)
    evidence = {
        "execution_mode": "automatic",
        "execution_decisions": {
            "mode": "automatic", "selected_count": 2, "launched_count": 1, "rejected_count": 0,
            "selected_targets": [{
                "scorecard_id": "card-a", "score_id": "score-a",
                "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
            }, {
                "scorecard_id": "card-b", "score_id": "score-b",
                "scorecard_name": "Second Portfolio", "score_name": "Priority Score",
            }], "rejected_targets": [],
        },
        "dispatch": {"children": [{
            "target": {"scorecard_id": "card-a", "score_id": "score-a"},
            "procedure_id": "procedure-1", "task_id": "task-1",
            "launch_state": {"phase": "terminal"},
        }]},
    }

    projected = _stakeholder_execution_projection(
        view, evidence, expected_execution_mode="automatic"
    )

    statuses = {
        row["scorecard_name"]: row["execution_status"] for row in projected["portfolio"]
    }
    assert statuses == {
        "Example Portfolio": "automatic_launched",
        "Second Portfolio": "automatic_selected",
    }
    assert projected["overview"]["execution_detail_coverage"] == "complete"


def test_missing_execution_fields_freeze_the_conservative_approval_required_mode(monkeypatch):
    uploaded: dict[str, bytes] = {}
    service = _service(monkeypatch)
    service._artifact_uploader = lambda task_id, name, content: (
        uploaded.__setitem__(name, content) or f"tasks/{task_id}/{name}"
    )
    service.start_or_resume({"scope": {}})
    service.publish_milestone("approval", {"coverage": {"complete": True}}, stakeholder_view=_safe_view())

    manifest = json.loads(uploaded["optimization-revision-r0001-test.json"])
    assert manifest["overview"]["execution_mode"] == "approval_required"


def test_approval_required_execution_mode_remains_explicit_without_automatic_decisions(monkeypatch):
    uploaded: dict[str, bytes] = {}
    service = _service(monkeypatch)
    service._artifact_uploader = lambda task_id, name, content: (
        uploaded.__setitem__(name, content) or f"tasks/{task_id}/{name}"
    )
    service.start_or_resume({"scope": {}, "execution_mode": "approval_required"})
    service.publish_milestone(
        "approval", {"execution_mode": "approval_required"}, stakeholder_view=_safe_view()
    )

    manifest = json.loads(uploaded["optimization-revision-r0001-test.json"])
    assert manifest["overview"]["execution_mode"] == "approval_required"
    assert "execution_selected_count" not in manifest["overview"]


@pytest.mark.parametrize(
    ("frozen_mode", "evidence_mode", "decision_mode"),
    [
        ("automatic", "approval_required", "automatic"),
        ("approval_required", "automatic", "approval_required"),
        ("automatic", "automatic", "approval_required"),
        ("approval_required", "approval_required", "automatic"),
    ],
)
def test_report_rejects_decision_evidence_that_conflicts_with_the_frozen_execution_mode(
    monkeypatch, frozen_mode, evidence_mode, decision_mode,
):
    from plexus.optimization import run_report

    service = _service(monkeypatch)
    state = service.start_or_resume({"scope": {}, "execution_mode": frozen_mode})

    with pytest.raises(run_report.OptimizationRunIntegrityError, match="execution mode"):
        service.publish_milestone(
            "approval",
            {
                "coverage": {"complete": True},
                "execution_mode": evidence_mode,
                "execution_decisions": {"mode": decision_mode},
            },
            stakeholder_view=_safe_view(),
        )

    assert state.task.status == "FAILED"


def test_publish_milestone_indexes_revisioned_scorecard_markdown_and_csv_without_attaching_each_child(monkeypatch):
    uploaded: dict[str, bytes] = {}
    service = _service(monkeypatch, dashboard_base_url="https://dashboard.example.com")
    service._artifact_uploader = lambda task_id, name, content: (
        uploaded.__setitem__(name, content) or f"tasks/{task_id}/{name}"
    )
    state = service.start_or_resume({"scope": {}})
    view = deepcopy(_safe_view())
    view["portfolio"][0]["scorecard_ref"] = "safe-ref-one"
    view["portfolio"][0]["score_name"] = "=Formula-like score"
    view["questions_and_issues"][0].update({
        "scorecard_ref": "safe-ref-one",
        "score_name": "=Formula-like score",
        "issue_flag": "feedback_rubric_contradiction",
        "finding": "Should this policy exception require stakeholder confirmation?",
        "next_action": "request_stakeholder_clarification",
    })
    view["portfolio"][0].update({
        "primary_disposition": "guideline_or_code_repair",
        "secondary_issue_flags": ["feedback_rubric_contradiction"],
        "secondary_issue_summary": "feedback_rubric_contradiction",
    })
    second_row = dict(view["portfolio"][0])
    second_row.update({
        "scorecard_ref": "safe-ref-two",
        "score_name": "Second Score",
        "valid_feedback_count": 75,
    })
    view["portfolio"].append(second_row)

    revision = service.publish_milestone(
        "assessment",
        {"coverage": {"complete": True}},
        stakeholder_view=view,
    )

    manifest = json.loads(uploaded["optimization-revision-r0001-test.json"])
    presentation_artifact = next(
        artifact for artifact in manifest["artifacts"]
        if artifact["kind"] == "stakeholder_presentation"
    )
    presentation_name = presentation_artifact["object_key"].rsplit("/", 1)[-1]
    presentation = json.loads(uploaded[presentation_name])
    assert presentation["questions_and_issues"] == view["questions_and_issues"]
    assert presentation["contradictions"] == view["questions_and_issues"]
    assert presentation["optimization_outcomes"] == view["optimization_outcomes"]
    assert sum(presentation["primary_decision_mix"].values()) == 2
    assert presentation["primary_disposition_counts"] == presentation["primary_decision_mix"]
    assert presentation["attention_queue"]
    assert set(presentation["attention_queue"][0]) == {
        "scorecard_name", "score_name", "primary_disposition", "evidence_count",
        "severity", "rationale", "next_action", "dashboard_url",
    }
    assert presentation["score_count"] == 2
    assert presentation["score_count"] == len(view["portfolio"])
    assert len(presentation["scorecards"]) == 2
    assert presentation["scorecards"][0]["score_count"] == 1
    assert presentation["top_priorities"][0]["opportunity"] == 70
    distribution = presentation["opportunity_distribution"]
    assert [row["score_name"] for row in distribution] == ["=Formula-like score", "Second Score"]
    assert all(
        row["primary_disposition"] == view["portfolio"][0]["primary_disposition"]
        for row in distribution
    )
    assert all(
        row["secondary_issue_flags"] == view["portfolio"][0]["secondary_issue_flags"]
        for row in distribution
    )
    assert presentation_artifact["object_key"] in state.task.attachedFiles
    status_envelope = json.loads(state.blocks["status"].output)
    assert status_envelope["preview"]["type"] == "optimization_run_status"
    assert status_envelope["preview"]["summary"]["presentation"] == presentation_artifact
    scorecard_artifacts = [
        artifact for artifact in manifest["artifacts"]
        if artifact["scope"] == "scorecard"
    ]
    assert {artifact["kind"] for artifact in scorecard_artifacts} == {
        "scorecard_summary",
        "scorecard_portfolio_csv",
        "scorecard_presentation",
    }
    assert len(scorecard_artifacts) == 6
    assert manifest["scorecard_count"] == 2
    assert manifest["score_count"] == 2
    assert all(artifact["source_revision"] == 1 for artifact in scorecard_artifacts)
    assert {artifact["scorecard_name"] for artifact in scorecard_artifacts} == {"Example Portfolio"}
    assert len({artifact["logical_id"] for artifact in scorecard_artifacts}) == 6
    assert all(artifact["sha256"] and artifact["size_bytes"] > 0 for artifact in scorecard_artifacts)
    assert all(artifact["task_id"] == state.task.id for artifact in scorecard_artifacts)
    assert all(
        artifact["dashboard_url"] == (
            f"https://dashboard.example.com/lab/reports/{state.report.id}"
            f"?revision=1&artifact={artifact['logical_id'].replace(':', '%3A')}"
        )
        for artifact in scorecard_artifacts
    )
    assert all(artifact["object_key"].startswith(f"tasks/{state.task.id}/") for artifact in scorecard_artifacts)
    assert all(artifact["object_key"] not in state.task.attachedFiles for artifact in scorecard_artifacts)
    assert len(revision.artifacts) == len(manifest["artifacts"])
    assert revision.row_counts["portfolio"] == len(view["portfolio"])
    assert revision.row_counts["questions_and_issues"] == len(view["questions_and_issues"])
    assert revision.row_counts["optimization_outcomes"] == len(view["optimization_outcomes"])

    score_artifacts = [
        artifact for artifact in manifest["artifacts"]
        if artifact["scope"] == "score"
    ]
    # Only the first score is represented in the priority and issue evidence.
    # The second remains fully visible in its scorecard's CSV, summary, and
    # presentation, but does not need an individual stakeholder brief.
    assert len(score_artifacts) == 1
    assert {artifact["kind"] for artifact in score_artifacts} == {"score_brief"}
    assert {artifact["score_name"] for artifact in score_artifacts} == {"=Formula-like score"}
    assert all(artifact["object_key"] not in state.task.attachedFiles for artifact in score_artifacts)

    csv_artifact = next(
        artifact for artifact in scorecard_artifacts
        if artifact["kind"] == "scorecard_portfolio_csv"
        and artifact["scorecard_name"] == "Example Portfolio"
    )
    csv_name = csv_artifact["object_key"].rsplit("/", 1)[-1]
    csv_rows = list(csv.DictReader(uploaded[csv_name].decode("utf-8-sig").splitlines()))
    assert len(csv_rows) == 1
    assert csv_rows[0]["Score"] == "'=Formula-like score"
    assert csv_rows[0]["Primary Disposition"] == view["portfolio"][0]["primary_disposition"]
    assert csv_rows[0]["Secondary Issues"] == view["portfolio"][0]["secondary_issue_summary"]

    summary_artifact = next(
        artifact for artifact in scorecard_artifacts
        if artifact["kind"] == "scorecard_summary"
        and artifact["scorecard_name"] == "Example Portfolio"
    )
    summary_name = summary_artifact["object_key"].rsplit("/", 1)[-1]
    summary = uploaded[summary_name].decode("utf-8")
    assert summary.startswith("# Example Portfolio")
    assert "repair_guidelines" in summary
    assert "Primary disposition" in summary
    assert "Secondary issue flags" in summary
    assert view["portfolio"][0]["primary_disposition"] in summary
    assert ", ".join(view["portfolio"][0]["secondary_issue_flags"]) in summary

    detail_artifact = next(
        artifact for artifact in scorecard_artifacts
        if artifact["kind"] == "scorecard_presentation"
    )
    detail_name = detail_artifact["object_key"].rsplit("/", 1)[-1]
    detail = json.loads(uploaded[detail_name])
    assert detail["scorecard_name"] == "Example Portfolio"
    assert len(detail["scores"]) == 1
    assert detail["scores"][0]["score_name"] == "=Formula-like score"
    assert detail["scores"][0]["primary_disposition"] == view["portfolio"][0]["primary_disposition"]
    assert detail["scores"][0]["secondary_issue_flags"] == view["portfolio"][0]["secondary_issue_flags"]
    brief_descriptor = detail["scores"][0]["artifacts"][0]
    assert brief_descriptor["kind"] == "score_brief"
    brief_name = brief_descriptor["object_key"].rsplit("/", 1)[-1]
    brief = uploaded[brief_name].decode("utf-8")
    assert brief.startswith("# =Formula-like score")
    assert "Should this policy exception require stakeholder confirmation?" in brief
    assert "request_stakeholder_clarification" in brief
    assert f"Primary disposition: {view['portfolio'][0]['primary_disposition']}" in brief
    assert (
        f"Secondary issue flags: {', '.join(view['portfolio'][0]['secondary_issue_flags'])}"
        in brief
    )


def test_score_briefs_are_bounded_to_rows_represented_in_stakeholder_findings_and_plan_matches(
    monkeypatch,
):
    """The detailed artifact fanout follows the same relevance contract as progress."""
    from plexus.optimization import run_report

    view = deepcopy(_safe_view())
    view["portfolio"] = [
        {
            **view["portfolio"][0],
            "scorecard_ref": "shared-scorecard",
            "score_ref": "score-a",
            "scorecard_name": "Shared Scorecard",
            "score_name": "Same display name",
        },
        {
            **view["portfolio"][0],
            "scorecard_ref": "shared-scorecard",
            "score_ref": "score-b",
            "scorecard_name": "Shared Scorecard",
            "score_name": "Same display name",
        },
        {
            **view["portfolio"][0],
            "scorecard_name": "Legacy Scorecard",
            "score_name": "Legacy name fallback",
        },
        {
            **view["portfolio"][0],
            "scorecard_name": "Unselected Scorecard",
            "score_name": "Unselected score",
        },
    ]
    # Opaque references distinguish duplicate display names.  The legacy
    # projection intentionally has no references and exercises the fallback.
    view["priorities"] = [{
        "scorecard_ref": "shared-scorecard",
        "score_ref": "score-b",
        "scorecard_name": "Shared Scorecard",
        "score_name": "Same display name",
    }]
    view["questions_and_issues"] = [
        {
            "scorecard_ref": "shared-scorecard",
            "score_ref": "score-a",
            "scorecard_name": "Shared Scorecard",
            "score_name": "Same display name",
            "finding": "Sibling opaque issue.",
        },
        {
            "scorecard_ref": "shared-scorecard",
            "score_ref": "score-b",
            "scorecard_name": "Shared Scorecard",
            "score_name": "Same display name",
            "finding": "Selected opaque issue.",
        },
        {
            "scorecard_name": "Legacy Scorecard",
            "score_name": "Legacy name fallback",
            "finding": "Clarification needed.",
        },
    ]
    view["optimization_outcomes"] = []
    uploaded: dict[str, bytes] = {}

    artifacts = run_report.build_scorecard_artifacts(
        view,
        revision_number=1,
        task_id="task-1",
        uploader=lambda task_id, filename, content: (
            uploaded.__setitem__(filename, content) or f"tasks/{task_id}/{filename}"
        ),
    )

    score_briefs = [artifact for artifact in artifacts if artifact["kind"] == "score_brief"]
    assert len(score_briefs) == 3
    assert {(artifact["scorecard_name"], artifact["score_name"]) for artifact in score_briefs} == {
        ("Shared Scorecard", "Same display name"),
        ("Legacy Scorecard", "Legacy name fallback"),
    }
    presentations = [artifact for artifact in artifacts if artifact["kind"] == "scorecard_presentation"]
    shared = next(
        artifact for artifact in presentations if artifact["scorecard_name"] == "Shared Scorecard"
    )
    shared_rows = json.loads(uploaded[shared["object_key"].rsplit("/", 1)[-1]])["scores"]
    assert all(row["artifacts"] for row in shared_rows)
    selected_brief = next(
        row["artifacts"][0] for row in shared_rows if row["score_ref"] == "score-b"
    )
    selected_brief_content = uploaded[selected_brief["object_key"].rsplit("/", 1)[-1]].decode("utf-8")
    assert "Selected opaque issue." in selected_brief_content
    assert "Sibling opaque issue." not in selected_brief_content
    assert len(json.loads(uploaded[next(
        artifact["object_key"].rsplit("/", 1)[-1]
        for artifact in presentations
        if artifact["scorecard_name"] == "Unselected Scorecard"
    )])["scores"]) == 1
    assert run_report._artifact_publication_plan(view)["score_briefs"] == {
        "completed": 0,
        "total": 3,
    }


def test_score_brief_relevance_uses_priority_cutoff_and_ignores_not_run_outcomes(monkeypatch):
    """Production-shaped full evidence projections cannot fan out every score brief."""
    from plexus.optimization import run_report

    view = deepcopy(_safe_view())
    template = view["portfolio"][0]
    view["overview"].update({
        "priority_cutoff_rank": 2,
        "priority_display_limit": 4,
    })
    view["portfolio"] = [
        {
            **template,
            "scorecard_ref": "production-shaped-scorecard",
            "score_ref": f"score-{rank}",
            "score_name": f"Score {rank}",
            "evidence_rank": rank,
        }
        for rank in range(1, 7)
    ]
    # The production projection retains every evidence-ranked row here.  Only
    # the report's explicit display cutoff makes a priority stakeholder-facing.
    view["priorities"] = [
        {
            "scorecard_ref": "production-shaped-scorecard",
            "score_ref": f"score-{rank}",
            "scorecard_name": "Example Portfolio",
            "score_name": f"Score {rank}",
            "evidence_rank": rank,
            "opportunity": 100 - rank,
        }
        for rank in range(1, 7)
    ]
    view["questions_and_issues"] = [{
        "scorecard_ref": "production-shaped-scorecard",
        "score_ref": "score-4",
        "scorecard_name": "Example Portfolio",
        "score_name": "Score 4",
        "finding": "A score-specific question remains material.",
    }]
    view["optimization_outcomes"] = [
        {
            "scorecard_ref": "production-shaped-scorecard",
            "score_ref": f"score-{rank}",
            "scorecard_name": "Example Portfolio",
            "score_name": f"Score {rank}",
            "outcome": "not_run",
        }
        for rank in range(1, 7)
    ] + [{
        "scorecard_ref": "production-shaped-scorecard",
        "score_ref": "score-5",
        "scorecard_name": "Example Portfolio",
        "score_name": "Score 5",
        "outcome": "optimization_in_progress",
    }]
    uploaded: dict[str, bytes] = {}

    artifacts = run_report.build_scorecard_artifacts(
        view,
        revision_number=1,
        task_id="task-1",
        uploader=lambda task_id, filename, content: (
            uploaded.__setitem__(filename, content) or f"tasks/{task_id}/{filename}"
        ),
    )

    briefs = [artifact for artifact in artifacts if artifact["kind"] == "score_brief"]
    assert {artifact["score_name"] for artifact in briefs} == {
        "Score 1", "Score 2", "Score 4", "Score 5",
    }
    assert run_report._artifact_publication_plan(view)["score_briefs"] == {
        "completed": 0,
        "total": 4,
    }


def test_a_later_milestone_creates_a_new_score_brief_when_it_first_becomes_relevant(monkeypatch):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    store = _ArtifactStore()
    publication_ids = iter(["ranking", "diagnosis"])
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="brief-later-relevant",
        report_configuration_id="config-1", artifact_store=store,
        publication_id_factory=lambda: next(publication_ids),
        task_lookup=lambda _: None, report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda task: [stage for stage in _TaskStage.created if stage.taskId == task.id],
    )
    initial = deepcopy(_safe_view())
    initial["priorities"] = []
    initial["questions_and_issues"] = []
    initial["optimization_outcomes"] = []
    service.start_or_resume({"scope": {}})
    first = service.publish_milestone(
        "ranking", {"coverage": {"complete": True}}, stakeholder_view=initial,
    )
    assert not [artifact for artifact in first.artifacts if artifact["kind"] == "score_brief"]

    later = deepcopy(initial)
    later["optimization_outcomes"] = [{
        "scorecard_name": "Example Portfolio",
        "score_name": "Priority Score",
        "outcome": "ready_for_review",
    }]
    second = service.publish_milestone(
        "diagnosis", {"coverage": {"complete": True}}, stakeholder_view=later,
    )
    score_briefs = [artifact for artifact in second.artifacts if artifact["kind"] == "score_brief"]
    assert len(score_briefs) == 1
    assert score_briefs[0]["score_name"] == "Priority Score"
    assert score_briefs[0]["source_revision"] == 2


def test_report_artifact_base_url_rejects_non_https_or_non_origin_values(monkeypatch):
    for value in (
        "http://dashboard.example.com",
        "https://dashboard.example.com/path",
        "https://user:secret@dashboard.example.com",
    ):
        with pytest.raises(ValueError, match="HTTPS origin"):
            _service(monkeypatch, dashboard_base_url=value)


def test_report_artifact_base_url_comes_from_existing_account_settings():
    from plexus.optimization.run_report import dashboard_base_url_from_account_settings

    assert dashboard_base_url_from_account_settings({
        "hiddenMenuItems": [],
        "reporting": {"dashboardBaseUrl": "https://dashboard.example.com/"},
    }) == "https://dashboard.example.com"
    assert dashboard_base_url_from_account_settings({"hiddenMenuItems": []}) is None
    assert dashboard_base_url_from_account_settings(None) is None


def test_multiple_milestones_preserve_full_revision_history(monkeypatch):
    service = _service(monkeypatch)
    state = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    service.publish_milestone("ranking", {"coverage": {"complete": True}}, stakeholder_view=_safe_view())
    service.publish_milestone("assessment", {"coverage": {"complete": True}}, stakeholder_view=_safe_view())

    history = state.report.parameters["optimization_run"]["revisions"]
    assert [revision["number"] for revision in history] == [1, 2]
    assert [revision["milestone"] for revision in history] == ["ranking", "assessment"]
    assert all(revision["evidence_path"].startswith("tasks/task-1/") for revision in history)
    assert len(state.blocks["workbook"].attachedFiles) == 3


def test_default_artifact_path_uses_only_task_graphql_tickets_without_direct_s3(monkeypatch):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    store = _ArtifactStore()
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(),
        account_id="account-1",
        run_key="daily-v1-2026-07-29",
        report_configuration_id="config-1",
        now=lambda: datetime(2026, 7, 29, 12, 0, tzinfo=timezone.utc),
        task_lookup=lambda _: None,
        report_lookup=lambda _: None,
        block_lookup=lambda _: [],
        stage_lookup=lambda _: [],
        artifact_store=store,
        verify_uploaded_artifacts=True,
    )

    state = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})
    service.publish_milestone("assessment", {"coverage": {"complete": True}}, stakeholder_view=_safe_view())

    requests = [request for request, _content in store.uploads]
    assert all(request.resource_type == "TASK" for request in requests)
    assert all(request.resource_id == state.task.id for request in requests)
    assert all(request.artifact_type == "TASK_ATTACHMENT" for request in requests)
    filenames = {request.filename for request in requests}
    assert filenames >= {
        "optimization-run-initial-status.json",
        "optimization-run-initial-evidence.json",
        "optimization-workbook-r0000.xlsx",
    }
    assert any(name.startswith("optimization-evidence-r0001-") for name in filenames)
    assert any(name.startswith("optimization-workbook-r0001-") for name in filenames)
    assert any(name.startswith("optimization-revision-r0001-") for name in filenames)
    assert {
        request.content_type for request in requests
        if request.filename.endswith((".md", ".csv"))
    } == {
        "text/markdown",
        "text/csv",
    }
    assert all(path.startswith(f"tasks/{state.task.id}/") for path in state.task.attachedFiles)
    child_paths = {
        f"tasks/{state.task.id}/{request.filename}"
        for request in requests
        if request.filename.startswith(("scorecard-", "score-"))
    }
    assert child_paths
    assert child_paths.isdisjoint(state.task.attachedFiles)
    assert len(state.task.attachedFiles) == len(store.uploads) - len(child_paths)
    assert all(
        json.loads(block.output)["output_attachment"].startswith(f"tasks/{state.task.id}/")
        for block in state.blocks.values()
    )
    assert len(store.downloads) == len(store.uploads)
    initial_workbook_upload = next(
        content for request, content in store.uploads
        if request.filename == "optimization-workbook-r0000.xlsx"
    )
    assert initial_workbook_upload.startswith(b"PK")
    source = inspect.getsource(run_report)
    assert "plexus.reports.s3_utils" not in source
    assert "upload_report_block_file" not in source
    assert "upload_procedure_file" not in source
    assert "REPORT_BLOCK_ATTACHMENT" not in source


def test_artifact_checksum_mismatch_is_fatal_and_marks_the_attempt_failed(monkeypatch):
    from plexus.optimization import run_report

    class _CorruptingStore(_ArtifactStore):
        def download_bytes(self, request):
            return super().download_bytes(request) + b"corrupt"

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="daily-v1-2026-07-29",
        report_configuration_id="config-1", task_lookup=lambda _: None,
        report_lookup=lambda _: None, block_lookup=lambda _: [], stage_lookup=lambda _: [],
        artifact_store=_CorruptingStore(), verify_uploaded_artifacts=True,
    )

    with pytest.raises(run_report.OptimizationRunIntegrityError, match="checksum"):
        service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    assert _Task.created[0].status == "FAILED"


def test_start_publishes_the_report_before_initializing_remote_task_stages(monkeypatch):
    from plexus.optimization import run_report

    events: list[str] = []

    class _TrackingTask(_Task):
        @classmethod
        def create(cls, **values):
            events.append("task")
            return super().create(**values)

    class _TrackingReport(_Report):
        @classmethod
        def create(cls, **values):
            events.append("report")
            return super().create(**values)

    class _TrackingStage(_TaskStage):
        @classmethod
        def create(cls, **values):
            events.append(f"stage:{values['name']}")
            return super().create(**values)

    monkeypatch.setattr(run_report, "Task", _TrackingTask)
    monkeypatch.setattr(run_report, "Report", _TrackingReport)
    monkeypatch.setattr(run_report, "TaskStage", _TrackingStage)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="startup-order",
        report_configuration_id="config-1", task_lookup=lambda _: None,
        report_lookup=lambda _: None, block_lookup=lambda _: [], stage_lookup=lambda _: [],
        artifact_store=_ArtifactStore(),
    )

    service.start_or_resume({"scope": {}})

    assert events.index("report") < events.index("stage:preflight")


def test_new_attempt_creates_its_report_without_scanning_report_history(monkeypatch):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="new-attempt",
        report_configuration_id="config-1", task_lookup=lambda _: None,
        report_lookup=lambda _task: (_ for _ in ()).throw(
            AssertionError("new attempts must not scan historical reports")
        ),
        block_lookup=lambda _: [], stage_lookup=lambda _: [],
        artifact_store=_ArtifactStore(),
    )

    state = service.start_or_resume({"scope": {}})

    assert state.report.taskId == state.task.id


def test_procedure_owned_task_is_reused_without_duplicate_task_or_stage_creation(monkeypatch):
    from plexus.optimization import run_report

    procedure_task = _Task(
        identifier="procedure-task",
        accountId="account-1",
        status="RUNNING",
        description="Account-wide optimization portfolio — All scorecards",
        metadata=json.dumps({"procedure_key": "preserve-me"}),
    )
    procedure_stages = [
        _TaskStage(
            f"procedure-stage-{index}", taskId=procedure_task.id, name=name.title(),
            order=index, status="RUNNING" if name == "analysis" else "PENDING",
        )
        for index, name in enumerate(
            ("preflight", "analysis", "approval", "optimization", "review", "finalization")
        )
    ]
    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="procedure-run",
        report_configuration_id="config-1", existing_task=procedure_task,
        task_lookup=lambda _: (_ for _ in ()).throw(
            AssertionError("an adopted Procedure Task must bypass run-key discovery")
        ),
        report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda _: procedure_stages, artifact_store=_ArtifactStore(),
    )

    state = service.start_or_resume({
        "scope": {"scorecard_name_prefixes": ["Example portfolio"]}
    })
    service.finalize(status="complete")

    assert state.task is procedure_task
    assert state.report.taskId == procedure_task.id
    assert _Task.created == []
    assert _TaskStage.created == []
    metadata = json.loads(procedure_task.metadata)
    assert metadata["procedure_key"] == "preserve-me"
    assert metadata["optimization_run_key"] == "procedure-run"
    assert metadata["optimization_run_final_status"] == "complete"
    assert metadata["operator_identity"] == {
        "kind": "scorecard_scoped_portfolio",
        "display_title": "Scorecard-scoped optimization portfolio",
        "display_scope": 'scorecard names beginning with "Example portfolio"',
    }
    assert procedure_task.description == (
        'Scorecard-scoped optimization portfolio — scorecard names beginning with '
        '"Example portfolio"'
    )
    assert not any(update.get("status") == "COMPLETED" for update in procedure_task.updates)


def test_progress_updates_the_existing_analysis_stage_and_live_cover_without_a_new_revision(monkeypatch):
    from plexus.optimization import run_report

    procedure_task = _Task(
        identifier="procedure-task",
        accountId="account-1",
        status="RUNNING",
        metadata=json.dumps({"procedure_key": "preserve-me"}),
    )
    procedure_stages = [
        _TaskStage(
            f"procedure-stage-{index}", taskId=procedure_task.id, name=name.title(),
            order=index, status="RUNNING" if name == "analysis" else "PENDING",
        )
        for index, name in enumerate(
            ("preflight", "analysis", "approval", "optimization", "review", "finalization")
        )
    ]
    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="progress-run",
        report_configuration_id="config-1", existing_task=procedure_task,
        report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda _: procedure_stages, artifact_store=_ArtifactStore(),
        now=lambda: datetime(2026, 7, 29, 12, 0, tzinfo=timezone.utc),
    )
    state = service.start_or_resume({"scope": {}})
    service.publish_milestone(
        "ranking", {"coverage": {"complete": True}}, stakeholder_view=_safe_view(),
    )
    revision_before = service._latest_revision_number(state.report)

    service.publish_progress(
        phase="assessment",
        current=30,
        total=100,
        message="Assessing 30 of 100 scores.",
    )

    analysis = next(stage for stage in procedure_stages if stage.name == "Analysis")
    assert analysis.processedItems == 30
    assert analysis.totalItems == 100
    assert analysis.statusMessage == "Assessing 30 of 100 scores."
    assert service._latest_revision_number(state.report) == revision_before
    status_envelope = json.loads(state.blocks["status"].output)
    assert status_envelope["preview"]["summary"]["live_progress"] == {
        "phase": "assessment",
        "current": 30,
        "total": 100,
        "message": "Assessing 30 of 100 scores.",
        "updated_at": "2026-07-29T12:00:00Z",
    }
    assert "Assessment: 30 of 100 scores assessed" in state.report.output


def test_ranking_progress_updates_the_existing_ranking_stage_without_an_artifact_revision(monkeypatch):
    from plexus.optimization import run_report

    procedure_task = _Task(
        identifier="procedure-task",
        accountId="account-1",
        status="RUNNING",
        metadata=json.dumps({"procedure_key": "preserve-me"}),
    )
    procedure_stages = [
        _TaskStage(
            f"procedure-stage-{index}", taskId=procedure_task.id, name=name.title(),
            order=index, status="RUNNING" if name == "ranking" else "PENDING",
        )
        for index, name in enumerate(
            ("preflight", "ranking", "assessment", "diagnosis", "approval", "optimization", "review", "finalization")
        )
    ]
    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="ranking-progress-run",
        report_configuration_id="config-1", existing_task=procedure_task,
        report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda _: procedure_stages, artifact_store=_ArtifactStore(),
        now=lambda: datetime(2026, 7, 31, 12, 0, tzinfo=timezone.utc),
    )
    state = service.start_or_resume({"scope": {}})
    revision_before = service._latest_revision_number(state.report)

    service.publish_progress(
        phase="ranking",
        subphase="inventory",
        current=12,
        total=None,
        unit="scorecards",
        state="retrying",
        elapsed_seconds=63,
        next_checkpoint="Retrying the scorecard inventory.",
        message="Inventory has inspected 12 scorecards; retrying a page.",
    )

    ranking = next(stage for stage in procedure_stages if stage.name == "Ranking")
    assert ranking.processedItems == 12
    assert ranking.totalItems is None
    assert ranking.statusMessage == "Inventory has inspected 12 scorecards; retrying a page."
    assert service._latest_revision_number(state.report) == revision_before
    status_envelope = json.loads(state.blocks["status"].output)
    assert status_envelope["preview"]["summary"]["live_progress"] == {
        "phase": "ranking",
        "subphase": "inventory",
        "current": 12,
        "total": None,
        "unit": "scorecards",
        "state": "retrying",
        "elapsed_seconds": 63,
        "next_checkpoint": "Retrying the scorecard inventory.",
        "message": "Inventory has inspected 12 scorecards; retrying a page.",
        "updated_at": "2026-07-31T12:00:00Z",
    }
    assert "Ranking / Inventory: 12 scorecards inspected" in state.report.output
    assert "Retrying the scorecard inventory." in state.report.output


def test_a_stale_ranking_progress_update_cannot_overlay_a_durable_ranking_or_later_milestone(monkeypatch):
    service = _service(monkeypatch)
    state = service.start_or_resume({"scope": {}})
    service.publish_milestone(
        "ranking", {"coverage": {"complete": True}}, stakeholder_view=_safe_view(),
    )
    output_after_ranking = state.blocks["status"].output
    revision_after_ranking = service._latest_revision_number(state.report)

    service.publish_progress(
        phase="ranking",
        subphase="feedback_analysis",
        current=5,
        total=10,
        unit="scorecards",
        message="A delayed feedback-progress callback arrived.",
    )

    assert state.blocks["status"].output == output_after_ranking
    assert service._latest_revision_number(state.report) == revision_after_ranking


def test_assessment_artifact_publication_has_a_distinct_safe_live_phase(monkeypatch):
    """A long artifact upload must not look like semantic diagnosis has begun."""
    service = _service(monkeypatch)
    state = service.start_or_resume({"scope": {}})
    updates: list[dict] = []
    original_publish_progress = service.publish_progress

    def record_progress(**progress):
        updates.append(progress)
        return original_publish_progress(**progress)

    service.publish_progress = record_progress
    service.publish_milestone(
        "assessment", {"coverage": {"complete": True}}, stakeholder_view=_safe_view(),
    )

    publication = [update for update in updates if update["phase"] == "publication"]
    assert publication
    assert service._latest_revision_number(state.report) == 1
    assert all(update["unit"] == "artifacts" for update in publication)
    assert publication[0]["current"] == 0
    assert publication[0]["total"] == 8
    assert publication[-1]["current"] == 8
    assert publication[-1]["next_checkpoint"] == "Publishing the assessment milestone."
    assert publication[-1]["artifact_counts"] == {
        "decision_evidence": {"completed": 1, "total": 1},
        "stakeholder_workbook": {"completed": 1, "total": 1},
        "score_briefs": {"completed": 1, "total": 1},
        "scorecard_summaries": {"completed": 1, "total": 1},
        "scorecard_spreadsheets": {"completed": 1, "total": 1},
        "scorecard_presentations": {"completed": 1, "total": 1},
        "stakeholder_presentation": {"completed": 1, "total": 1},
        "revision_manifest": {"completed": 1, "total": 1},
    }
    assert all("object_key" not in str(update) for update in publication)
    assert all("diagnosis" not in update["message"].lower() for update in publication)


def test_large_artifact_publication_advances_durable_progress_at_a_bounded_cadence(monkeypatch):
    service = _service(monkeypatch)
    service.start_or_resume({"scope": {}})
    view = deepcopy(_safe_view())
    template = view["portfolio"][0]
    view["portfolio"] = [
        {**template, "score_name": f"Score {index:02d}"}
        for index in range(30)
    ]
    view["priorities"] = [
        {
            **view["priorities"][0],
            "score_name": row["score_name"],
        }
        for row in view["portfolio"]
    ]
    updates: list[dict] = []
    original_publish_progress = service.publish_progress

    def record_progress(**progress):
        updates.append(deepcopy(progress))
        return original_publish_progress(**progress)

    service.publish_progress = record_progress
    service.publish_milestone(
        "assessment",
        {"coverage": {"complete": True}},
        stakeholder_view=view,
    )

    score_brief_counts = [
        update["artifact_counts"]["score_briefs"]["completed"]
        for update in updates
        if update["phase"] == "publication"
    ]
    assert 25 in score_brief_counts
    assert score_brief_counts[-1] == 30
    assert sorted(set(score_brief_counts)) == [0, 25, 30]


def test_unchanged_score_artifacts_are_reused_from_the_latest_committed_revision(monkeypatch):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    store = _ArtifactStore()
    publication_ids = iter(["ranking", "assessment"])
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(),
        account_id="account-1",
        run_key="reuse-unchanged-artifacts",
        report_configuration_id="config-1",
        artifact_store=store,
        publication_id_factory=lambda: next(publication_ids),
        task_lookup=lambda _: None,
        report_lookup=lambda _: None,
        block_lookup=lambda _: [],
        stage_lookup=lambda task: [
            stage for stage in _TaskStage.created if stage.taskId == task.id
        ],
    )
    service.start_or_resume({"scope": {}})
    first = service.publish_milestone(
        "ranking",
        {"run_key": "reuse-unchanged-artifacts", "coverage": {"complete": True}},
        stakeholder_view=_safe_view(),
    )
    uploads_after_first = len(store.uploads)

    second = service.publish_milestone(
        "assessment",
        {
            "run_key": "reuse-unchanged-artifacts",
            "coverage": {"complete": True},
            "assessments": [],
        },
        stakeholder_view=_safe_view(),
    )

    reusable_kinds = {
        "score_brief",
        "scorecard_summary",
        "scorecard_portfolio_csv",
        "scorecard_presentation",
        "stakeholder_presentation",
    }
    first_by_id = {
        artifact["logical_id"]: artifact
        for artifact in first.artifacts
        if artifact["kind"] in reusable_kinds
    }
    second_by_id = {
        artifact["logical_id"]: artifact
        for artifact in second.artifacts
        if artifact["kind"] in reusable_kinds
    }
    assert second_by_id.keys() == first_by_id.keys()
    assert {
        logical_id: artifact["object_key"]
        for logical_id, artifact in second_by_id.items()
    } == {
        logical_id: artifact["object_key"]
        for logical_id, artifact in first_by_id.items()
    }
    assert all(artifact["source_revision"] == 1 for artifact in second_by_id.values())

    second_upload_names = [
        request.filename for request, _content in store.uploads[uploads_after_first:]
    ]
    assert not any(name.startswith("score-") for name in second_upload_names)
    assert not any(name.startswith("scorecard-") for name in second_upload_names)
    assert not any(name.startswith("optimization-presentation-") for name in second_upload_names)


def test_changed_score_content_republishes_dependent_presentation_but_reuses_unchanged_csv(monkeypatch):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    store = _ArtifactStore()
    publication_ids = iter(["ranking", "diagnosis"])
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="dependency-reuse",
        report_configuration_id="config-1", artifact_store=store,
        publication_id_factory=lambda: next(publication_ids),
        task_lookup=lambda _: None, report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda task: [
            stage for stage in _TaskStage.created if stage.taskId == task.id
        ],
    )
    service.start_or_resume({"scope": {}})
    first = service.publish_milestone(
        "ranking", {"coverage": {"complete": True}}, stakeholder_view=_safe_view(),
    )
    changed_view = deepcopy(_safe_view())
    changed_view["questions_and_issues"][0]["finding"] = "A clarified stakeholder question."
    second = service.publish_milestone(
        "diagnosis", {"coverage": {"complete": True}}, stakeholder_view=changed_view,
    )

    first_by_kind = {artifact["kind"]: artifact for artifact in first.artifacts}
    second_by_kind = {artifact["kind"]: artifact for artifact in second.artifacts}
    assert (
        second_by_kind["scorecard_portfolio_csv"]["object_key"]
        == first_by_kind["scorecard_portfolio_csv"]["object_key"]
    )
    assert (
        second_by_kind["score_brief"]["object_key"]
        != first_by_kind["score_brief"]["object_key"]
    )
    assert (
        second_by_kind["scorecard_presentation"]["object_key"]
        != first_by_kind["scorecard_presentation"]["object_key"]
    )


def test_corrupt_reuse_candidate_is_replaced_instead_of_becoming_current(monkeypatch):
    from plexus.optimization import run_report

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    monkeypatch.setattr(run_report, "TaskStage", _TaskStage)
    store = _ArtifactStore()
    first_service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="corrupt-reuse",
        report_configuration_id="config-1", artifact_store=store,
        publication_id_factory=lambda: "ranking",
        task_lookup=lambda _: None, report_lookup=lambda _: None, block_lookup=lambda _: [],
        stage_lookup=lambda task: [
            stage for stage in _TaskStage.created if stage.taskId == task.id
        ],
    )
    state = first_service.start_or_resume({"scope": {}})
    first = first_service.publish_milestone(
        "ranking", {"coverage": {"complete": True}}, stakeholder_view=_safe_view(),
    )
    first_score_brief = next(
        artifact for artifact in first.artifacts if artifact["kind"] == "score_brief"
    )
    store.content_by_key[first_score_brief["object_key"]] = b"corrupt"

    recovered = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="corrupt-reuse",
        report_configuration_id="config-1", artifact_store=store,
        publication_id_factory=lambda: "assessment",
        task_lookup=lambda _: state.task,
        report_lookup=lambda _: state.report,
        block_lookup=lambda _: list(_Block.created),
        stage_lookup=lambda task: [
            stage for stage in _TaskStage.created if stage.taskId == task.id
        ],
    )
    recovered.start_or_resume({"scope": {}})
    second = recovered.publish_milestone(
        "assessment", {"coverage": {"complete": True}}, stakeholder_view=_safe_view(),
    )
    second_score_brief = next(
        artifact for artifact in second.artifacts if artifact["kind"] == "score_brief"
    )

    assert second_score_brief["object_key"] != first_score_brief["object_key"]
    assert second_score_brief["source_revision"] == 2
    assert store.content_by_key[second_score_brief["object_key"]] != b"corrupt"


def test_failed_assessment_artifact_publication_remains_distinct_from_diagnosis(monkeypatch):
    service = _service(monkeypatch)
    state = service.start_or_resume({"scope": {}})
    original_uploader = service._artifact_uploader

    def fail_workbook(task_id, name, content):
        if name.endswith(".xlsx"):
            raise RuntimeError("upload is unavailable")
        return original_uploader(task_id, name, content)

    service._artifact_uploader = fail_workbook
    with pytest.raises(Exception, match="assessment"):
        service.publish_milestone(
            "assessment", {"coverage": {"complete": True}}, stakeholder_view=_safe_view(),
        )

    progress = json.loads(state.blocks["status"].output)["preview"]["summary"]["live_progress"]
    assert progress["phase"] == "publication"
    assert progress["state"] == "failed"
    assert progress["artifact_counts"]["stakeholder_workbook"] == {"completed": 0, "total": 1}
    assert progress["failure"] == {
        "exception_class": "RuntimeError",
        "operation_category": "stakeholder_workbook",
        "retry_classification": "retryable",
        "completed": progress["current"],
        "total": progress["total"],
    }
    assert "upload is unavailable" not in json.dumps(progress)
    assert "stakeholder workbook" in progress["message"].lower()
    assert "diagnosis" not in state.report.output.lower()


def test_procedure_owned_task_rejects_a_conflicting_run_claim(monkeypatch):
    from plexus.optimization import run_report

    procedure_task = _Task(
        identifier="procedure-task",
        metadata=json.dumps({
            "optimization_run_key": "another-run",
            "run_spec": {"scope": {}},
            "attempt_id": "attempt-1",
        }),
    )
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="requested-run",
        report_configuration_id="config-1", existing_task=procedure_task,
        artifact_store=_ArtifactStore(),
    )

    with pytest.raises(ValueError, match="already claimed"):
        service.start_or_resume({"scope": {}})


def test_stage_initialization_interruption_keeps_the_attempt_active_after_publishing_the_report(monkeypatch):
    from plexus.optimization import run_report

    class _FailingStage(_TaskStage):
        @classmethod
        def create(cls, **_values):
            raise RuntimeError("stage service unavailable")

    monkeypatch.setattr(run_report, "Task", _Task)
    monkeypatch.setattr(run_report, "Report", _Report)
    monkeypatch.setattr(run_report, "TaskStage", _FailingStage)
    monkeypatch.setattr(run_report, "ReportBlock", _Block)
    service = run_report.OptimizationRunReportService(
        client=SimpleNamespace(), account_id="account-1", run_key="startup-failure",
        report_configuration_id="config-1", task_lookup=lambda _: None,
        report_lookup=lambda _: None, block_lookup=lambda _: [], stage_lookup=lambda _: [],
        artifact_store=_ArtifactStore(),
    )

    with pytest.raises(run_report.OptimizationRunPublicationError, match="initialize"):
        service.start_or_resume({"scope": {}})

    assert len(_Report.created) == 1
    assert _Task.created[0].status == "RUNNING"
    assert "Status: running" in _Report.created[0].output


def test_publish_interruption_keeps_task_active_and_raises_without_silent_progress(monkeypatch):
    class _FailingBlock(_Block):
        def update(self, **values):
            if self.name == "Decision Evidence" and "r0001" in str(values.get("output", "")):
                raise RuntimeError("attachment envelope write failed")
            return super().update(**values)

    service = _service(monkeypatch, block_class=_FailingBlock)
    state = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    from plexus.optimization.run_report import OptimizationRunPublicationError

    with pytest.raises(OptimizationRunPublicationError, match="assessment"):
        service.publish_milestone("assessment", {"coverage": {"complete": True}}, stakeholder_view=_safe_view())

    assert state.task.status == "RUNNING"
    assert not any(update.get("status") == "FAILED" for update in state.task.updates)
    assert all(stage.status != "FAILED" for stage in _TaskStage.created)


def test_initialization_write_interruption_keeps_the_same_attempt_active(monkeypatch):
    from plexus.optimization import run_report

    class _InterruptedInitialBlock(_Block):
        def update(self, **values):
            if self.name == "Run Status":
                raise RuntimeError("temporary authorization interruption")
            return super().update(**values)

    service = _service(monkeypatch, block_class=_InterruptedInitialBlock)
    with pytest.raises(run_report.OptimizationRunRetryablePublicationError):
        service.start_or_resume({"scope": {}})

    assert service._state is not None
    assert service._state.task.status == "RUNNING"
    assert json.loads(service._state.task.metadata).get("optimization_run_final_status") is None


def test_finalization_write_interruption_keeps_the_same_attempt_active(monkeypatch):
    from plexus.optimization import run_report

    service = _service(monkeypatch)
    state = service.start_or_resume({"scope": {}})
    state.report.update = lambda **_values: (_ for _ in ()).throw(
        RuntimeError("temporary authorization interruption")
    )

    with pytest.raises(run_report.OptimizationRunRetryablePublicationError):
        service.finalize(status="COMPLETED")

    assert state.task.status == "RUNNING"
    assert json.loads(state.task.metadata).get("optimization_run_final_status") is None


def test_finalize_marks_the_single_task_complete_after_the_latest_revision(monkeypatch):
    service = _service(monkeypatch)
    state = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})
    service.publish_milestone("summary", {"coverage": {"complete": True}}, stakeholder_view=_safe_view())

    completed = service.finalize()

    assert completed is state
    assert state.task.status == "COMPLETED"
    assert "Status: completed" in state.report.output


def test_finalize_records_an_honest_incomplete_terminal_state_without_reclassifying_it_as_a_failure(monkeypatch):
    service = _service(monkeypatch)
    state = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    service.finalize(status="incomplete")

    assert state.task.status == "COMPLETED"
    assert __import__("json").loads(state.task.metadata)["optimization_run_final_status"] == "incomplete"
    assert "Status: incomplete" in state.report.output


@pytest.mark.parametrize(
    ("requested", "task_status"),
    [
        ("complete", "COMPLETED"),
        ("completed", "COMPLETED"),
        ("complete_with_unresolved_actions", "COMPLETED"),
        ("completed_with_unresolved_actions", "COMPLETED"),
        ("incomplete", "COMPLETED"),
        ("blocked", "COMPLETED"),
    ],
)
def test_finalize_supports_every_nonfailure_lifecycle_state(monkeypatch, requested, task_status):
    service = _service(monkeypatch)
    state = service.start_or_resume({"window": {"start": "2026-04-30T00:00:00Z"}})

    service.finalize(status=requested)

    assert state.task.status == task_status
    assert json.loads(state.task.metadata)["optimization_run_final_status"] == requested
    assert f"Status: {requested}" in state.report.output


def test_workbook_is_deterministic_macro_free_formula_safe_and_has_all_stakeholder_sheets():
    from plexus.optimization.run_report import build_stakeholder_workbook

    generated_at = datetime(2026, 7, 29, 12, 0, tzinfo=timezone.utc)
    first = build_stakeholder_workbook(_safe_view(), revision_number=1, generated_at=generated_at)
    second = build_stakeholder_workbook(_safe_view(), revision_number=1, generated_at=generated_at)

    assert first.checksum == second.checksum
    assert first.content == second.content
    workbook = load_workbook(BytesIO(first.content), data_only=False)
    assert workbook.sheetnames == [
        "Overview",
        "Portfolio",
        "Priorities",
        "Feedback Investment",
        "Questions and Issues",
        "Optimization Outcomes",
        "Run Log",
        "Definitions",
    ]
    assert first.row_counts["priorities"] == 1
    assert first.row_counts == {
        "portfolio": 1,
        "priorities": 1,
        "feedback_investment": 1,
        "questions_and_issues": 1,
        "optimization_outcomes": 0,
        "run_log": 1,
        "definitions": 1,
    }
    priority_headers = [cell.value for cell in workbook["Priorities"][1]]
    rationale_cell = workbook["Priorities"].cell(2, priority_headers.index("Rationale") + 1)
    assert rationale_cell.data_type != "f"
    assert rationale_cell.value.startswith("'")
    with ZipFile(BytesIO(first.content)) as archive:
        names = archive.namelist()
        assert not any(name.lower().endswith("vbaProject.bin".lower()) for name in names)
        core_properties = archive.read("docProps/core.xml")
        assert core_properties.count(b"2026-07-29T12:00:00Z") == 2
        content = b"".join(archive.read(name) for name in names)
        assert b"opaque-id" not in content
        assert b"do not expose" not in content
    assert all(sheet.sheet_state == "visible" for sheet in workbook.worksheets)
    assert not workbook._external_links
    assert all(cell.data_type != "f" for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row)
    assert {table.name for sheet in workbook.worksheets for table in sheet.tables.values()} >= {
        "PortfolioTable", "PrioritiesTable", "FeedbackInvestmentTable",
        "QuestionsAndIssuesTable", "OptimizationOutcomesTable", "RunLogTable", "DefinitionsTable",
    }
    assert workbook["Portfolio"].freeze_panes == "A2"
    assert workbook["Portfolio"].auto_filter.ref
    assert workbook["Portfolio"].column_dimensions["A"].width > 10


def test_workbook_overview_is_sized_and_wrapped_for_default_opening():
    from plexus.optimization.run_report import build_stakeholder_workbook

    artifact = build_stakeholder_workbook(
        _safe_view(),
        revision_number=1,
        generated_at=datetime(2026, 7, 29, 12, 0, tzinfo=timezone.utc),
    )
    overview = load_workbook(BytesIO(artifact.content))["Overview"]

    assert overview.column_dimensions["A"].width >= len("Optimization Run Overview") + 2
    assert overview.column_dimensions["B"].width >= 40
    assert all(cell.alignment.wrap_text for cell in overview["B"])


def test_workbook_rejects_unsafe_view_keys_instead_of_copying_raw_evidence():
    from plexus.optimization.run_report import build_stakeholder_workbook

    unsafe = _safe_view()
    unsafe["portfolio"][0]["score_id"] = "opaque-id"

    with pytest.raises(ValueError, match="not allowed"):
        build_stakeholder_workbook(unsafe, revision_number=1, generated_at=datetime(2026, 7, 29, tzinfo=timezone.utc))
