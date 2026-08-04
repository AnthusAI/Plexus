"""Outside-in specifications for the reported optimization portfolio runner."""

from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass, field
from hashlib import sha256
from typing import Any

import pytest


def _assessment(scorecard_id: str, score_id: str) -> dict[str, Any]:
    """A minimal, complete assessment packet accepted by the existing runner."""
    return {
        "scope": {"scorecard_id": scorecard_id, "score_id": score_id},
        "coverage": {"complete": True, "failures": []},
        "states": {"optimization": "ready_to_optimize"},
        "champion_version": f"champion-{score_id}",
        "feedback_watermark": "2026-07-01T00:00:00Z",
        "evidence_fingerprint": f"fingerprint-{score_id}",
        "fingerprint": f"fingerprint-{score_id}",
        "evidence": {
            "score_activity": {
                "policy_version": "score-activity-cooldown-v1",
                "recent": False,
                "complete": True,
            }
        },
    }


@dataclass
class _ReportService:
    started: list[dict[str, Any]] = field(default_factory=list)
    milestones: list[tuple[str, dict[str, Any], dict[str, Any]]] = field(default_factory=list)
    terminal: list[str] = field(default_factory=list)
    failures: list[str] = field(default_factory=list)
    progress_updates: list[dict[str, Any]] = field(default_factory=list)
    latest_checkpoint: dict[str, Any] | None = None
    semantic_ledger: dict[str, Any] | None = None
    semantic_commits: list[dict[str, Any]] = field(default_factory=list)
    report: Any = field(default_factory=lambda: type("Report", (), {"id": "report-1"})())

    def start_or_resume(self, run_spec):
        self.started.append(dict(run_spec))
        return type("State", (), {"report": self.report})()

    def publish_milestone(self, milestone, evidence, *, stakeholder_view):
        self.milestones.append((
            milestone,
            deepcopy(dict(evidence)),
            deepcopy(dict(stakeholder_view)),
        ))
        return object()

    def load_latest_checkpoint(self):
        return self.latest_checkpoint

    def load_semantic_budget_ledger(self):
        return deepcopy(self.semantic_ledger)

    def persist_semantic_budget_ledger(self, value):
        self.semantic_ledger = deepcopy(dict(value))
        self.semantic_commits.append(deepcopy(dict(value)))

    def publish_progress(self, **progress):
        self.progress_updates.append(dict(progress))

    def finalize(self, *, status="COMPLETED"):
        self.terminal.append(status)

    def fail(self, message):
        self.failures.append(str(message))


def _dependencies(
    *, rank, assess, diagnose, summary, dispatch, review, report, human_review,
    create_action=None, publish_update=None, optimizer_child_step=None,
    optimizer_child_request=None, diagnosis_preflight=None,
):
    from plexus.optimization.portfolio_run import PortfolioRunDependencies

    return PortfolioRunDependencies(
        rank=rank,
        assess=assess,
        diagnose=diagnose,
        summary=summary,
        dispatch=dispatch,
        review=review,
        report_service=lambda _run_key, _request: report,
        human_review=human_review,
        create_action=create_action,
        publish_update=publish_update,
        optimizer_child_step=optimizer_child_step,
        optimizer_child_request=optimizer_child_request,
        diagnosis_preflight=diagnosis_preflight,
    )


def test_required_diagnosis_evidence_preflight_blocks_semantic_spend_and_optimizer_dispatch():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    calls: list[str] = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": "score"},
        ]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnosis_preflight=lambda _request: calls.append("preflight") or {
            "complete": False,
            "failure_category": "required_evidence_unavailable",
            "message": "Required rubric-memory evidence is unavailable; refresh worker authorization and resume.",
        },
        diagnose=lambda _request: calls.append("diagnose") or {},
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: calls.append("dispatch") or {},
        review=lambda _request: {}, report=report, human_review=lambda _request: {},
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "preflight-failure",
        "max_semantic_diagnoses": 1, "max_semantic_cost_usd": "1",
        "execution_mode": "automatic",
    })

    assert result["status"] == "INCOMPLETE", result.get("error")
    assert calls == ["preflight"]
    assert result["diagnosis_coverage"]["prerequisite_failure_count"] == 1
    assert result["diagnosis_coverage"]["failure_category"] == "required_evidence_unavailable"
    assert any("rubric-memory evidence" in blocker for blocker in result["diagnosis_coverage"]["blockers"])
    assert [milestone for milestone, _evidence, _view in report.milestones] == [
        "started", "ranking", "assessment", "diagnosis", "finalization",
    ]
    final_view = report.milestones[-1][2]
    assert final_view["overview"]["analysis_coverage_status"] == "incomplete"
    assert final_view["overview"]["diagnosis_prerequisite_failure_count"] == 1
    assert final_view["portfolio"][0]["next_action"] == "refresh_worker_authorization_and_resume"
    assert final_view["portfolio"][0]["semantic_diagnosis_status"] == "required_evidence_unavailable"
    assert any(
        issue.get("issue_flag") == "required_evidence_unavailable"
        for issue in final_view["questions_and_issues"]
    )


def test_diagnosis_evidence_preflight_is_not_called_without_a_semantic_target():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    calls: list[str] = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": []},
        assess=lambda _request: {}, diagnose=lambda _request: {},
        diagnosis_preflight=lambda _request: calls.append("preflight") or {"complete": True},
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({"account_id": "account-1", "run_key": "no-semantic-target"})

    assert result["status"] == "COMPLETED"
    assert calls == []


def test_portfolio_run_publishes_idempotent_operator_milestones_after_report_updates():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    updates: list[dict[str, Any]] = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": []},
        assess=lambda _request: {},
        diagnose=lambda _request: {},
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: {},
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: {},
        publish_update=lambda update: updates.append(update) or {"created": True},
    ))

    result = runner.run({"account_id": "account-1", "run_key": "daily-run"})

    assert result.get("error") is None
    assert result["status"] == "COMPLETED"
    assert [update["event_key"] for update in updates] == [
        "optimization:daily-run:started",
        "optimization:daily-run:analysis_ready",
        "optimization:daily-run:completed",
    ]
    assert updates[0]["milestone"] == "STARTED"
    assert updates[1]["milestone"] == "COMPLETED"
    assert updates[2]["milestone"] == "COMPLETED"
    terminal_view = report.milestones[-1][2]
    assert terminal_view["overview"]["lifecycle_status"] == "completed"
    assert "complete" in terminal_view["overview"]["current_activity"].lower()
    assert "Finalizing" not in terminal_view["overview"]["current_activity"]
    assert [milestone for milestone, _evidence, _view in report.milestones] == [
        "started", "ranking", "assessment", "diagnosis", "approval", "finalization",
    ]
    assert all(update["resource_refs"] == [{
        "system": "plexus", "kind": "report", "id": "report-1",
        "relation": "optimization_run",
    }] for update in updates)


def test_portfolio_run_routes_private_ranking_progress_to_the_existing_report_path():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()

    def rank(request):
        callback = request.get("_optimization_rank_progress")
        assert callable(callback)
        callback({
            "phase": "ranking",
            "subphase": "inventory",
            "state": "retrying",
            "current": 12,
            "total": None,
            "unit": "scorecards",
            "elapsed_seconds": 63,
            "next_checkpoint": "Retrying the inventory page.",
            "message": "Inventory has inspected 12 scorecards; retrying a page.",
        })
        callback({
            "phase": "ranking",
            "subphase": "activity_evidence",
            "state": "active",
            "current": 12,
            "total": 12,
            "unit": "scorecards",
            "message": "Score activity evidence is complete.",
        })
        callback({
            "phase": "ranking",
            "subphase": "feedback_analysis",
            "state": "active",
            "current": 5,
            "total": 12,
            "unit": "scorecards",
            "message": "Feedback analysis is running.",
        })
        return {"coverage": {"complete": True}, "ranked": []}

    result = OptimizationPortfolioRunner(_dependencies(
        rank=rank,
        assess=lambda _request: {},
        diagnose=lambda _request: {},
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: {},
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: {},
    )).run({"account_id": "account-1", "run_key": "ranking-progress-run"})

    assert result["status"] == "COMPLETED"
    assert [update["subphase"] for update in report.progress_updates[:3]] == [
        "inventory", "activity_evidence", "feedback_analysis",
    ]
    assert report.progress_updates[0] == {
        "phase": "ranking",
        "subphase": "inventory",
        "current": 12,
        "total": None,
        "message": "Inventory has inspected 12 scorecards; retrying a page.",
        "unit": "scorecards",
        "state": "retrying",
        "elapsed_seconds": 63,
        "next_checkpoint": "Retrying the inventory page.",
        "heartbeat_interval_seconds": None,
    }


@pytest.mark.parametrize("task_terminal", [False, True])
def test_finalization_recovery_uses_durable_state_and_replays_one_terminal_update(
    task_terminal,
):
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService(latest_checkpoint={
        "milestone": "finalization",
        "task_terminal": task_terminal,
        "evidence": {
            "run_key": "recovery-run",
            "run_spec": {"account_id": "account-1", "run_key": "recovery-run"},
            "terminal_status": "INCOMPLETE",
            "coverage": {"complete": True},
            "rank": {"coverage": {"complete": True}, "ranked": []},
            "assessments": [],
            "diagnoses": [],
            "diagnosis_coverage": {"selected_scope_complete": False},
            "actions": [],
            "approval_requests": [],
            "approved_targets": [],
            "dispatch": None,
            "reviews": [],
            "summary": {"coverage": {"complete": False}},
        },
    })
    calls = {"rank": 0, "assess": 0, "diagnose": 0}

    def unexpected(name):
        def call(_request):
            calls[name] += 1
            raise AssertionError(f"{name} must not run after a durable finalization checkpoint")
        return call

    updates: list[dict[str, Any]] = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=unexpected("rank"), assess=unexpected("assess"), diagnose=unexpected("diagnose"),
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("dispatch must not run")),
        review=lambda _request: (_ for _ in ()).throw(AssertionError("review must not run")),
        report=report, human_review=lambda _request: {},
        publish_update=lambda update: updates.append(update) or {"created": len(updates) == 1},
    ))

    result = runner.run({"account_id": "account-1", "run_key": "recovery-run"})

    assert result["status"] == "INCOMPLETE"
    assert calls == {"rank": 0, "assess": 0, "diagnose": 0}
    assert report.milestones == []
    assert report.terminal == ([] if task_terminal else ["INCOMPLETE"])
    assert [update["event_key"] for update in updates] == [
        "optimization:recovery-run:completed"
    ]


def test_finalization_recovery_write_interruption_defers_the_current_attempt():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.run_report import OptimizationRunRetryablePublicationError

    report = _ReportService(latest_checkpoint={
        "milestone": "finalization", "task_terminal": False,
        "evidence": {
            "run_key": "finalize-retry", "terminal_status": "COMPLETED",
            "coverage": {"complete": True}, "rank": {"coverage": {"complete": True}, "ranked": []},
            "assessments": [], "diagnoses": [], "diagnosis_coverage": {},
            "actions": [], "approval_requests": [], "approved_targets": [],
            "dispatch": None, "reviews": [], "promotion_candidates": [], "summary": {},
        },
    })
    report.finalize = lambda **_kwargs: (_ for _ in ()).throw(
        OptimizationRunRetryablePublicationError("temporary finalization interruption")
    )
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: (_ for _ in ()).throw(AssertionError("rank must not run")),
        assess=lambda _request: {}, diagnose=lambda _request: {}, summary=lambda _request: {},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("dispatch must not run")),
        review=lambda _request: (_ for _ in ()).throw(AssertionError("review must not run")),
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({"account_id": "account-1", "run_key": "finalize-retry"})

    assert result["status"] == "RETRYABLE_PUBLICATION"
    assert result["retry"]["reason"] == "retryable_report_publication"
    assert report.failures == []


def test_assessment_recovery_continues_with_diagnosis_without_repeating_rank_or_assess():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    rank_packet = {
        "coverage": {"complete": True},
        "window": {"start": "2026-04-01T00:00:00Z", "end": "2026-07-01T00:00:00Z"},
        "ranked": [{
            "scorecard_id": "opaque-card", "score_id": "opaque-score",
            "scorecard_name": "Example Portfolio", "score_name": "Example Score",
        }],
    }
    assessment = _assessment("opaque-card", "opaque-score")
    assessment["scorecard_name"] = "Example Portfolio"
    assessment["score_name"] = "Example Score"
    report = _ReportService(latest_checkpoint={
        "milestone": "assessment",
        "task_terminal": False,
        "evidence": {
            "run_key": "assessment-recovery",
            "coverage": {"complete": True},
            "rank": rank_packet,
            "assessments": [assessment],
            "diagnoses": [],
            "diagnosis_coverage": {
                "scheduled_count": 1, "completed_count": 0, "failed_count": 0,
                "deferred_by_cap_count": 0, "blockers": [],
            },
            "actions": [], "approval_requests": [], "approved_targets": [],
            "dispatch": None, "reviews": [], "summary": None,
        },
    })
    calls = {"rank": 0, "assess": 0, "diagnose": 0}

    def diagnose(request):
        calls["diagnose"] += 1
        return {**request["assessment"], "states": {"optimization": "monitoring_candidate"}}

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: calls.__setitem__("rank", calls["rank"] + 1),
        assess=lambda _request: calls.__setitem__("assess", calls["assess"] + 1),
        diagnose=diagnose,
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("dispatch must not run")),
        review=lambda _request: (_ for _ in ()).throw(AssertionError("review must not run")),
        report=report, human_review=lambda _request: {"decisions": []},
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "assessment-recovery",
        "max_semantic_cost_usd": "1",
    })

    assert result["status"] in {"COMPLETED", "COMPLETED_WITH_UNRESOLVED_ACTIONS"}
    assert calls == {"rank": 0, "assess": 0, "diagnose": 1}
    assert report.milestones[0][0] == "diagnosis"


def test_review_checkpoint_recovery_only_publishes_the_missing_finalization():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    rank_packet = {
        "coverage": {"complete": True},
        "ranked": [{
            "scorecard_id": "opaque-card", "score_id": "opaque-score",
            "scorecard_name": "Example Portfolio", "score_name": "Example Score",
        }],
    }
    assessment = _assessment("opaque-card", "opaque-score")
    diagnosis = {**assessment, "states": {"optimization": "ready_to_optimize"}}
    report = _ReportService(latest_checkpoint={
        "milestone": "optimization_review",
        "task_terminal": False,
        "evidence": {
            "run_key": "review-recovery", "coverage": {"complete": True},
            "rank": rank_packet, "assessments": [assessment], "diagnoses": [diagnosis],
            "diagnosis_coverage": {
                "scheduled_count": 1, "completed_count": 1, "failed_count": 0,
                "deferred_by_cap_count": 0, "blockers": [],
                "scheduled_scope_complete": True, "selected_scope_complete": True,
            },
            "actions": [], "approval_requests": [],
            "approved_targets": [{"scorecard_id": "opaque-card", "score_id": "opaque-score"}],
            "dispatch": {
                "batches": [{
                    "dispatches": [{"status": "dispatched", "procedure_id": "procedure-1"}],
                    "rejected": [],
                }],
                "rejected": [],
            },
            "reviews": [{"procedure_id": "procedure-1", "promotion_ready": False}],
            "promotion_candidates": [],
            "summary": {"coverage": {"complete": True}},
        },
    })

    def must_not_run(name):
        return lambda _request: (_ for _ in ()).throw(
            AssertionError(f"{name} must not repeat after the review checkpoint")
        )

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=must_not_run("rank"), assess=must_not_run("assess"),
        diagnose=must_not_run("diagnose"), summary=must_not_run("summary"),
        dispatch=must_not_run("dispatch"), review=must_not_run("review"),
        report=report, human_review=must_not_run("human review"),
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "review-recovery",
        "max_semantic_cost_usd": "1",
    })

    assert result.get("error") is None
    assert result["status"] == "COMPLETED"
    assert [milestone for milestone, _evidence, _view in report.milestones] == ["finalization"]
    assert report.terminal == ["COMPLETED"]


def test_retryable_publication_interruption_emits_a_sanitized_scheduled_retry():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.run_report import OptimizationRunRetryablePublicationError

    report = _ReportService()
    def interrupted(*_args, **_kwargs):
        raise OptimizationRunRetryablePublicationError("temporary publication interruption")
    report.publish_milestone = interrupted
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": []},
        assess=lambda _request: {}, diagnose=lambda _request: {},
        summary=lambda _request: {}, dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({"account_id": "account-1", "run_key": "retryable-publication"})

    assert result["status"] == "RETRYABLE_PUBLICATION"
    assert "error" not in result
    assert result["retry"] == {
        "key": "optimization-report-publication",
        "reason": "retryable_report_publication",
        "resume_at": result["retry"]["resume_at"],
    }
    assert result["retry"]["resume_at"].endswith("Z")
    assert "temporary publication interruption" not in str(result)
    assert report.failures == []
    assert report.terminal == []


def test_retryable_milestone_publication_failure_is_nonterminal_at_the_portfolio_boundary():
    """A caller must defer a behind-Report attempt rather than complete it."""
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.run_report import OptimizationRunRetryablePublicationError

    report = _ReportService()

    def interrupted(*_args, **_kwargs):
        raise OptimizationRunRetryablePublicationError(
            "temporary publication interruption"
        )

    report.publish_milestone = interrupted
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": []},
        assess=lambda _request: {}, diagnose=lambda _request: {},
        summary=lambda _request: {}, dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({"account_id": "account-1", "run_key": "fatal-publication"})

    assert result["status"] == "RETRYABLE_PUBLICATION"
    assert result["retry"]["reason"] == "retryable_report_publication"
    assert "error" not in result
    assert report.failures == []
    assert report.terminal == []


def test_retryable_initialization_interruption_defers_without_analysis():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.run_report import OptimizationRunRetryablePublicationError

    report = _ReportService()
    report.start_or_resume = lambda _spec: (_ for _ in ()).throw(
        OptimizationRunRetryablePublicationError("temporary initialization interruption")
    )
    rank_calls = 0

    def rank(_request):
        nonlocal rank_calls
        rank_calls += 1
        raise AssertionError("rank must not run before initialization is durable")

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=rank, assess=lambda _request: {}, diagnose=lambda _request: {},
        summary=lambda _request: {}, dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({"account_id": "account-1", "run_key": "init-retry"})

    assert result["status"] == "RETRYABLE_PUBLICATION"
    assert result["retry"]["key"] == "optimization-report-publication"
    assert rank_calls == 0
    assert report.failures == []


@pytest.mark.parametrize(
    "error_type_name",
    ["ArtifactTicketError", "ArtifactAuthorizationError", "ArtifactTransferError"],
)
def test_retryable_checkpoint_read_defers_current_attempt_and_later_replay_uses_checkpoint(
    error_type_name,
):
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.storage import graphql_artifact_store

    checkpoint = {
        "milestone": "finalization", "task_terminal": True,
        "evidence": {
            "run_key": "read-retry", "terminal_status": "COMPLETED",
            "coverage": {"complete": True},
            "rank": {"coverage": {"complete": True}, "ranked": []},
            "assessments": [], "diagnoses": [], "diagnosis_coverage": {},
            "actions": [], "approval_requests": [], "approved_targets": [],
            "dispatch": None, "reviews": [], "promotion_candidates": [], "summary": {},
        },
    }
    report = _ReportService()
    reads = [getattr(graphql_artifact_store, error_type_name)("temporary read interruption"), checkpoint]

    def load_latest_checkpoint():
        result = reads.pop(0)
        if isinstance(result, Exception):
            from plexus.optimization.run_report import OptimizationRunRetryablePublicationError

            raise OptimizationRunRetryablePublicationError(
                "mapped retryable checkpoint read"
            ) from result
        return result

    report.load_latest_checkpoint = load_latest_checkpoint
    analysis_calls = 0

    def analysis_must_not_run(_request):
        nonlocal analysis_calls
        analysis_calls += 1
        raise AssertionError("analysis must not run while replaying a committed finalization")

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=analysis_must_not_run, assess=analysis_must_not_run,
        diagnose=analysis_must_not_run, summary=analysis_must_not_run,
        dispatch=analysis_must_not_run, review=analysis_must_not_run,
        report=report, human_review=lambda _request: {},
    ))

    first = runner.run({"account_id": "account-1", "run_key": "read-retry"})
    second = runner.run({"account_id": "account-1", "run_key": "read-retry"})

    assert first["status"] == "RETRYABLE_PUBLICATION"
    assert first["retry"]["reason"] == "retryable_report_publication"
    assert second["status"] == "COMPLETED"
    assert analysis_calls == 0
    assert report.failures == []


def test_long_analysis_exposes_incremental_progress_without_extra_report_revisions(monkeypatch):
    import plexus.optimization.portfolio_run as portfolio_run
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    # Keep this synthetic progress run bounded to two semantic diagnoses while
    # retaining a larger deterministic assessment batch.  The production
    # policy diagnoses the top ten; the fixture deliberately uses a smaller
    # policy-sized batch so its assertions cover progress through both phases.
    monkeypatch.setattr(portfolio_run, "MAX_PRIORITY_DIAGNOSES", 2)

    report = _ReportService()
    ranked = [
        {
            "scorecard_id": "card-1",
            "score_id": f"score-{index}",
            "scorecard_name": "Example Portfolio",
            "score_name": f"Example Score {index}",
        }
        for index in range(12)
    ]

    def assess(request):
        packet = _assessment(request["scorecard_id"], request["score_id"])
        if request["score_id"] in {"score-0", "score-1"}:
            packet["states"]["optimization"] = "insufficient_evidence"
        else:
            packet["states"] = {
                "optimization": "repair_required",
                "guideline_health": "invalid",
            }
        return packet

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": ranked},
        assess=assess,
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: {},
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: {"decisions": []},
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "progress-run",
        "max_semantic_diagnoses": 2,
        "max_semantic_cost_usd": "1",
    })

    assert result["status"] == "COMPLETED"
    assert [
        (row["phase"], row["current"], row["total"])
        for row in report.progress_updates
    ] == [
        ("assessment", 0, 12),
        ("assessment", 10, 12),
        ("assessment", 12, 12),
        ("diagnosis", 12, 14),
        ("diagnosis", 13, 14),
        ("diagnosis", 14, 14),
    ]
    assert "Example Score 9" in report.progress_updates[1]["message"]
    assert "semantic diagnosis" in report.progress_updates[-1]["message"].lower()
    assert [milestone for milestone, _, _ in report.milestones] == [
        "started", "ranking", "assessment", "diagnosis", "approval", "finalization",
    ]


def test_portfolio_run_creates_the_living_report_before_analysis_and_only_launches_independently_approved_exact_targets():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.optimizer_dispatch import OptimizerTaskDispatchService

    report = _ReportService()
    backend = _OptimizerChildBackend()
    child_service = OptimizerTaskDispatchService(backend)
    dispatches: list[dict[str, Any]] = []
    review_requests: list[dict[str, Any]] = []
    rank_packet = {
        "coverage": {"complete": True, "failures": []},
        "window": {"start": "2026-04-01T00:00:00Z", "end": "2026-07-01T00:00:00Z"},
        "ranked": [
            {"scorecard_id": "opaque-card", "score_id": "opaque-one", "scorecard_name": "Example", "score_name": "One"},
            {"scorecard_id": "opaque-card", "score_id": "opaque-two", "scorecard_name": "Example", "score_name": "Two"},
        ],
    }

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda request: rank_packet,
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: {**request["assessment"], "states": {"optimization": "ready_to_optimize"}},
        summary=lambda request: {"coverage": {"complete": True}, "per_score_outcomes": []},
        dispatch=lambda request: dispatches.append(request) or {
            "accepted": True,
            "accepted_targets": list(request["targets"]),
            "rejected": [],
        },
        review=lambda request: {"coverage": {"complete": True}, "post_run_state": "promotion_ready", "promotion_ready": True},
        report=report,
        human_review=lambda request: review_requests.append(request) or {
            "decisions": [
                {"scorecard_id": "opaque-card", "score_id": "opaque-one", "decision": "approve", "comment": "go"},
                {"scorecard_id": "opaque-card", "score_id": "opaque-two", "decision": "reject", "comment": "not today"},
            ]
        },
        optimizer_child_step=child_service.step,
        optimizer_child_request=lambda request: {
            **request,
            "optimizer_yaml": "name: optimizer\n",
            "stages": [{"name": "Optimize", "order": 1, "status": "PENDING"}],
        },
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "daily-account-1-2026-07-01",
        "toolchain_version": "test-v1",
        "max_cost_usd": 5.0,
        "max_samples": 50,
        "max_iterations": 2,
        "max_concurrency": 1,
        "max_semantic_cost_usd": "1",
    })

    assert len(report.started) == 1
    assert report.milestones[0][0] == "started"
    assert [milestone[0] for milestone in report.milestones[:6]] == [
        "started", "ranking", "assessment", "diagnosis", "approval", "optimization",
    ]
    assert review_requests[0]["action_key"] == "optimization-approval:daily-account-1-2026-07-01:1"
    assert review_requests[0]["response_schema"]["type"] == "object"
    assert review_requests[0]["expires_in_seconds"] == 24 * 60 * 60
    assert len(review_requests[0]["targets"]) == 2
    assert review_requests[0]["preconditions"]["targets"][0]["scorecard_name"] == "Example"
    assert review_requests[0]["preconditions"]["targets"][0]["score_name"] == "One"
    assert review_requests[0]["preconditions"]["limits"] == {
        "max_cost_usd": 5.0,
        "max_samples": 50,
        "max_iterations": 2,
        "max_concurrency": 1,
    }
    assert review_requests[0]["resource_refs"][0]["kind"] == "report"
    assert dispatches[0]["approved"] is True
    assert [(row["scorecard_id"], row["score_id"]) for row in dispatches[0]["targets"]] == [("opaque-card", "opaque-one")]
    assert dispatches[0]["max_samples"] == 50
    assert dispatches[0]["run_key"] == "daily-account-1-2026-07-01"
    assert result["status"] == "WAITING_FOR_CHILDREN"
    assert result["dispatch"]["children"][0]["target"] == {
        "scorecard_id": "opaque-card", "score_id": "opaque-one",
    }
    assert result["promotion_candidates"] == []
    assert report.terminal == []


class _OptimizerChildBackend:
    """In-memory contract double for the pure optimizer child coordinator."""

    def __init__(self) -> None:
        self.procedures: list[dict[str, Any]] = []
        self.tasks: list[dict[str, Any]] = []
        self.artifacts: dict[str, bytes] = {}
        self.stages: dict[str, list[dict[str, Any]]] = {}
        self.create_procedure_calls = 0
        self.create_task_calls = 0
        self.release_calls = 0

    def procedure_pages_for_account(self, _account_id):
        return iter([{"items": [dict(row) for row in self.procedures], "next_token": None}])

    def create_procedure(self, record):
        self.create_procedure_calls += 1
        row = {"id": f"procedure-{len(self.procedures) + 1}", **dict(record)}
        self.procedures.append(row)
        return dict(row)

    def get_procedure(self, procedure_id):
        return next(dict(row) for row in self.procedures if row["id"] == procedure_id)

    def upload_and_verify_procedure_yaml(self, procedure, optimizer_yaml, metadata):
        key = f"procedures/{procedure['id']}/code.tac"
        self.artifacts[key] = optimizer_yaml.encode("utf-8")
        for row in self.procedures:
            if row["id"] == procedure["id"]:
                row["metadata"] = {
                    **dict(metadata),
                    "code_artifact": {
                        "key": key,
                        "sha256": metadata["optimizer_yaml_sha256"],
                    },
                }
                return dict(row)
        raise AssertionError("procedure must exist before attachment upload")

    def read_procedure_artifact(self, key):
        return self.artifacts[key]

    def task_pages_for_account(self, _account_id):
        return iter([{"items": [dict(row) for row in self.tasks], "next_token": None}])

    def create_task(self, record):
        self.create_task_calls += 1
        assert record["dispatchStatus"] == "HELD"
        row = {"id": f"task-{len(self.tasks) + 1}", **dict(record)}
        self.tasks.append(row)
        return dict(row)

    def get_task(self, task_id):
        return next(dict(row) for row in self.tasks if row["id"] == task_id)

    def task_stage_pages_for_task(self, task_id):
        return iter([{
            "items": [dict(row) for row in self.stages.get(task_id, [])],
            "next_token": None,
        }])

    def reconcile_task_stages(self, task_id, stages):
        existing = self.stages.setdefault(task_id, [])
        for stage in stages:
            if not any(
                row["name"] == stage["name"] and row["order"] == stage["order"]
                for row in existing
            ):
                existing.append({
                    "id": f"stage-{len(existing) + 1}",
                    "taskId": task_id,
                    **dict(stage),
                })
        return [dict(row) for row in existing]

    def release_held_task(self, task_id):
        self.release_calls += 1
        task = next(row for row in self.tasks if row["id"] == task_id)
        assert task["dispatchStatus"] == "HELD"
        task["dispatchStatus"] = "PENDING"


def _optimizer_child_dependencies(report, backend, *, review, dispatch=None, rank=None):
    from plexus.optimization.optimizer_dispatch import OptimizerTaskDispatchService

    service = OptimizerTaskDispatchService(backend)
    return _dependencies(
        rank=rank or (lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{"scorecard_id": "card", "score_id": "score"}],
        }),
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=dispatch or (lambda request: {
            "accepted": True,
            "accepted_targets": list(request["targets"]),
            "rejected": [],
        }),
        review=review,
        report=report,
        human_review=lambda _request: {
            "decisions": [{
                "scorecard_id": "card",
                "score_id": "score",
                "decision": "approve",
            }],
        },
        optimizer_child_step=service.step,
        optimizer_child_request=lambda request: {
            **request,
            "optimizer_yaml": "name: optimizer\n",
            "stages": [{"name": "Optimize", "order": 1, "status": "PENDING"}],
        },
    )


def test_approved_optimizer_child_is_reported_phase_by_phase_then_waits_without_review():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    reviews: list[dict[str, Any]] = []
    runner = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report,
        backend,
        review=lambda request: reviews.append(request) or {},
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "durable-child-run",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    })

    assert result["status"] == "WAITING_FOR_CHILDREN"
    assert reviews == []
    assert report.terminal == []
    assert "finalization" not in [
        milestone for milestone, _evidence, _view in report.milestones
    ]
    child = result["dispatch"]["children"][0]
    assert child["procedure_id"] == "procedure-1"
    assert child["task_id"] == "task-1"
    assert child["launch_state"]["phase"] == "waiting"
    assert child["launch_state"]["launch_spec"]["run_key"] == "durable-child-run"
    assert child["launch_state"]["launch_spec"]["scorecard_id"] == "card"
    assert child["launch_state"]["launch_spec"]["score_id"] == "score"
    assert child["launch_state"]["launch_spec"]["assessment_fingerprint"] == "fingerprint-score"
    assert child["launch_state"]["launch_spec"]["limits"] == {
        "max_cost_usd": 1.0,
        "max_samples": 20,
        "max_iterations": 2,
        "max_concurrency": 1,
    }
    assert result["execution_decisions"]["launched_count"] == 1
    assert result["execution_decisions"]["selected_targets"][0]["launch_status"] == "launched"
    optimization_states = [
        evidence["dispatch"]["children"][0]["launch_state"]["phase"]
        for milestone, evidence, _view in report.milestones
        if milestone == "optimization"
        and isinstance(evidence.get("dispatch"), dict)
        and evidence["dispatch"].get("children")
        and evidence["dispatch"]["children"][0].get("launch_state")
    ]
    assert optimization_states[:9] == [
        "planned",
        "procedure_create_attempted",
        "procedure_record_observed",
        "procedure_provisioned",
        "task_create_attempted",
        "task_record_observed",
        "task_held",
        "release_attempted",
        "waiting",
    ]
    assert optimization_states[-1] == "waiting"
    assert backend.create_procedure_calls == 1
    assert backend.create_task_calls == 1
    assert backend.release_calls == 1


def test_approved_zero_cost_limit_is_reported_incomplete_without_dispatching_or_creating_children():
    import csv
    from datetime import datetime, timezone
    from io import BytesIO
    import json

    from openpyxl import load_workbook

    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.run_report import (
        build_scorecard_artifacts,
        build_stakeholder_presentation,
        build_stakeholder_workbook,
    )

    report = _ReportService()
    backend = _OptimizerChildBackend()
    runner = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report,
        backend,
        dispatch=lambda _request: (_ for _ in ()).throw(
            AssertionError("invalid limits must not reach dispatch")
        ),
        review=lambda _request: (_ for _ in ()).throw(
            AssertionError("no child may be reviewed")
        ),
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{
                "scorecard_id": "card",
                "score_id": "score",
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "valid_feedback_count": 240,
                "disagreement_rate": 0.25,
            }],
        },
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "zero-cost-approved-run",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 0,
            "max_samples": 1,
            "max_iterations": 1,
            "max_concurrency": 1,
        },
    })

    assert result["status"] == "INCOMPLETE"
    assert result["dispatch"]["phase"] == "incomplete"
    assert result["dispatch"]["rejected"] == [{
        "reason": "invalid_run_limits",
        "invalid_fields": ["max_cost_usd"],
    }]
    assert result["dispatch"]["children"] == []
    assert result["dispatch"]["processed_child_keys"] == []
    assert backend.create_procedure_calls == 0
    assert backend.create_task_calls == 0
    assert backend.release_calls == 0
    assert report.terminal == ["INCOMPLETE"]

    final_view = next(
        view
        for milestone, _evidence, view in reversed(report.milestones)
        if milestone == "finalization"
    )
    issue = final_view["questions_and_issues"][0]
    assert issue["affected_disagreement_rate"] == 0.25
    assert "disagreement_rate" not in issue

    workbook = build_stakeholder_workbook(
        final_view,
        revision_number=1,
        generated_at=datetime(2026, 7, 30, tzinfo=timezone.utc),
    )
    reopened = load_workbook(BytesIO(workbook.content), data_only=True)
    portfolio_headers = [cell.value for cell in reopened["Portfolio"][1]]
    portfolio_values = [cell.value for cell in reopened["Portfolio"][2]]
    portfolio_record = dict(zip(portfolio_headers, portfolio_values))
    assert portfolio_record["Primary Disposition"] == "failed_or_incomplete"
    assert portfolio_record["Coverage"] == "incomplete"
    assert portfolio_record["Next Action"] == "provide_valid_run_limits"
    assert "invalid_run_limits" in portfolio_record["Dispatch Rejection"]
    issue_headers = [cell.value for cell in reopened["Questions and Issues"][1]]
    issue_values = [cell.value for cell in reopened["Questions and Issues"][2]]
    issue_record = dict(zip(issue_headers, issue_values))
    assert issue_record["Affected Disagreement Rate"] == 0.25

    uploaded: dict[str, bytes] = {}

    def upload(_task_id, filename, content):
        uploaded[filename] = content
        return f"tasks/task-1/{filename}"

    artifacts = build_scorecard_artifacts(
        final_view,
        revision_number=1,
        task_id="task-1",
        uploader=upload,
    )
    brief = next(
        content.decode("utf-8")
        for filename, content in uploaded.items()
        if "-brief-" in filename
    )
    summary = next(
        content.decode("utf-8")
        for filename, content in uploaded.items()
        if "-summary-" in filename
    )
    portfolio_csv = next(
        content.decode("utf-8-sig")
        for filename, content in uploaded.items()
        if "-portfolio-" in filename
    )
    scorecard_presentation = json.loads(next(
        content
        for filename, content in uploaded.items()
        if "-presentation-" in filename
    ))
    assert "Primary disposition: failed_or_incomplete" in brief
    assert "Next action: provide_valid_run_limits" in brief
    assert "not launched because its run limits are invalid" in brief
    assert "failed_or_incomplete" in summary
    assert "provide_valid_run_limits" in summary
    csv_row = next(csv.DictReader(portfolio_csv.splitlines()))
    assert csv_row["Primary Disposition"] == "failed_or_incomplete"
    assert csv_row["Coverage"] == "incomplete"
    assert csv_row["Next Action"] == "provide_valid_run_limits"
    assert "invalid_run_limits" in csv_row["Dispatch Rejection"]
    score_row = scorecard_presentation["scores"][0]
    assert score_row["primary_disposition"] == "failed_or_incomplete"
    assert score_row["next_action"] == "provide_valid_run_limits"
    assert score_row["dispatch_rejection"]["reason"] == "invalid_run_limits"
    assert scorecard_presentation["questions_and_issues"][0] == issue

    presentation = build_stakeholder_presentation(
        final_view,
        scorecard_artifacts=artifacts,
    )
    assert presentation["primary_disposition_counts"] == {
        "failed_or_incomplete": 1,
    }
    assert presentation["optimization_outcomes"][0]["dispatch_rejection"] == {
        "reason": "invalid_run_limits",
        "invalid_fields": ["max_cost_usd"],
    }


def test_optimizer_child_replay_adopts_reported_refs_and_never_recreates_or_rereleases():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    runner = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report, backend, review=lambda _request: {},
    ))
    request = {
        "account_id": "account-1",
        "run_key": "durable-child-replay",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }

    first = runner.run(request)
    final_optimization = next(
        row for row in reversed(report.milestones) if row[0] == "optimization"
    )
    report.latest_checkpoint = {
        "milestone": "optimization",
        "evidence": final_optimization[1],
    }
    before = (
        backend.create_procedure_calls,
        backend.create_task_calls,
        backend.release_calls,
    )

    replay = runner.run(request)

    assert first["status"] == replay["status"] == "WAITING_FOR_CHILDREN"
    assert first["execution_decisions"]["launched_count"] == 1
    assert replay["execution_decisions"]["launched_count"] == 1
    assert first["execution_decisions"]["selected_targets"][0]["launch_status"] == "launched"
    assert replay["execution_decisions"]["selected_targets"][0]["launch_status"] == "launched"
    assert replay["dispatch"]["children"][0]["procedure_id"] == "procedure-1"
    assert replay["dispatch"]["children"][0]["task_id"] == "task-1"
    assert (
        backend.create_procedure_calls,
        backend.create_task_calls,
        backend.release_calls,
    ) == before


def test_optimizer_child_resume_records_exact_wait_snapshot_but_uses_backend_terminal_readback():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    reviews = []
    runner = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report,
        backend,
        review=lambda request: reviews.append(request) or {
            "coverage": {"complete": True},
            "promotion_ready": False,
        },
    ))
    request = {
        "account_id": "account-1",
        "run_key": "durable-child-snapshot",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }
    first = runner.run(request)
    report.latest_checkpoint = {
        "milestone": report.milestones[-1][0],
        "evidence": report.milestones[-1][1],
    }
    backend.tasks[0]["status"] = "COMPLETED"
    snapshot = [{
        "id": first["dispatch"]["children"][0]["launch_state"]["launch_spec"]["identity"],
        "task_id": "task-1",
        "procedure_id": "procedure-1",
        "scorecard_id": "card",
        "score_id": "score",
        "terminal": True,
        "status": "COMPLETED",
    }]

    resumed = runner.run({**request, "optimizer_child_snapshots": snapshot})

    assert first["status"] == "WAITING_FOR_CHILDREN"
    assert resumed["dispatch"]["last_wait_snapshot"] == snapshot
    assert reviews == [{
        "account_id": "account-1",
        "procedure_id": "procedure-1",
        "persist": False,
    }]


def test_optimizer_child_resume_fails_closed_for_mismatched_wait_snapshot():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    runner = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report, backend, review=lambda _request: {},
    ))
    request = {
        "account_id": "account-1",
        "run_key": "durable-child-mismatched-snapshot",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }
    runner.run(request)
    report.latest_checkpoint = {
        "milestone": report.milestones[-1][0],
        "evidence": report.milestones[-1][1],
    }

    resumed = runner.run({
        **request,
        "optimizer_child_snapshots": [{
            "id": "different-task",
            "task_id": "different-task",
            "procedure_id": "procedure-1",
            "scorecard_id": "card",
            "score_id": "score",
            "terminal": True,
        }],
    })

    assert resumed["status"] == "FAILED"
    assert "does not match durable optimizer children" in resumed["error"]
    assert report.failures


def test_optimizer_child_replay_rejects_changed_limits_before_another_mutation():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    runner = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report, backend, review=lambda _request: {},
    ))
    request = {
        "account_id": "account-1",
        "run_key": "durable-child-frozen-limits",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }

    first = runner.run(request)
    assert first["status"] == "WAITING_FOR_CHILDREN"
    final_optimization = next(
        row for row in reversed(report.milestones) if row[0] == "optimization"
    )
    report.latest_checkpoint = {
        "milestone": "optimization",
        "evidence": final_optimization[1],
    }
    before = (
        backend.create_procedure_calls,
        backend.create_task_calls,
        backend.release_calls,
    )

    changed = runner.run({
        **request,
        "limits": {**request["limits"], "max_samples": 21},
    })

    assert changed["status"] == "FAILED"
    assert "frozen request" in changed["error"]
    assert (
        backend.create_procedure_calls,
        backend.create_task_calls,
        backend.release_calls,
    ) == before


def test_terminal_child_is_reviewed_only_after_a_later_report_resume():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    reviews: list[dict[str, Any]] = []
    runner = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report,
        backend,
        review=lambda request: reviews.append(request) or {
            "coverage": {"complete": True},
            "promotion_ready": False,
        },
    ))
    request = {
        "account_id": "account-1",
        "run_key": "terminal-child-review",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }

    first = runner.run(request)
    assert first["status"] == "WAITING_FOR_CHILDREN"
    assert reviews == []
    final_optimization = next(
        row for row in reversed(report.milestones) if row[0] == "optimization"
    )
    report.latest_checkpoint = {
        "milestone": "optimization",
        "evidence": final_optimization[1],
    }
    backend.tasks[0]["status"] = "COMPLETED"
    backend.tasks[0]["dispatchStatus"] = "DISPATCHED"

    completed = runner.run(request)

    assert completed["status"] == "COMPLETED"
    assert completed["execution_decisions"]["launched_count"] == 1
    assert completed["execution_decisions"]["selected_targets"][0]["launch_status"] == "launched"
    assert reviews == [{
        "account_id": "account-1",
        "procedure_id": "procedure-1",
        "persist": False,
    }]
    assert report.terminal == ["COMPLETED"]


def test_completed_child_procedure_is_reviewed_despite_stale_dispatcher_task_failure():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    reviews: list[dict[str, Any]] = []
    runner = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report,
        backend,
        review=lambda request: reviews.append(request) or {
            "coverage": {"complete": True},
            "post_run_state": "no_safe_improvement",
            "promotion_ready": False,
        },
    ))
    request = {
        "account_id": "account-1",
        "run_key": "completed-procedure-stale-task",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }
    first = runner.run(request)
    report.latest_checkpoint = {
        "milestone": report.milestones[-1][0],
        "evidence": report.milestones[-1][1],
    }
    backend.tasks[0]["status"] = "FAILED"
    backend.tasks[0]["dispatchStatus"] = "ERROR"
    backend.procedures[0]["status"] = "COMPLETED"

    completed = runner.run(request)

    assert first["status"] == "WAITING_FOR_CHILDREN"
    assert completed["status"] == "COMPLETED"
    assert reviews == [{
        "account_id": "account-1",
        "procedure_id": "procedure-1",
        "persist": False,
    }]
    child = completed["dispatch"]["children"][0]
    assert child["launch_state"]["task"]["status"] == "FAILED"
    assert child["launch_state"]["procedure"]["status"] == "COMPLETED"
    assert completed["reviews"][0]["post_run_state"] == "no_safe_improvement"


@pytest.mark.parametrize("dispatch", [
    {
        "phase": "incomplete",
        "children": [],
        "rejected": [{"reason": "stale_assessment"}],
    },
    {
        "phase": "incomplete",
        "children": [{"target": {"scorecard_id": "card", "score_id": "score"}}],
        "rejected": [],
    },
    {
        "phase": "incomplete",
        "children": [{
            "target": {"scorecard_id": "card", "score_id": "score"},
            "procedure_id": "procedure-1",
            "task_id": "task-1",
            "launch_state": {
                "phase": "dispatch_outcome_unknown",
                "task": {"status": "RUNNING"},
            },
        }],
        "rejected": [],
    },
    {
        "phase": "children_terminal",
        "children": [{
            "target": {"scorecard_id": "card", "score_id": "score"},
            "procedure_id": "procedure-1",
            "task_id": "task-1",
            "launch_state": {
                "phase": "terminal",
                "task": {"status": "FAILED"},
            },
        }],
        "rejected": [],
    },
])
def test_terminal_status_fails_closed_for_incomplete_unknown_malformed_or_failed_dispatch(
    dispatch,
):
    from plexus.optimization.portfolio_run import _terminal_status

    state = {
        "rank": {"coverage": {"complete": True}},
        "diagnosis_coverage": {"selected_count": 0},
        "approved_targets": [{"scorecard_id": "card", "score_id": "score"}],
        "dispatch": dispatch,
    }

    assert _terminal_status(state, has_unresolved_actions=False) == "INCOMPLETE"


def test_blocked_dispatch_without_a_child_skips_optimizer_and_review_revisions():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{"scorecard_id": "card", "score_id": "score"}],
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: {
            "accepted": False,
            "accepted_targets": [],
            "rejected": [{"reason": "stale_assessment"}],
        },
        review=lambda _request: (_ for _ in ()).throw(
            AssertionError("blocked target must not be reviewed")
        ),
        report=report,
        human_review=lambda _request: {
            "decisions": [{
                "scorecard_id": "card",
                "score_id": "score",
                "decision": "approve",
            }],
        },
    ))

    result = runner.run({"account_id": "account-1", "run_key": "blocked-child"})

    assert result["status"] == "INCOMPLETE"
    milestones = [milestone for milestone, _evidence, _view in report.milestones]
    assert "optimization" not in milestones
    assert "optimization_review" not in milestones
    assert milestones[-1] == "finalization"


def test_mixed_children_persist_each_terminal_review_once_while_other_children_run():
    from plexus.optimization.optimizer_dispatch import OptimizerTaskDispatchService
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    child_service = OptimizerTaskDispatchService(backend)
    review_calls: list[str] = []
    targets = [
        {"scorecard_id": "card", "score_id": "one"},
        {"scorecard_id": "card", "score_id": "two"},
        {"scorecard_id": "card", "score_id": "rejected"},
    ]

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [
                {**target, "scorecard_name": "Example", "score_name": target["score_id"]}
                for target in targets
            ],
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda request: {
            "accepted": True,
            "accepted_targets": list(request["targets"][:2]),
            "rejected": [{
                "target": dict(request["targets"][2]),
                "reason": "stale_assessment",
            }],
        },
        review=lambda request: review_calls.append(request["procedure_id"]) or {
            "coverage": {"complete": True},
            "promotion_ready": False,
        },
        report=report,
        human_review=lambda _request: {
            "decisions": [
                {**target, "decision": "approve"}
                for target in targets
            ],
        },
        optimizer_child_step=child_service.step,
        optimizer_child_request=lambda request: {
            **request,
            "optimizer_yaml": "name: optimizer\n",
            "stages": [{"name": "Optimize", "order": 1, "status": "PENDING"}],
        },
    ))
    request = {
        "account_id": "account-1",
        "run_key": "mixed-child-review",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }

    first = runner.run(request)
    assert first["status"] == "WAITING_FOR_CHILDREN"
    assert report.terminal == []
    report.latest_checkpoint = {
        "milestone": report.milestones[-1][0],
        "evidence": report.milestones[-1][1],
    }
    backend.tasks[0]["status"] = "COMPLETED"
    backend.tasks[0]["dispatchStatus"] = "DISPATCHED"
    backend.tasks[1]["status"] = "RUNNING"
    backend.tasks[1]["dispatchStatus"] = "DISPATCHED"

    first_wait_snapshot = [
        {
            "id": child["launch_state"]["launch_spec"]["identity"],
            "task_id": child["task_id"],
            "procedure_id": child["procedure_id"],
            "scorecard_id": child["target"]["scorecard_id"],
            "score_id": child["target"]["score_id"],
            "terminal": child["task_id"] == "task-1",
            "status": "COMPLETED" if child["task_id"] == "task-1" else "RUNNING",
        }
        for child in first["dispatch"]["children"]
    ]
    partially_complete = runner.run({
        **request,
        "optimizer_child_snapshots": first_wait_snapshot,
    })

    assert partially_complete["status"] == "WAITING_FOR_CHILDREN"
    assert review_calls == ["procedure-1"]
    assert len(partially_complete["reviews"]) == 1
    assert len(partially_complete["dispatch"]["processed_child_keys"]) == 1
    assert partially_complete["dispatch"]["last_wait_snapshot"] == first_wait_snapshot
    assert partially_complete["execution_decisions"]["launched_count"] == 2
    assert {
        row["score_id"]: row["launch_status"]
        for row in partially_complete["execution_decisions"]["selected_targets"]
    } == {"one": "launched", "two": "launched", "rejected": "selected"}
    assert report.milestones[-1][0] == "optimization_review"
    assert report.terminal == []

    report.latest_checkpoint = {
        "milestone": report.milestones[-1][0],
        "evidence": report.milestones[-1][1],
    }
    unchanged = runner.run(request)
    assert unchanged["status"] == "WAITING_FOR_CHILDREN"
    assert review_calls == ["procedure-1"]

    report.latest_checkpoint = {
        "milestone": report.milestones[-1][0],
        "evidence": report.milestones[-1][1],
    }
    backend.tasks[1]["status"] = "COMPLETED"
    second_wait_snapshot = [{
        "id": first["dispatch"]["children"][1]["launch_state"]["launch_spec"]["identity"],
        "task_id": "task-2",
        "procedure_id": "procedure-2",
        "scorecard_id": "card",
        "score_id": "two",
        "terminal": True,
        "status": "COMPLETED",
    }]
    completed = runner.run({
        **request,
        "optimizer_child_snapshots": second_wait_snapshot,
    })

    assert completed["status"] == "INCOMPLETE"
    assert review_calls == ["procedure-1", "procedure-2"]
    assert len(completed["reviews"]) == 2
    assert len(completed["dispatch"]["processed_child_keys"]) == 2
    assert completed["dispatch"]["last_wait_snapshot"] == second_wait_snapshot
    assert report.terminal == ["INCOMPLETE"]


def test_terminal_child_retries_failed_review_and_replaces_it_when_indexed_evidence_becomes_conclusive():
    from plexus.optimization.optimizer_dispatch import OptimizerTaskDispatchService
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    child_service = OptimizerTaskDispatchService(backend)
    review_calls: list[str] = []
    review_outcomes = iter([
        {
            "coverage": {"complete": False},
            "post_run_state": "failed_or_incomplete",
            "promotion_ready": False,
        },
        {
            "coverage": {"complete": True},
            "post_run_state": "no_safe_improvement",
            "promotion_ready": False,
        },
    ])
    runner = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report,
        backend,
        review=lambda request: (
            review_calls.append(request["procedure_id"])
            or next(review_outcomes)
        ),
    ))
    request = {
        "account_id": "account-1",
        "run_key": "retry-inconclusive-terminal-review",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }

    first = runner.run(request)
    assert first["status"] == "WAITING_FOR_CHILDREN"
    report.latest_checkpoint = {
        "milestone": report.milestones[-1][0],
        "evidence": report.milestones[-1][1],
    }
    backend.tasks[0]["status"] = "COMPLETED"
    backend.tasks[0]["dispatchStatus"] = "DISPATCHED"

    inconclusive = runner.run(request)

    assert inconclusive["status"] == "INCOMPLETE"
    assert review_calls == ["procedure-1"]
    assert len(inconclusive["reviews"]) == 1
    assert inconclusive["reviews"][0]["post_run_state"] == "failed_or_incomplete"
    assert inconclusive["dispatch"]["processed_child_keys"] == []
    assert not any(
        action.get("kind") == "promotion_approval"
        for action in inconclusive["actions"]
    )
    assert report.terminal == ["INCOMPLETE"]

    report.latest_checkpoint = {
        "milestone": report.milestones[-1][0],
        "evidence": report.milestones[-1][1],
    }
    completed = runner.run(request)

    assert completed["status"] == "COMPLETED"
    assert review_calls == ["procedure-1", "procedure-1"]
    assert len(completed["reviews"]) == 1
    assert completed["reviews"][0]["post_run_state"] == "no_safe_improvement"
    assert len(completed["dispatch"]["processed_child_keys"]) == 1
    assert not any(
        action.get("kind") == "promotion_approval"
        for action in completed["actions"]
    )
    assert [
        milestone for milestone, _evidence, _view in report.milestones
    ].count("optimization_review") == 2
    assert report.terminal == ["INCOMPLETE", "COMPLETED"]


def test_terminal_failed_child_is_reviewed_once_and_remains_fail_closed_on_resume():
    from plexus.optimization.optimizer_dispatch import OptimizerTaskDispatchService
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    child_service = OptimizerTaskDispatchService(backend)
    review_calls: list[str] = []
    targets = [
        {"scorecard_id": "card", "score_id": "success-one", "score_name": "Success one"},
        {"scorecard_id": "card", "score_id": "success-two", "score_name": "Success two"},
        {"scorecard_id": "card", "score_id": "failure", "score_name": "Failure"},
    ]
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": targets,
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda request: {
            "accepted": True,
            "accepted_targets": list(request["targets"]),
            "rejected": [],
        },
        review=lambda request: review_calls.append(request["procedure_id"]) or {
            "coverage": {"complete": True},
            "post_run_state": (
                "failed_or_incomplete"
                if request["procedure_id"] == "procedure-3"
                else "no_safe_improvement"
            ),
            "promotion_ready": False,
        },
        report=report,
        human_review=lambda _request: {
            "decisions": [{**target, "decision": "approve"} for target in targets],
        },
        optimizer_child_step=child_service.step,
        optimizer_child_request=lambda request: {
            **request,
            "optimizer_yaml": "name: optimizer\n",
            "stages": [{"name": "Optimize", "order": 1, "status": "PENDING"}],
        },
    ))
    request = {
        "account_id": "account-1",
        "run_key": "failed-sibling-review",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }
    first = runner.run(request)
    report.latest_checkpoint = {
        "milestone": report.milestones[-1][0],
        "evidence": report.milestones[-1][1],
    }
    for task in backend.tasks[:2]:
        task["status"] = "COMPLETED"
        task["dispatchStatus"] = "DISPATCHED"
    backend.tasks[2]["status"] = "FAILED"
    backend.tasks[2]["dispatchStatus"] = "DISPATCHED"
    backend.procedures[2]["status"] = "FAILED"

    completed = runner.run(request)

    assert first["status"] == "WAITING_FOR_CHILDREN"
    assert completed["status"] == "INCOMPLETE"
    assert review_calls == ["procedure-1", "procedure-2", "procedure-3"]
    assert len(completed["reviews"]) == 3
    assert len(completed["dispatch"]["processed_child_keys"]) == 3
    failure = next(
        row for row in report.milestones[-1][2]["optimization_outcomes"]
        if row["score_name"] == "Failure"
    )
    assert failure["outcome"] == "failed_or_incomplete"
    assert failure["outcome"] != "awaiting_optimizer_review"

    final_review = next(
        row for row in reversed(report.milestones) if row[0] == "optimization_review"
    )
    report.latest_checkpoint = {"milestone": "optimization_review", "evidence": final_review[1]}
    replay = runner.run(request)

    assert replay["status"] == "INCOMPLETE"
    assert review_calls == ["procedure-1", "procedure-2", "procedure-3"]
    assert len(replay["dispatch"]["processed_child_keys"]) == 3


def test_incomplete_finalization_replaces_legacy_terminal_reviews_once_without_recreating_children():
    from plexus.optimization.optimizer_dispatch import OptimizerTaskDispatchService
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    backend = _OptimizerChildBackend()
    child_service = OptimizerTaskDispatchService(backend)
    review_calls: list[str] = []
    targets = [
        {"scorecard_id": "card", "score_id": score_id}
        for score_id in ("one", "two", "failed")
    ]
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": targets},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda request: {
            "accepted": True, "accepted_targets": list(request["targets"]), "rejected": [],
        },
        review=lambda request: review_calls.append(request["procedure_id"]) or {
            "coverage": {"complete": True},
            "post_run_state": (
                "failed_or_incomplete"
                if request["procedure_id"] == "procedure-3"
                else "no_safe_improvement"
            ),
            "promotion_ready": False,
        },
        report=report,
        human_review=lambda _request: {
            "decisions": [{**target, "decision": "approve"} for target in targets],
        },
        optimizer_child_step=child_service.step,
        optimizer_child_request=lambda request: {
            **request,
            "optimizer_yaml": "name: optimizer\n",
            "stages": [{"name": "Optimize", "order": 1, "status": "PENDING"}],
        },
    ))
    request = {
        "account_id": "account-1",
        "run_key": "legacy-terminal-review-repair",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    }

    first = runner.run(request)
    checkpoint = deepcopy(next(
        row for row in reversed(report.milestones) if row[0] == "optimization"
    )[1])
    checkpoint["terminal_status"] = "INCOMPLETE"
    checkpoint["dispatch"]["phase"] = "children_terminal"
    legacy_reviews = []
    for index, child in enumerate(checkpoint["dispatch"]["children"]):
        task = backend.tasks[index]
        procedure = backend.procedures[index]
        failed = index == 2
        task["status"] = "FAILED" if failed else "COMPLETED"
        task["dispatchStatus"] = "DISPATCHED"
        procedure["status"] = "FAILED" if failed else "COMPLETED"
        child["launch_state"] = {
            **child["launch_state"],
            "phase": "terminal",
            "task": dict(task),
            "procedure": dict(procedure),
        }
        if index < 2:
            legacy_reviews.append({
                "scope": dict(child["target"]),
                "procedure_id": child["procedure_id"],
                "optimizer_child_key": child["launch_state"]["launch_spec"]["identity"],
                "coverage": {"complete": True},
                "post_run_state": "no_safe_improvement",
                "promotion_ready": False,
            })
    checkpoint["reviews"] = legacy_reviews
    checkpoint["dispatch"]["processed_child_keys"] = [
        row["optimizer_child_key"] for row in legacy_reviews
    ]
    report.latest_checkpoint = {
        "milestone": "finalization", "task_terminal": False, "evidence": checkpoint,
    }
    create_counts = (
        backend.create_procedure_calls, backend.create_task_calls, backend.release_calls,
    )

    repaired = runner.run(request)

    assert first["status"] == "WAITING_FOR_CHILDREN"
    assert repaired["status"] == "INCOMPLETE"
    assert review_calls == ["procedure-1", "procedure-2", "procedure-3"]
    assert len(repaired["reviews"]) == 3
    assert len(repaired["dispatch"]["processed_child_keys"]) == 3
    assert create_counts == (
        backend.create_procedure_calls, backend.create_task_calls, backend.release_calls,
    )

    finalization = next(
        row for row in reversed(report.milestones) if row[0] == "finalization"
    )
    report.latest_checkpoint = {
        "milestone": "finalization", "task_terminal": True, "evidence": finalization[1],
    }
    replay = runner.run(request)

    assert replay["status"] == "INCOMPLETE"
    assert review_calls == ["procedure-1", "procedure-2", "procedure-3"]


def test_report_publication_failure_before_optimizer_mutation_prevents_child_creation():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.run_report import OptimizationRunPublicationError

    class FailingReport(_ReportService):
        def publish_milestone(self, milestone, evidence, *, stakeholder_view):
            child = ((evidence.get("dispatch") or {}).get("children") or [{}])[0]
            phase = (child.get("launch_state") or {}).get("phase")
            if milestone == "optimization" and phase == "procedure_create_attempted":
                raise OptimizationRunPublicationError("phase publication failed")
            return super().publish_milestone(
                milestone, evidence, stakeholder_view=stakeholder_view,
            )

    report = FailingReport()
    backend = _OptimizerChildBackend()
    result = OptimizationPortfolioRunner(_optimizer_child_dependencies(
        report, backend, review=lambda _request: {},
    )).run({
        "account_id": "account-1",
        "run_key": "durable-child-publication-failure",
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 20,
            "max_iterations": 2,
            "max_concurrency": 1,
        },
    })

    assert result["status"] == "FAILED"
    assert backend.create_procedure_calls == 0
    assert backend.create_task_calls == 0
    assert backend.release_calls == 0


def test_incomplete_ranking_is_published_and_finalized_incomplete_without_assessment_or_optimizer_dispatch():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    calls = {"assess": 0, "review": 0, "dispatch": 0}
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": False, "failures": ["page failed"]}, "ranked": []},
        assess=lambda _request: calls.__setitem__("assess", calls["assess"] + 1),
        diagnose=lambda _request: None,
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: calls.__setitem__("dispatch", calls["dispatch"] + 1),
        review=lambda _request: calls.__setitem__("review", calls["review"] + 1),
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("must not ask for approval")),
    ))

    result = runner.run({"account_id": "account-1", "run_key": "incomplete", "limits": {"max_cost_usd": 1.0, "max_samples": 1, "max_iterations": 1, "max_concurrency": 1}})

    assert result["status"] == "INCOMPLETE"
    assert calls == {"assess": 0, "review": 0, "dispatch": 0}
    assert report.terminal == ["INCOMPLETE"]
    assert [row[0] for row in report.milestones] == ["started", "ranking", "finalization"]
    assert report.milestones[-1][2]["overview"]["lifecycle_status"] == "incomplete"


def test_ranking_and_all_assessments_are_published_as_distinct_milestones_before_semantic_diagnosis_begins():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    rank_packet = {
        "coverage": {"complete": True},
        "ranked": [
            {"scorecard_id": "card", "score_id": "one"},
            {"scorecard_id": "card", "score_id": "two"},
        ],
    }
    assessed: list[str] = []

    def assess(request):
        if not assessed:
            assert [row[0] for row in report.milestones] == ["started", "ranking"]
            ranking_evidence = report.milestones[-1][1]
            assert ranking_evidence["rank"]["ranked"] == rank_packet["ranked"]
            assert ranking_evidence["assessments"] == []
        assessed.append(request["score_id"])
        return _assessment(request["scorecard_id"], request["score_id"])

    def diagnose(request):
        assert assessed == ["one", "two"]
        assert [row[0] for row in report.milestones] == ["started", "ranking", "assessment"]
        published = report.milestones[-1][1]
        assert len(published["assessments"]) == 2
        assert published["diagnoses"] == []
        return {
            **request["assessment"],
            "states": {"optimization": "monitoring_candidate"},
        }

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: rank_packet,
        assess=assess,
        diagnose=diagnose,
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("must not dispatch")),
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("must not request approval")),
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "assessment-first",
        "max_semantic_cost_usd": "1",
    })

    assert result["status"] == "COMPLETED"
    assert [row[0] for row in report.milestones[:4]] == [
        "started", "ranking", "assessment", "diagnosis"
    ]


def test_semantic_diagnosis_skips_incomplete_assessments_but_preserves_them_in_report():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    ranked = [
        {
            "scorecard_id": "card",
            "score_id": score_id,
            "scorecard_name": "Example Portfolio",
            "score_name": score_name,
        }
        for score_id, score_name in (
            ("stale", "Stale Score"),
            ("complete", "Complete Score"),
        )
    ]
    diagnosed: list[str] = []

    def assess(request):
        packet = _assessment(request["scorecard_id"], request["score_id"])
        packet["states"] = {"optimization": "monitoring_candidate"}
        if request["score_id"] == "stale":
            packet["coverage"] = {
                "complete": False,
                "failures": ["champion changed since frozen ranking"],
            }
        return packet

    def diagnose(request):
        diagnosed.append(request["score_id"])
        return dict(request["assessment"])

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": ranked},
        assess=assess,
        diagnose=diagnose,
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: (_ for _ in ()).throw(
            AssertionError("monitoring candidates must not dispatch")
        ),
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: {},
    ))

    runner.run({
        "account_id": "account-1",
        "run_key": "skip-incomplete-diagnosis",
        "max_semantic_diagnoses": 2,
        "max_semantic_cost_usd": "1",
    })

    assert diagnosed == ["complete"]
    diagnosis_evidence, diagnosis_view = next(
        (evidence, view)
        for milestone, evidence, view in report.milestones
        if milestone == "diagnosis"
    )
    assert len(diagnosis_evidence["assessments"]) == 2
    assert diagnosis_evidence["diagnosis_coverage"]["selected_count"] == 1
    assert diagnosis_evidence["diagnosis_coverage"]["scheduled_count"] == 1
    assert diagnosis_evidence["diagnosis_coverage"]["incomplete_assessment_count"] == 1
    assert {row["score_name"] for row in diagnosis_view["portfolio"]} == {
        "Stale Score", "Complete Score",
    }


def test_one_slot_semantic_diagnosis_prefers_an_actionable_score_over_known_repair():
    """A deterministic repair case stays visible without spending the one model slot."""
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    diagnosed: list[str] = []
    ranked = [
        {
            "scorecard_id": "card", "score_id": "blocked", "evidence_rank": 1,
            "scorecard_name": "Example Portfolio", "score_name": "Known Repair",
        },
        {
            "scorecard_id": "card", "score_id": "actionable", "evidence_rank": 2,
            "scorecard_name": "Example Portfolio", "score_name": "Actionable Score",
        },
    ]

    def assess(request):
        packet = _assessment(request["scorecard_id"], request["score_id"])
        if request["score_id"] == "blocked":
            packet["states"] = {
                "optimization": "repair_required",
                "guideline_health": "missing",
            }
            packet["primary_next_action"] = "repair_guidelines"
        return packet

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": ranked},
        assess=assess,
        diagnose=lambda request: diagnosed.append(request["score_id"]) or request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("must wait for approval")),
        review=lambda _request: {}, report=report,
        human_review=lambda _request: {"decisions": []},
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "one-slot-actionable-diagnosis",
        "wait_for_human": True, "max_semantic_diagnoses": 1,
        "max_semantic_cost_usd": "1",
    })

    assert diagnosed == ["actionable"]
    assert result["diagnosis_coverage"]["scheduled_count"] == 1
    assert result["diagnosis_coverage"]["deterministic_repair_blocker_count"] == 1
    diagnosis_view = next(
        view for milestone, _evidence, view in report.milestones if milestone == "diagnosis"
    )
    rows = {row["score_name"]: row for row in diagnosis_view["portfolio"]}
    assert rows["Known Repair"]["evidence_rank"] == 1
    assert rows["Known Repair"]["readiness"] == "repair_required"
    assert rows["Known Repair"]["next_action"] == "repair_guidelines"
    assert rows["Known Repair"]["semantic_diagnosis_status"] == "not_selected"


def test_semantic_diagnosis_keeps_actionable_candidates_in_evidence_order():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    diagnosed: list[str] = []
    ranked = [
        {"scorecard_id": "card", "score_id": "blocked", "evidence_rank": 1},
        {"scorecard_id": "card", "score_id": "first", "evidence_rank": 2},
        {"scorecard_id": "card", "score_id": "second", "evidence_rank": 3},
    ]

    def assess(request):
        packet = _assessment(request["scorecard_id"], request["score_id"])
        if request["score_id"] == "blocked":
            packet["states"] = {
                "optimization": "repair_required",
                "guideline_health": "invalid",
            }
        return packet

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": ranked},
        assess=assess,
        diagnose=lambda request: diagnosed.append(request["score_id"]) or request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("must wait for approval")),
        review=lambda _request: {}, report=report,
        human_review=lambda _request: {"decisions": []},
    ))

    runner.run({
        "account_id": "account-1", "run_key": "actionable-evidence-order",
        "wait_for_human": True, "max_semantic_diagnoses": 2,
        "max_semantic_cost_usd": "1",
    })

    assert diagnosed == ["first", "second"]


def test_semantic_diagnosis_continues_past_top_ten_in_rank_order():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    ranked = [
        {
            "scorecard_id": "card",
            "score_id": f"score-{index:02d}",
            "scorecard_name": "Example Portfolio",
            "score_name": f"Score {index:02d}",
        }
        for index in range(14)
    ]
    diagnosed: list[str] = []

    def assess(request):
        packet = _assessment(request["scorecard_id"], request["score_id"])
        if request["score_id"] in {"score-04", "score-11"}:
            packet["states"] = {"optimization": "monitoring_candidate"}
        return packet

    def diagnose(request):
        diagnosed.append(request["score_id"])
        return dict(request["assessment"])

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": ranked},
        assess=assess,
        diagnose=diagnose,
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("must wait for approval")),
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: {"decisions": []},
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "bounded-diagnosis",
        "wait_for_human": True,
        "max_semantic_diagnoses": 14,
        "max_semantic_cost_usd": "1",
    })

    assert diagnosed == [
        "score-00", "score-01", "score-02", "score-03", "score-04",
        "score-05", "score-06", "score-07", "score-08", "score-09",
        "score-10", "score-11", "score-12", "score-13",
    ]
    diagnosis_evidence = next(
        evidence for milestone, evidence, _view in report.milestones
        if milestone == "diagnosis"
    )
    assert diagnosis_evidence["diagnosis_coverage"] == {
        "policy_version": "portfolio-diagnosis-scope-v3",
        "ranked_count": 14,
        "top_priority_count": 10,
        "monitoring_candidate_count": 2,
        "overlap_count": 1,
            "selected_count": 14,
            "incomplete_assessment_count": 0,
            "deterministic_repair_blocker_count": 0,
            "scheduled_count": 14,
        "deferred_by_cap_count": 0,
        "deferred_by_budget_count": 0,
        "deferred_after_failure_count": 0,
        "completed_count": 14,
        "failed_count": 0,
        "budget_exhausted_count": 0,
        "outcome_unknown_count": 0,
        "authority_publication_failure_count": 0,
        "skipped_count": 0,
        "max_semantic_diagnoses": 14,
        "scheduled_scope_complete": True,
        "selected_scope_complete": True,
        "portfolio_semantic_complete": True,
        "blockers": [],
    }
    approval_target_ids = {
        target["score_id"]
        for request in result["approval_requests"]
        for target in request["targets"]
    }
    assert approval_target_ids == {
        "score-00", "score-01", "score-02", "score-03",
        "score-05", "score-06", "score-07", "score-08", "score-09",
        "score-10", "score-12", "score-13",
    }
    diagnosis_view = next(
        view for milestone, _evidence, view in report.milestones
        if milestone == "diagnosis"
    )
    assert diagnosis_view["portfolio"][10]["semantic_diagnosis_status"] == "complete"
    assert diagnosis_view["portfolio"][10]["readiness"] == "ready_to_optimize"
    assert diagnosis_view["portfolio"][10]["next_action"] == "review"


def test_semantic_diagnosis_limit_runs_the_highest_priority_subset_and_reports_deferred_coverage():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    diagnosed: list[str] = []
    ranked = [
        {"scorecard_id": "card", "score_id": f"score-{index:02d}"}
        for index in range(10)
    ]
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": ranked},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: diagnosed.append(request["score_id"]) or {
            **request["assessment"],
            "states": {"optimization": "repair_required"},
        },
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("must not dispatch")),
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("must not request approval")),
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "semantic-limit",
        "max_semantic_diagnoses": 2,
        "max_semantic_cost_usd": "1",
    })

    assert result["status"] == "INCOMPLETE"
    assert diagnosed == ["score-00", "score-01"]
    assert report.terminal == ["INCOMPLETE"]
    coverage = result["diagnosis_coverage"]
    assert coverage["selected_count"] == 10
    assert coverage["scheduled_count"] == 2
    assert coverage["deferred_by_cap_count"] == 8
    assert coverage["completed_count"] == 2
    assert coverage["scheduled_scope_complete"] is True
    assert coverage["selected_scope_complete"] is False
    assert coverage["blockers"] == []

    diagnosis_view = next(
        view for milestone, _evidence, view in report.milestones
        if milestone == "diagnosis"
    )
    assert diagnosis_view["overview"]["diagnosis_coverage"] == (
        "2 of 2 scheduled diagnoses returned; 0 incomplete results; "
        "0 execution failures; 8 deferred by the configured diagnosis limit"
    )
    assert diagnosis_view["overview"]["diagnosis_scheduled_count"] == 2
    assert diagnosis_view["overview"]["diagnosis_deferred_count"] == 8
    assert diagnosis_view["overview"]["diagnosis_limit_reached"] is True
    assert diagnosis_view["overview"]["diagnosis_limit_type"] == "configured_count_limit"
    assert diagnosis_view["overview"]["analysis_incomplete_reason"] == "configured_count_limit"
    from plexus.optimization.run_report import build_stakeholder_presentation

    presentation = build_stakeholder_presentation(
        diagnosis_view,
        scorecard_artifacts=[],
    )
    assert presentation["decision_summary"]["headline"] == (
        "The configured run limit left 8 candidates unanalyzed"
    )


def test_zero_semantic_diagnosis_limit_defers_selected_work_without_invoking_a_model():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    diagnosed: list[str] = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{"scorecard_id": "card", "score_id": "score-00"}],
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: diagnosed.append(request["score_id"]) or request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("must not dispatch")),
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("must not request approval")),
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "zero-semantic-limit",
        "max_semantic_diagnoses": 0,
        "max_semantic_cost_usd": "0",
    })

    assert result["status"] == "INCOMPLETE"
    assert diagnosed == []
    assert result["diagnosis_coverage"]["scheduled_count"] == 0
    assert result["diagnosis_coverage"]["deferred_by_cap_count"] == 1
    assert result["diagnosis_coverage"]["failed_count"] == 0


def test_positive_semantic_scope_with_missing_budget_fails_before_any_diagnosis_call():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    calls = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{"scorecard_id": "card", "score_id": "score"}],
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: calls.append(request) or request["assessment"],
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "missing-semantic-budget",
        "max_semantic_diagnoses": 1,
    })

    assert result["status"] == "INCOMPLETE"
    assert calls == []
    assert "positive semantic diagnosis scope" in result["diagnosis_coverage"]["blockers"][0]


def test_run_spec_records_loaded_toolchain_versions_when_caller_omits_identity():
    from plexus.optimization.portfolio_run import _run_spec

    spec = _run_spec({}, account_id="account-1", run_key="daily-run")

    assert spec["toolchain_version"] != "unknown"
    assert "plexus/" in spec["toolchain_version"]
    assert "tactus/" in spec["toolchain_version"]


def test_run_spec_includes_immutable_build_revision_when_available(monkeypatch):
    from plexus.optimization.portfolio_run import _run_spec

    monkeypatch.setenv("AWS_COMMIT_ID", "0123456789abcdef")

    spec = _run_spec({}, account_id="account-1", run_key="daily-run")

    assert "build/0123456789abcdef" in spec["toolchain_version"]


def test_default_positive_semantic_scope_with_omitted_cap_and_budget_never_calls_diagnose():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    calls = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{"scorecard_id": "card", "score_id": "score"}],
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: calls.append(request) or request["assessment"],
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "default-missing-semantic-budget",
    })

    assert result["status"] == "INCOMPLETE"
    assert calls == []
    assert result["diagnosis_coverage"]["scheduled_count"] == 1
    assert "positive semantic diagnosis scope" in result["diagnosis_coverage"]["blockers"][0]


def test_returned_incomplete_diagnosis_is_not_counted_successful_or_ready():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    updates = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{"scorecard_id": "card", "score_id": "score"}],
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: {
            **request["assessment"],
            "status": "incomplete",
            "coverage": {"complete": False, "failures": ["semantic outcome unknown"]},
            "states": {"readiness": "incomplete", "optimization": "incomplete"},
        },
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
        publish_update=lambda update: updates.append(update) or {"created": True},
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "returned-incomplete-diagnosis",
        "max_semantic_diagnoses": 1,
        "max_semantic_cost_usd": "1",
    })

    assert result["status"] == "INCOMPLETE"
    assert result["diagnosis_coverage"]["completed_count"] == 0
    assert result["diagnosis_coverage"]["incomplete_count"] == 1
    assert result["diagnosis_coverage"]["selected_scope_complete"] is False
    assert result["diagnoses"][0]["status"] == "incomplete"
    assert result["approval_requests"] == []
    assert not any(update["event_key"].endswith(":analysis_ready") for update in updates)


def test_budget_exhaustion_preserves_partial_evidence_and_never_emits_analysis_ready():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.semantic_budget import SemanticUsage

    report = _ReportService()
    updates = []
    contacted_targets = []

    def diagnose(request):
        contacted_targets.append(request["score_id"])
        coordinator = request["_semantic_budget_coordinator"]
        view = coordinator.view(
            target_id=f"card:{request['score_id']}", call_site="test-call", max_attempts=1
        )
        plan = view.direct_plan(
            attempt=1, max_input_tokens=200, max_output_tokens=20,
            request_payload={"target": request["score_id"]},
        )
        decision = view.reserve_direct(plan)
        view.settle_direct(
            decision.reservation_id,
            SemanticUsage(input_tokens=200, output_tokens=20),
            output_text="{}",
        )
        return request["assessment"]

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [
                {"scorecard_id": "card", "score_id": "one"},
                {"scorecard_id": "card", "score_id": "two"},
            ],
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=diagnose,
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
        publish_update=lambda update: updates.append(update) or {"created": True},
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "semantic-exhaustion",
        "max_semantic_diagnoses": 2,
        "max_semantic_cost_usd": "0.00009",
    })

    assert result["status"] == "INCOMPLETE"
    assert contacted_targets == ["one", "two"]
    assert result["diagnosis_coverage"]["completed_count"] == 1
    assert result["diagnosis_coverage"]["failed_count"] == 1
    assert result["diagnosis_coverage"]["budget_exhausted_count"] == 1
    assert result["diagnosis_coverage"]["outcome_unknown_count"] == 0
    assert result["diagnosis_coverage"]["deferred_by_budget_count"] == 0
    assert result["diagnoses"][0]["scope"]["score_id"] == "one"
    assert result["diagnoses"][1]["status"] == "budget_exhausted"
    assert not any(update["event_key"].endswith(":analysis_ready") for update in updates)
    assert result["approval_requests"] == []


def test_diagnosis_failure_preserves_the_published_assessment_milestone():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{
                "scorecard_id": "card",
                "score_id": "score",
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
            }],
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda _request: (_ for _ in ()).throw(RuntimeError("diagnosis unavailable")),
        summary=lambda _request: {},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("must not dispatch")),
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("must not request approval")),
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "diagnosis-failure",
        "max_semantic_cost_usd": "1",
    })

    assert result["status"] == "INCOMPLETE"
    assert result.get("error") is None
    assert [row[0] for row in report.milestones] == [
        "started", "ranking", "assessment", "diagnosis", "approval", "finalization"
    ]
    assert report.milestones[2][1]["assessments"]
    assert result["diagnoses"][0]["status"] == "incomplete"
    assert result["diagnosis_coverage"]["failed_count"] == 1
    assert report.failures == []


def test_stale_or_rejected_targets_are_reported_and_never_receive_terminal_review_or_promotion():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    reviews: list[dict[str, Any]] = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "window": {"start": "a", "end": "b"}, "ranked": [{"scorecard_id": "card", "score_id": "score"}]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: {"accepted": False, "rejected": [{"reason": "stale_assessment"}], "dispatches": []},
        review=lambda request: reviews.append(request) or {},
        report=report,
        human_review=lambda _request: {"decisions": [{"scorecard_id": "card", "score_id": "score", "decision": "approve"}]},
    ))

    result = runner.run({"account_id": "account-1", "run_key": "stale", "limits": {"max_cost_usd": 1.0, "max_samples": 1, "max_iterations": 1, "max_concurrency": 1}})

    assert result["status"] == "INCOMPLETE"
    assert result["promotion_candidates"] == []
    assert reviews == []
    assert report.terminal == ["INCOMPLETE"]


def test_tactus_checkpoint_mode_leaves_the_same_report_running_until_an_authoritative_review_response_arrives():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "window": {"start": "a", "end": "b"}, "ranked": [{"scorecard_id": "card", "score_id": "score"}]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("dispatch must wait for Human.review")),
        review=lambda _request: None,
        report=report,
        human_review=lambda _request: {"decisions": []},
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "wait",
        "wait_for_human": True,
        "max_semantic_cost_usd": "1",
        "limits": {"max_cost_usd": 1.0, "max_samples": 1, "max_iterations": 1, "max_concurrency": 1},
    })

    assert result["status"] == "WAITING_FOR_APPROVAL"
    assert report.terminal == []
    assert result["approval_requests"][0]["action_key"] == "optimization-approval:wait:1"


def test_automatic_mode_dispatches_a_complete_target_without_an_optimization_approval():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    dispatches: list[dict[str, Any]] = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": "score"},
        ]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda request: dispatches.append(request) or {
            "accepted": True,
            "accepted_targets": list(request["targets"]),
            "rejected": [],
        },
        review=lambda _request: {}, report=report,
        human_review=lambda _request: (_ for _ in ()).throw(
            AssertionError("automatic mode must not request optimization approval")
        ),
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "automatic-dispatch",
        "execution_mode": "automatic", "wait_for_human": True,
        "max_semantic_cost_usd": "1",
        "limits": {"max_cost_usd": 1.0, "max_samples": 1,
                   "max_iterations": 1, "max_concurrency": 1},
    })

    assert len(dispatches) == 1
    assert dispatches[0]["authorization"] == {
        "mode": "automatic", "source": "deterministic_policy",
    }
    assert result["status"] != "WAITING_FOR_APPROVAL"
    assert result["approval_requests"] == []
    decisions = result["execution_decisions"]
    assert decisions["mode"] == "automatic"
    assert decisions["selected_count"] == 1
    assert decisions["launched_count"] == 0
    assert decisions["selected_targets"][0]["launch_status"] == "selected"
    assert decisions["rejected_count"] >= 1  # durable child authority is absent in this unit seam
    assert decisions["selected_targets"][0]["scorecard_id"] == "card"
    assert decisions["selected_targets"][0]["authorization_source"] == "deterministic_policy"
    assert result["promotion_candidates"] == []
    assert not any(action.get("kind") == "promotion_approval" for action in result["actions"])


def test_automatic_mode_diagnoses_multiple_candidates_but_obeys_frozen_top_k_execution_limit():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    dispatches: list[dict[str, Any]] = []
    report = _ReportService()
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": "first"},
            {"scorecard_id": "card", "score_id": "second"},
        ]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda request: dispatches.append(request) or {
            "accepted": True,
            "accepted_targets": list(request["targets"]),
            "rejected": [],
        },
        review=lambda _request: {}, report=report,
        human_review=lambda _request: (_ for _ in ()).throw(
            AssertionError("automatic mode must not request optimization approval")
        ),
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "automatic-top-one",
        "execution_mode": "automatic",
        "max_execution_targets": 1,
        "max_semantic_diagnoses": 2,
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 1,
            "max_iterations": 1,
            "max_concurrency": 1,
        },
    })

    assert [[target["score_id"] for target in call["targets"]] for call in dispatches] == [["first"]]
    assert result["execution_decisions"]["selected_count"] == 1
    assert any(
        row["score_id"] == "second" and row["reason"] == "execution_target_limit"
        for row in result["execution_decisions"]["rejected_targets"]
    )
    assert report.started[0]["max_execution_targets"] == 1


def test_automatic_dispatch_backfills_fresh_targets_after_ranked_freshness_rejections():
    """Feature: accepted optimizer children fill the target limit, not attempts.

    Scenario: higher ranked candidates become stale during validation
      Given five rank-ordered eligible targets and a two-target execution cap
      When the first validation rejects one stale target
      Then validation continues in evidence order until two fresh targets are accepted
      And later unused candidates are recorded as execution_target_limit
      And a resume reuses the persisted dispatch evidence without duplicate validation
    """
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    dispatches: list[dict[str, Any]] = []
    child_steps: list[str] = []

    def dispatch(request):
        dispatches.append(request)
        target = request["targets"][0]
        if target["score_id"] == "rank-one":
            return {
                "accepted": False,
                "accepted_targets": [],
                "rejected": [{"target": target, "reason": "stale_assessment"}],
            }
        return {"accepted": True, "accepted_targets": [target], "rejected": []}

    def child_step(request, _state, *, may_mutate):
        child_steps.append(request["score_id"])
        return {
            "phase": "waiting",
            "procedure_id": f"procedure-{request['score_id']}",
            "task_id": f"task-{request['score_id']}",
        }

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": f"rank-{number}"}
            for number in ("one", "two", "three", "four", "five")
        ]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=dispatch,
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("no approval")),
        optimizer_child_step=child_step,
        optimizer_child_request=lambda request: request,
    ))
    request = {
        "account_id": "account-1",
        "run_key": "automatic-freshness-backfill",
        "execution_mode": "automatic",
        "max_execution_targets": 2,
        "max_semantic_diagnoses": 5,
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 1,
            "max_iterations": 1,
            "max_concurrency": 1,
        },
    }

    first = runner.run(request)

    assert [[row["score_id"] for row in call["targets"]] for call in dispatches] == [
        ["rank-one"], ["rank-two"], ["rank-three"],
    ]
    assert all(len(call["targets"]) <= 5 for call in dispatches)
    assert [child["target"]["score_id"] for child in first["dispatch"]["children"]] == [
        "rank-two", "rank-three",
    ]
    reasons = {
        row["score_id"]: row["reason"]
        for row in first["execution_decisions"]["rejected_targets"]
    }
    assert reasons == {
        "rank-one": "stale_assessment",
        "rank-four": "execution_target_limit",
        "rank-five": "execution_target_limit",
    }

    final_optimization = next(
        row for row in reversed(report.milestones) if row[0] == "optimization"
    )
    assert (
        first["execution_decisions"]["selected_count"]
        == len(final_optimization[1]["approved_targets"])
        == len(first["dispatch"]["children"])
        == 2
    )
    report.latest_checkpoint = {"milestone": "optimization", "evidence": final_optimization[1]}
    first_child_step_counts = {
        score_id: child_steps.count(score_id) for score_id in ("rank-two", "rank-three")
    }
    replay = runner.run(request)

    assert len(dispatches) == 3
    assert [child["target"]["score_id"] for child in replay["dispatch"]["children"]] == [
        "rank-two", "rank-three",
    ]
    replay_optimization = next(
        row for row in reversed(report.milestones) if row[0] == "optimization"
    )
    assert (
        replay["execution_decisions"]["selected_count"]
        == len(replay_optimization[1]["approved_targets"])
        == len(replay["dispatch"]["children"])
        == 2
    )
    assert child_steps.count("rank-two") == first_child_step_counts["rank-two"] + 1
    assert child_steps.count("rank-three") == first_child_step_counts["rank-three"] + 1


@pytest.mark.parametrize("value", [0, -1, 1.5, 6, True, "not-a-number"])
def test_execution_target_limit_rejects_values_outside_one_through_five(value):
    from plexus.optimization.portfolio_run import _max_execution_targets

    with pytest.raises(ValueError, match="one through five"):
        _max_execution_targets({"max_execution_targets": value})


@pytest.mark.parametrize(("phase", "procedure_id", "task_id", "expected"), [
    (None, None, None, 0),
    ("dispatch_outcome_unknown", "procedure-1", "task-1", 0),
    ("waiting", "procedure-1", "task-1", 1),
    ("running", "procedure-1", "task-1", 1),
    ("terminal", "procedure-1", "task-1", 1),
    ("waiting", "procedure-1", None, 0),
])
def test_launched_count_requires_exact_owned_durable_child_evidence(
    phase, procedure_id, task_id, expected,
):
    from plexus.optimization.portfolio_run import (
        _empty_execution_decisions,
        _reconcile_execution_launch_evidence,
    )

    child = {
        "target": {"scorecard_id": "card", "score_id": "score"},
        "procedure_id": procedure_id,
        "task_id": task_id,
        "launch_state": {"phase": phase} if phase is not None else None,
    }

    decisions = _empty_execution_decisions("automatic")
    decisions["selected_targets"] = [{
        "scorecard_id": "card", "score_id": "score",
        "reason": "eligible_for_launch",
        "authorization_source": "deterministic_policy",
    }]
    _reconcile_execution_launch_evidence(decisions, [child])

    assert decisions["launched_count"] == expected
    assert decisions["selected_targets"][0]["launch_status"] == (
        "launched" if expected else "selected"
    )


@pytest.mark.parametrize("field", [
    "procedure_id", "task_id", "scorecard_id", "score_id",
])
def test_whitespace_only_child_identity_is_not_durable_launch_authority(field):
    from plexus.optimization.portfolio_run import (
        _empty_execution_decisions,
        _reconcile_execution_launch_evidence,
    )

    child = {
        "target": {"scorecard_id": "card", "score_id": "score"},
        "procedure_id": "procedure-1",
        "task_id": "task-1",
        "launch_state": {"phase": "waiting"},
    }
    if field in child["target"]:
        child["target"][field] = " \t "
    else:
        child[field] = " \t "
    decisions = _empty_execution_decisions("automatic")
    decisions["selected_targets"] = [{
        "scorecard_id": child["target"]["scorecard_id"],
        "score_id": child["target"]["score_id"],
        "launch_status": "selected",
    }]

    _reconcile_execution_launch_evidence(decisions, [child])

    assert decisions["launched_count"] == 0
    assert decisions["selected_targets"][0]["launch_status"] == "selected"


@pytest.mark.parametrize(("phase", "include_ids"), [
    ("dispatch_outcome_unknown", True),
    ("waiting", False),
])
def test_runner_output_does_not_mark_uncertain_or_unowned_child_launched(
    phase, include_ids,
):
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()

    def child_step(_request, _state, *, may_mutate):
        return {
            "phase": phase,
            **({"procedure_id": "procedure-1", "task_id": "task-1"} if include_ids else {}),
        }

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": "score"},
        ]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"], summary=lambda _request: {},
        dispatch=lambda request: {
            "accepted": True, "accepted_targets": list(request["targets"]), "rejected": [],
        },
        review=lambda _request: {}, report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("no approval")),
        optimizer_child_step=child_step,
        optimizer_child_request=lambda request: request,
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": f"unlaunched-{phase}-{include_ids}",
        "execution_mode": "automatic", "max_semantic_cost_usd": "1",
        "limits": {"max_cost_usd": 1.0, "max_samples": 1,
                   "max_iterations": 1, "max_concurrency": 1},
    })

    assert result["execution_decisions"]["launched_count"] == 0
    assert result["execution_decisions"]["selected_targets"][0]["launch_status"] == "selected"


def test_execution_selection_preserves_only_supplied_safe_target_names():
    from plexus.optimization.portfolio_run import (
        _execution_selection, _execution_target_row,
    )

    named = _assessment("card-named", "score-named")
    named.update({"scorecard_name": "Safe Portfolio", "score_name": "Safe Score"})
    unnamed = _assessment("card-unnamed", "score-unnamed")
    unnamed["states"] = {"optimization": "incomplete"}
    selected, decisions = _execution_selection(
        "automatic", "promotion_ready", [named, unnamed], [named, unnamed],
        ranked_rows=[
            {"scorecard_id": "card-named", "score_id": "score-named", "evidence_rank": 1},
            {"scorecard_id": "card-unnamed", "score_id": "score-unnamed", "evidence_rank": 2},
        ],
        max_samples=1,
    )
    decisions["selected_targets"] = [
        _execution_target_row(row, reason="eligible_for_launch") for row in selected
    ]

    selected_row = decisions["selected_targets"][0]
    assert selected_row["scorecard_name"] == "Safe Portfolio"
    assert selected_row["score_name"] == "Safe Score"
    rejected_row = decisions["rejected_targets"][0]
    assert rejected_row["scorecard_id"] == "card-unnamed"
    assert "scorecard_name" not in rejected_row
    assert "score_name" not in rejected_row


def test_automatic_mode_launches_complete_target_while_incomplete_sibling_is_rejected():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    dispatches: list[dict[str, Any]] = []

    def diagnose(request):
        if request["score_id"] == "deferred":
            return {
                "scope": {"scorecard_id": "card", "score_id": "deferred"},
                "coverage": {"complete": False, "failures": ["semantic deferred"]},
                "states": {"optimization": "incomplete"},
            }
        return request["assessment"]

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": "complete"},
            {"scorecard_id": "card", "score_id": "deferred"},
        ]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=diagnose, summary=lambda _request: {},
        dispatch=lambda request: dispatches.append(request) or {
            "accepted": True, "accepted_targets": list(request["targets"]), "rejected": [],
        },
        review=lambda _request: {}, report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("no approval")),
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "automatic-sibling",
        "execution_mode": "automatic", "max_semantic_cost_usd": "1",
        "limits": {"max_cost_usd": 1.0, "max_samples": 1,
                   "max_iterations": 1, "max_concurrency": 1},
    })

    assert [[target["score_id"] for target in call["targets"]] for call in dispatches] == [["complete"]]
    rejected = result["execution_decisions"]["rejected_targets"]
    assert any(row["score_id"] == "deferred" and row["reason"] == "incomplete_diagnosis" for row in rejected)


def test_ordinary_incomplete_diagnosis_does_not_prevent_later_independent_target():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    diagnosed: list[str] = []
    dispatches: list[dict[str, Any]] = []

    def diagnose(request):
        diagnosed.append(request["score_id"])
        if request["score_id"] == "inconclusive-first":
            return {
                "scope": {
                    "scorecard_id": request["scorecard_id"],
                    "score_id": request["score_id"],
                },
                "coverage": {"complete": False, "failures": ["inconclusive evidence"]},
                "states": {"optimization": "incomplete"},
                "status": "incomplete",
            }
        return request["assessment"]

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": "inconclusive-first"},
            {"scorecard_id": "card", "score_id": "complete-second"},
        ]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=diagnose,
        summary=lambda _request: {},
        dispatch=lambda request: dispatches.append(request) or {
            "accepted": True,
            "accepted_targets": list(request["targets"]),
            "rejected": [],
        },
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("no approval")),
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "continue-after-inconclusive",
        "execution_mode": "automatic",
        "max_semantic_diagnoses": 2,
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 1,
            "max_iterations": 1,
            "max_concurrency": 1,
        },
    })

    assert diagnosed == ["inconclusive-first", "complete-second"]
    assert [[row["score_id"] for row in call["targets"]] for call in dispatches] == [
        ["complete-second"]
    ]
    assert result["diagnosis_coverage"]["incomplete_count"] == 1
    assert result["diagnosis_coverage"]["completed_count"] == 1
    assert result["diagnosis_coverage"]["scheduled_scope_complete"] is False


def test_resume_from_partial_diagnosis_checkpoint_attempts_only_remaining_target():
    from plexus.optimization.portfolio_run import (
        OptimizationPortfolioRunner,
        _run_spec,
    )

    request = {
        "account_id": "account-1",
        "run_key": "resume-partial-diagnosis",
        "execution_mode": "automatic",
        "max_semantic_diagnoses": 2,
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0,
            "max_samples": 1,
            "max_iterations": 1,
            "max_concurrency": 1,
        },
    }
    first = _assessment("card", "inconclusive-first")
    second = _assessment("card", "complete-second")
    incomplete = {
        "scope": {"scorecard_id": "card", "score_id": "inconclusive-first"},
        "coverage": {"complete": False, "failures": ["inconclusive evidence"]},
        "states": {"optimization": "incomplete"},
        "status": "incomplete",
    }
    run_spec = _run_spec(
        request,
        account_id="account-1",
        run_key="resume-partial-diagnosis",
    )
    report = _ReportService(latest_checkpoint={
        "milestone": "diagnosis",
        "task_terminal": False,
        "evidence": {
            "run_key": "resume-partial-diagnosis",
            "run_spec": run_spec,
            "rank": {
                "coverage": {"complete": True},
                "ranked": [
                    {"scorecard_id": "card", "score_id": "inconclusive-first"},
                    {"scorecard_id": "card", "score_id": "complete-second"},
                ],
            },
            "assessments": [first, second],
            "diagnoses": [incomplete],
            "diagnosis_coverage": {
                "selected_count": 2,
                "scheduled_count": 2,
                "completed_count": 0,
                "incomplete_count": 1,
                "failed_count": 0,
                "deferred_by_cap_count": 0,
                "deferred_after_incomplete_count": 1,
                "scheduled_scope_complete": False,
                "selected_scope_complete": False,
                "blockers": [],
            },
            "actions": [],
            "approval_requests": [],
            "approved_targets": [],
            "execution_mode": "automatic",
            "execution_decisions": {"mode": "automatic", "selected_targets": [], "rejected_targets": [], "selected_count": 0, "rejected_count": 0, "launched_count": 0},
            "dispatch": None,
            "reviews": [],
            "promotion_candidates": [],
            "summary": {},
        },
    })
    diagnosed: list[str] = []
    dispatches: list[dict[str, Any]] = []

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: (_ for _ in ()).throw(AssertionError("rank must not rerun")),
        assess=lambda _request: (_ for _ in ()).throw(AssertionError("assess must not rerun")),
        diagnose=lambda diagnosis_request: (
            diagnosed.append(diagnosis_request["score_id"])
            or diagnosis_request["assessment"]
        ),
        summary=lambda _request: {},
        dispatch=lambda dispatch_request: dispatches.append(dispatch_request) or {
            "accepted": True,
            "accepted_targets": list(dispatch_request["targets"]),
            "rejected": [],
        },
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("no approval")),
    ))

    result = runner.run(request)

    assert diagnosed == ["complete-second"]
    assert [[row["score_id"] for row in call["targets"]] for call in dispatches] == [
        ["complete-second"]
    ]
    assert [row["scope"]["score_id"] for row in result["diagnoses"]] == [
        "inconclusive-first",
        "complete-second",
    ]


def test_execution_mode_is_frozen_in_run_identity_fingerprint_and_recovery():
    from plexus.optimization.portfolio_run import (
        OptimizationPortfolioRunner, OptimizationRunPublicationError,
        _portfolio_evidence_fingerprint, _run_key, _run_spec,
    )

    base = {"account_id": "account-1", "max_semantic_cost_usd": "1"}
    assert _run_key({**base, "execution_mode": "automatic"}) != _run_key({
        **base, "execution_mode": "approval_required",
    })
    automatic = _run_spec({**base, "execution_mode": "automatic"}, account_id="account-1", run_key="same")
    approval = _run_spec({**base, "execution_mode": "approval_required"}, account_id="account-1", run_key="same")
    assert automatic["execution_mode"] == "automatic"
    assert _run_spec(base, account_id="account-1", run_key="same")["execution_mode"] == "approval_required"
    assert _portfolio_evidence_fingerprint({"run_key": "same", "run_spec": automatic}) != _portfolio_evidence_fingerprint({"run_key": "same", "run_spec": approval})

    report = _ReportService(latest_checkpoint={
        "milestone": "ranking", "evidence": {
            "run_key": "same", "run_spec": automatic, "rank": {"coverage": {"complete": True}, "ranked": []},
            "assessments": [], "diagnoses": [], "diagnosis_coverage": {}, "actions": [],
            "approval_requests": [], "approved_targets": [], "dispatch": None, "reviews": [],
            "promotion_candidates": [], "summary": {},
        },
    })
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {}, assess=lambda _request: {}, diagnose=lambda _request: {},
        summary=lambda _request: {}, dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))
    result = runner.run({**base, "run_key": "same", "execution_mode": "approval_required"})
    assert result["status"] == "FAILED"
    assert "execution mode" in result["error"].lower()


def test_execution_candidate_policy_is_frozen_in_run_identity_spec_and_fingerprint():
    from plexus.optimization.portfolio_run import (
        _portfolio_evidence_fingerprint, _run_key, _run_spec,
    )

    base = {"account_id": "account-1", "max_semantic_cost_usd": "1"}
    assert _run_key({**base, "execution_candidate_policy": "promotion_ready"}) != _run_key({
        **base,
        "execution_candidate_policy": "promotion_ready_plus_bounded_diagnostic",
    })
    default_spec = _run_spec(base, account_id="account-1", run_key="same")
    diagnostic_spec = _run_spec(
        {**base, "execution_candidate_policy": "promotion_ready_plus_bounded_diagnostic"},
        account_id="account-1", run_key="same",
    )
    assert default_spec["execution_candidate_policy"] == "promotion_ready"
    assert diagnostic_spec["execution_candidate_policy"] == (
        "promotion_ready_plus_bounded_diagnostic"
    )
    assert _portfolio_evidence_fingerprint({"run_key": "same", "run_spec": default_spec}) != (
        _portfolio_evidence_fingerprint({"run_key": "same", "run_spec": diagnostic_spec})
    )


def _bounded_diagnostic_assessment(score_id: str, *, valid_feedback_count: int = 50) -> dict[str, Any]:
    """A structurally safe, non-ready candidate for a bounded experiment."""
    packet = _assessment("card", score_id)
    first_class_count = valid_feedback_count // 2
    packet.update({
        "readiness_state": "insufficient_evidence",
        "primary_next_action": "collect_targeted_classes",
        "guideline_state": "consistent",
        "feedback_rubric_state": "consistent",
        "class_counts": {
            "positive": first_class_count,
            "negative": valid_feedback_count - first_class_count,
        },
        "blockers": ["reachable class below minimum: negative"],
    })
    packet["states"] = {
        "optimization": "insufficient_evidence",
        "guideline_health": "consistent",
        "feedback_rubric_health": "consistent",
    }
    return packet


def _launchable_diagnosis(score_id: str) -> dict[str, Any]:
    return {
        "scope": {"scorecard_id": "card", "score_id": score_id},
        "coverage": {"complete": True, "failures": []},
        "states": {
            "optimization": "insufficient_evidence",
            "guideline_health": "consistent",
            "feedback_rubric_health": "consistent",
        },
        "blockers": [],
        "stakeholder_questions": [],
    }


def test_bounded_diagnostic_policy_preserves_the_full_safe_pool_for_freshness_validation():
    from plexus.optimization.portfolio_run import _execution_selection

    assessments = [
        _bounded_diagnostic_assessment("rank-four"),
        _bounded_diagnostic_assessment("rank-one"),
        _bounded_diagnostic_assessment("rank-three"),
        _bounded_diagnostic_assessment("rank-two"),
    ]
    ranked = [
        {"scorecard_id": "card", "score_id": score_id, "evidence_rank": rank}
        for rank, score_id in enumerate(
            ("rank-one", "rank-two", "rank-three", "rank-four"), start=1
        )
    ]
    diagnoses = [_launchable_diagnosis(row["scope"]["score_id"]) for row in assessments]

    selected, decisions = _execution_selection(
        "automatic",
        "promotion_ready_plus_bounded_diagnostic",
        assessments,
        diagnoses,
        ranked_rows=ranked,
        max_samples=50,
    )

    assert [row["score_id"] for row in selected] == [
        "rank-one", "rank-two", "rank-three", "rank-four",
    ]
    assert all(row["candidate_kind"] == "bounded_diagnostic" for row in selected)
    assert decisions["rejected_targets"] == []


def test_automatic_bounded_diagnostic_dispatch_caps_accepted_children_without_blocking_later_promotion_ready_target():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    dispatches: list[dict[str, Any]] = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": f"diagnostic-{index}"}
            for index in range(1, 5)
        ] + [{"scorecard_id": "card", "score_id": "promotion-ready"}]},
        assess=lambda request: (
            _assessment(request["scorecard_id"], request["score_id"])
            if request["score_id"] == "promotion-ready"
            else _bounded_diagnostic_assessment(request["score_id"])
        ),
        diagnose=lambda request: (
            request["assessment"]
            if request["score_id"] == "promotion-ready"
            else _launchable_diagnosis(request["score_id"])
        ),
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda request: dispatches.append(request) or {
            "accepted": True, "accepted_targets": list(request["targets"]), "rejected": [],
        },
        review=lambda _request: {}, report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("no approval")),
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "automatic-bounded-diagnostic-limit",
        "execution_mode": "automatic",
        "execution_candidate_policy": "promotion_ready_plus_bounded_diagnostic",
        "max_execution_targets": 4,
        "max_semantic_diagnoses": 5,
        "max_semantic_cost_usd": "1",
        "limits": {
            "max_cost_usd": 1.0, "max_samples": 50,
            "max_iterations": 1, "max_concurrency": 1,
        },
    })

    assert [[row["score_id"] for row in call["targets"]] for call in dispatches] == [
        ["diagnostic-1"], ["diagnostic-2"], ["diagnostic-3"], ["promotion-ready"],
    ]
    assert [row["target"]["score_id"] for row in result["dispatch"]["children"]] == [
        "diagnostic-1", "diagnostic-2", "diagnostic-3", "promotion-ready",
    ]
    assert any(
        row["score_id"] == "diagnostic-4"
        and row["reason"] == "bounded_diagnostic_target_limit"
        for row in result["execution_decisions"]["rejected_targets"]
    )


def test_bounded_diagnostic_replays_the_live_revision_shape_and_selects_three_safe_targets():
    """Regression fixture mirrors the restricted latest decision-revision shape."""
    from plexus.optimization.portfolio_run import _execution_selection

    def assessment(score_id: str, counts: dict[str, int]) -> dict[str, Any]:
        packet = _bounded_diagnostic_assessment(score_id, valid_feedback_count=sum(counts.values()))
        packet["class_counts"] = counts
        packet["evidence"].pop("configuration_readable", None)
        packet["evidence"].pop("terminal_classes_resolved", None)
        return packet

    candidates = {
        "rank-51": assessment("rank-51", {"yes": 31, "no": 19}),
        "rank-56": assessment("rank-56", {"yes": 25, "no": 25}),
        "rank-84": assessment("rank-84", {"yes": 30, "no": 20}),
        "rank-86": assessment("rank-86", {"yes": 28, "no": 22}),
        "rank-153": assessment("rank-153", {"yes": 0, "no": 50}),
    }
    diagnoses = {
        score_id: _launchable_diagnosis(score_id)
        for score_id in candidates
    }
    diagnoses["rank-84"]["states"]["guideline_health"] = "inconclusive"
    ranked = [
        {"scorecard_id": "card", "score_id": score_id, "evidence_rank": evidence_rank}
        for evidence_rank, score_id in (
            (51, "rank-51"), (56, "rank-56"), (84, "rank-84"),
            (86, "rank-86"), (153, "rank-153"),
        )
    ]

    selected, decisions = _execution_selection(
        "automatic",
        "promotion_ready_plus_bounded_diagnostic",
        list(candidates.values()),
        list(diagnoses.values()),
        ranked_rows=ranked,
        max_samples=50,
    )

    assert [target["score_id"] for target in selected] == ["rank-51", "rank-56", "rank-86"]
    reasons = {row["score_id"]: row["reason"] for row in decisions["rejected_targets"]}
    assert reasons["rank-84"] == "bounded_diagnostic_guideline_not_consistent"
    assert reasons["rank-153"] == "bounded_diagnostic_insufficient_observed_classes"


@pytest.mark.parametrize(
    ("description", "change"),
    [
        ("missing champion", lambda packet, diagnosis: packet.pop("champion_version")),
        ("unreadable configuration", lambda packet, diagnosis: packet.update({"configuration_readable": False})),
        ("missing guidelines", lambda packet, diagnosis: packet.update({"guideline_state": "missing", "states": {**packet["states"], "guideline_health": "missing"}})),
        ("code conflict", lambda packet, diagnosis: packet.update({"guideline_state": "potential_code_conflict", "states": {**packet["states"], "guideline_health": "potential_code_conflict"}})),
        ("stakeholder question", lambda packet, diagnosis: diagnosis.update({"stakeholder_questions": ["clarify"]})),
        ("cooldown", lambda packet, diagnosis: packet["evidence"]["score_activity"].update({"recent": True})),
        ("feedback conflict", lambda packet, diagnosis: packet.update({"feedback_rubric_state": "inconsistent", "states": {**packet["states"], "feedback_rubric_health": "inconsistent"}})),
    ],
)
def test_bounded_diagnostic_policy_keeps_each_hard_blocker_out_of_optimizer_dispatch(description, change):
    from plexus.optimization.portfolio_run import _execution_selection

    assessment = _bounded_diagnostic_assessment("blocked")
    diagnosis = _launchable_diagnosis("blocked")
    change(assessment, diagnosis)

    selected, decisions = _execution_selection(
        "automatic",
        "promotion_ready_plus_bounded_diagnostic",
        [assessment], [diagnosis],
        ranked_rows=[{"scorecard_id": "card", "score_id": "blocked", "evidence_rank": 1}],
        max_samples=50,
    )

    assert selected == [], description
    assert decisions["rejected_targets"][0]["reason"].startswith("bounded_diagnostic_")


def test_bounded_diagnostic_policy_requires_feedback_for_the_frozen_max_samples():
    from plexus.optimization.portfolio_run import _execution_selection

    assessment = _bounded_diagnostic_assessment("too-small", valid_feedback_count=49)
    selected, decisions = _execution_selection(
        "automatic",
        "promotion_ready_plus_bounded_diagnostic",
        [assessment], [_launchable_diagnosis("too-small")],
        ranked_rows=[{"scorecard_id": "card", "score_id": "too-small", "evidence_rank": 1}],
        max_samples=50,
    )

    assert selected == []
    assert decisions["rejected_targets"][0]["reason"] == "bounded_diagnostic_insufficient_samples"


def test_default_execution_candidate_policy_remains_promotion_ready_only():
    from plexus.optimization.portfolio_run import _execution_selection

    assessment = _bounded_diagnostic_assessment("diagnostic")
    selected, decisions = _execution_selection(
        "automatic", "promotion_ready", [assessment], [_launchable_diagnosis("diagnostic")],
        ranked_rows=[{"scorecard_id": "card", "score_id": "diagnostic", "evidence_rank": 1}],
        max_samples=50,
    )

    assert selected == []
    assert decisions["rejected_targets"][0]["reason"] == "not_ready"


def test_automatic_mode_keeps_validator_freshness_rejection_visible_and_non_launching():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": "stale"},
        ]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"], summary=lambda _request: {},
        dispatch=lambda _request: {"accepted": False, "accepted_targets": [], "rejected": [{
            "target": {"scorecard_id": "card", "score_id": "stale"},
            "reason": "stale_assessment",
        }]},
        review=lambda _request: (_ for _ in ()).throw(AssertionError("must not review")),
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("must not ask")),
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "automatic-stale",
        "execution_mode": "automatic", "max_semantic_cost_usd": "1",
        "limits": {"max_cost_usd": 1.0, "max_samples": 1,
                   "max_iterations": 1, "max_concurrency": 1},
    })

    assert result["approval_requests"] == []
    assert result["dispatch"]["children"] == []
    assert result["execution_decisions"]["launched_count"] == 0
    assert result["execution_decisions"]["rejected_targets"] == [{
        "scorecard_id": "card", "score_id": "stale",
        "assessment_fingerprint": "fingerprint-stale", "reason": "stale_assessment",
        "authorization_source": "deterministic_policy",
    }]


def test_tactus_suspension_happens_only_after_pending_approval_is_durable_in_the_report():
    import pytest
    from tactus.core.exceptions import ProcedureWaitingForHuman
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{
                "scorecard_id": "card",
                "score_id": "score",
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
            }],
        },
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("must not dispatch")),
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(
            ProcedureWaitingForHuman("procedure-1", "message-1")
        ),
    ))

    with pytest.raises(ProcedureWaitingForHuman):
        runner.run({
            "account_id": "account-1",
            "run_key": "suspend",
            "max_semantic_cost_usd": "1",
        })

    assert report.milestones[-1][0] == "approval"
    approval_evidence = report.milestones[-1][1]
    assert approval_evidence["approval_requests"][0]["action_key"] == "optimization-approval:suspend:1"
    assert report.terminal == []


def test_saved_approval_cannot_authorize_recomputed_evidence():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    dispatches: list[dict[str, Any]] = []
    fingerprint = {"value": "fingerprint-before-review"}
    human_response = {"value": {"decisions": []}}

    def assess(request):
        packet = _assessment(request["scorecard_id"], request["score_id"])
        packet["evidence_fingerprint"] = fingerprint["value"]
        packet["fingerprint"] = fingerprint["value"]
        return packet

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{"scorecard_id": "card", "score_id": "score"}],
        },
        assess=assess,
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {},
        dispatch=lambda request: dispatches.append(request) or {},
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: human_response["value"],
    ))

    prepared = runner.run({
        "account_id": "account-1",
        "run_key": "frozen-approval",
        "wait_for_human": True,
        "max_semantic_cost_usd": "1",
    })
    original_request = prepared["approval_requests"][0]
    fingerprint["value"] = "fingerprint-after-review"
    human_response["value"] = {
        "decisions": [
            {"scorecard_id": "card", "score_id": "score", "decision": "approve"}
        ]
    }

    resumed = runner.run({
        "account_id": "account-1",
        "run_key": "frozen-approval",
        "wait_for_human": True,
        "max_semantic_cost_usd": "1",
        "approval_responses": {
            original_request["action_key"]: {
                "request": original_request,
                "response": {
                    "decisions": [
                        {"scorecard_id": "card", "score_id": "score", "decision": "approve"}
                    ]
                },
            }
        },
    })

    assert resumed["status"] == "WAITING_FOR_APPROVAL"
    assert dispatches == []
    assert resumed["approval_requests"][0]["preconditions"]["targets"][0][
        "assessment_fingerprint"
    ] == "fingerprint-after-review"


def test_every_five_target_approval_batch_is_resolved_before_any_dispatch():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    dispatches: list[dict[str, Any]] = []
    ranked = [
        {"scorecard_id": "card", "score_id": f"score-{index}"}
        for index in range(6)
    ]
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": ranked},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: request["assessment"],
        summary=lambda _request: {},
        dispatch=lambda request: dispatches.append(request) or {
            "accepted": True,
            "dispatches": [],
        },
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: {"decisions": []},
    ))

    prepared = runner.run({
        "account_id": "account-1",
        "run_key": "two-batches",
        "wait_for_human": True,
        "max_semantic_cost_usd": "1",
    })
    first, second = prepared["approval_requests"]
    first_response = {
        "decisions": [
            {
                "scorecard_id": row["scorecard_id"],
                "score_id": row["score_id"],
                "decision": "approve" if index == 0 else "reject",
            }
            for index, row in enumerate(first["targets"])
        ]
    }
    approvals = {
        first["action_key"]: {"request": first, "response": first_response},
    }

    after_first = runner.run({
        "account_id": "account-1",
        "run_key": "two-batches",
        "wait_for_human": True,
        "max_semantic_cost_usd": "1",
        "approval_responses": approvals,
    })

    assert after_first["status"] == "WAITING_FOR_APPROVAL"
    assert [row["action_key"] for row in after_first["approval_requests"]] == [
        second["action_key"]
    ]
    assert dispatches == []

    approvals[second["action_key"]] = {
        "request": second,
        "response": {
            "decisions": [
                {
                    "scorecard_id": row["scorecard_id"],
                    "score_id": row["score_id"],
                    "decision": "reject",
                }
                for row in second["targets"]
            ]
        },
    }
    completed = runner.run({
        "account_id": "account-1",
        "run_key": "two-batches",
        "wait_for_human": True,
        "max_semantic_cost_usd": "1",
        "approval_responses": approvals,
    })

    assert completed["status"] == "INCOMPLETE"
    assert dispatches == []
    assert completed["dispatch"]["rejected"] == [{
        "reason": "invalid_run_limits",
        "invalid_fields": [
            "max_cost_usd",
            "max_samples",
            "max_iterations",
            "max_concurrency",
        ],
    }]
    milestones = [milestone for milestone, _evidence, _view in report.milestones]
    assert "optimization" not in milestones
    assert "optimization_review" not in milestones


def test_report_publication_failure_stops_the_run_before_any_optimizer_dispatch():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.run_report import OptimizationRunPublicationError

    class _FailingReport(_ReportService):
        def publish_milestone(self, milestone, evidence, *, stakeholder_view):
            raise OptimizationRunPublicationError(f"cannot publish {milestone}")

    report = _FailingReport()
    dispatched = []
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": []},
        assess=lambda _request: None,
        diagnose=lambda _request: None,
        summary=lambda _request: {},
        dispatch=lambda request: dispatched.append(request),
        review=lambda _request: None,
        report=report,
        human_review=lambda _request: None,
    ))

    result = runner.run({"account_id": "account-1", "run_key": "publication-failure", "limits": {"max_cost_usd": 1.0, "max_samples": 1, "max_iterations": 1, "max_concurrency": 1}})

    assert result["status"] == "FAILED"
    assert dispatched == []
    assert report.failures


def test_nonblocking_findings_are_persisted_as_existing_chat_messages_and_report_their_authoritative_state():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    report = _ReportService()
    action_inputs: list[dict[str, Any]] = []
    assessment = _assessment("card", "score")
    assessment["scorecard_name"] = "Example Portfolio"
    assessment["score_name"] = "Priority Score"
    diagnosis = {
        **assessment,
        "states": {"optimization": "stakeholder_clarification_required"},
        "stakeholder_questions": ["Which policy should apply?"],
    }
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {
            "coverage": {"complete": True},
            "ranked": [{"scorecard_id": "card", "score_id": "score"}],
        },
        assess=lambda _request: assessment,
        diagnose=lambda _request: diagnosis,
        summary=lambda _request: {"coverage": {"complete": True}},
        dispatch=lambda _request: (_ for _ in ()).throw(AssertionError("must not dispatch")),
        review=lambda _request: {},
        report=report,
        human_review=lambda _request: (_ for _ in ()).throw(AssertionError("must not request launch approval")),
        create_action=lambda action: action_inputs.append(action) or {
            "action": {"id": "chat-message-1", "responseStatus": "COMPLETED", "responseOwner": "response-1"},
            "created": False,
            "resolution": {"response_message_id": "response-1", "response": {"response": "Use the documented policy."}},
        },
    ))

    result = runner.run({
        "account_id": "account-1",
        "run_key": "questions",
        "max_semantic_cost_usd": "1",
    })

    assert action_inputs[0]["action_key"].startswith("stakeholder_clarification:questions:")
    assert action_inputs[0]["title"] == "Clarify policy for Priority Score"
    assert action_inputs[0]["message"] == "Which policy should apply?"
    assert action_inputs[0]["payload"]["scorecard_name"] == "Example Portfolio"
    assert action_inputs[0]["payload"]["score_name"] == "Priority Score"
    assert action_inputs[0]["resource_refs"][1]["label"] == "Example Portfolio — Priority Score"
    assert result["actions"][0]["message_id"] == "chat-message-1"
    assert result["actions"][0]["response_status"] == "COMPLETED"
    assert result["actions"][0]["response_message_id"] == "response-1"


def test_stakeholder_projection_preserves_available_counts_states_trends_and_actions_without_opaque_ids():
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {
            "coverage": {"complete": True},
            "window": {"start": "2026-04-01T00:00:00Z", "end": "2026-07-01T00:00:00Z"},
            "ranked": [{
                "scorecard_id": "opaque-card",
                "score_id": "opaque-score",
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "valid_feedback_count": 240,
                "reviewed_disagreements": 48,
                "disagreement_rate": 0.2,
                "reviewed_error_opportunity": 48,
                "dashboard_url": "https://dashboard.example/scores/summary",
            }],
        },
        "assessments": [{
            "scope": {"scorecard_id": "opaque-card", "score_id": "opaque-score"},
            "coverage": {"complete": True},
            "states": {
                "optimization": "ready_to_optimize",
                "feedback_collection": "continue_broad_collection",
                "guideline_health": "consistent",
                "feedback_rubric_health": "consistent",
                "promotion_readiness": "not_evaluated",
            },
            "weekly_stability": {
                "weekly_disagreement_range": 0.03,
                "weekly_ac1_range": 0.04,
                "weekly_bucket_counts": [20, 24, 22, 26],
            },
            "rationale": "Reviewed errors show a safe improvement opportunity.",
            "primary_next_action": "request_optimization_approval",
        }],
        "diagnoses": [],
        "reviews": [],
        "dispatch": None,
    }, milestone="assessment")

    row = view["portfolio"][0]
    assert row["coverage_status"] == "complete"
    assert row["collection_state"] == "continue_broad_collection"
    assert row["guideline_state"] == "consistent"
    assert row["feedback_rubric_state"] == "consistent"
    assert row["promotion_readiness"] == "not_evaluated"
    assert row["rationale"] == "Reviewed errors show a safe improvement opportunity."
    assert row["next_action"] == "request_optimization_approval"
    assert "20, 24, 22, 26" in row["trend"]
    assert row["dashboard_url"].startswith("https://")
    assert "opaque-card" not in str(view)
    assert "opaque-score" not in str(view)
    assert view["priorities"][0]["evidence_count"] == 240
    assert view["feedback_investment"][0]["coverage_status"] == "complete"


def test_stakeholder_projection_uses_stable_distinct_score_refs_without_exporting_opaque_score_ids():
    """Duplicate display names must stay distinguishable without leaking IDs."""
    from datetime import datetime, timezone
    from io import BytesIO

    from openpyxl import load_workbook

    from plexus.optimization.portfolio_run import _stakeholder_view
    from plexus.optimization.run_report import build_stakeholder_workbook

    opaque_first = "opaque-score-id-first"
    opaque_second = "opaque-score-id-second"
    state = {
        "rank": {
            "coverage": {"complete": True},
            "ranked": [
                {
                    "scorecard_id": "opaque-card-id",
                    "score_id": opaque_first,
                    "scorecard_name": "Example Portfolio",
                    "score_name": "Repeated Score Name",
                    "valid_feedback_count": 20,
                    "reviewed_disagreements": 10,
                    "disagreement_rate": 0.5,
                    "reviewed_error_opportunity": 10,
                    "evidence_rank": 1,
                },
                {
                    "scorecard_id": "opaque-card-id",
                    "score_id": opaque_second,
                    "scorecard_name": "Example Portfolio",
                    "score_name": "Repeated Score Name",
                    "valid_feedback_count": 10,
                    "reviewed_disagreements": 5,
                    "disagreement_rate": 0.5,
                    "reviewed_error_opportunity": 5,
                    "evidence_rank": 2,
                },
            ],
        },
        "assessments": [
            {
                "scope": {"scorecard_id": "opaque-card-id", "score_id": opaque_first},
                "coverage": {"complete": True},
                "states": {"optimization": "repair_required", "guideline_health": "missing"},
            },
            {
                "scope": {"scorecard_id": "opaque-card-id", "score_id": opaque_second},
                "coverage": {"complete": True},
                "states": {"optimization": "ready_to_optimize", "feedback_collection": "continue_broad_collection"},
            },
        ],
    }

    view = _stakeholder_view(state, milestone="diagnosis")
    expected_refs = {
        sha256(opaque_first.encode("utf-8")).hexdigest()[:16],
        sha256(opaque_second.encode("utf-8")).hexdigest()[:16],
    }
    for section in ("portfolio", "priorities", "feedback_investment", "questions_and_issues", "optimization_outcomes"):
        rows = view[section]
        assert rows
        assert {row["score_ref"] for row in rows}.issubset(expected_refs)
    assert {row["score_ref"] for row in view["portfolio"]} == expected_refs
    assert {row["score_ref"] for row in view["priorities"]} == expected_refs
    assert {row["score_ref"] for row in view["optimization_outcomes"]} == expected_refs
    assert view["questions_and_issues"][0]["score_ref"] == sha256(opaque_first.encode("utf-8")).hexdigest()[:16]
    assert next(
        row["score_ref"]
        for row in view["feedback_investment"]
        if row["state"] == "continue_broad_collection"
    ) == sha256(opaque_second.encode("utf-8")).hexdigest()[:16]
    stakeholder_values = str(view)
    assert opaque_first not in stakeholder_values
    assert opaque_second not in stakeholder_values

    artifact = build_stakeholder_workbook(
        view, revision_number=1, generated_at=datetime(2026, 7, 31, tzinfo=timezone.utc),
    )
    workbook = load_workbook(BytesIO(artifact.content), data_only=False)
    exported_values = str([
        cell.value
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    ])
    assert opaque_first not in exported_values
    assert opaque_second not in exported_values


def test_stakeholder_overview_explains_current_work_and_next_durable_checkpoint():
    from plexus.optimization.portfolio_run import _stakeholder_view

    started = _stakeholder_view({
        "rank": None,
        "assessments": [],
        "diagnoses": [],
        "reviews": [],
        "approval_requests": [],
        "diagnosis_coverage": {"selected_count": 0, "completed_count": 0, "failed_count": 0},
    }, milestone="started")
    assert started["overview"]["coverage_status"] == "pending"
    assert started["overview"]["inventory_coverage_status"] == "pending"
    assert started["overview"]["analysis_coverage_status"] == "pending"
    assert "Enumerating every scorecard" in started["overview"]["current_activity"]
    assert "ranked portfolio" in started["overview"]["next_checkpoint"]

    ranked = _stakeholder_view({
        "rank": {
            "coverage": {
                "complete": True,
                "scope": {
                    "total_scorecards_inspected": 12,
                    "matched_scorecard_count": 1,
                },
                "activity": {"recent_activity_excluded_count": 3},
            },
            "ranked": [{"scorecard_id": "card", "score_id": "score"}],
            "unranked": [{}, {}],
            "window": {"start": "a", "end": "b"},
        },
        "assessments": [],
        "diagnoses": [],
        "reviews": [],
        "approval_requests": [],
        "diagnosis_coverage": {"selected_count": 1, "completed_count": 0, "failed_count": 0},
    }, milestone="ranking")
    assert ranked["overview"]["scorecards_inspected"] == 12
    assert ranked["overview"]["scorecards_in_scope"] == 1
    assert ranked["overview"]["ranked_score_count"] == 1
    assert ranked["overview"]["unranked_score_count"] == 2
    assert ranked["overview"]["cooldown_excluded_count"] == 3
    assert ranked["overview"]["assessment_progress"] == "0 of 1 eligible candidates assessed"
    assert "deterministic readiness" in ranked["overview"]["current_activity"]


def test_stakeholder_overview_does_not_imply_optimizer_work_when_no_targets_were_approved():
    from plexus.optimization.portfolio_run import _stakeholder_view

    state = {
        "rank": {"coverage": {"complete": True}, "ranked": []},
        "assessments": [],
        "diagnoses": [],
        "reviews": [],
        "approved_targets": [],
        "dispatch": {"batches": [], "rejected": []},
        "approval_requests": [],
        "diagnosis_coverage": {
            "selected_count": 0,
            "scheduled_count": 0,
            "completed_count": 0,
            "failed_count": 0,
            "deferred_by_cap_count": 0,
        },
    }

    optimization = _stakeholder_view(state, milestone="optimization")["overview"]
    assert "no optimizations were launched" in optimization["current_activity"].lower()
    assert "evaluations will be reviewed" not in optimization["next_checkpoint"].lower()

    review = _stakeholder_view(state, milestone="optimization_review")["overview"]
    assert "no optimization results" in review["current_activity"].lower()
    assert "reviewing completed optimizer" not in review["current_activity"].lower()


def test_automatic_stakeholder_narrative_describes_policy_selection_without_human_approval():
    from plexus.optimization.portfolio_run import _stakeholder_view

    state = {
        "execution_mode": "automatic",
        "run_spec": {"execution_mode": "automatic"},
        "rank": {"coverage": {"complete": True}, "ranked": []},
        "assessments": [],
        "diagnoses": [],
        "reviews": [],
        "approved_targets": [],
        "dispatch": {"batches": [], "rejected": []},
        "approval_requests": [],
        "diagnosis_coverage": {
            "selected_count": 0,
            "scheduled_count": 0,
            "completed_count": 0,
            "failed_count": 0,
            "deferred_by_cap_count": 0,
        },
    }

    diagnosis = _stakeholder_view(state, milestone="diagnosis")["overview"]
    assert "automatic execution policy" in diagnosis["current_activity"].lower()
    assert "human decisions" not in diagnosis["current_activity"].lower()
    assert "explicit decision" not in diagnosis["next_checkpoint"].lower()

    optimization = _stakeholder_view(state, milestone="optimization")["overview"]
    assert "passed the automatic execution policy" in optimization["current_activity"].lower()
    assert "approved" not in optimization["current_activity"].lower()


def test_approval_required_stakeholder_narrative_retains_explicit_human_decisions():
    from plexus.optimization.portfolio_run import _stakeholder_view

    state = {
        "execution_mode": "approval_required",
        "run_spec": {"execution_mode": "approval_required"},
        "rank": {"coverage": {"complete": True}, "ranked": []},
        "assessments": [],
        "diagnoses": [],
        "reviews": [],
        "approved_targets": [],
        "dispatch": {"batches": [], "rejected": []},
        "approval_requests": [],
        "diagnosis_coverage": {
            "selected_count": 0,
            "scheduled_count": 0,
            "completed_count": 0,
            "failed_count": 0,
            "deferred_by_cap_count": 0,
        },
    }

    diagnosis = _stakeholder_view(state, milestone="diagnosis")["overview"]
    assert "human decisions" in diagnosis["current_activity"].lower()
    assert "explicit decision" in diagnosis["next_checkpoint"].lower()

    optimization = _stakeholder_view(state, milestone="optimization")["overview"]
    assert "no targets were approved" in optimization["current_activity"].lower()


def test_stakeholder_overview_retains_active_narration_for_dispatched_optimizer_work():
    from plexus.optimization.portfolio_run import _stakeholder_view

    state = {
        "rank": {"coverage": {"complete": True}, "ranked": []},
        "assessments": [],
        "diagnoses": [],
        "reviews": [{"procedure_id": "procedure-1"}],
        "approved_targets": [{"scorecard_id": "card", "score_id": "score"}],
        "dispatch": {
            "batches": [{
                "dispatches": [{"status": "dispatched", "procedure_id": "procedure-1"}],
            }],
            "rejected": [],
        },
        "approval_requests": [],
        "diagnosis_coverage": {
            "selected_count": 0,
            "scheduled_count": 0,
            "completed_count": 0,
            "failed_count": 0,
            "deferred_by_cap_count": 0,
        },
    }

    optimization = _stakeholder_view(state, milestone="optimization")["overview"]
    assert "approved optimization" in optimization["current_activity"].lower()
    review = _stakeholder_view(state, milestone="optimization_review")["overview"]
    assert "optimizer and evaluation evidence" in review["current_activity"].lower()


def test_stakeholder_view_reports_waiting_optimizer_child_as_active_work():
    from plexus.optimization.portfolio_run import _stakeholder_view

    state = {
        "rank": {
            "coverage": {"complete": True},
            "ranked": [{
                "scorecard_id": "card",
                "score_id": "score",
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "valid_feedback_count": 240,
            }, {
                "scorecard_id": "card",
                "score_id": "other-score",
                "scorecard_name": "Example Portfolio",
                "score_name": "Other Score",
                "valid_feedback_count": 200,
            }],
        },
        "assessments": [],
        "diagnoses": [],
        "reviews": [],
        "approved_targets": [{"scorecard_id": "card", "score_id": "score"}],
        "dispatch": {
            "phase": "waiting_for_children",
            "batches": [],
            "rejected": [],
            "children": [{
                "target": {"scorecard_id": "card", "score_id": "score"},
                "procedure_id": "procedure-1",
                "task_id": "task-1",
                "launch_state": {"phase": "waiting"},
            }],
        },
        "approval_requests": [],
        "diagnosis_coverage": {
            "selected_count": 0,
            "scheduled_count": 0,
            "completed_count": 0,
            "failed_count": 0,
            "deferred_by_cap_count": 0,
        },
    }

    view = _stakeholder_view(state, milestone="optimization")

    assert view["overview"]["dispatched_optimizer_count"] == 1
    assert "approved optimization" in view["overview"]["current_activity"].lower()
    assert view["optimization_outcomes"][0]["outcome"] == "optimization_in_progress"
    assert view["optimization_outcomes"][0]["next_action"] == "wait_for_optimizer_completion"
    assert view["optimization_outcomes"][1]["outcome"] == "not_run"
    assert "procedure-1" not in str(view)
    assert "task-1" not in str(view)


def test_stakeholder_view_surfaces_invalid_run_limits_for_approved_unlaunched_target():
    from plexus.optimization.portfolio_run import _stakeholder_view

    assessment = _assessment("card", "score")
    rejection = {
        "reason": "invalid_run_limits",
        "invalid_fields": ["max_cost_usd"],
    }
    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card",
            "score_id": "score",
            "scorecard_name": "Example Portfolio",
            "score_name": "Priority Score",
            "valid_feedback_count": 240,
        }]},
        "assessments": [assessment],
        "diagnoses": [assessment],
        "reviews": [],
        "approved_targets": [{"scorecard_id": "card", "score_id": "score"}],
        "dispatch": {
            "phase": "incomplete",
            "batches": [{"rejected": [rejection]}],
            "rejected": [rejection],
            "children": [],
        },
        "terminal_status": "INCOMPLETE",
    }, milestone="finalization")

    portfolio = view["portfolio"][0]
    outcome = view["optimization_outcomes"][0]
    assert portfolio["primary_disposition"] == "failed_or_incomplete"
    assert portfolio["next_action"] == "provide_valid_run_limits"
    assert portfolio["dispatch_rejection"] == rejection
    assert outcome["outcome"] == "failed_or_incomplete"
    assert outcome["next_action"] == "provide_valid_run_limits"
    assert outcome["dispatch_rejection"] == rejection
    assert view["overview"]["invalid_run_limit_target_count"] == 1
    assert "run limits" in view["overview"]["current_activity"].lower()
    assert "valid run limits" in view["overview"]["next_checkpoint"].lower()


@pytest.mark.parametrize("phase", [
    "planned",
    "procedure_create_attempted",
    "procedure_record_observed",
    "procedure_provisioned",
    "task_create_attempted",
    "task_record_observed",
    "task_stage_reconcile_attempted",
    "task_held",
    "release_attempted",
])
def test_stakeholder_view_treats_every_durable_pre_observation_child_phase_as_launching(phase):
    """A persisted launch boundary is active work, never an unselected score."""
    from plexus.optimization.portfolio_run import _stakeholder_view

    state = {
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
            "valid_feedback_count": 240,
        }]},
        "assessments": [],
        "diagnoses": [],
        "reviews": [],
        "dispatch": {"children": [{
            "target": {"scorecard_id": "card", "score_id": "score"},
            "launch_state": {"phase": phase},
        }]},
    }

    view = _stakeholder_view(state, milestone="optimization")

    assert view["portfolio"][0]["primary_disposition"] == "optimizer_launching"
    assert view["optimization_outcomes"][0]["outcome"] == "optimizer_launching"


def test_stakeholder_view_projects_outcomes_for_the_entire_evidence_row_universe():
    """Policy-deferred evidence stays visible as a not-run optimization outcome."""
    from plexus.optimization.portfolio_run import _stakeholder_view

    state = {
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "eligible",
            "scorecard_name": "Example Portfolio", "score_name": "Eligible Score",
            "valid_feedback_count": 100, "evidence_rank": 2,
        }], "unranked": [{
            "scorecard_id": "card", "score_id": "cooldown",
            "scorecard_name": "Example Portfolio", "score_name": "Cooldown Score",
            "valid_feedback_count": 150, "evidence_rank": 1,
            "policy_disposition": "cooldown", "policy_reason": "recent_score_activity",
            "eligible_for_optimization": False,
        }, {
            "scorecard_id": "card", "score_id": "policy",
            "scorecard_name": "Example Portfolio", "score_name": "Policy Deferred Score",
            "valid_feedback_count": 50, "evidence_rank": 3,
            "policy_disposition": "blocked", "policy_reason": "disabled",
            "eligible_for_optimization": False,
        }]},
        "assessments": [], "diagnoses": [], "reviews": [],
    }

    view = _stakeholder_view(state, milestone="optimization")

    assert [row["score_name"] for row in view["optimization_outcomes"]] == [
        row["score_name"] for row in view["portfolio"]
    ] == ["Cooldown Score", "Eligible Score", "Policy Deferred Score"]
    outcomes = {row["score_name"]: row for row in view["optimization_outcomes"]}
    assert outcomes["Cooldown Score"]["outcome"] == "not_run"
    assert outcomes["Policy Deferred Score"]["outcome"] == "not_run"


def test_stakeholder_overview_separates_complete_inventory_from_incomplete_diagnosis_results():
    from plexus.optimization.portfolio_run import _stakeholder_view
    from plexus.optimization.run_report import _validate_view

    base_state = {
        "rank": {
            "coverage": {"complete": True},
            "ranked": [{
                "scorecard_id": "card",
                "score_id": "score",
                "scorecard_name": "Example Portfolio",
                "score_name": "Priority Score",
                "valid_feedback_count": 240,
                "reviewed_disagreements": 48,
                "disagreement_rate": 0.2,
                "reviewed_error_opportunity": 48,
            }],
        },
        "assessments": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {"complete": True},
            "states": {"optimization": "ready_to_optimize"},
        }],
        "reviews": [],
        "approval_requests": [],
        "diagnosis_coverage": {
            "selected_count": 1,
            "scheduled_count": 1,
            "completed_count": 1,
            "failed_count": 0,
            "deferred_by_cap_count": 0,
        },
    }
    incomplete = _stakeholder_view({
        **base_state,
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "states": {"optimization": "incomplete", "readiness": "incomplete"},
            "outcome": "incomplete",
            "failures": ["Required evidence could not be read."],
        }],
        "terminal_status": "INCOMPLETE",
    }, milestone="finalization")

    overview = incomplete["overview"]
    assert overview["coverage_status"] == "complete"
    assert overview["inventory_coverage_status"] == "complete"
    assert overview["analysis_coverage_status"] == "incomplete"
    assert overview["diagnosis_incomplete_count"] == 1
    assert overview["diagnosis_coverage"] == (
        "1 of 1 scheduled diagnoses returned; 1 incomplete result; "
        "0 execution failures; 0 deferred by the configured diagnosis limit"
    )
    assert overview["analysis_incomplete_reason"] == "incomplete_diagnosis_evidence"
    _validate_view(incomplete)

    complete = _stakeholder_view({
        **base_state,
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {"complete": True},
            "states": {"optimization": "ready_to_optimize"},
            "outcome": "ready_to_optimize",
        }],
    }, milestone="diagnosis")
    assert complete["overview"]["inventory_coverage_status"] == "complete"
    assert complete["overview"]["analysis_coverage_status"] == "complete"
    assert complete["overview"]["diagnosis_incomplete_count"] == 0


def test_stakeholder_view_and_workbook_publish_reconciled_semantic_budget_evidence():
    from datetime import datetime, timezone
    from io import BytesIO

    from openpyxl import load_workbook

    from plexus.optimization.portfolio_run import _stakeholder_view
    from plexus.optimization.run_report import build_stakeholder_workbook

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "private-card", "score_id": "private-score",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
            "valid_feedback_count": 240, "reviewed_disagreements": 48,
            "disagreement_rate": 0.2, "reviewed_error_opportunity": 48,
        }]},
        "assessments": [{
            "scope": {"scorecard_id": "private-card", "score_id": "private-score"},
            "coverage": {"complete": True},
            "states": {"optimization": "ready_to_optimize"},
        }],
        "diagnoses": [{
            "scope": {"scorecard_id": "private-card", "score_id": "private-score"},
            "coverage": {"complete": True},
            "states": {"optimization": "ready_to_optimize"},
        }],
        "diagnosis_coverage": {"selected_count": 1, "scheduled_count": 1, "completed_count": 1},
        "semantic_budget_evidence": {
            "policy_version": "semantic-budget-policy-v1",
            "budget_spec_schema_version": "semantic-budget-v1",
            "ledger_schema_version": "semantic-budget-ledger-v1",
            "provider": "openai", "model": "gpt-5-mini-2025-08-07",
            "pricing_version": "openai-2025-08-07-v1",
            "authorized_max_usd": "1", "settled_actual_usd": "0.000045",
            "held_reserved_usd": "0.0009", "available_usd": "0.999055",
            "reservation_count": 4, "reserved_count": 1, "settled_count": 1,
            "unknown_count": 1, "cancelled_count": 1, "target_count": 3,
            "call_site_coverage": [{"call_site": "rubric_consistency", "count": 3}],
            "ledger_revision": 7, "evidence_reference": "semantic-budget-ledger:r000007",
            "evidence_digest": "a" * 64,
        },
    }, milestone="diagnosis")

    assert view["overview"]["semantic_budget_authorized_usd"] == "1"
    assert view["overview"]["semantic_budget_available_usd"] == "0.999055"
    assert view["overview"]["semantic_budget_spec_schema_version"] == "semantic-budget-v1"
    assert view["overview"]["semantic_budget_ledger_schema_version"] == "semantic-budget-ledger-v1"
    assert view["overview"]["semantic_budget_model"] == "gpt-5-mini-2025-08-07"
    assert view["overview"]["semantic_budget_reserved_count"] == 1
    assert view["overview"]["semantic_budget_unknown_count"] == 1
    assert view["portfolio"][0]["semantic_diagnosis_status"] == "complete"
    assert view["portfolio"][0]["semantic_budget_evidence_reference"] == "semantic-budget-ledger:r000007"

    artifact = build_stakeholder_workbook(
        view, revision_number=7, generated_at=datetime(2026, 7, 30, tzinfo=timezone.utc),
    )
    workbook = load_workbook(BytesIO(artifact.content), data_only=False)
    overview_values = [cell.value for row in workbook["Overview"].iter_rows() for cell in row]
    portfolio_headers = [cell.value for cell in workbook["Portfolio"][1]]
    portfolio_values = [cell.value for row in workbook["Portfolio"].iter_rows() for cell in row]
    priorities_headers = [cell.value for cell in workbook["Priorities"][1]]
    issues_headers = [cell.value for cell in workbook["Questions and Issues"][1]]
    run_log_headers = [cell.value for cell in workbook["Run Log"][1]]
    run_log_values = [cell.value for cell in workbook["Run Log"][2]]
    definition_terms = [row[0].value for row in workbook["Definitions"].iter_rows(min_row=2)]
    assert "Semantic Budget Authorized Usd" in overview_values
    assert "Semantic Diagnosis Status" in portfolio_headers
    assert "Semantic Budget Evidence" in priorities_headers
    assert "Semantic Diagnosis Status" in issues_headers
    assert "Semantic Evidence" in run_log_headers
    assert "Semantic Calls" in run_log_headers
    assert "semantic-budget-ledger:r000007" in portfolio_values
    assert "semantic-budget-ledger:r000007" in run_log_values
    assert any(
        "total 4; reserved 1; settled 1; unknown 1; cancelled 1" in str(value)
        for value in run_log_values
    )
    assert "Semantic budget" in definition_terms
    assert "private-card" not in str(overview_values + portfolio_values)
    assert not any(cell.data_type == "f" for sheet in workbook for row in sheet.iter_rows() for cell in row)


def test_per_score_semantic_status_uses_actual_selection_and_result_coverage():
    from plexus.optimization.portfolio_run import _stakeholder_view

    def row(index):
        return {
            "scorecard_id": "card", "score_id": f"score-{index}",
            "scorecard_name": "Example Portfolio", "score_name": f"Score {index}",
            "valid_feedback_count": 100, "reviewed_disagreements": 10,
            "disagreement_rate": 0.1, "reviewed_error_opportunity": 10,
        }

    ranked = [row(index) for index in range(11)]
    assessments = [{
        "scope": {"scorecard_id": "card", "score_id": f"score-{index}"},
        "coverage": {"complete": True},
        "states": {"optimization": "ready_to_optimize"},
    } for index in range(11)]
    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": ranked},
        "assessments": assessments,
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score-0"},
            "coverage": {"complete": True},
            "states": {"optimization": "ready_to_optimize"},
        }],
        "diagnosis_coverage": {
            "selected_count": 10, "scheduled_count": 10, "completed_count": 1,
            "max_semantic_diagnoses": 10,
        },
    }, milestone="diagnosis")

    statuses = {row["score_name"]: row["semantic_diagnosis_status"] for row in view["portfolio"]}
    assert statuses["Score 0"] == "complete"
    assert statuses["Score 1"] == "incomplete"
    assert statuses["Score 10"] == "deferred"


def test_outcome_unknown_overrides_generic_rationale_with_safe_budget_ambiguity():
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Unknown Score",
            "valid_feedback_count": 100, "reviewed_disagreements": 10,
            "disagreement_rate": 0.1, "reviewed_error_opportunity": 10,
        }]},
        "assessments": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {"complete": True}, "states": {"optimization": "ready_to_optimize"},
            "rationale": "=FORMULA_SENTINEL",
        }],
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {"complete": False}, "states": {"optimization": "incomplete"},
            "semantic_failure_category": "outcome_unknown",
            "status": "outcome_unknown", "failures": ["contradictory budget exhausted prose"],
            "rationale": "=FORMULA_SENTINEL",
        }],
        "diagnosis_coverage": {"selected_count": 1, "scheduled_count": 1, "failed_count": 1},
    }, milestone="finalization")

    row = view["portfolio"][0]
    assert row["semantic_diagnosis_status"] == "outcome_unknown"
    assert row["next_action"] == "review_semantic_budget"
    assert "provider outcome is unknown" in row["rationale"].lower()
    assert "FORMULA_SENTINEL" not in row["rationale"]


def test_legacy_nested_unknown_failure_text_fails_closed_without_typed_category():
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Unknown Score",
            "valid_feedback_count": 100, "reviewed_disagreements": 10,
            "disagreement_rate": 0.1, "reviewed_error_opportunity": 10,
        }]},
        "assessments": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {"complete": True}, "states": {"optimization": "ready_to_optimize"},
            "rationale": "+GENERIC_SENTINEL",
        }],
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {"complete": False, "failures": ["provider outcome unknown"]},
            "states": {"optimization": "incomplete"}, "rationale": "+GENERIC_SENTINEL",
        }],
        "diagnosis_coverage": {
            "selected_count": 1, "scheduled_count": 1, "failed_count": 1,
            "outcome_unknown_count": 1, "deferred_by_budget_count": 0,
        },
        "semantic_budget_evidence": {
            "policy_version": "semantic-budget-policy-v1",
            "budget_spec_schema_version": "semantic-budget-v1",
            "ledger_schema_version": "semantic-budget-ledger-v1",
            "provider": "openai", "model": "gpt-5-mini-2025-08-07",
            "pricing_version": "openai-2025-08-07-v1",
            "authorized_max_usd": "1", "settled_actual_usd": "0",
            "held_reserved_usd": "0.00045", "available_usd": "0.99955",
            "reservation_count": 1, "reserved_count": 0, "settled_count": 0,
            "unknown_count": 1, "cancelled_count": 0, "target_count": 1,
            "call_site_coverage": [], "ledger_revision": 2,
            "evidence_reference": "semantic-budget-ledger:r000002",
            "evidence_digest": "b" * 64,
        },
    }, milestone="finalization")

    row = view["portfolio"][0]
    assert row["semantic_diagnosis_status"] == "incomplete"
    assert row["next_action"] == "review"
    assert "GENERIC_SENTINEL" not in row["rationale"]
    assert view["overview"]["semantic_budget_deferred_count"] == 0


def test_provider_outcome_unknown_defers_remaining_work_without_counting_budget_exhaustion():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.semantic_authority import SemanticOutcomeUnknown

    report = _ReportService()
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [
            {"scorecard_id": "card", "score_id": "one"},
            {"scorecard_id": "card", "score_id": "two"},
        ]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda _request: (_ for _ in ()).throw(
            SemanticOutcomeUnknown("provider outcome unknown")
        ),
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "unknown-outcome",
        "max_semantic_diagnoses": 2, "max_semantic_cost_usd": "1",
    })

    assert result["status"] == "INCOMPLETE"
    assert result["diagnosis_coverage"]["outcome_unknown_count"] == 1
    assert result["diagnosis_coverage"]["deferred_after_failure_count"] == 1
    assert result["diagnosis_coverage"]["deferred_by_budget_count"] == 0


@pytest.mark.parametrize(
    ("failure_category", "expected_action", "rationale_fragment"),
    [
        ("budget_exhausted", "review_semantic_budget", "frozen semantic budget"),
        ("outcome_unknown", "review_semantic_budget", "provider outcome is unknown"),
        (
            "authority_publication_failure",
            "repair_semantic_authority_publication",
            "semantic authority evidence could not be published",
        ),
    ],
)
def test_actual_runner_semantic_failure_state_reaches_stakeholder_and_workbook_without_collapse(
    failure_category, expected_action, rationale_fragment,
):
    from datetime import datetime, timezone
    from io import BytesIO

    from openpyxl import load_workbook

    from plexus.optimization.operator_identity import optimization_operator_identity
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.run_report import (
        OptimizationRunReportService,
        build_stakeholder_workbook,
    )
    from plexus.optimization.semantic_authority import (
        SemanticAuthorityPublicationError,
        SemanticOutcomeUnknown,
    )
    from plexus.optimization.semantic_budget import SemanticBudgetExceeded

    exception_type = {
        "budget_exhausted": SemanticBudgetExceeded,
        "outcome_unknown": SemanticOutcomeUnknown,
        "authority_publication_failure": SemanticAuthorityPublicationError,
    }[failure_category]
    report = _ReportService()
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
            "valid_feedback_count": 100, "reviewed_disagreements": 10,
            "disagreement_rate": 0.1, "reviewed_error_opportunity": 10,
        }]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda _request: (_ for _ in ()).throw(
            exception_type(f"safe {failure_category} sentinel")
        ),
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": f"projection-{failure_category}",
        "max_semantic_diagnoses": 1, "max_semantic_cost_usd": "1",
    })

    final_view = report.milestones[-1][2]
    final_evidence = report.milestones[-1][1]
    portfolio_row = final_view["portfolio"][0]
    outcome_row = final_view["optimization_outcomes"][0]
    assert result["diagnoses"][0]["status"] == failure_category
    assert result["diagnoses"][0]["semantic_failure_category"] == failure_category
    assert (
        result["diagnoses"][0]["coverage"]["semantic_failure_category"]
        == failure_category
    )
    assert result["diagnosis_coverage"]["semantic_failure_category"] == failure_category
    assert (
        final_evidence["diagnoses"][0]["semantic_failure_category"]
        == failure_category
    )
    assert (
        final_evidence["diagnosis_coverage"]["semantic_failure_category"]
        == failure_category
    )
    assert portfolio_row["semantic_diagnosis_status"] == failure_category
    assert portfolio_row["next_action"] == expected_action
    assert rationale_fragment in portfolio_row["rationale"].lower()
    assert "sentinel" not in portfolio_row["rationale"].lower()
    assert outcome_row["semantic_diagnosis_status"] == failure_category
    assert outcome_row["next_action"] == expected_action
    assert rationale_fragment in outcome_row["rationale"].lower()
    assert result["summary"]["semantic_budget_next_action"] == expected_action

    artifact = build_stakeholder_workbook(
        final_view, revision_number=1,
        generated_at=datetime(2026, 7, 31, tzinfo=timezone.utc),
    )
    workbook = load_workbook(BytesIO(artifact.content), data_only=False)
    headers = [cell.value for cell in workbook["Portfolio"][1]]
    values = [cell.value for cell in workbook["Portfolio"][2]]
    workbook_row = dict(zip(headers, values))
    assert workbook_row["Semantic Diagnosis Status"] == failure_category
    assert workbook_row["Next Action"] == expected_action
    assert rationale_fragment in workbook_row["Rationale"].lower()

    cover = OptimizationRunReportService._render_report_manifest(
        "incomplete",
        {"number": 1, "milestone": "finalization", **final_view},
        identity=optimization_operator_identity(scope={}),
    )
    assert "Semantic diagnosis issues" in cover
    assert failure_category in cover
    assert expected_action in cover
    assert rationale_fragment in cover.lower()


def test_incomplete_only_semantic_result_has_safe_summary_and_cover_issue():
    from plexus.optimization.operator_identity import optimization_operator_identity
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner
    from plexus.optimization.run_report import OptimizationRunReportService

    report = _ReportService()
    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "opaque-card-id", "score_id": "opaque-score-id",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
        }]},
        assess=lambda request: _assessment(request["scorecard_id"], request["score_id"]),
        diagnose=lambda request: {
            **request["assessment"],
            "coverage": {"complete": False, "failures": ["private raw error"]},
            "states": {"optimization": "incomplete"},
            "rationale": "private raw rationale",
        },
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=report, human_review=lambda _request: {},
    ))

    result = runner.run({
        "account_id": "account-1", "run_key": "generic-incomplete-cover",
        "max_semantic_diagnoses": 1, "max_semantic_cost_usd": "1",
    })
    final_view = report.milestones[-1][2]
    assert result["diagnosis_coverage"]["incomplete_count"] == 1
    assert result["diagnosis_coverage"]["failed_count"] == 0
    assert result["summary"]["semantic_budget_next_action"] == "review"
    assert "without a recognized structured failure category" in (
        result["summary"]["semantic_budget_failure"]
    )

    cover = OptimizationRunReportService._render_report_manifest(
        "incomplete",
        {"number": 1, "milestone": "finalization", **final_view},
        identity=optimization_operator_identity(scope={}),
    )
    assert "Semantic diagnosis issues" in cover
    assert "Priority Score" in cover
    assert "Status: incomplete" in cover
    assert "Next action: review" in cover
    assert "without a recognized structured failure category" in cover
    assert "opaque-card-id" not in cover
    assert "opaque-score-id" not in cover
    assert "private raw" not in cover


@pytest.mark.parametrize(
    ("failure_category", "expected_action"),
    [
        ("budget_exhausted", "review_semantic_budget"),
        ("outcome_unknown", "review_semantic_budget"),
        ("authority_publication_failure", "repair_semantic_authority_publication"),
    ],
)
def test_nested_diagnosis_coverage_preserves_exact_semantic_failure_state(
    failure_category, expected_action,
):
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
            "valid_feedback_count": 100, "reviewed_disagreements": 10,
            "disagreement_rate": 0.1, "reviewed_error_opportunity": 10,
        }]},
        "assessments": [_assessment("card", "score")],
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {
                "complete": False,
                "semantic_failure_category": failure_category,
            },
            "states": {"optimization": "incomplete"},
            "rationale": "generic rationale must not survive",
        }],
        "diagnosis_coverage": {
            "selected_count": 1, "scheduled_count": 1, "failed_count": 1,
        },
    }, milestone="finalization")

    assert view["portfolio"][0]["semantic_diagnosis_status"] == failure_category
    assert view["portfolio"][0]["next_action"] == expected_action
    assert "generic rationale" not in view["portfolio"][0]["rationale"]
    assert view["optimization_outcomes"][0]["semantic_diagnosis_status"] == failure_category
    assert view["optimization_outcomes"][0]["next_action"] == expected_action


@pytest.mark.parametrize("failure_text", [
    "semantic reservation could not be durably published before provider contact",
    "semantic reservation could not be persisted before provider contact",
    "semantic ledger commit failed before provider contact",
    "semantic durable-write failed before provider contact",
])
def test_legacy_failure_prose_never_selects_authority_publication_category(
    failure_text,
):
    from datetime import datetime, timezone
    from io import BytesIO

    from openpyxl import load_workbook

    from plexus.optimization.portfolio_run import (
        OptimizationPortfolioRunner,
        _stakeholder_view,
    )
    from plexus.optimization.run_report import build_stakeholder_workbook

    diagnosis = {
        "scope": {"scorecard_id": "card", "score_id": "score"},
        "coverage": {"complete": False, "failures": [failure_text]},
        "states": {"optimization": "incomplete"},
        "rationale": "generic rationale must not survive",
    }
    state = {
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
            "valid_feedback_count": 100, "reviewed_disagreements": 10,
            "disagreement_rate": 0.1, "reviewed_error_opportunity": 10,
        }]},
        "assessments": [_assessment("card", "score")],
        "diagnoses": [diagnosis],
        "diagnosis_coverage": {
            "selected_count": 1, "scheduled_count": 1, "failed_count": 1,
        },
        "semantic_budget_evidence": {"unknown_count": 0},
    }
    view = _stakeholder_view(state, milestone="finalization")

    row = view["portfolio"][0]
    assert row["semantic_diagnosis_status"] == "incomplete"
    assert row["next_action"] == "review"
    assert "generic rationale" not in row["rationale"]

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {}, assess=lambda _request: {}, diagnose=lambda _request: {},
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=_ReportService(), human_review=lambda _request: {},
    ))
    summary = runner._summary(state)
    assert summary["semantic_budget_next_action"] == "review"
    assert "incomplete" in summary["semantic_budget_failure"].lower()

    artifact = build_stakeholder_workbook(
        view, revision_number=1,
        generated_at=datetime(2026, 7, 31, tzinfo=timezone.utc),
    )
    workbook = load_workbook(BytesIO(artifact.content), data_only=False)
    headers = [cell.value for cell in workbook["Portfolio"][1]]
    values = [cell.value for cell in workbook["Portfolio"][2]]
    workbook_row = dict(zip(headers, values))
    assert workbook_row["Semantic Diagnosis Status"] == "incomplete"
    assert workbook_row["Next Action"] == "review"


@pytest.mark.parametrize("failure_text", [
    "semantic reservation requires operator review",
    "semantic remaining cost requires operator review",
    "semantic reservation requires an exceedingly careful review",
    "semantic cost requires operator review",
    "semantic remaining evidence requires operator review",
])
def test_generic_budget_vocabulary_is_not_misclassified_as_budget_exhaustion(
    failure_text,
):
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
        }]},
        "assessments": [_assessment("card", "score")],
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {
                "complete": False,
                "failures": [failure_text],
            },
            "states": {"optimization": "incomplete"},
        }],
        "diagnosis_coverage": {
            "selected_count": 1, "scheduled_count": 1, "failed_count": 1,
        },
    }, milestone="finalization")

    row = view["portfolio"][0]
    assert row["semantic_diagnosis_status"] == "incomplete"
    assert row["next_action"] == "review"
    assert "frozen semantic budget" not in row["rationale"].lower()


@pytest.mark.parametrize(
    ("failure_category", "contradictory_failure_text", "expected_action"),
    [
        ("budget_exhausted", "authority publication failed", "review_semantic_budget"),
        ("outcome_unknown", "budget exhausted", "review_semantic_budget"),
        (
            "authority_publication_failure", "no remaining budget",
            "repair_semantic_authority_publication",
        ),
    ],
)
def test_structured_semantic_category_wins_over_arbitrary_contradictory_prose(
    failure_category, contradictory_failure_text, expected_action,
):
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
        }]},
        "assessments": [_assessment("card", "score")],
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {
                "complete": False,
                "semantic_failure_category": failure_category,
                "failures": [contradictory_failure_text],
            },
            "states": {"optimization": "incomplete"},
        }],
        "diagnosis_coverage": {
            "selected_count": 1, "scheduled_count": 1, "failed_count": 1,
        },
    }, milestone="finalization")

    row = view["portfolio"][0]
    assert row["semantic_diagnosis_status"] == failure_category
    assert row["next_action"] == expected_action
    assert contradictory_failure_text not in row["rationale"].lower()


@pytest.mark.parametrize("invalid_category", [None, "", "unknown", "budget_exhausted_typo"])
def test_missing_or_unknown_structured_semantic_category_fails_closed(invalid_category):
    from plexus.optimization.portfolio_run import _stakeholder_view

    coverage = {
        "complete": False,
        "failures": ["budget exhausted; authority publication failed; outcome unknown"],
    }
    if invalid_category is not None:
        coverage["semantic_failure_category"] = invalid_category
    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
        }]},
        "assessments": [_assessment("card", "score")],
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": coverage,
            "states": {"optimization": "incomplete"},
        }],
        "diagnosis_coverage": {
            "selected_count": 1, "scheduled_count": 1, "failed_count": 1,
        },
    }, milestone="finalization")

    assert view["portfolio"][0]["semantic_diagnosis_status"] == "incomplete"
    assert view["portfolio"][0]["next_action"] == "review"


@pytest.mark.parametrize(
    ("top_category", "nested_category", "expected_status", "expected_action"),
    [
        ("budget_exhausted", "outcome_unknown", "incomplete", "review"),
        ("invalid_category", "outcome_unknown", "incomplete", "review"),
        (
            "authority_publication_failure", "authority_publication_failure",
            "authority_publication_failure", "repair_semantic_authority_publication",
        ),
    ],
)
def test_conflicting_or_invalid_typed_categories_fail_closed_without_precedence(
    top_category, nested_category, expected_status, expected_action,
):
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Priority Score",
        }]},
        "assessments": [_assessment("card", "score")],
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "semantic_failure_category": top_category,
            "coverage": {
                "complete": False,
                "semantic_failure_category": nested_category,
            },
            "states": {"optimization": "incomplete"},
        }],
        "diagnosis_coverage": {
            "selected_count": 1, "scheduled_count": 1, "failed_count": 1,
        },
    }, milestone="finalization")

    assert view["portfolio"][0]["semantic_diagnosis_status"] == expected_status
    assert view["portfolio"][0]["next_action"] == expected_action


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ({}, "incomplete"),
        ({"semantic_failure_category": "invalid"}, "incomplete"),
        ({
            "diagnoses": [{
                "semantic_failure_category": "budget_exhausted",
                "coverage": {"semantic_failure_category": "outcome_unknown"},
            }],
            "diagnosis_coverage": {
                "semantic_failure_category": "budget_exhausted",
            },
        }, "incomplete"),
        ({
            "diagnoses": [{
                "semantic_failure_category": "outcome_unknown",
                "coverage": {"semantic_failure_category": "outcome_unknown"},
            }],
            "diagnosis_coverage": {
                "semantic_failure_category": "outcome_unknown",
            },
        }, "outcome_unknown"),
    ],
)
def test_one_normalizer_collects_all_nested_and_aggregate_typed_categories(value, expected):
    from plexus.optimization.portfolio_run import _normalize_semantic_failure_category

    assert _normalize_semantic_failure_category(value) == expected


def test_summary_conflicting_aggregate_category_counts_fail_closed_generic():
    from plexus.optimization.portfolio_run import OptimizationPortfolioRunner

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {}, assess=lambda _request: {}, diagnose=lambda _request: {},
        summary=lambda _request: {"coverage": {"complete": False}},
        dispatch=lambda _request: {}, review=lambda _request: {},
        report=_ReportService(), human_review=lambda _request: {},
    ))
    summary = runner._summary({
        "semantic_budget_evidence": {"unknown_count": 0},
        "diagnoses": [], "assessments": [], "reviews": [],
        "diagnosis_coverage": {
            "failed_count": 2,
            "budget_exhausted_count": 1,
            "authority_publication_failure_count": 1,
        },
    })

    assert summary["semantic_budget_next_action"] == "review"
    assert "without a recognized structured failure category" in (
        summary["semantic_budget_failure"]
    )


@pytest.mark.parametrize(
    ("exception_kind", "expected_category"),
    [
        ("semantic_budget", "budget_exhausted"),
        ("semantic_unknown", "outcome_unknown"),
        ("tactus_unknown", "outcome_unknown"),
        ("semantic_authority", "authority_publication_failure"),
        ("semantic_authority_generic", "incomplete"),
        ("report_publication", "authority_publication_failure"),
        ("report_integrity", "authority_publication_failure"),
        ("generic", "incomplete"),
    ],
)
def test_semantic_failure_category_is_derived_from_exception_type_not_message(
    exception_kind, expected_category,
):
    from plexus.optimization.portfolio_run import _semantic_exception_category
    from plexus.optimization.run_report import (
        OptimizationRunIntegrityError,
        OptimizationRunPublicationError,
    )
    from plexus.optimization.semantic_authority import (
        SemanticAuthorityError,
        SemanticAuthorityPublicationError,
        SemanticOutcomeUnknown,
    )
    from plexus.optimization.semantic_budget import SemanticBudgetExceeded
    from tactus.protocols.model_attempt import ModelAttemptOutcomeUnknown

    exception_type = {
        "semantic_budget": SemanticBudgetExceeded,
        "semantic_unknown": SemanticOutcomeUnknown,
        "tactus_unknown": ModelAttemptOutcomeUnknown,
        "semantic_authority": SemanticAuthorityPublicationError,
        "semantic_authority_generic": SemanticAuthorityError,
        "report_publication": OptimizationRunPublicationError,
        "report_integrity": OptimizationRunIntegrityError,
        "generic": RuntimeError,
    }[exception_kind]
    contradictory_message = (
        "budget exhausted; outcome unknown; authority publication failed"
    )

    assert _semantic_exception_category(
        exception_type(contradictory_message)
    ) == expected_category


def test_budget_deferred_semantic_diagnosis_is_never_presented_as_optimization_ready():
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Deferred Score",
            "valid_feedback_count": 200, "reviewed_disagreements": 20,
            "disagreement_rate": 0.1, "reviewed_error_opportunity": 20,
        }]},
        "assessments": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {"complete": True},
            "states": {"optimization": "ready_to_optimize"},
        }],
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {
                "complete": False,
                "semantic_failure_category": "budget_exhausted",
            },
            "semantic_failure_category": "budget_exhausted",
            "states": {"optimization": "incomplete"},
            "failures": ["semantic call worst-case reservation exceeds the frozen run budget"],
        }],
        "diagnosis_coverage": {
            "selected_count": 1, "scheduled_count": 1, "completed_count": 0,
            "failed_count": 1, "deferred_by_budget_count": 1,
        },
    }, milestone="finalization")

    row = view["portfolio"][0]
    assert row["readiness"] == "incomplete"
    assert row["next_action"] == "review_semantic_budget"
    assert "not an optimization-ready target" in row["rationale"]
    assert row["primary_disposition"] == "insufficient_evidence"


def test_portfolio_fingerprint_binds_frozen_semantic_policy_and_ledger_proof_without_changing_target_identity():
    from plexus.optimization.portfolio_run import (
        OptimizationPortfolioRunner,
        _approval_request,
        _evidence_snapshot,
        _portfolio_evidence_fingerprint,
    )

    state = {
        "run_key": "run-1",
        "run_spec": {"semantic_budget": {
            "schema_version": "semantic-budget-v1", "max_cost_usd": "1",
            "pricing_version": "openai-2025-08-07-v1",
        }},
        "semantic_budget_evidence": {
            "ledger_revision": 7, "evidence_digest": "a" * 64,
            "authorized_max_usd": "1", "settled_actual_usd": "0",
        },
    }
    first = _portfolio_evidence_fingerprint(state)
    second = _portfolio_evidence_fingerprint({
        **state,
        "semantic_budget_evidence": {
            **state["semantic_budget_evidence"], "ledger_revision": 8,
            "evidence_digest": "b" * 64, "settled_actual_usd": "0.1",
        },
    })
    snapshot = _evidence_snapshot(state)
    target = {
        "scorecard_id": "card", "score_id": "score", "score_name": "Score",
        "assessment_fingerprint": "assessment-fingerprint",
        "champion_version": "champion", "feedback_watermark": "watermark",
    }
    request = _approval_request(
        run_key="run-1", account_id="account-1", batch_number=1,
        targets=[target], report_ref={"kind": "report", "id": "report-1"},
        limits={}, portfolio_evidence_fingerprint=first,
    )

    assert first != second
    assert snapshot["portfolio_evidence_fingerprint"] == first
    assert snapshot["semantic_budget_evidence"]["ledger_revision"] == 7
    assert request["preconditions"]["portfolio_evidence_fingerprint"] == first
    assert request["targets"][0]["assessment_fingerprint"] == "assessment-fingerprint"

    runner = OptimizationPortfolioRunner(_dependencies(
        rank=lambda _request: {}, assess=lambda _request: {}, diagnose=lambda _request: {},
        summary=lambda _request: {"coverage": {"complete": False}}, dispatch=lambda _request: {},
        review=lambda _request: {}, report=_ReportService(), human_review=lambda _request: {},
    ))
    summary = runner._summary({
        **state,
        "diagnosis_coverage": {"failed_count": 0, "deferred_by_budget_count": 2},
    })
    assert summary["semantic_budget_next_action"] == "review_semantic_budget"
    assert "2 deferred diagnoses" in summary["semantic_budget_failure"]


def test_stakeholder_overview_explains_ranking_and_semantic_diagnosis_cutoffs():
    from plexus.optimization.portfolio_run import _stakeholder_view
    from plexus.optimization.run_report import _validate_view

    ranked_rows = [
        {
            "scorecard_id": "card",
            "score_id": f"score-{index:02d}",
            "scorecard_name": "Example Portfolio",
            "score_name": f"Score {index:02d}",
            "valid_feedback_count": 200 + index,
            "reviewed_disagreements": 30 - index,
            "disagreement_rate": (30 - index) / (200 + index),
            "reviewed_error_opportunity": 30 - index,
        }
        for index in range(12)
    ]
    view = _stakeholder_view({
        "rank": {
            "coverage": {"complete": True},
            "ranked": ranked_rows,
            "unranked": [],
            "window": {"start": "a", "end": "b"},
        },
        "assessments": [],
        "diagnoses": [],
        "reviews": [],
        "approval_requests": [],
        "diagnosis_coverage": {
            "top_priority_count": 10,
            "monitoring_candidate_count": 1,
            "selected_count": 11,
            "scheduled_count": 5,
            "deferred_by_cap_count": 6,
            "completed_count": 0,
            "failed_count": 0,
            "skipped_count": 1,
            "max_semantic_diagnoses": 5,
        },
    }, milestone="assessment")

    overview = view["overview"]
    assert overview["ranking_cutoff"] == "none"
    assert overview["priority_display_limit"] == 10
    assert overview["priority_displayed_count"] == 10
    assert overview["priority_cutoff_rank"] == 10
    assert overview["priority_cutoff_opportunity"] == 21
    assert overview["ranked_below_priority_cutoff"] == 2
    assert overview["diagnosis_selected_count"] == 11
    assert overview["diagnosis_scheduled_count"] == 5
    assert overview["assessed_score_count"] == 0
    assert overview["diagnosis_completed_count"] == 0
    assert overview["optimizer_review_count"] == 0
    assert overview["diagnosis_deferred_count"] == 6
    assert overview["diagnosis_skipped_count"] == 1
    assert overview["diagnosis_max_count"] == 5
    assert overview["diagnosis_coverage"] == (
        "0 of 5 scheduled diagnoses returned; 0 incomplete results; "
        "0 execution failures; 6 deferred by the configured diagnosis limit"
    )
    assert view["priorities"][0]["rank"] == 1
    assert view["priorities"][0]["disagreement_rate"] == 0.15
    assert view["feedback_investment"][0]["rank"] == 1
    assert view["optimization_outcomes"][0]["rank"] == 1
    _validate_view(view)

    zero_cap_view = _stakeholder_view({
        "rank": {
            "coverage": {"complete": True},
            "ranked": ranked_rows,
            "unranked": [],
        },
        "diagnosis_coverage": {
            "selected_count": 10,
            "scheduled_count": 0,
            "deferred_by_cap_count": 10,
            "max_semantic_diagnoses": 0,
        },
    }, milestone="assessment")
    assert zero_cap_view["overview"]["diagnosis_max_count"] == 0
    assert zero_cap_view["overview"]["diagnosis_scheduled_count"] == 0
    assert zero_cap_view["overview"]["diagnosis_deferred_count"] == 10


def test_stakeholder_view_preserves_pre_policy_rank_and_visible_cooldown_disposition():
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {
            "coverage": {
                "complete": True,
                "activity": {"recent_activity_excluded_count": 1},
            },
            "ranked": [{
                "scorecard_id": "card",
                "score_id": "eligible",
                "scorecard_name": "Example Portfolio",
                "score_name": "Eligible Score",
                "valid_feedback_count": 100,
                "reviewed_disagreements": 20,
                "disagreement_rate": 0.2,
                "reviewed_error_opportunity": 20,
                "evidence_rank": 2,
                "candidate_rank": 1,
                "policy_disposition": "eligible",
                "policy_reason": "meets_rank_policy",
                "eligible_for_optimization": True,
            }],
            "unranked": [{
                "scorecard_id": "card",
                "score_id": "cooldown",
                "scorecard_name": "Example Portfolio",
                "score_name": "Recently Changed Score",
                "valid_feedback_count": 100,
                "reviewed_disagreements": 80,
                "disagreement_rate": 0.8,
                "reviewed_error_opportunity": 80,
                "evidence_rank": 1,
                "policy_disposition": "cooldown",
                "policy_reason": "recent_score_activity",
                "eligible_for_optimization": False,
                "score_activity": {"eligible_at": "2026-08-05T00:00:00Z"},
                "unranked_reason": "recent_score_activity",
            }],
        },
        "assessments": [],
        "diagnoses": [],
        "reviews": [],
        "approval_requests": [],
        "diagnosis_coverage": {"selected_count": 0, "max_semantic_diagnoses": 10},
    }, milestone="ranking")

    assert [row["score_name"] for row in view["priorities"]] == [
        "Recently Changed Score",
        "Eligible Score",
    ]
    cooldown = view["priorities"][0]
    assert cooldown["rank"] == 1
    assert cooldown["candidate_rank"] is None
    assert cooldown["policy_disposition"] == "cooldown"
    assert cooldown["policy_reason"] == "recent_score_activity"
    assert cooldown["next_action"] == "wait_for_cooldown"
    assert cooldown["eligibility_timestamp"] == "2026-08-05T00:00:00Z"
    assert view["overview"]["evidence_ranked_score_count"] == 2
    assert view["overview"]["ranked_score_count"] == 1


def test_stakeholder_projection_uses_explicit_outcome_precedence_and_overlapping_issue_flags():
    """Terminal results win over active work and secondary issues remain visible."""
    from plexus.optimization.portfolio_run import _stakeholder_view

    def ranked(score_id, name, rank):
        return {
            "scorecard_id": "card", "score_id": score_id,
            "scorecard_name": "Example Portfolio", "score_name": name,
            "valid_feedback_count": 100, "reviewed_disagreements": 25,
            "disagreement_rate": 0.25, "reviewed_error_opportunity": 25,
            "evidence_rank": rank,
        }

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [
            ranked("terminal", "Reviewed Score", 1),
            ranked("active", "Running Score", 2),
            ranked("question", "Question Score", 3),
        ]},
        "assessments": [
            {"scope": {"scorecard_id": "card", "score_id": "terminal"}, "coverage": {"complete": True},
             "states": {"optimization": "feedback_curation_review", "guideline_health": "missing", "feedback_rubric_health": "inconsistent"}},
            {"scope": {"scorecard_id": "card", "score_id": "active"}, "coverage": {"complete": True},
             "states": {"optimization": "ready_to_optimize"}},
            {"scope": {"scorecard_id": "card", "score_id": "question"}, "coverage": {"complete": True},
             "states": {"optimization": "stakeholder_clarification_required"}},
        ],
        "diagnoses": [{
            "scope": {"scorecard_id": "card", "score_id": "question"}, "coverage": {"complete": True},
            "states": {"optimization": "stakeholder_clarification_required", "guideline_health": "potential_code_conflict"},
            "stakeholder_questions": ["Which documented policy applies?"],
        }],
        "reviews": [{
            "scope": {"scorecard_id": "card", "score_id": "terminal"}, "coverage": {"complete": True},
            "states": {"post_run": "promotion_ready"}, "post_run_state": "promotion_ready",
            "primary_next_action": "request_promotion_approval",
            "rationale": "Validated terminal evidence supports a promotion decision.",
        }],
        "dispatch": {"children": [{
            "target": {"scorecard_id": "card", "score_id": "active"},
            "launch_state": {"phase": "running"},
        }]},
    }, milestone="optimization")

    rows = {row["score_name"]: row for row in view["portfolio"]}
    assert rows["Reviewed Score"]["primary_disposition"] == "promotion_ready"
    assert rows["Reviewed Score"]["next_action"] == "request_promotion_approval"
    assert rows["Reviewed Score"]["rationale"] == (
        "Validated terminal evidence supports a promotion decision."
    )
    assert rows["Running Score"]["primary_disposition"] == "optimization_in_progress"
    assert rows["Question Score"]["primary_disposition"] == "stakeholder_clarification_required"
    assert rows["Reviewed Score"]["secondary_issue_flags"] == [
        "missing_guidelines", "feedback_rubric_contradiction",
    ]
    assert rows["Question Score"]["secondary_issue_flags"] == [
        "potential_code_conflict", "stakeholder_question",
    ]
    assert view["overview"]["primary_disposition_counts"]["promotion_ready"] == 1
    assert view["overview"]["secondary_issue_counts"]["stakeholder_question"] == 1


def test_stakeholder_projection_retains_validated_improvement_without_promotion_action():
    from plexus.optimization.portfolio_run import _stakeholder_view

    view = _stakeholder_view({
        "rank": {"coverage": {"complete": True}, "ranked": [{
            "scorecard_id": "card", "score_id": "score",
            "scorecard_name": "Example Portfolio", "score_name": "Reviewed Score",
        }]},
        "assessments": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {"complete": True}, "states": {"optimization": "ready_to_optimize"},
        }],
        "reviews": [{
            "scope": {"scorecard_id": "card", "score_id": "score"},
            "coverage": {"complete": True},
            "post_run_state": "validated_improvement",
            "promotion_ready": False,
            "primary_next_action": "complete_promotion_evidence",
            "alignment_evidence": {
                "recent": {"baseline": 0.826, "candidate": 0.942, "delta": 0.116},
                "regression": {"baseline": 0.310, "candidate": 0.450, "delta": 0.140},
            },
        }],
    }, milestone="optimization_review")

    row = view["portfolio"][0]
    outcome = view["optimization_outcomes"][0]
    assert row["primary_disposition"] == "validated_improvement"
    assert row["next_action"] == "complete_promotion_evidence"
    assert outcome["outcome"] == "validated_improvement"
    assert outcome["primary_disposition"] == "validated_improvement"
    assert outcome["alignment_evidence"] == {
        "recent": {"baseline": 0.826, "candidate": 0.942, "delta": 0.116},
        "regression": {"baseline": 0.310, "candidate": 0.450, "delta": 0.140},
    }


def test_stakeholder_projection_orders_questions_by_severity_then_evidence_then_rank():
    from plexus.optimization.portfolio_run import _stakeholder_view

    def packet(score_id, name, rank, evidence, guideline, questions=()):
        return {
            "rank": {"coverage": {"complete": True}, "ranked": [{
                "scorecard_id": "card", "score_id": score_id,
                "scorecard_name": "Example Portfolio", "score_name": name,
                "valid_feedback_count": evidence, "reviewed_disagreements": evidence // 2,
                "disagreement_rate": 0.5, "reviewed_error_opportunity": evidence // 2,
                "evidence_rank": rank,
            }]},
            "assessment": {"scope": {"scorecard_id": "card", "score_id": score_id}, "coverage": {"complete": True},
                           "states": {"optimization": "repair_required", "guideline_health": guideline}},
            "diagnosis": {"scope": {"scorecard_id": "card", "score_id": score_id}, "coverage": {"complete": True},
                          "states": {"optimization": "repair_required", "guideline_health": guideline},
                          "stakeholder_questions": list(questions)},
        }

    first = packet("missing", "Missing", 2, 20, "missing")
    second = packet("conflict", "Conflict", 3, 200, "potential_code_conflict", ["Resolve policy."])
    second["diagnosis"]["guideline_code_conflict_claim"] = (
        "The guideline requires an explicit confirmation, but the score code accepts an implied answer."
    )
    second["diagnosis"]["evidence_ids"] = ["restricted-semantic-evidence-1"]
    second["diagnosis"]["evidence_fingerprint"] = "immutable-diagnosis-evidence"
    state = {
        "rank": {"coverage": {"complete": True}, "ranked": first["rank"]["ranked"] + second["rank"]["ranked"]},
        "assessments": [first["assessment"], second["assessment"]],
        "diagnoses": [first["diagnosis"], second["diagnosis"]],
        "reviews": [],
    }
    issues = _stakeholder_view(state, milestone="diagnosis")["questions_and_issues"]

    assert [row["issue_flag"] for row in issues] == [
        "missing_guidelines", "potential_code_conflict", "stakeholder_question",
    ]
    assert issues[0]["affected_evidence_count"] == 20
    assert issues[1]["affected_evidence_count"] == 200
    assert issues[1]["finding"] == (
        "The guideline requires an explicit confirmation, but the score code accepts an implied answer."
    )
    assert issues[1]["next_action"] == "review_and_repair_guideline_code_alignment"
    diagnosis_alias = "semantic-diagnosis-" + sha256(
        b"immutable-diagnosis-evidence"
    ).hexdigest()[:16]
    assert issues[1]["evidence_references"].startswith(
        f"semantic diagnosis packet; {diagnosis_alias}; semantic-evidence-"
    )
    assert issues[1]["evidence_reference_tokens"] == [
        diagnosis_alias,
        "semantic-evidence-" + sha256(b"restricted-semantic-evidence-1").hexdigest()[:16]
    ]
    assert issues[2]["finding"] == "Resolve policy."
    assert all(row["scorecard_ref"] != "card" for row in issues)
